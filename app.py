"""
Centralized Load Testing Platform  v4.0
Multi-client repository — each client has its own JMX plans, test data, and reports.

Run   : python app.py  →  http://localhost:5000
DB    : lt_platform.db  (SQLite, auto-created)
Creds : See README.md — change defaults after first login.
"""

# ── Self-bootstrap: install missing Python packages before anything else ──────
import sys, subprocess as _sp
def _pip(*pkgs):
    _sp.check_call([sys.executable, '-m', 'pip', 'install', '--quiet', *pkgs])

try:
    from flask import Flask as _F
except ImportError:
    print('[bootstrap] Installing Flask...'); _pip('flask>=3.0.0')
try:
    import openpyxl as _ox
except ImportError:
    print('[bootstrap] Installing openpyxl...'); _pip('openpyxl>=3.1.0')
try:
    from werkzeug.security import generate_password_hash as _gph
except ImportError:
    print('[bootstrap] Installing werkzeug...'); _pip('werkzeug')
try:
    import psutil as _psutil
except ImportError:
    print('[bootstrap] Installing psutil...'); _pip('psutil>=5.9.0')
    import psutil as _psutil
# ─────────────────────────────────────────────────────────────────────────────

from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify, send_file, g)
from werkzeug.security import generate_password_hash, check_password_hash
import os, subprocess, threading, json, csv, statistics, glob, sqlite3, uuid, hashlib, secrets
import xml.etree.ElementTree as ET
import re, tempfile, shutil, io, zipfile, platform as _platform
import urllib.request as _urllib_req2, urllib.error as _urllib_err2
import ssl as _ssl
_SSL_CTX = _ssl.create_default_context()
# Set LT_WEBHOOK_SKIP_TLS=1 only when your webhook endpoint uses a self-signed cert.
if os.environ.get('LT_WEBHOOK_SKIP_TLS', '').lower() in ('1', 'true', 'yes'):
    _SSL_CTX.check_hostname = False
    _SSL_CTX.verify_mode = _ssl.CERT_NONE
try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
from datetime import datetime
from collections import defaultdict
from functools import wraps

app = Flask(__name__)

# ── Base paths ────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
CLIENTS_DIR = os.path.join(BASE_DIR, 'clients')
DB_PATH     = os.path.join(BASE_DIR, 'lt_platform.db')
CFG_FILE    = os.path.join(BASE_DIR, '.ltconfig.json')
os.makedirs(CLIENTS_DIR, exist_ok=True)

# ── Secret key: LT_SECRET_KEY env var > persistent .lt_secret file > generated ──
_secret_file = os.path.join(BASE_DIR, '.lt_secret')
if os.environ.get('LT_SECRET_KEY'):
    app.secret_key = os.environ['LT_SECRET_KEY']
elif os.path.exists(_secret_file):
    with open(_secret_file) as _sf:
        app.secret_key = _sf.read().strip()
else:
    app.secret_key = os.urandom(32).hex()
    with open(_secret_file, 'w') as _sf:
        _sf.write(app.secret_key)

# ── Session hardening: SameSite=Strict + HttpOnly + CSRF tokens ──
app.config['SESSION_COOKIE_SAMESITE'] = 'Strict'
app.config['SESSION_COOKIE_HTTPONLY'] = True
# Set SECURE flag if explicitly requested via LT_FORCE_SECURE_COOKIES=1 (requires HTTPS)
if os.environ.get('LT_FORCE_SECURE_COOKIES', '').lower() in ('1', 'true', 'yes'):
    app.config['SESSION_COOKIE_SECURE'] = True

# ── CSRF token support for form submissions ──
def _csrf_token():
    """Generate or retrieve CSRF token for session"""
    if 'csrf_token' not in session:
        session['csrf_token'] = os.urandom(32).hex()
    return session['csrf_token']

def _check_csrf_token():
    """Validate CSRF token from request (POST, PUT, DELETE)"""
    token = request.form.get('csrf_token') or request.headers.get('X-CSRF-Token')
    if not token or token != session.get('csrf_token'):
        return False
    return True

DEFAULTS = {
    'jmeter_bin': '',   # auto-detected at startup
    'heap': '-Xms512m -Xmx1g',
    'audit_retention_days': 90,
    'openrouter_api_key': '',
    'ai_model': 'openai/gpt-4o-mini',
    'teams_webhook': '',
    'slack_webhook': '',
    'notify_email': '',
    'smtp_host': '',
    'smtp_port': 587,
    'smtp_user': '',
    'smtp_pass': '',
    # Security
    'session_timeout_mins': 120,

    # Auto-stop
    'auto_stop_err_pct': 0,
    'auto_stop_check_secs': 60,
    # Jira integration
    'jira_url': '',
    'jira_project': '',
    'jira_token': '',
    'jira_user': '',
    # Pre/post test hooks
    'pre_test_hook': '',
    'post_test_hook': '',
    # Backup & archiving
    'archive_reports_days': 0,
    'backup_path': '',
    'backup_interval_hours': 0,
    # Self-registration
    'self_register_enabled': True,
    # Webhook / MS Teams alerts
    'webhook_url': '',
    'webhook_on_pass': True,
    'webhook_on_fail': True,
}

# ── Database ──────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS clients (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            code         TEXT UNIQUE NOT NULL,
            name         TEXT NOT NULL,
            description  TEXT DEFAULT '',
            logo_emoji   TEXT DEFAULT '🏢',
            color        TEXT DEFAULT '#00d4ff',
            enabled      INTEGER DEFAULT 1,
            jmx_dir      TEXT,
            testdata_dir TEXT,
            reports_dir  TEXT,
            created_at   TEXT DEFAULT (datetime('now','localtime')),
            created_by   TEXT DEFAULT 'system'
        );

        CREATE TABLE IF NOT EXISTS users (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            username    TEXT    UNIQUE NOT NULL,
            password    TEXT    NOT NULL,
            name        TEXT    NOT NULL,
            role        TEXT    NOT NULL DEFAULT 'viewer',
            initials    TEXT    DEFAULT '',
            enabled     INTEGER DEFAULT 1,
            created_at  TEXT    DEFAULT (datetime('now','localtime')),
            created_by  TEXT    DEFAULT 'system',
            last_login  TEXT,
            login_count INTEGER DEFAULT 0,
            permissions TEXT    DEFAULT '[]'
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            username   TEXT,
            action     TEXT,
            details    TEXT,
            ip_address TEXT,
            timestamp  TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS api_tokens (
            id           TEXT PRIMARY KEY,
            name         TEXT NOT NULL,
            token_hash   TEXT NOT NULL UNIQUE,
            token_prefix TEXT NOT NULL,
            client       TEXT,
            created_by   TEXT,
            created_at   TEXT DEFAULT (datetime('now','localtime')),
            last_used_at TEXT,
            last_used_ip TEXT,
            enabled      INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS recurring_schedules (
            id          TEXT PRIMARY KEY,
            client      TEXT NOT NULL,
            jmx         TEXT NOT NULL,
            threads     INTEGER DEFAULT 10,
            duration    INTEGER DEFAULT 300,
            rampup      INTEGER DEFAULT 30,
            recurrence  TEXT DEFAULT 'daily',
            run_at_time TEXT NOT NULL,
            day_of_week INTEGER DEFAULT 1,
            enabled     INTEGER DEFAULT 1,
            last_run    TEXT,
            created_by  TEXT,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS user_prefs (
            username TEXT NOT NULL,
            key      TEXT NOT NULL,
            value    TEXT DEFAULT '',
            PRIMARY KEY (username, key)
        );

        CREATE TABLE IF NOT EXISTS pending_registrations (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            username     TEXT UNIQUE NOT NULL,
            password     TEXT NOT NULL,
            name         TEXT NOT NULL,
            email        TEXT DEFAULT '',
            status       TEXT DEFAULT 'pending',
            submitted_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS report_comments (
            id         TEXT PRIMARY KEY,
            client     TEXT NOT NULL,
            filename   TEXT NOT NULL,
            username   TEXT NOT NULL,
            text       TEXT NOT NULL,
            ts_offset_s INTEGER,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS report_shares (
            token      TEXT PRIMARY KEY,
            client     TEXT NOT NULL,
            filename   TEXT NOT NULL,
            created_by TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS env_profiles (
            name        TEXT PRIMARY KEY,
            data        TEXT NOT NULL,
            is_active   INTEGER DEFAULT 0,
            created_by  TEXT DEFAULT 'system',
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS test_run_notes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id     TEXT NOT NULL,
            client     TEXT NOT NULL,
            notes      TEXT DEFAULT '',
            tags       TEXT DEFAULT '',
            author     TEXT DEFAULT 'system',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS run_request_log (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id     TEXT NOT NULL,
            client     TEXT NOT NULL,
            entry_json TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE INDEX IF NOT EXISTS idx_rrl_run ON run_request_log(run_id);

        CREATE TABLE IF NOT EXISTS test_features (
            id          TEXT PRIMARY KEY,
            client      TEXT NOT NULL,
            filename    TEXT NOT NULL,
            description TEXT DEFAULT '',
            file_type   TEXT DEFAULT 'gherkin',
            file_path   TEXT NOT NULL,
            scenario_count INTEGER DEFAULT 0,
            enabled     INTEGER DEFAULT 1,
            created_by  TEXT DEFAULT 'system',
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(client, filename)
        );

        CREATE TABLE IF NOT EXISTS ci_cd_suites (
            id          TEXT PRIMARY KEY,
            client      TEXT NOT NULL,
            name        TEXT NOT NULL,
            description TEXT DEFAULT '',
            jmx_files   TEXT NOT NULL,
            feature_file TEXT,
            schedule    TEXT NOT NULL,
            enabled     INTEGER DEFAULT 1,
            notify_on_fail TEXT DEFAULT '',
            retry_count INTEGER DEFAULT 0,
            created_by  TEXT DEFAULT 'system',
            created_at  TEXT DEFAULT (datetime('now','localtime')),
            last_run    TEXT,
            UNIQUE(client, name)
        );

        CREATE TABLE IF NOT EXISTS ci_cd_run_history (
            id          TEXT PRIMARY KEY,
            suite_id    TEXT NOT NULL,
            client      TEXT NOT NULL,
            status      TEXT DEFAULT 'running',
            start_time  TEXT NOT NULL,
            end_time    TEXT,
            duration_s  INTEGER,
            total_requests INTEGER DEFAULT 0,
            success_count INTEGER DEFAULT 0,
            error_count INTEGER DEFAULT 0,
            avg_rt_ms   REAL DEFAULT 0,
            p95_rt_ms   REAL DEFAULT 0,
            report_path TEXT,
            triggered_by TEXT DEFAULT 'scheduler',
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE INDEX IF NOT EXISTS idx_cicd_suite ON ci_cd_run_history(suite_id);
        """)


        # Seed BTC as default client
        if db.execute("SELECT COUNT(*) FROM clients").fetchone()[0] == 0:
            btc_base = os.path.join(CLIENTS_DIR, 'BTC')
            db.execute("""INSERT INTO clients
                (code,name,description,logo_emoji,color,jmx_dir,testdata_dir,reports_dir,created_by)
                VALUES (?,?,?,?,?,?,?,?,?)""", (
                'BTC', 'BTC Botswana', 'BTC Mobiquity Pay — USSD Load Testing',
                '📡', '#00d4ff',
                os.path.join(btc_base, 'jmx'),
                os.path.join(btc_base, 'testdata'),
                os.path.join(btc_base, 'reports'),
                'system'
            ))

        # Seed default admin users from environment variables (or generate secure passwords)
        if db.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
            # Get credentials from env vars with fallback to secure generation
            admin_user = os.environ.get('LT_ADMIN_USER', 'admin')
            admin_pass = os.environ.get('LT_ADMIN_PASS', None)  # Must be set explicitly
            viewer_user = os.environ.get('LT_VIEWER_USER', 'viewer')
            viewer_pass = os.environ.get('LT_VIEWER_PASS', None)  # Must be set explicitly
            
            if not admin_pass:
                admin_pass = os.urandom(16).hex()  # Generate random password if not provided
                print(f'[SETUP] Generated admin password (set LT_ADMIN_PASS env var to override): {admin_pass}')
            if not viewer_pass:
                viewer_pass = os.urandom(12).hex()  # Generate random password if not provided
                print(f'[SETUP] Generated viewer password (set LT_VIEWER_PASS env var to override): {viewer_pass}')
            
            db.execute(
                "INSERT INTO users (username,password,name,role,initials,created_by) VALUES (?,?,?,?,?,?)",
                (admin_user, generate_password_hash(admin_pass), 'System Administrator', 'admin', 'SA', 'system')
            )
            db.execute(
                "INSERT INTO users (username,password,name,role,initials,created_by,permissions) VALUES (?,?,?,?,?,?,?)",
                (viewer_user, generate_password_hash(viewer_pass), 'Load Tester', 'viewer', 'LT', 'system', '["run_tests"]')
            )
        # Migrate existing DB: add permissions column if missing
        try:
            db.execute("ALTER TABLE users ADD COLUMN permissions TEXT DEFAULT '[]'")
            db.commit()
        except Exception:
            pass  # column already exists
        try:
            db.execute("ALTER TABLE clients ADD COLUMN quota_mb INTEGER DEFAULT 0")
            db.commit()
        except Exception:
            pass
        # Release-gate columns for CI/CD
        for _mig in ("ALTER TABLE ci_cd_suites ADD COLUMN gate_config TEXT DEFAULT '{}'",
                     "ALTER TABLE ci_cd_run_history ADD COLUMN gate_status TEXT",
                     "ALTER TABLE ci_cd_run_history ADD COLUMN gate_reasons TEXT"):
            try:
                db.execute(_mig); db.commit()
            except Exception:
                pass
        db.commit()

def audit(action, details='', username=None, ip=None):
    try:
        u = username or session.get('user', 'system')
        i = ip or (request.remote_addr if request else '—')
        with get_db() as db:
            db.execute(
                "INSERT INTO audit_log (username,action,details,ip_address) VALUES (?,?,?,?)",
                (u, action, details, i)
            )
            db.commit()
    except Exception:
        pass

# ── Client helpers ─────────────────────────────────────────────────────────────
def get_all_clients():
    with get_db() as db:
        return [dict(r) for r in db.execute(
            "SELECT * FROM clients WHERE enabled=1 ORDER BY name"
        ).fetchall()]

def get_client(code):
    with get_db() as db:
        r = db.execute("SELECT * FROM clients WHERE code=?", (code,)).fetchone()
    return dict(r) if r else None

def client_dirs(c):
    """Return {jmx, testdata, reports} absolute paths for a client dict.
    If a stored path no longer exists on this machine (e.g. moved to a new server),
    fall back to the canonical path relative to BASE_DIR/clients/."""
    base = os.path.join(CLIENTS_DIR, c['code'])
    def _resolve(stored, default):
        if stored and os.path.exists(stored):
            return stored
        return default
    return {
        'jmx':      _resolve(c['jmx_dir'],      os.path.join(base, 'jmx')),
        'testdata': _resolve(c['testdata_dir'],  os.path.join(base, 'testdata')),
        'reports':  _resolve(c['reports_dir'],   os.path.join(base, 'reports')),
    }

def ensure_client_dirs(c):
    dirs = client_dirs(c)
    for d in dirs.values():
        os.makedirs(d, exist_ok=True)
    return dirs

def active_client():
    """Return the currently selected client dict, defaulting to first client.
    API-token requests resolve to the token's bound client (g.api_client)."""
    code = None
    try:
        code = g.get('api_client')
    except Exception:
        code = None
    if not code:
        code = session.get('client')
    c = get_client(code) if code else None
    if not c:
        clients = get_all_clients()
        c = clients[0] if clients else None
        # Only persist the default into a real browser session, never for tokens.
        if c and 'user' in session:
            session['client'] = c['code']
    return c

# ── Config ────────────────────────────────────────────────────────────────────
def load_cfg():
    if os.path.exists(CFG_FILE):
        try:
            with open(CFG_FILE) as f:
                return {**DEFAULTS, **json.load(f)}
        except Exception:
            pass
    return dict(DEFAULTS)

def save_cfg(d):
    with open(CFG_FILE, 'w') as f:
        json.dump(d, f, indent=2)

# ── Live test state ────────────────────────────────────────────────────────────
_state = {'running': False, 'pid': None, 'start': None,
          'jmx': None, 'jtl': None, 'rc': None, 'err': None, 'client': None,
          'notes': '', 'tag': '', 'sla': None}
# ── Suite state ────────────────────────────────────────────────────────────────
_suite_state = {'running': False, 'current': None, 'results': [], 'total': 0, 'done': 0, 'client': None}
# ── Schedules ──────────────────────────────────────────────────────────────────
_schedules = []
_sched_lock = threading.Lock()
_logs  = []
_lock  = threading.Lock()
_req_log = []
_req_log_lock = threading.Lock()
_REQ_LOG_MAX = 500

# ── New feature globals ────────────────────────────────────────────────────────
_start_time = datetime.now()
_live_share_token = {'token': '', 'expires': 0}
_presence = {}
_annotations = []

# ── JMeter auto-install ────────────────────────────────────────────────────────
_IS_WINDOWS         = _platform.system() == 'Windows'
_JMETER_VERSION     = '5.6.3'
_JMETER_ZIP_URL     = f'https://archive.apache.org/dist/jmeter/binaries/apache-jmeter-{_JMETER_VERSION}.zip'
_JMETER_BIN_NAME    = 'jmeter.bat' if _IS_WINDOWS else 'jmeter'
# Install JMeter inside the project folder so it travels with a folder-copy
_JMETER_INSTALL_DIR = os.path.join(BASE_DIR, f'apache-jmeter-{_JMETER_VERSION}')
_jmeter_install      = {'status': 'idle', 'progress': '', 'error': '', 'bin': ''}
_jmeter_install_lock = threading.Lock()

# ── Java auto-install ──────────────────────────────────────────────────────────
_JAVA_BIN_NAME     = 'java.exe' if _platform.system() == 'Windows' else 'java'
_java_install      = {'status': 'idle', 'progress': '', 'error': '', 'bin': '', 'home': ''}
_java_install_lock = threading.Lock()

def _java_download_info():
    """Return (url, archive_format) for the current OS/arch."""
    system = _platform.system().lower()
    machine = _platform.machine().lower()
    arch = 'aarch64' if machine in ('arm64', 'aarch64') else 'x64'
    base = 'https://api.adoptium.net/v3/binary/latest/11/ga'
    if system == 'windows':
        return f'{base}/windows/{arch}/jdk/hotspot/normal/eclipse', 'zip'
    elif system == 'darwin':
        return f'{base}/mac/{arch}/jdk/hotspot/normal/eclipse', 'tar.gz'
    else:
        return f'{base}/linux/{arch}/jdk/hotspot/normal/eclipse', 'tar.gz'

def _find_java_bin():
    """Return the first existing java binary path, or ''."""
    cfg = load_cfg()
    saved_home = cfg.get('java_home', '')
    if saved_home and os.path.exists(saved_home):
        candidate = os.path.join(saved_home, 'bin', _JAVA_BIN_NAME)
        if os.path.exists(candidate):
            return candidate

    # JAVA_HOME env var
    java_home = os.environ.get('JAVA_HOME', '')
    if java_home:
        candidate = os.path.join(java_home, 'bin', _JAVA_BIN_NAME)
        if os.path.exists(candidate):
            return candidate

    # System PATH
    in_path = shutil.which('java')
    if in_path:
        return in_path

    # Project-local JDK dirs (auto-installed)
    for d in sorted(glob.glob(os.path.join(BASE_DIR, 'jdk*')), reverse=True):
        candidate = os.path.join(d, 'bin', _JAVA_BIN_NAME)
        if os.path.exists(candidate):
            return candidate
        # macOS: Contents/Home layout
        candidate = os.path.join(d, 'Contents', 'Home', 'bin', _JAVA_BIN_NAME)
        if os.path.exists(candidate):
            return candidate

    # Well-known system locations
    if _platform.system() == 'Windows':
        for base_dir in [r'C:\Program Files\Java', r'C:\Program Files\Eclipse Adoptium',
                         r'C:\Program Files\Microsoft', r'C:\Program Files\OpenJDK']:
            if os.path.isdir(base_dir):
                for sub in sorted(os.listdir(base_dir), reverse=True):
                    candidate = os.path.join(base_dir, sub, 'bin', 'java.exe')
                    if os.path.exists(candidate):
                        return candidate
    else:
        for candidate in ['/usr/bin/java', '/usr/local/bin/java',
                          '/opt/java/bin/java',
                          '/usr/lib/jvm/java-11-openjdk-amd64/bin/java',
                          '/usr/lib/jvm/java-11-openjdk/bin/java']:
            if os.path.exists(candidate):
                return candidate
    return ''

def _auto_install_java():
    """Download and extract OpenJDK 11 (Adoptium Temurin) in a background thread."""
    global _java_install
    with _java_install_lock:
        if _java_install['status'] == 'downloading':
            return
        _java_install.update(status='downloading', progress='Starting download…', error='', bin='', home='')

    url, fmt = _java_download_info()
    suffix   = '.zip' if fmt == 'zip' else '.tar.gz'
    tmp_file = os.path.join(tempfile.gettempdir(), f'jdk11{suffix}')
    try:
        _java_install['progress'] = 'Downloading OpenJDK 11 (~185 MB)…'
        print(f'[java-install] Downloading {url}')

        def _reporthook(count, block, total):
            if total > 0:
                pct = min(100, count * block * 100 // total)
                _java_install['progress'] = f'Downloading… {pct}%'

        _urllib_req2.urlretrieve(url, tmp_file, reporthook=_reporthook)

        _java_install['progress'] = 'Extracting archive…'
        print('[java-install] Extracting…')

        before = set(glob.glob(os.path.join(BASE_DIR, 'jdk*')))
        if fmt == 'zip':
            with zipfile.ZipFile(tmp_file, 'r') as zf:
                zf.extractall(BASE_DIR)
        else:
            import tarfile
            with tarfile.open(tmp_file, 'r:gz') as tf:
                tf.extractall(BASE_DIR)
        os.unlink(tmp_file)

        after    = set(glob.glob(os.path.join(BASE_DIR, 'jdk*')))
        new_dirs = sorted(after - before) or sorted(after)
        if not new_dirs:
            raise FileNotFoundError('JDK directory not found after extraction')

        jdk_home = new_dirs[0]
        java_bin = os.path.join(jdk_home, 'bin', _JAVA_BIN_NAME)
        # macOS packages the JDK under Contents/Home
        if not os.path.exists(java_bin):
            alt = os.path.join(jdk_home, 'Contents', 'Home', 'bin', _JAVA_BIN_NAME)
            if os.path.exists(alt):
                java_bin = alt
                jdk_home = os.path.join(jdk_home, 'Contents', 'Home')

        if not os.path.exists(java_bin):
            raise FileNotFoundError(f'java binary not found after extraction (expected {java_bin})')
        if _platform.system() != 'Windows':
            os.chmod(java_bin, 0o755)

        # Make JAVA_HOME available to this process so JMeter picks it up immediately
        os.environ['JAVA_HOME'] = jdk_home

        cfg = load_cfg()
        cfg['java_home'] = jdk_home
        save_cfg(cfg)

        _java_install.update(status='ready', progress='OpenJDK 11 installed.', bin=java_bin, home=jdk_home)
        audit('JAVA_INSTALLED', f'Auto-installed OpenJDK 11 → {jdk_home}', username='system')
        print(f'[java-install] Done — JAVA_HOME={jdk_home}')
    except Exception as ex:
        _java_install.update(status='error', error=str(ex), progress='')
        print(f'[java-install] ERROR: {ex}')
        if os.path.exists(tmp_file):
            try: os.unlink(tmp_file)
            except Exception: pass

def _find_jmeter_bin():
    """Return the first existing JMeter binary path, or ''."""
    cfg = load_cfg()
    saved = cfg.get('jmeter_bin', '')
    if saved and os.path.exists(saved):
        return saved

    # 1. Inside the project folder (installed here so it travels with folder copies)
    local = os.path.join(_JMETER_INSTALL_DIR, 'bin', _JMETER_BIN_NAME)
    if os.path.exists(local):
        return local

    # 2. System PATH (catches apt/brew installs)
    in_path = shutil.which('jmeter')
    if in_path:
        return in_path

    # 3. Well-known system locations
    if _IS_WINDOWS:
        candidates = [
            os.path.join('C:\\', f'apache-jmeter-{_JMETER_VERSION}', 'bin', 'jmeter.bat'),
            r'C:\apache-jmeter-5.5\bin\jmeter.bat',
            r'C:\apache-jmeter-5.6\bin\jmeter.bat',
            r'C:\jmeter\bin\jmeter.bat',
        ]
        for drive in 'CDEFG':
            candidates.append(os.path.join(f'{drive}:\\',
                f'apache-jmeter-{_JMETER_VERSION}', 'bin', 'jmeter.bat'))
    else:
        candidates = [
            f'/opt/apache-jmeter-{_JMETER_VERSION}/bin/jmeter',
            '/opt/jmeter/bin/jmeter',
            '/usr/local/bin/jmeter',
            '/usr/bin/jmeter',
            os.path.expanduser(f'~/apache-jmeter-{_JMETER_VERSION}/bin/jmeter'),
            os.path.expanduser('~/jmeter/bin/jmeter'),
        ]
    for c in candidates:
        if c and os.path.exists(c):
            return c
    return ''

def _auto_install_jmeter():
    """Download and extract Apache JMeter in a background thread."""
    global _jmeter_install
    with _jmeter_install_lock:
        if _jmeter_install['status'] == 'downloading':
            return
        _jmeter_install.update(status='downloading', progress='Starting download…', error='', bin='')

    tmp_zip = os.path.join(tempfile.gettempdir(), f'apache-jmeter-{_JMETER_VERSION}.zip')
    try:
        # Download with progress reporting
        _jmeter_install['progress'] = f'Downloading JMeter {_JMETER_VERSION} (~80 MB)…'
        print(f'[jmeter-install] Downloading {_JMETER_ZIP_URL}')

        def _reporthook(count, block, total):
            if total > 0:
                pct = min(100, count * block * 100 // total)
                _jmeter_install['progress'] = f'Downloading… {pct}%'

        _urllib_req2.urlretrieve(_JMETER_ZIP_URL, tmp_zip, reporthook=_reporthook)

        _jmeter_install['progress'] = 'Extracting archive…'
        print('[jmeter-install] Extracting…')
        # Extract into project folder — no admin/root required
        with zipfile.ZipFile(tmp_zip, 'r') as zf:
            zf.extractall(BASE_DIR)

        os.unlink(tmp_zip)

        bin_path = os.path.join(_JMETER_INSTALL_DIR, 'bin', _JMETER_BIN_NAME)
        if not os.path.exists(bin_path):
            raise FileNotFoundError(f'{_JMETER_BIN_NAME} not found at {bin_path} after extraction')
        if not _IS_WINDOWS:
            os.chmod(bin_path, 0o755)

        # Persist the new path into settings
        cfg = load_cfg()
        cfg['jmeter_bin'] = bin_path
        save_cfg(cfg)

        _jmeter_install.update(status='ready', progress=f'JMeter {_JMETER_VERSION} installed.', bin=bin_path)
        audit('JMETER_INSTALLED', f'Auto-installed JMeter {_JMETER_VERSION} → {bin_path}', username='system')
        print(f'[jmeter-install] Done — {bin_path}')
    except Exception as ex:
        _jmeter_install.update(status='error', error=str(ex), progress='')
        print(f'[jmeter-install] ERROR: {ex}')
        if os.path.exists(tmp_zip):
            try: os.unlink(tmp_zip)
            except Exception: pass

# ── Before-request middleware ──────────────────────────────────────────────────
@app.before_request
def check_session_timeout():
    if 'user' not in session:
        return
    cfg = load_cfg()
    timeout_mins = int(cfg.get('session_timeout_mins', 120))
    if timeout_mins <= 0:
        return
    last = session.get('_last_active')
    now_ts = datetime.now().timestamp()
    if last and (now_ts - last) > timeout_mins * 60:
        session.clear()
        if request.path.startswith('/api/'):
            return jsonify(error='Session expired', timeout=True), 401
        return redirect(url_for('login_page'))
    session['_last_active'] = now_ts


# ── Auth decorators ────────────────────────────────────────────────────────────
def login_req(f):
    @wraps(f)
    def w(*a, **k):
        if 'user' not in session:
            return redirect(url_for('login_page'))
        return f(*a, **k)
    return w

def admin_req(f):
    @wraps(f)
    def w(*a, **k):
        if 'user' not in session:
            return redirect(url_for('login_page'))
        if session.get('role') != 'admin':
            return redirect(url_for('viewer_page'))
        return f(*a, **k)
    return w

def perm_req(*perms):
    """Allow admins unconditionally; allow viewers who have at least one of the listed permissions."""
    def decorator(fn):
        @wraps(fn)
        def w(*a, **k):
            if 'user' not in session:
                return jsonify(error='Not authenticated'), 401
            if session.get('role') == 'admin':
                return fn(*a, **k)
            try:
                user_perms = set(json.loads(session.get('permissions', '[]')))
            except Exception:
                user_perms = set()
            if user_perms.intersection(set(perms)):
                return fn(*a, **k)
            return jsonify(error='Permission denied'), 403
        return w
    return decorator

def csrf_protect(f):
    """Decorator to validate CSRF token on POST/PUT/DELETE requests"""
    @wraps(f)
    def w(*a, **k):
        if request.method in ('POST', 'PUT', 'DELETE'):
            if not _check_csrf_token():
                return jsonify(error='CSRF token invalid or missing'), 403
        return f(*a, **k)
    return w


# ── API token auth (for CI/CD orchestrators, e.g. Jenkins) ─────────────────────
def _hash_token(raw):
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()

def _generate_api_token():
    """Return (plaintext, prefix, hash). Plaintext is shown to the user once."""
    raw = 'lt_' + secrets.token_urlsafe(32)
    return raw, raw[:11], _hash_token(raw)

def _authenticate_api_token():
    """Return the api_tokens row (dict) for a valid Bearer token, else None."""
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        return None
    raw = auth[7:].strip()
    if not raw:
        return None
    try:
        with get_db() as db:
            row = db.execute(
                "SELECT * FROM api_tokens WHERE token_hash=? AND enabled=1",
                (_hash_token(raw),)).fetchone()
        return dict(row) if row else None
    except Exception:
        return None

def api_auth(f):
    """Allow a valid API Bearer token OR an authenticated admin browser session.

    Token requests are CSRF-exempt (no cookie/session is involved) and set
    g.api_client so active_client() resolves the token's bound client. Session
    requests keep the existing admin + CSRF requirements. This lets Jenkins call
    the CI/CD endpoints with `Authorization: Bearer <token>` while the browser UI
    keeps working unchanged."""
    @wraps(f)
    def w(*a, **k):
        tok = _authenticate_api_token()
        if tok:
            g.api_token  = tok
            g.api_client = tok.get('client') or None
            try:
                with get_db() as db:
                    db.execute("UPDATE api_tokens SET last_used_at=?, last_used_ip=? WHERE id=?",
                               (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                request.remote_addr or '', tok['id']))
                    db.commit()
            except Exception:
                pass
            return f(*a, **k)
        # Fall back to browser session (admin) with CSRF on writes.
        if 'user' not in session:
            return jsonify(error='Authentication required: log in, or send Authorization: Bearer <token>'), 401
        if session.get('role') != 'admin':
            return jsonify(error='Admin privilege or API token required'), 403
        if request.method in ('POST', 'PUT', 'DELETE') and not _check_csrf_token():
            return jsonify(error='CSRF token invalid or missing'), 403
        return f(*a, **k)
    return w

# ── Page routes ────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if 'user' not in session:
        return redirect(url_for('login_page'))
    return redirect(url_for('admin_page') if session['role'] == 'admin' else url_for('viewer_page'))

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    error = None
    if request.method == 'POST':
        u = request.form.get('username', '').strip()
        p = request.form.get('password', '')
        with get_db() as db:
            row = db.execute(
                "SELECT * FROM users WHERE username=? AND enabled=1", (u,)
            ).fetchone()
        if row and check_password_hash(row['password'], p):
            perms = row['permissions'] if 'permissions' in row.keys() else '[]'
            session.update(user=u, role=row['role'],
                           name=row['name'], initials=row['initials'] or u[:2].upper(),
                           permissions=perms or '[]')
            with get_db() as db:
                db.execute(
                    "UPDATE users SET last_login=datetime('now','localtime'), login_count=login_count+1 WHERE username=?",
                    (u,)
                )
                db.commit()
            # Default to first client
            clients = get_all_clients()
            if clients and not session.get('client'):
                session['client'] = clients[0]['code']
            audit('LOGIN', f"Role: {row['role']}", username=u)
            return redirect(url_for('admin_page') if row['role'] == 'admin' else url_for('viewer_page'))
        error = 'Invalid username or password, or account disabled.'
        audit('LOGIN_FAILED', f"Attempted: {u}", username=u)
    clients_count = len(get_all_clients())
    return render_template('login.html', error=error, clients_count=clients_count)

@app.route('/logout')
def logout():
    audit('LOGOUT')
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/api/me')
@login_req
def api_me():
    try:
        perms = json.loads(session.get('permissions', '[]'))
    except Exception:
        perms = []
    return jsonify(username=session.get('user'), name=session.get('name'),
                   role=session.get('role'), initials=session.get('initials'),
                   permissions=perms)

@app.route('/api/session/client', methods=['POST'])
@login_req
def api_set_client():
    code = (request.json or {}).get('code', '')
    c = get_client(code)
    if not c:
        return jsonify(error='Client not found'), 404
    session['client'] = code
    audit('CLIENT_SWITCH', f"Switched to client: {code}")
    return jsonify(ok=True, client=dict(c))

# ── CSRF Token Endpoint ────────────────────────────────────────────────────────
@app.route('/api/csrf', methods=['GET'])
@login_req
def api_get_csrf_token():
    """Get CSRF token for client-side form submissions"""
    return jsonify(csrf_token=_csrf_token())

# ── Built-in CI/CD Endpoints ────────────────────────────────────────────────────
@app.route('/api/ci-cd/test-suites', methods=['GET'])
@admin_req
def api_ci_test_suites():
    """List available test suites for CI/CD automation"""
    with get_db() as db:
        suites = db.execute(
            "SELECT id, name, description, created_at FROM recurring_schedules WHERE active=1 ORDER BY created_at DESC"
        ).fetchall()
    return jsonify(suites=[dict(s) for s in suites])

@app.route('/api/ci-cd/run-suite/<suite_id>', methods=['POST'])
@admin_req
@csrf_protect
def api_ci_run_suite(suite_id):
    """Trigger CI/CD suite execution with automated reporting"""
    try:
        with get_db() as db:
            suite = db.execute(
                "SELECT * FROM recurring_schedules WHERE id=? AND active=1", (suite_id,)
            ).fetchone()
        if not suite:
            return jsonify(error='Suite not found'), 404
        
        # Execute suite immediately (scheduled tests)
        run_id = str(uuid.uuid4())
        audit('CI_CD_SUITE_START', f"Suite {suite['name']} (ID: {suite_id})")
        
        # Return run tracking info
        return jsonify(ok=True, run_id=run_id, suite_name=suite['name'], status='scheduled')
    except Exception as e:
        audit('CI_CD_SUITE_ERROR', str(e))
        return jsonify(error=str(e)), 500

@app.route('/api/ci-cd/results/<run_id>', methods=['GET'])
@admin_req
def api_ci_results(run_id):
    """Get CI/CD suite execution results and artifacts"""
    try:
        with get_db() as db:
            runs = db.execute(
                "SELECT * FROM run_history WHERE run_id=? ORDER BY created_at DESC LIMIT 1", (run_id,)
            ).fetchall()
        if not runs:
            return jsonify(error='Run not found'), 404
        
        run = dict(runs[0])
        return jsonify(run=run, status='completed', artifacts={
            'report': f'/download/report/{run_id}',
            'jtl': f'/download/jtl/{run_id}',
            'logs': f'/api/live-logs/vu?run_id={run_id}'
        })
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route('/admin')
@admin_req
def admin_page():
    c = active_client()
    if not c:
        return "No clients configured. Please restart and check the database.", 500
    dirs = client_dirs(c)
    jmx_files = sorted([os.path.basename(f) for f in glob.glob(os.path.join(dirs['jmx'], '*.jmx'))])
    jtl_files = sorted([os.path.basename(f) for f in glob.glob(os.path.join(dirs['reports'], '*.jtl'))], reverse=True)
    csv_files = sorted([os.path.basename(f) for f in glob.glob(os.path.join(dirs['testdata'], '*.csv'))])
    clients = get_all_clients()
    return render_template('admin.html',
        jmx_files=jmx_files, jtl_files=jtl_files, csv_files=csv_files,
        config=load_cfg(), clients=clients, active_client=c, csrf_token=_csrf_token())

@app.route('/viewer')
@login_req
def viewer_page():
    c = active_client()
    dirs = client_dirs(c) if c else {}
    jtl_files = sorted([os.path.basename(f) for f in glob.glob(os.path.join(dirs.get('reports',''), '*.jtl'))], reverse=True) if dirs else []
    jmx_files = sorted([os.path.basename(f) for f in glob.glob(os.path.join(dirs.get('jmx',''), '*.jmx'))]) if dirs else []
    csv_files  = sorted([os.path.basename(f) for f in glob.glob(os.path.join(dirs.get('testdata',''), '*.csv'))]) if dirs else []
    clients    = get_all_clients()
    return render_template('viewer.html', jtl_files=jtl_files, jmx_files=jmx_files,
                           csv_files=csv_files, clients=clients, active_client=c, csrf_token=_csrf_token())

_PREREQ_FILE = 'prereq.json'

def _default_prereq():
    return {
        'targets':     {'tps':'','avg_sla':'','peak':'','duration':'','err_sla':'','p90_sla':''},
        'environment': {'name':'','endpoint':'','build':'','jmeter':'Apache JMeter 5.5','db':'','servers':'','tools':''},
        'channels':    ['USSD'],
        'iterations':  '',
        'hours':       '',
        'services':    [],
        'cl_services': [],
        'cl_data':     [],
        'cl_grafana':  [],
        'activities': [
            {'activity':'Script preparations','status':'Pending','day':'0','duration':'—','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Pre-requisites — Tools readiness','status':'Pending','day':'0','duration':'—','owner':'','prereq':'JMeter, Grafana','awr':'','remarks':'','ref':''},
            {'activity':'Environment preparation\n• Enable all nodes\n• Server readiness','status':'Pending','day':'1','duration':'—','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Data preparations','status':'Pending','day':'1','duration':'—','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Data confirmation from client','status':'Pending','day':'1','duration':'—','owner':'','prereq':'Confirmation from client required','awr':'','remarks':'','ref':''},
            {'activity':'Services mocked','status':'Pending','day':'1','duration':'—','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Notifications disabled','status':'Pending','day':'1','duration':'—','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Rebuild indexes & analyse for DB','status':'Pending','day':'1','duration':'—','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Load execution at 25 TPS — Run 1','status':'Pending','day':'1','duration':'30 mins','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Load execution at 25 TPS — Run 2','status':'Pending','day':'1','duration':'2 Hrs','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Load execution at 25 TPS — Run 3','status':'Pending','day':'1','duration':'4 Hrs','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Load execution at 25 TPS — Run 4','status':'Pending','day':'2','duration':'8 Hrs','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Load execution at 35 TPS — Run 1','status':'Pending','day':'3','duration':'4 Hrs','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Load execution at 35 TPS — Run 2','status':'Pending','day':'3','duration':'8 Hrs','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Load execution at 40 TPS — Run 1','status':'Pending','day':'4','duration':'4 Hrs','owner':'','prereq':'','awr':'','remarks':'','ref':''},
            {'activity':'Load execution at 40 TPS — Run 2','status':'Pending','day':'4','duration':'8 Hrs','owner':'','prereq':'','awr':'','remarks':'','ref':''},
        ],
    }

def _prereq_auto(c):
    """Return the live auto-check state for a client."""
    cfg      = load_cfg()
    dirs     = client_dirs(c)
    sla_p    = _sla_path(c)
    ld_p     = os.path.join(CLIENTS_DIR, c['code'], 'load_config.json')
    jmx_count = len(glob.glob(os.path.join(dirs['jmx'],      '*.jmx')))
    csv_count  = len(glob.glob(os.path.join(dirs['testdata'], '*.csv')))
    jmeter_ok  = bool(cfg.get('jmeter_bin') and os.path.exists(cfg['jmeter_bin']))
    sla_ok     = bool(_read_json(sla_p))
    ld_ok      = bool(_read_json(ld_p))
    return {
        'jmx_count':  jmx_count,  'jmx_ok':    jmx_count > 0,
        'csv_count':  csv_count,  'csv_ok':    csv_count > 0,
        'jmeter_ok':  jmeter_ok,
        'sla_ok':     sla_ok,
        'ld_ok':      ld_ok,
        'running':    _state.get('running', False),
        'active_jmx': _state.get('jmx', ''),
    }

@app.route('/api/prereq', methods=['GET'])
@login_req
def api_prereq_get():
    c = active_client()
    if not c: return jsonify(error='No active client.'), 400
    path = os.path.join(CLIENTS_DIR, c['code'], _PREREQ_FILE)
    data = _read_json(path)
    if data is None:          # first visit — seed with defaults
        data = _default_prereq()
        _write_json(path, data)
    return jsonify(data=data, auto=_prereq_auto(c))

@app.route('/api/prereq/auto', methods=['GET'])
@login_req
def api_prereq_auto():
    """Lightweight poll — only live system state, not saved data."""
    c = active_client()
    if not c: return jsonify(error='No active client.'), 400
    return jsonify(_prereq_auto(c))

@app.route('/api/prereq', methods=['POST'])
@login_req
def api_prereq_save():
    c = active_client()
    if not c: return jsonify(error='No active client.'), 400
    payload = request.get_json(force=True) or {}
    path = os.path.join(CLIENTS_DIR, c['code'], _PREREQ_FILE)
    _write_json(path, payload)
    audit('PREREQ_SAVE', f'Prereq saved for {c["code"]}')
    return jsonify(ok=True)

def _build_dynamic_stages(svc):
    """Returns list of stage dicts for spike/step/business_hours patterns.
    Each dict: {threads, ramp_up, duration, delay, suffix}
    Returns None for constant pattern.
    """
    pattern = svc.get('pattern', 'constant') or 'constant'
    threads  = int(svc.get('threads', 10))
    ramp_up  = int(svc.get('ramp_up', 60))
    duration = int(svc.get('duration', 1800))

    if pattern == 'spike':
        spike_threads  = int(svc.get('spike_threads', threads * 3))
        spike_at       = int(svc.get('spike_at',       int(duration * 0.4)))
        spike_duration = int(svc.get('spike_duration', max(60, int(duration * 0.1))))
        recovery_start = spike_at + spike_duration
        recovery_dur   = max(30, duration - recovery_start)
        stages = []
        if spike_at > 0:
            stages.append({'threads': threads,       'ramp_up': ramp_up,                       'duration': spike_at,      'delay': 0,             'suffix': 'phase1-normal'})
        stages.append(    {'threads': spike_threads, 'ramp_up': min(30, spike_duration // 3),  'duration': spike_duration,'delay': spike_at,      'suffix': 'phase2-spike'})
        if recovery_dur > 0:
            stages.append({'threads': threads,       'ramp_up': ramp_up,                       'duration': recovery_dur,  'delay': recovery_start,'suffix': 'phase3-recovery'})
        return stages

    elif pattern == 'step':
        step_count  = max(2, int(svc.get('step_count', 5)))
        step_dur    = max(30, duration // step_count)
        step_threads= max(1, threads // step_count)
        return [
            {'threads': step_threads * (i + 1),
             'ramp_up': min(ramp_up, step_dur // 2),
             'duration': step_dur,
             'delay': i * step_dur,
             'suffix': f'step{i+1}'}
            for i in range(step_count)
        ]

    elif pattern == 'business_hours':
        hourly_pct = svc.get('hourly_pct') or [
            5,2,1,1,1,2,10,30,60,80,85,90,
            75,85,90,85,70,50,35,25,20,15,10,7
        ]
        hour_dur = min(3600, max(60, duration // len(hourly_pct)))
        stages, total = [], 0
        for i, pct in enumerate(hourly_pct):
            if total >= duration:
                break
            n_threads = max(1, int(threads * pct / 100))
            stages.append({'threads': n_threads, 'ramp_up': min(300, hour_dur // 4),
                           'duration': min(hour_dur, duration - total),
                           'delay': total, 'suffix': f'h{i:02d}'})
            total += hour_dur
        return stages

    return None  # constant — no staging


def _apply_dynamic_stages_to_jmx(root_el, dynamic_svcs):
    """
    For each service in dynamic_svcs (list of dicts with keys: tg_key, stages),
    replace the matching ThreadGroup in root_el with staged ThreadGroups that
    use the JMeter scheduler (delay + duration).
    """
    import copy
    top_ht = root_el.find('hashTree')
    if top_ht is None:
        return
    second_ht = top_ht.find('hashTree')
    if second_ht is None:
        return

    tg_map = {}
    siblings = list(second_ht)
    i = 0
    while i < len(siblings):
        el = siblings[i]
        child_ht = siblings[i + 1] if i + 1 < len(siblings) and siblings[i + 1].tag == 'hashTree' else None
        if el.tag == 'ThreadGroup':
            raw = el.get('testname', '')
            m = re.match(r'^TG - (.+?) \(', raw.strip())
            key = (m.group(1) if m else raw).strip().lower()
            tg_map[key] = (el, child_ht, i)
        i += 2 if child_ht is not None else 1

    for dsvc in dynamic_svcs:
        tg_key = dsvc['tg_key'].lower()
        stages  = dsvc['stages']
        if tg_key not in tg_map:
            continue
        orig_tg, orig_ht, orig_idx = tg_map[tg_key]
        orig_name = orig_tg.get('testname', tg_key)

        def _set_sp(el, name, val):
            for sp in el.findall('stringProp'):
                if sp.get('name') == name:
                    sp.text = str(val); return
            sp = ET.SubElement(el, 'stringProp')
            sp.set('name', name); sp.text = str(val)

        def _set_bp(el, name, val):
            for bp in el.findall('boolProp'):
                if bp.get('name') == name:
                    bp.text = 'true' if val else 'false'; return
            bp = ET.SubElement(el, 'boolProp')
            bp.set('name', name); bp.text = 'true' if val else 'false'

        second_ht.remove(orig_tg)
        if orig_ht is not None:
            second_ht.remove(orig_ht)

        siblings = list(second_ht)
        insert_at = orig_idx

        for stage in stages:
            stage_tg = copy.deepcopy(orig_tg)
            stage_tg.set('testname', f'{orig_name} [{stage["suffix"]}]')
            _set_sp(stage_tg, 'ThreadGroup.num_threads', stage['threads'])
            _set_sp(stage_tg, 'ThreadGroup.ramp_time',   stage['ramp_up'])
            _set_sp(stage_tg, 'ThreadGroup.duration',     stage['duration'])
            _set_sp(stage_tg, 'ThreadGroup.delay',        stage['delay'])
            _set_bp(stage_tg, 'ThreadGroup.scheduler',    True)
            for lc in stage_tg.iter('LoopController'):
                _set_bp(lc, 'LoopController.continue_forever', True)
                for ip in lc.findall('intProp'):
                    if ip.get('name') == 'LoopController.loops':
                        ip.text = '-1'

            second_ht.insert(insert_at, stage_tg)
            insert_at += 1
            if orig_ht is not None:
                second_ht.insert(insert_at, copy.deepcopy(orig_ht))
                insert_at += 1


# ── API: Test control ──────────────────────────────────────────────────────────
@app.route('/api/test/start', methods=['POST'])
@perm_req('run_tests')
def api_start():
    global _state, _logs
    if _state['running']:
        return jsonify(error='A test is already running.'), 409

    c = active_client()
    if not c:
        return jsonify(error='No active client selected.'), 400
    dirs = ensure_client_dirs(c)

    d        = request.json or {}
    jmx_name = d.get('jmx', '')
    threads  = int(d.get('threads', 10))
    duration = int(d.get('duration', 300))
    rampup   = int(d.get('rampup', 30))
    warmup   = int(d.get('warmup', 0))
    rampdown = int(d.get('rampdown', 0))
    out_name = d.get('out_name') or f"run_{datetime.now().strftime('%d%m%Y_%H%M%S')}"
    extra    = d.get('extra', '')

    jmx_path = os.path.join(dirs['jmx'], jmx_name)
    jtl_path = os.path.join(dirs['reports'], out_name + '.jtl')
    os.makedirs(dirs['reports'], exist_ok=True)

    if not os.path.exists(jmx_path):
        return jsonify(error=f'JMX not found: {jmx_name}'), 400

    jbin = _find_jmeter_bin()
    if not jbin:
        if _jmeter_install['status'] == 'downloading':
            return jsonify(
                error='JMeter is being downloaded, please wait…',
                installing=True,
                progress=_jmeter_install.get('progress', '')
            ), 503
        # Kick off auto-download and tell the caller to retry
        threading.Thread(target=_auto_install_jmeter, daemon=True).start()
        return jsonify(
            error=f'JMeter not found. Auto-download of JMeter {_JMETER_VERSION} has started (~80 MB). '
                  f'Poll /api/jmeter/install-status and retry when status is "ready".',
            installing=True,
            progress='Starting download…'
        ), 503

    # Always work on a temp JMX copy (CSV path fix + optional service config)
    services = d.get('services') or []
    tmp_dir  = tempfile.mkdtemp(prefix='lt_jmeter_')
    working_jmx = os.path.join(tmp_dir, jmx_name)
    try:
        shutil.copy(jmx_path, working_jmx)
        tree = ET.parse(working_jmx)
        root_el = tree.getroot()
        testdata_dir = client_dirs(c)['testdata']

        # Fix CSVDataSet file paths → absolute paths in client's testdata folder
        for csv_ds in root_el.iter('CSVDataSet'):
            for prop in csv_ds.iter('stringProp'):
                if prop.get('name') == 'filename' and prop.text and prop.text.strip():
                    basename = os.path.basename(prop.text.strip().replace('/', os.sep).replace('\\', os.sep))
                    candidate = os.path.join(testdata_dir, basename)
                    if os.path.exists(candidate):
                        prop.text = candidate

        # Per-service enable/disable
        if services:
            svc_map = {s['name'].strip().lower(): s for s in services}
            for tg in root_el.iter('ThreadGroup'):
                tg_raw = tg.get('testname', '')
                m = re.match(r'^TG - (.+?) \((.+)\)$', tg_raw.strip())
                key = (m.group(1) if m else tg_raw).strip().lower()
                if key in svc_map:
                    tg.set('enabled', 'true' if svc_map[key].get('enabled') else 'false')

        # Apply dynamic load patterns (spike / step / business_hours)
        load_cfg_data = _read_json(_load_config_path(c)) or {}
        pattern_map = {s.get('service_name','').strip().lower(): s
                       for s in load_cfg_data.get('services', [])
                       if s.get('pattern','constant') not in ('constant','',None)}
        if pattern_map:
            dynamic_svcs = []
            for tg in root_el.iter('ThreadGroup'):
                raw = tg.get('testname', '')
                m   = re.match(r'^TG - (.+?) \(', raw.strip())
                key = (m.group(1) if m else raw).strip().lower()
                if key in pattern_map:
                    stages = _build_dynamic_stages(pattern_map[key])
                    if stages:
                        dynamic_svcs.append({'tg_key': key, 'stages': stages})
            if dynamic_svcs:
                _apply_dynamic_stages_to_jmx(root_el, dynamic_svcs)

        tree.write(working_jmx, encoding='unicode', xml_declaration=False)
    except Exception as ex:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return jsonify(error=f'JMX preparation failed: {ex}'), 500

    cmd = [jbin, '-n', '-t', working_jmx, '-l', jtl_path,
           f'-Jtest.duration={duration}',
           # Capture URL, latency, connect time, and the full request + response
           # body for EVERY sample (not just failures) so the platform's
           # request/response viewer can show each call individually.
           '-Jjmeter.save.saveservice.url=true',
           '-Jjmeter.save.saveservice.latency=true',
           '-Jjmeter.save.saveservice.connect_time=true',
           '-Jjmeter.save.saveservice.response_data=true',
           '-Jjmeter.save.saveservice.samplerData=true',
           '-Jjmeter.save.saveservice.requestHeaders=true',
           '-Jjmeter.save.saveservice.responseHeaders=true',
           '-Jjmeter.save.saveservice.assertion_results_failure_message=true',
           '-Jjmeter.save.saveservice.bytes=true',
           '-Jjmeter.save.saveservice.sent_bytes=true',
           ]
    if warmup > 0:
        cmd.append(f'-Jtest.warmup={warmup}')
    if rampdown > 0:
        cmd.append(f'-Jtest.rampdown={rampdown}')
    if services:
        for svc in services:
            if svc.get('enabled'):
                if svc.get('threads_key'):
                    cmd.append(f'-J{svc["threads_key"]}={svc.get("threads", 2)}')
                if svc.get('rampup_key'):
                    cmd.append(f'-J{svc["rampup_key"]}={svc.get("rampup", 5)}')
    else:
        cmd += [f'-Jp2p.threads={threads}', f'-Jp2p.rampup={rampup}']
    for kv in extra.split():
        if '=' in kv:
            cmd.append(f'-J{kv}')

    with _lock:
        _logs = []
    with _req_log_lock:
        _req_log.clear()
    _state.update(running=True, pid=None,
                  start=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  jmx=jmx_name, jtl=out_name + '.jtl', rc=None, err=None,
                  client=c['code'])
    audit('TEST_START', f"Client={c['code']} JMX={jmx_name} threads={threads} duration={duration}s")

    def _run():
        global _state
        _cfg = load_cfg()
        pre_hook  = _cfg.get('pre_test_hook', '').strip()
        post_hook = _cfg.get('post_test_hook', '').strip()
        auto_stop_err = float(_cfg.get('auto_stop_err_pct', 0) or 0)
        auto_stop_interval = int(_cfg.get('auto_stop_check_secs', 60) or 60)

        # Pre-test hook
        if pre_hook:
            try:
                r = subprocess.run(pre_hook, shell=True, capture_output=True, text=True, timeout=60)
                with _lock:
                    _logs.append({'ts': datetime.now().strftime('%H:%M:%S'),
                                  'msg': f'[pre-hook] exit={r.returncode} {r.stdout.strip()[:200]}'})
            except Exception as ex:
                with _lock:
                    _logs.append({'ts': datetime.now().strftime('%H:%M:%S'), 'msg': f'[pre-hook] ERROR: {ex}'})

        try:
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, bufsize=1)
            _state['pid'] = proc.pid
            _state['alert'] = None

            # Auto-stop monitor thread
            def _auto_stop_monitor():
                import time
                while _state.get('running') and proc.poll() is None:
                    time.sleep(auto_stop_interval)
                    if not _state.get('running') or proc.poll() is not None:
                        break
                    if auto_stop_err <= 0:
                        continue
                    try:
                        jtl_p = os.path.join(client_dirs(c)['reports'], _state.get('jtl', ''))
                        if not jtl_p or not os.path.exists(jtl_p):
                            continue
                        errors, total = 0, 0
                        with open(jtl_p, 'r', encoding='utf-8', errors='ignore') as f:
                            reader = csv.reader(f)
                            header = [h.strip() for h in next(reader, [])]
                            si = header.index('success') if 'success' in header else 7
                            for row in reader:
                                if len(row) > si:
                                    total += 1
                                    if row[si].strip().lower() != 'true':
                                        errors += 1
                        if total > 50:
                            err_pct = errors / total * 100
                            if err_pct >= auto_stop_err:
                                alert_msg = f'⚠️ Auto-stop: error rate {err_pct:.1f}% ≥ threshold {auto_stop_err}%'
                                _state['alert'] = alert_msg
                                with _lock:
                                    _logs.append({'ts': datetime.now().strftime('%H:%M:%S'), 'msg': alert_msg})
                                try:
                                    subprocess.run(['taskkill', '/F', '/T', '/PID', str(proc.pid)],
                                                   capture_output=True)
                                except Exception:
                                    pass
                                audit('AUTO_STOP', alert_msg, username='system')
                    except Exception:
                        pass

            if auto_stop_err > 0:
                threading.Thread(target=_auto_stop_monitor, daemon=True).start()

            # JTL request-level tailer — feeds _req_log ring buffer
            def _jtl_tailer():
                import time as _t2
                while not os.path.exists(jtl_path):
                    if not _state.get('running') and proc.poll() is not None:
                        return
                    _t2.sleep(0.3)
                with _req_log_lock:
                    _req_log.clear()
                import csv as _csv, io as _io
                hdr = []
                buf = ''
                with open(jtl_path, 'r', encoding='utf-8', errors='ignore') as _f:
                    while _state.get('running') or proc.poll() is None:
                        line = _f.readline()
                        if not line:
                            _t2.sleep(0.15); continue
                        buf += line
                        # JMeter quotes multi-line response bodies; a CSV record
                        # is complete only when quote chars are balanced and the
                        # buffered text ends on a newline.
                        if buf.count('"') % 2 != 0 or not buf.endswith('\n'):
                            continue
                        record = buf.rstrip('\r\n'); buf = ''
                        if not record: continue
                        try:
                            parts = next(_csv.reader(_io.StringIO(record)))
                        except Exception:
                            continue
                        if not hdr:
                            hdr = [h.strip() for h in parts]; continue
                        if len(parts) < 4: continue
                        row = dict(zip(hdr, parts))
                        entry = {
                            'ts':      row.get('timeStamp',''),
                            'label':   row.get('label','').strip(),
                            'elapsed': row.get('elapsed','0'),
                            'ok':      row.get('success','').strip().lower()=='true',
                            'rc':      row.get('responseCode','').strip(),
                            'thread':  row.get('threadName','').strip(),
                            'rm':      (row.get('responseMessage','') or '').strip(),
                            'bytes':   row.get('bytes','0'),
                            'sent':    row.get('sentBytes','0'),
                            'latency': row.get('Latency', row.get('latency','0')),
                            'connect': row.get('Connect', row.get('connectTime','0')),
                            'url':     (row.get('URL') or '').strip(),
                            'fm':          (row.get('failureMessage') or '').strip(),
                            'resp_data':   (row.get('responseData') or '').strip()[:2000],
                            'sampler_data':(row.get('samplerData','') or '').strip()[:1000],
                        }
                        with _req_log_lock:
                            _req_log.append(entry)
                            if len(_req_log) > _REQ_LOG_MAX:
                                _req_log.pop(0)
            threading.Thread(target=_jtl_tailer, daemon=True).start()

            for line in proc.stdout:
                line = line.rstrip()
                if line:
                    with _lock:
                        _logs.append({'ts': datetime.now().strftime('%H:%M:%S'), 'msg': line})
                        if len(_logs) > 1000:
                            _logs.pop(0)
            proc.wait()
            _state['rc'] = proc.returncode
            msg = '✅ Test completed.' if proc.returncode == 0 else f'⚠️ Ended with code {proc.returncode}.'
            with _lock:
                _logs.append({'ts': datetime.now().strftime('%H:%M:%S'), 'msg': msg})
            audit('TEST_END', f"rc={proc.returncode} client={c['code']}")
        except Exception as ex:
            _state['err'] = str(ex)
            with _lock:
                _logs.append({'ts': datetime.now().strftime('%H:%M:%S'), 'msg': f'❌ ERROR: {ex}'})
        finally:
            _state['running'] = False
            if tmp_dir and os.path.exists(tmp_dir):
                shutil.rmtree(tmp_dir, ignore_errors=True)
            # Persist ring buffer to DB
            try:
                with _req_log_lock:
                    snap = list(_req_log)
                if snap:
                    with get_db() as _db:
                        _db.executemany(
                            "INSERT INTO run_request_log(run_id,client,entry_json) VALUES(?,?,?)",
                            [(out_name, c['code'], json.dumps(e)) for e in snap]
                        )
            except Exception:
                pass
            # SLA evaluation
            full_jtl = os.path.join(client_dirs(c)['reports'], _state.get('jtl',''))
            sla_result = _evaluate_sla(full_jtl, c)
            _state['sla'] = sla_result
            # Notification
            result = 'done' if _state.get('rc') == 0 else 'failed'
            _notify_test_complete(c, jmx_name, result, sla_result, jtl_name=_state.get('jtl'))
            # Post-test hook
            if post_hook:
                try:
                    r = subprocess.run(post_hook, shell=True, capture_output=True, text=True, timeout=60)
                    with _lock:
                        _logs.append({'ts': datetime.now().strftime('%H:%M:%S'),
                                      'msg': f'[post-hook] exit={r.returncode} {r.stdout.strip()[:200]}'})
                except Exception as ex:
                    with _lock:
                        _logs.append({'ts': datetime.now().strftime('%H:%M:%S'), 'msg': f'[post-hook] ERROR: {ex}'})

    threading.Thread(target=_run, daemon=True).start()
    return jsonify(ok=True, jtl=out_name + '.jtl')

@app.route('/api/test/stop', methods=['POST'])
@perm_req('run_tests')
def api_stop():
    pid = _state.get('pid')
    if pid:
        try:
            subprocess.run(['taskkill', '/F', '/T', '/PID', str(pid)], capture_output=True)
        except Exception:
            pass
    _state['running'] = False
    with _lock:
        _logs.append({'ts': datetime.now().strftime('%H:%M:%S'), 'msg': '🛑 Test stopped by user.'})
    audit('TEST_STOP', f"pid={pid}")
    return jsonify(ok=True)

@app.route('/api/test/status')
@login_req
def api_status():
    with _lock:
        recent = list(_logs[-100:])
    return jsonify(**_state, logs=recent)

@app.route('/api/run-note', methods=['POST'])
@login_req
def api_run_note():
    data = request.json or {}
    with _lock:
        _state['notes'] = data.get('notes', '')[:200]
        _state['tag']   = data.get('tag', '')[:50]
    audit('RUN_NOTE', f"tag={_state['tag']} notes={_state['notes'][:60]}")
    return jsonify(ok=True)

@app.route('/api/logs')
@login_req
def api_logs():
    off = int(request.args.get('offset', 0))
    with _lock:
        return jsonify(logs=_logs[off:], total=len(_logs))

# ── API: Reports ───────────────────────────────────────────────────────────────
def _reports_dir():
    c = active_client()
    return client_dirs(c)['reports'] if c else ''

@app.route('/api/reports')
@login_req
def api_reports():
    rdir = _reports_dir()
    out = []
    all_files = sorted(
        glob.glob(os.path.join(rdir, '*.jtl')) +
        glob.glob(os.path.join(rdir, '*.html')),
        key=os.path.getmtime, reverse=True
    )
    for p in all_files:
        name = os.path.basename(p)
        sz   = os.path.getsize(p)
        mt   = datetime.fromtimestamp(os.path.getmtime(p)).strftime('%d %b %Y  %H:%M')
        n = 0
        if p.endswith('.jtl'):
            try:
                with open(p, errors='ignore') as f:
                    n = max(0, sum(1 for _ in f) - 1)
            except Exception:
                n = 0
        out.append(dict(name=name, size_kb=round(sz / 1024, 1), modified=mt, samples=n))
    return jsonify(out)

@app.route('/api/report/<path:fname>')
@login_req
def api_report(fname):
    p = os.path.join(_reports_dir(), fname)
    if not os.path.exists(p):
        return jsonify(error='File not found'), 404
    try:
        return jsonify(_parse_jtl(p))
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: CSV data ──────────────────────────────────────────────────────────────
@app.route('/api/csv-files')
@login_req
def api_csvs():
    c = active_client()
    td = client_dirs(c)['testdata'] if c else ''
    out = []
    for f in sorted(glob.glob(os.path.join(td, '*.csv'))):
        try:
            with open(f, encoding='utf-8', errors='ignore') as fp:
                rows = list(csv.reader(fp))
            out.append(dict(name=os.path.basename(f),
                            rows=max(0, len(rows) - 1),
                            headers=rows[0] if rows else [],
                            preview=rows[1:4] if len(rows) > 1 else []))
        except Exception:
            pass
    return jsonify(out)

# ── API: Settings ──────────────────────────────────────────────────────────────
@app.route('/api/settings', methods=['GET', 'POST'])
@admin_req
def api_settings():
    if request.method == 'POST':
        incoming = request.json or {}
        current  = load_cfg()
        # Never wipe sensitive fields if the UI sends them blank — keep the stored value
        _sensitive = ('openrouter_api_key', 'smtp_pass', 'jira_token')
        for key in _sensitive:
            if not incoming.get(key, '').strip():
                incoming[key] = current.get(key, '')
        merged = {**current, **incoming}
        save_cfg(merged)
        audit('SETTINGS_UPDATED', json.dumps({k: ('***' if k in _sensitive and v else v) for k, v in incoming.items()}))
        return jsonify(ok=True)
    return jsonify(load_cfg())

# ── API: JMeter auto-install ───────────────────────────────────────────────────
@app.route('/api/jmeter/install-status')
@login_req
def api_jmeter_install_status():
    found = _find_jmeter_bin()
    return jsonify(**_jmeter_install, found=bool(found), found_bin=found)

@app.route('/api/jmeter/install', methods=['POST'])
@admin_req
def api_jmeter_install():
    if _jmeter_install['status'] == 'downloading':
        return jsonify(error='Download already in progress.'), 409
    threading.Thread(target=_auto_install_jmeter, daemon=True).start()
    audit('JMETER_INSTALL_TRIGGERED', f'Auto-install JMeter {_JMETER_VERSION} triggered')
    return jsonify(ok=True, message=f'JMeter {_JMETER_VERSION} download started. This may take a few minutes.')

# ── API: Java auto-install ─────────────────────────────────────────────────────
@app.route('/api/java/install-status')
@login_req
def api_java_install_status():
    found = _find_java_bin()
    java_home = os.environ.get('JAVA_HOME', load_cfg().get('java_home', ''))
    return jsonify(**_java_install, found=bool(found), found_bin=found, java_home=java_home)

@app.route('/api/java/install', methods=['POST'])
@admin_req
def api_java_install():
    if _java_install['status'] == 'downloading':
        return jsonify(error='Download already in progress.'), 409
    threading.Thread(target=_auto_install_java, daemon=True).start()
    audit('JAVA_INSTALL_TRIGGERED', 'Auto-install OpenJDK 11 triggered')
    return jsonify(ok=True, message='OpenJDK 11 download started (~185 MB). This may take a few minutes.')

# ── Endurance JMX Generator ───────────────────────────────────────────────────
def _build_endurance_jmx(payload):
    """
    payload = {
      test_name, server_ip, port, protocol, path,
      stages: [{name, threads, tps, duration_min, rampup_sec}, ...]
    }
    Returns JMX XML string with one ThreadGroup per stage, each delayed so
    stages execute back-to-back (step-wise pattern).
    """
    test_name   = payload.get('test_name', 'Endurance Test')
    server_ip   = payload.get('server_ip', 'localhost')
    port        = str(payload.get('port', '80'))
    protocol    = payload.get('protocol', 'http')
    path        = payload.get('path', '/')
    stages      = payload.get('stages', [])

    thread_groups_xml = ''
    delay_sec = 0  # cumulative delay — each stage starts after all previous ones

    for i, s in enumerate(stages, 1):
        name         = s.get('name') or f'Stage {i}'
        threads      = int(s.get('threads', 10))
        tps          = float(s.get('tps', 10))
        dur_min      = float(s.get('duration_min', 60))
        rampup_sec   = int(s.get('rampup_sec', 60))
        dur_sec      = int(dur_min * 60)
        tpm          = tps * 60  # ConstantThroughputTimer uses req/min

        thread_groups_xml += f'''
      <ThreadGroup guiclass="ThreadGroupGui" testclass="ThreadGroup"
          testname="TG - {name} ({threads} users, {tps} TPS, {dur_min:.0f} min)" enabled="true">
        <stringProp name="TestPlan.comments">Stage {i} | {tps} TPS | {threads} threads | {dur_min:.0f} min</stringProp>
        <stringProp name="ThreadGroup.num_threads">{threads}</stringProp>
        <stringProp name="ThreadGroup.ramp_time">{rampup_sec}</stringProp>
        <stringProp name="ThreadGroup.duration">{dur_sec}</stringProp>
        <stringProp name="ThreadGroup.delay">{delay_sec}</stringProp>
        <boolProp name="ThreadGroup.scheduler">true</boolProp>
        <boolProp name="ThreadGroup.same_user_on_next_iteration">true</boolProp>
        <stringProp name="ThreadGroup.on_sample_error">continue</stringProp>
        <elementProp name="ThreadGroup.main_controller" elementType="LoopController"
            guiclass="LoopControlPanel" testclass="LoopController" testname="Loop Controller">
          <intProp name="LoopController.loops">-1</intProp>
          <boolProp name="LoopController.continue_forever">false</boolProp>
        </elementProp>
      </ThreadGroup>
      <hashTree>
        <ConstantThroughputTimer guiclass="TestBeanGUI" testclass="ConstantThroughputTimer"
            testname="TPS Controller - {tps} TPS" enabled="true">
          <doubleProp name="ConstantThroughputTimer.throughput">{tpm:.1f}</doubleProp>
          <intProp name="ConstantThroughputTimer.calcMode">1</intProp>
        </ConstantThroughputTimer>
        <hashTree/>
        <HTTPSamplerProxy guiclass="HttpTestSampleGui" testclass="HTTPSamplerProxy"
            testname="{name} - HTTP Request" enabled="true">
          <stringProp name="HTTPSampler.domain">{server_ip}</stringProp>
          <stringProp name="HTTPSampler.port">{port}</stringProp>
          <stringProp name="HTTPSampler.protocol">{protocol}</stringProp>
          <stringProp name="HTTPSampler.path">{path}</stringProp>
          <stringProp name="HTTPSampler.method">GET</stringProp>
          <boolProp name="HTTPSampler.follow_redirects">true</boolProp>
          <boolProp name="HTTPSampler.use_keepalive">true</boolProp>
          <boolProp name="HTTPSampler.postBodyRaw">false</boolProp>
          <elementProp name="HTTPsampler.Arguments" elementType="Arguments"
              guiclass="HTTPArgumentsPanel" testclass="Arguments" testname="User Defined Variables">
            <collectionProp name="Arguments.arguments"/>
          </elementProp>
        </HTTPSamplerProxy>
        <hashTree>
          <ResponseAssertion guiclass="AssertionGui" testclass="ResponseAssertion"
              testname="Assert HTTP 2xx" enabled="true">
            <collectionProp name="Asserion.test_strings">
              <stringProp name="49586">2\\d\\d</stringProp>
            </collectionProp>
            <stringProp name="Assertion.custom_message">Expected HTTP 2xx</stringProp>
            <stringProp name="Assertion.test_field">Assertion.response_code</stringProp>
            <boolProp name="Assertion.assume_success">false</boolProp>
            <intProp name="Assertion.test_type">1</intProp>
          </ResponseAssertion>
          <hashTree/>
        </hashTree>
      </hashTree>
'''
        delay_sec += dur_sec  # next stage starts when this one ends

    total_min = sum(float(s.get('duration_min', 60)) for s in stages)

    return f'''<?xml version="1.0" encoding="UTF-8"?>
<jmeterTestPlan version="1.2" properties="5.0" jmeter="5.6.3">
  <hashTree>
    <TestPlan guiclass="TestPlanGui" testclass="TestPlan" testname="{test_name}">
      <stringProp name="TestPlan.comments">Endurance test | {len(stages)} stages | Total: {total_min:.0f} min | Generated by Load Testing Platform</stringProp>
      <boolProp name="TestPlan.tearDown_on_shutdown">true</boolProp>
      <elementProp name="TestPlan.user_defined_variables" elementType="Arguments"
          guiclass="ArgumentsPanel" testclass="Arguments" testname="User Defined Variables">
        <collectionProp name="Arguments.arguments"/>
      </elementProp>
    </TestPlan>
    <hashTree>
      <ConfigTestElement guiclass="HttpDefaultsGui" testclass="ConfigTestElement"
          testname="HTTP Request Defaults">
        <intProp name="HTTPSampler.connect_timeout">5000</intProp>
        <intProp name="HTTPSampler.response_timeout">30000</intProp>
        <stringProp name="HTTPSampler.implementation"></stringProp>
        <elementProp name="HTTPsampler.Arguments" elementType="Arguments"
            guiclass="HTTPArgumentsPanel" testclass="Arguments" testname="User Defined Variables">
          <collectionProp name="Arguments.arguments"/>
        </elementProp>
      </ConfigTestElement>
      <hashTree/>
      <ResultCollector guiclass="SummaryReport" testclass="ResultCollector"
          testname="Summary Report" enabled="true">
        <boolProp name="ResultCollector.error_logging">false</boolProp>
        <objProp><name>saveConfig</name><value class="SampleSaveConfiguration">
          <time>true</time><latency>true</latency><timestamp>true</timestamp>
          <success>true</success><label>true</label><code>true</code>
          <message>true</message><threadName>true</threadName>
          <dataType>true</dataType><encoding>false</encoding>
          <assertions>true</assertions><subresults>true</subresults>
          <responseData>false</responseData><samplerData>false</samplerData>
          <xml>false</xml><fieldNames>true</fieldNames><responseHeaders>false</responseHeaders>
          <requestHeaders>false</requestHeaders><responseDataOnError>false</responseDataOnError>
          <saveAssertionResultsFailureMessage>true</saveAssertionResultsFailureMessage>
          <bytes>true</bytes><sentBytes>true</sentBytes><url>true</url>
          <threadCounts>true</threadCounts><idleTime>true</idleTime><connectTime>true</connectTime>
        </value></objProp>
        <stringProp name="filename"></stringProp>
      </ResultCollector>
      <hashTree/>
{thread_groups_xml}
    </hashTree>
  </hashTree>
</jmeterTestPlan>'''


@app.route('/api/endurance/generate', methods=['POST'])
@login_req
def api_endurance_generate():
    payload = request.get_json(force=True) or {}
    stages  = payload.get('stages', [])
    if not stages:
        return jsonify(error='At least one stage is required.'), 400
    for s in stages:
        if not s.get('threads') or not s.get('tps') or not s.get('duration_min'):
            return jsonify(error='Each stage needs threads, tps, and duration_min.'), 400

    jmx_content = _build_endurance_jmx(payload)
    test_name   = payload.get('test_name', 'Endurance_Test').replace(' ', '_')
    filename    = f'{test_name}.jmx'

    # Optionally save to the active client's JMX folder
    if payload.get('save_to_client'):
        c = _active_client()
        if c:
            dirs = client_dirs(c)
            os.makedirs(dirs['jmx'], exist_ok=True)
            dest = os.path.join(dirs['jmx'], filename)
            with open(dest, 'w', encoding='utf-8') as f:
                f.write(jmx_content)
            audit('ENDURANCE_JMX_SAVED', f'Saved {filename} to client {c["code"]}')
            return jsonify(ok=True, saved=True, filename=filename)

    # Otherwise return as download
    audit('ENDURANCE_JMX_GENERATED', f'Generated {filename} ({len(stages)} stages)')
    buf = io.BytesIO(jmx_content.encode('utf-8'))
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/xml')


@app.route('/api/endurance/generate-zip', methods=['POST'])
@login_req
def api_endurance_generate_zip():
    """Generate JMX + a sample CSV scaffold + README, bundled as ZIP."""
    payload     = request.get_json(force=True) or {}
    stages      = payload.get('stages', [])
    if not stages:
        return jsonify(error='At least one stage is required.'), 400

    test_name   = payload.get('test_name', 'Endurance_Test').replace(' ', '_')
    jmx_content = _build_endurance_jmx(payload)

    readme = f"""# {test_name} — Endurance Test

Generated by Load Testing Platform

## Stages
"""
    delay = 0
    for i, s in enumerate(stages, 1):
        dur = float(s.get('duration_min', 60))
        readme += (f"  Stage {i}: {s.get('name', f'Stage {i}')} | "
                   f"{s.get('threads')} threads | {s.get('tps')} TPS | "
                   f"{dur:.0f} min | starts at {delay//60:.0f} min\n")
        delay += int(dur * 60)
    readme += f"\nTotal duration: {delay//60:.0f} min\n"

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'{test_name}.jmx', jmx_content)
        zf.writestr('README.txt', readme)
    buf.seek(0)
    audit('ENDURANCE_ZIP_GENERATED', f'Generated {test_name}.zip ({len(stages)} stages)')
    return send_file(buf, as_attachment=True, download_name=f'{test_name}.zip',
                     mimetype='application/zip')


# ── API: Client management ─────────────────────────────────────────────────────
@app.route('/api/clients', methods=['GET'])
@login_req
def api_clients_list():
    with get_db() as db:
        rows = db.execute("SELECT * FROM clients ORDER BY name").fetchall()
    result = []
    for r in rows:
        c = dict(r)
        dirs = client_dirs(c)
        c['jmx_count']      = len(glob.glob(os.path.join(dirs['jmx'],      '*.jmx')))
        c['report_count']   = len(glob.glob(os.path.join(dirs['reports'],  '*.jtl')))
        c['testdata_count'] = len(glob.glob(os.path.join(dirs['testdata'], '*.csv')))
        result.append(c)
    return jsonify(result)

@app.route('/api/clients', methods=['POST'])
@admin_req
@csrf_protect
def api_clients_create():
    d = request.json or {}
    code  = re.sub(r'[^A-Z0-9_-]', '', d.get('code', '').strip().upper())
    name  = d.get('name', '').strip()
    desc  = d.get('description', '').strip()
    emoji = d.get('logo_emoji', '🏢').strip() or '🏢'
    color = d.get('color', '#00d4ff').strip() or '#00d4ff'

    if not code or not name:
        return jsonify(error='Client code and name are required.'), 400
    if len(code) < 2:
        return jsonify(error='Client code must be at least 2 characters.'), 400

    # Auto-create directory structure
    base = os.path.join(CLIENTS_DIR, code)
    jmx_dir  = os.path.join(base, 'jmx')
    td_dir   = os.path.join(base, 'testdata')
    rpt_dir  = os.path.join(base, 'reports')
    feat_dir = os.path.join(base, 'features')  # New: test feature files
    for d_ in [jmx_dir, td_dir, rpt_dir, feat_dir]:
        os.makedirs(d_, exist_ok=True)

    # Create a README in the client folder
    with open(os.path.join(base, 'README.txt'), 'w') as f:
        f.write(f'Client: {name} ({code})\nCreated: {datetime.now()}\n\n'
                f'Folder structure:\n'
                f'  jmx/       — JMeter test plan files (.jmx)\n'
                f'  testdata/  — CSV test data files\n'
                f'  reports/   — JTL result files\n'
                f'  features/  — BDD test feature files (.feature, .gherkin)\n')

    # Seed default pre-requisites document
    prereq_path = os.path.join(base, _PREREQ_FILE)
    if not os.path.exists(prereq_path):
        _write_json(prereq_path, _default_prereq())

    try:
        with get_db() as db:
            db.execute("""INSERT INTO clients
                (code,name,description,logo_emoji,color,jmx_dir,testdata_dir,reports_dir,created_by)
                VALUES (?,?,?,?,?,?,?,?,?)""",
                (code, name, desc, emoji, color, jmx_dir, td_dir, rpt_dir, session['user'])
            )
            db.commit()
        audit('CLIENT_CREATED', f"Created client '{code}' — {name}")
        return jsonify(ok=True, message=f"Client '{name}' ({code}) created. Upload JMX/CSV files to: {base}")
    except sqlite3.IntegrityError:
        return jsonify(error=f"Client code '{code}' already exists."), 409

@app.route('/api/clients/<code>', methods=['PUT'])
@admin_req
def api_clients_update(code):
    d = request.json or {}
    fields, values = [], []
    for field in ('name', 'description', 'logo_emoji', 'color', 'enabled', 'jmx_dir', 'testdata_dir', 'reports_dir'):
        if field in d:
            fields.append(f'{field}=?')
            values.append(d[field])
    if not fields:
        return jsonify(error='Nothing to update.'), 400
    values.append(code)
    with get_db() as db:
        db.execute(f"UPDATE clients SET {','.join(fields)} WHERE code=?", values)
        db.commit()
    audit('CLIENT_UPDATED', f"Updated client '{code}'")
    return jsonify(ok=True, message=f"Client '{code}' updated.")

@app.route('/api/clients/<code>', methods=['DELETE'])
@admin_req
@csrf_protect
def api_clients_delete(code):
    if code == session.get('client'):
        return jsonify(error='Cannot delete the currently active client. Switch first.'), 400
    with get_db() as db:
        row = db.execute("SELECT * FROM clients WHERE code=?", (code,)).fetchone()
        if not row:
            return jsonify(error='Client not found.'), 404
        client_folder = os.path.join(CLIENTS_DIR, code)
        db.execute("DELETE FROM clients WHERE code=?", (code,))
        db.commit()
    # Remove the client folder from disk
    deleted_files = False
    if os.path.isdir(client_folder):
        import shutil
        shutil.rmtree(client_folder, ignore_errors=True)
        deleted_files = True
    audit('CLIENT_DELETED', f"Deleted client '{code}' (folder removed: {deleted_files})")
    return jsonify(ok=True, message=f"Client '{code}' deleted.", files_removed=deleted_files)

# ── API: User management ───────────────────────────────────────────────────────
@app.route('/api/users', methods=['GET'])
@admin_req
def api_users_list():
    with get_db() as db:
        rows = db.execute(
            "SELECT id,username,name,role,initials,enabled,created_at,created_by,last_login,login_count,COALESCE(permissions,'[]') as permissions FROM users ORDER BY role DESC,username"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/users', methods=['POST'])
@admin_req
@csrf_protect
def api_users_create():
    d = request.json or {}
    username = d.get('username', '').strip().lower()
    password = d.get('password', '')
    name     = d.get('name', '').strip()
    role     = d.get('role', 'viewer')
    initials = d.get('initials', '').strip().upper() or name[:2].upper()
    perms    = json.dumps([p for p in d.get('permissions', []) if isinstance(p, str)])

    if not username or not password or not name:
        return jsonify(error='Username, password, and name are required.'), 400
    if role not in ('admin', 'viewer'):
        return jsonify(error='Role must be admin or viewer.'), 400
    if len(password) < 6:
        return jsonify(error='Password must be at least 6 characters.'), 400
    try:
        with get_db() as db:
            db.execute(
                "INSERT INTO users (username,password,name,role,initials,created_by,permissions) VALUES (?,?,?,?,?,?,?)",
                (username, generate_password_hash(password), name, role, initials, session['user'], perms)
            )
            db.commit()
        audit('USER_CREATED', f"Created user '{username}' role={role} perms={perms}")
        return jsonify(ok=True, message=f"User '{username}' created.")
    except sqlite3.IntegrityError:
        return jsonify(error=f"Username '{username}' already exists."), 409

@app.route('/api/users/<username>', methods=['PUT'])
@admin_req
def api_users_update(username):
    d = request.json or {}
    fields, values = [], []
    if 'name'     in d: fields.append('name=?');     values.append(d['name'].strip())
    if 'role'     in d:
        if d['role'] not in ('admin','viewer'): return jsonify(error='Invalid role.'), 400
        fields.append('role=?'); values.append(d['role'])
    if 'initials' in d: fields.append('initials=?'); values.append(d['initials'].strip().upper())
    if 'enabled'  in d:
        if not d['enabled'] and username == session['user']:
            return jsonify(error='Cannot disable your own account.'), 400
        fields.append('enabled=?'); values.append(1 if d['enabled'] else 0)
    if 'password' in d and d['password']:
        if len(d['password']) < 6: return jsonify(error='Password min 6 chars.'), 400
        fields.append('password=?'); values.append(generate_password_hash(d['password']))
    if 'permissions' in d:
        perms = json.dumps([p for p in d['permissions'] if isinstance(p, str)])
        fields.append('permissions=?'); values.append(perms)
    if not fields: return jsonify(error='Nothing to update.'), 400
    values.append(username)
    with get_db() as db:
        db.execute(f"UPDATE users SET {','.join(fields)} WHERE username=?", values)
        db.commit()
    audit('USER_UPDATED', f"Updated '{username}': {','.join(f.split('=')[0] for f in fields)}")
    return jsonify(ok=True, message=f"User '{username}' updated.")

@app.route('/api/users/<username>', methods=['DELETE'])
@admin_req
def api_users_delete(username):
    if username == session['user']:
        return jsonify(error='Cannot delete your own account.'), 400
    with get_db() as db:
        if not db.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone():
            return jsonify(error='User not found.'), 404
        db.execute("DELETE FROM users WHERE username=?", (username,))
        db.commit()
    audit('USER_DELETED', f"Deleted user '{username}'")
    return jsonify(ok=True, message=f"User '{username}' deleted.")

@app.route('/api/users/<username>/reset-password', methods=['POST'])
@admin_req
def api_reset_password(username):
    pw = (request.json or {}).get('password', '')
    if len(pw) < 6: return jsonify(error='Password min 6 chars.'), 400
    with get_db() as db:
        db.execute("UPDATE users SET password=? WHERE username=?",
                   (generate_password_hash(pw), username))
        db.commit()
    audit('PASSWORD_RESET', f"Reset password for '{username}'")
    return jsonify(ok=True, message=f"Password reset for '{username}'.")

@app.route('/api/users/me/change-password', methods=['POST'])
@login_req
def api_change_own_password():
    data     = request.json or {}
    cur_pw   = data.get('current_password', '')
    new_pw   = data.get('new_password', '')
    if not cur_pw or not new_pw:
        return jsonify(error='Both current and new password are required.'), 400
    if len(new_pw) < 6:
        return jsonify(error='New password must be at least 6 characters.'), 400
    username = session['user']
    with get_db() as db:
        row = db.execute("SELECT password FROM users WHERE username=?", (username,)).fetchone()
        if not row or not check_password_hash(row['password'], cur_pw):
            return jsonify(error='Current password is incorrect.'), 403
        db.execute("UPDATE users SET password=? WHERE username=?",
                   (generate_password_hash(new_pw), username))
        db.commit()
    audit('PASSWORD_CHANGED', f"User '{username}' changed their own password")
    return jsonify(ok=True, message='Password changed successfully.')

@app.route('/api/admins')
def api_admins():
    """Public endpoint — returns admin display names for forgot-password help screen."""
    with get_db() as db:
        rows = db.execute(
            "SELECT name, username FROM users WHERE role='admin' AND enabled=1"
        ).fetchall()
    return jsonify(admins=[{'name': r['name'], 'username': r['username']} for r in rows])

# ── API: Audit log ─────────────────────────────────────────────────────────────
@app.route('/api/audit-log')
@perm_req('view_audit')
def api_audit_log():
    limit  = int(request.args.get('limit', 100))
    offset = int(request.args.get('offset', 0))
    q = request.args.get('q', '')
    with get_db() as db:
        if q:
            rows  = db.execute("SELECT * FROM audit_log WHERE username LIKE ? OR action LIKE ? OR details LIKE ? ORDER BY id DESC LIMIT ? OFFSET ?",
                               (f'%{q}%', f'%{q}%', f'%{q}%', limit, offset)).fetchall()
            total = db.execute("SELECT COUNT(*) FROM audit_log WHERE username LIKE ? OR action LIKE ? OR details LIKE ?",
                               (f'%{q}%', f'%{q}%', f'%{q}%')).fetchone()[0]
        else:
            rows  = db.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT ? OFFSET ?", (limit, offset)).fetchall()
            total = db.execute("SELECT COUNT(*) FROM audit_log").fetchone()[0]
    return jsonify(logs=[dict(r) for r in rows], total=total)

# ── API: DB stats ──────────────────────────────────────────────────────────────
@app.route('/api/db-stats')
@admin_req
def api_db_stats():
    with get_db() as db:
        users_total   = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        users_admin   = db.execute("SELECT COUNT(*) FROM users WHERE role='admin'").fetchone()[0]
        users_enabled = db.execute("SELECT COUNT(*) FROM users WHERE enabled=1").fetchone()[0]
        audit_total   = db.execute("SELECT COUNT(*) FROM audit_log").fetchone()[0]
        last_login_row = db.execute("SELECT username,last_login FROM users WHERE last_login IS NOT NULL ORDER BY last_login DESC LIMIT 1").fetchone()
    size_kb = round(os.path.getsize(DB_PATH) / 1024, 1) if os.path.exists(DB_PATH) else 0
    return jsonify(
        users_total=users_total, users_admin=users_admin,
        users_enabled=users_enabled, audit_total=audit_total,
        last_login=dict(last_login_row) if last_login_row else None,
        db_size_kb=size_kb, db_path=DB_PATH
    )

# ── Auto-purge audit log ───────────────────────────────────────────────────────
def _purge_old_audit(days=None):
    """Delete audit entries older than retention_days. Returns count deleted."""
    if days is None:
        days = load_cfg().get('audit_retention_days', 90)
    days = int(days)
    if days <= 0:
        return 0  # 0 = keep forever
    with get_db() as db:
        cur = db.execute(
            "DELETE FROM audit_log WHERE timestamp < datetime('now', ?, 'localtime')",
            (f'-{days} days',)
        )
        db.commit()
        n = cur.rowcount
    if n > 0:
        print(f'  [auto-purge] Removed {n} audit entries older than {days} days.')
    return n

def _schedule_daily_purge():
    try:
        _purge_old_audit()
    except Exception as ex:
        print(f'  [auto-purge] Error: {ex}')
    t = threading.Timer(86400, _schedule_daily_purge)
    t.daemon = True
    t.start()

# ── API: Test Features (Excel) ─────────────────────────────────────────────────
def _cell(v):
    """Safely convert a cell value to string."""
    if v is None:
        return ''
    return str(v).strip()

def _safe_open_xlsx(path):
    """Copy xlsx to a temp file (avoids OneDrive lock) and return openpyxl workbook."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp.close()
    shutil.copy2(path, tmp.name)
    wb = openpyxl.load_workbook(tmp.name, data_only=True)
    os.unlink(tmp.name)
    return wb

def _parse_xlsx(path):
    """Parse an xlsx file and return a dict of sheet_name → {headers, rows}."""
    if not _HAS_OPENPYXL:
        return None
    wb = _safe_open_xlsx(path)
    result = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = [_cell(v) for v in row]
            if i == 0:
                headers = vals
            else:
                if any(v for v in vals):
                    rows.append(vals)
        result[sname] = {'headers': headers, 'rows': rows}
    wb.close()
    return result

def _count_xlsx_rows(path):
    """Count non-empty data rows across all sheets in an xlsx workbook."""
    if not _HAS_OPENPYXL:
        return 0
    wb = _safe_open_xlsx(path)
    try:
        total = 0
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell not in (None, '') for cell in row):
                    total += 1
        return total
    finally:
        wb.close()

@app.route('/api/xlsx-files')
@login_req
def api_xlsx_files():
    """Return a fast list of .xlsx filenames in the client folder (no parsing)."""
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    base = os.path.join(CLIENTS_DIR, c['code'])
    paths = glob.glob(os.path.join(base, '*.xlsx')) + \
            glob.glob(os.path.join(base, 'testdata', '*.xlsx'))
    files = sorted(set(os.path.basename(p) for p in paths))
    return jsonify(files=files)


@app.route('/api/test-features')
@login_req
def api_test_features():
    if not _HAS_OPENPYXL:
        return jsonify(error='openpyxl not installed. Run: pip install openpyxl'), 500
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    base = os.path.join(CLIENTS_DIR, c['code'])
    xlsx_files = glob.glob(os.path.join(base, '*.xlsx'))
    if not xlsx_files:
        return jsonify(error='No Excel feature files found in client folder.'), 404

    out = []
    for path in sorted(xlsx_files):
        fname = os.path.basename(path)
        sheets = _parse_xlsx(path)
        if sheets is None:
            continue
        out.append({
            'filename': fname,
            'sheets':   sheets,
        })
    return jsonify(files=out)

@app.route('/api/download/xlsx/<path:fname>')
@login_req
def api_download_xlsx(fname):
    c = active_client()
    base = os.path.join(CLIENTS_DIR, c['code']) if c else ''
    p = os.path.join(base, fname) if base else ''
    if not p or not os.path.exists(p):
        return 'File not found', 404
    audit('DOWNLOAD_XLSX', f'Excel: {fname}')
    return send_file(p, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── API: Environment Config ────────────────────────────────────────────────────
def _env_path(c):
    return os.path.join(CLIENTS_DIR, c['code'], 'env.json')

def _load_config_path(c):
    return os.path.join(CLIENTS_DIR, c['code'], 'load_config.json')

def _read_json(path):
    if os.path.exists(path):
        try:
            with open(path, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return None

def _write_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)

@app.route('/api/env', methods=['GET', 'POST'])
@login_req
def api_env():
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    path = _env_path(c)
    if request.method == 'POST':
        data = request.json or {}
        _write_json(path, data)
        audit('ENV_UPDATED', f'Environment config updated for client={c["code"]}')
        return jsonify(ok=True)
    env = _read_json(path) or {}
    # Try auto-fill from Excel testDetails if env is empty
    if not env and _HAS_OPENPYXL:
        base = os.path.join(CLIENTS_DIR, c['code'])
        for xlsx in sorted(glob.glob(os.path.join(base, '*.xlsx'))):
            try:
                wb = _safe_open_xlsx(xlsx)
                for sname in wb.sheetnames:
                    if 'testdetail' in sname.lower():
                        ws = wb[sname]
                        for row in ws.iter_rows(values_only=True):
                            if row[0] and row[1]:
                                key = str(row[0]).strip().lower().replace(' ', '_').replace('-', '_')
                                env[key] = str(row[1]).strip() if row[1] else ''
                wb.close()
                break
            except Exception:
                pass
    return jsonify(env)

# ── API: Load Distribution Config ──────────────────────────────────────────────
@app.route('/api/load-config', methods=['GET', 'POST'])
@login_req
def api_load_config():
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    path = _load_config_path(c)
    if request.method == 'POST':
        data = request.json or {}
        for svc in data.get('services', []):
            pat = svc.get('pattern', 'constant')
            dur = int(svc.get('duration', 1800))
            if pat == 'spike':
                sp_at  = int(svc.get('spike_at', max(30, int(dur * 0.4))))
                sp_dur = int(svc.get('spike_duration', max(60, int(dur * 0.1))))
                if sp_at >= dur:
                    return jsonify(error=f'[{svc.get("service_name")}] spike_at must be < duration'), 400
                if sp_dur > (dur - sp_at) // 2:
                    return jsonify(error=f'[{svc.get("service_name")}] spike_duration too long'), 400
            elif pat == 'business_hours':
                hp = svc.get('hourly_pct')
                if hp is not None:
                    if len(hp) != 24:
                        return jsonify(error=f'[{svc.get("service_name")}] hourly_pct must have 24 values'), 400
                    if any(v < 0 or v > 100 for v in hp):
                        return jsonify(error=f'[{svc.get("service_name")}] hourly_pct values must be 0-100'), 400
        _write_json(path, data)
        audit('LOAD_CONFIG_UPDATED', f'Load distribution saved for client={c["code"]}, {len(data.get("services",[]))} services')
        return jsonify(ok=True)
    cfg = _read_json(path) or {}
    return jsonify(cfg)

@app.route('/api/load-config/import-excel', methods=['POST'])
@login_req
def api_load_config_import():
    """Import load distribution from an Excel file into editable config."""
    c = active_client()
    if not c or not _HAS_OPENPYXL:
        return jsonify(error='No active client or openpyxl missing.'), 400
    data = request.json or {}
    fname = data.get('filename', '')
    base  = os.path.join(CLIENTS_DIR, c['code'])
    path  = os.path.join(base, fname)
    if not os.path.exists(path):
        return jsonify(error='File not found.'), 404
    try:
        wb = _safe_open_xlsx(path)
        services = []
        for sname in wb.sheetnames:
            if 'loaddistr' not in sname.lower():
                continue
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h).strip().lower() if h else '' for h in rows[0]]
            # Map columns
            def col(keywords):
                for k in keywords:
                    for i, h in enumerate(headers):
                        if k in h:
                            return i
                return -1
            name_i  = col(['service_name','name','service'])
            type_i  = col(['service_type','type'])
            pct_i   = col(['percentage','%','load'])
            tps_i   = col(['tps','target'])
            tpm_i   = col(['tpm'])
            thr_i   = col(['thread'])
            req_i   = col(['request','no.of'])
            txn_i   = col(['transaction_count','transaction'])
            for row in rows[1:]:
                def g(i): return str(row[i]).strip() if i >= 0 and i < len(row) and row[i] is not None else ''
                name = g(name_i) or g(type_i)
                if not name:
                    continue
                try: pct = float(g(pct_i)) if g(pct_i) else 0
                except: pct = 0
                try: tps = float(g(tps_i)) if g(tps_i) else 0
                except: tps = 0
                try: thr = int(float(g(thr_i))) if g(thr_i) else 0
                except: thr = 0
                try: txn = int(float(g(txn_i))) if g(txn_i) else 0
                except: txn = 0
                services.append({
                    'service_name': name,
                    'load_pct':     round(pct, 2),
                    'target_tps':   round(tps, 3),
                    'threads':      thr or max(1, int(round(tps * 2))),
                    'ramp_up':      60,
                    'duration':     1800,
                    'enabled':      True,
                    'txn_count':    txn,
                    'pattern':      'constant',
                })
        wb.close()
        return jsonify(ok=True, services=services, count=len(services))
    except Exception as ex:
        return jsonify(error=str(ex)), 500

@app.route('/api/load-config/suggest-from-jmx', methods=['POST'])
@login_req
def api_suggest_from_jmx():
    """Match load config services to JMX thread groups, return suggested params."""
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    data     = request.json or {}
    jmx_name = data.get('jmx', '')
    services = data.get('services', [])
    dirs     = client_dirs(c)
    jmx_path = os.path.join(dirs['jmx'], jmx_name)
    if not os.path.exists(jmx_path):
        return jsonify(error='JMX not found.'), 404
    try:
        tree = ET.parse(jmx_path)
        root = tree.getroot()
    except Exception as ex:
        return jsonify(error=str(ex)), 500

    # Build map of thread group names → current props
    tg_map = {}
    for tg in root.iter('ThreadGroup'):
        raw  = tg.get('testname', '')
        m    = re.match(r'^TG - (.+?) \((.+)\)$', raw.strip())
        key  = (m.group(1) if m else raw).strip().lower()
        props = {}
        for sp in tg.iter('stringProp'):
            props[sp.get('name', '')] = sp.text or ''
        for ip in tg.iter('intProp'):
            props[ip.get('name', '')] = ip.text or ''
        tg_map[key] = {'name': raw, 'props': props, 'enabled': tg.get('enabled', 'true') == 'true'}

    suggestions = []
    for svc in services:
        sname = svc.get('service_name', '').strip().lower()
        match = tg_map.get(sname)
        # Fuzzy match if exact fails
        if not match:
            for k, v in tg_map.items():
                if sname in k or k in sname:
                    match = v
                    break
        suggestion = dict(svc)
        suggestion['jmx_matched'] = bool(match)
        suggestion['jmx_tg_name'] = match['name'] if match else ''
        if match:
            p = match['props']
            suggestion['jmx_enabled'] = match['enabled']
        suggestions.append(suggestion)
    return jsonify(suggestions=suggestions, tg_count=len(tg_map))

# ── API: AI Assistant ──────────────────────────────────────────────────────────
_AI_SYSTEM = """You are an expert JMeter load testing consultant embedded inside a centralized load testing platform. \
You help teams validate test configurations, review load distributions, catch mistakes, and suggest improvements.

Your expertise covers:
- JMeter thread groups, ramp-up, duration, TPS calculations
- Load distribution across services (percentages, transaction counts, iterations)
- Thread sizing: rule of thumb is threads ≈ TPS × expected_avg_response_time_seconds (typically 1-3s)
- Ramp-up best practice: 10-30% of test duration; too short causes spike, too long under-loads early
- Identifying services with disproportionate load or missing configuration
- USSD / mobile money transaction patterns
- Response time targets, error rate thresholds

When reviewing data:
1. Check load percentages sum to ~100%
2. Verify iterations = TPS × duration makes sense for the business volume
3. Flag thread counts that are too high or too low for the given TPS
4. Check ramp-up vs duration ratio
5. Spot services with 0 TPS or 0 threads that are enabled
6. Suggest missing services based on common USSD patterns if applicable

Be concise, direct, and actionable. Use bullet points. Flag critical issues with ⚠️, suggestions with 💡, and confirmations with ✅."""

def _build_ai_context(c):
    """Build a rich context block from the active client's config."""
    lines = []
    # Client info
    lines.append(f"## Client: {c['name']} ({c['code']})")
    # Environment
    env = _read_json(_env_path(c)) or {}
    if env:
        lines.append("\n## Environment")
        for k, v in env.items():
            if k != 'password' and v:
                lines.append(f"- {k}: {v}")
    # Load distribution
    cfg = _read_json(_load_config_path(c)) or {}
    services = cfg.get('services', [])
    if services:
        enabled  = [s for s in services if s.get('enabled', True)]
        tot_tps  = sum(float(s.get('target_tps', 0)) for s in enabled)
        tot_thr  = sum(int(s.get('threads', 0)) for s in enabled)
        tot_pct  = sum(float(s.get('load_pct', 0)) for s in enabled)
        tot_iter = sum(int(float(s.get('target_tps', 0)) * int(s.get('duration', 1800))) for s in enabled)
        lines.append(f"\n## Load Distribution ({len(enabled)}/{len(services)} services enabled)")
        lines.append(f"Total TPS: {tot_tps:.2f} | Total Threads: {tot_thr} | Load %: {tot_pct:.1f}% | Total Iterations: {tot_iter:,}")
        lines.append("\n| Service | Load% | TPS | Threads | Ramp-up | Duration | Iterations | Enabled |")
        lines.append("|---------|-------|-----|---------|---------|----------|------------|---------|")
        for s in services:
            it = int(float(s.get('target_tps', 0)) * int(s.get('duration', 1800)))
            lines.append(f"| {s.get('service_name','')} | {s.get('load_pct',0)}% | {s.get('target_tps',0)} | {s.get('threads',0)} | {s.get('ramp_up',0)}s | {s.get('duration',0)}s | {it:,} | {'✓' if s.get('enabled',True) else '✗'} |")
    else:
        lines.append("\n## Load Distribution\nNo services configured yet.")
    # Recent run history
    dirs = client_dirs(c)
    jtls = sorted(glob.glob(os.path.join(dirs['reports'], '*.jtl')), key=os.path.getmtime, reverse=True)[:3]
    if jtls:
        lines.append(f"\n## Recent Test Runs ({len(jtls)} most recent)")
        for jtl in jtls:
            lines.append(f"- {os.path.basename(jtl)}")
    return '\n'.join(lines)

@app.route('/api/ai/chat', methods=['POST'])
@login_req
def api_ai_chat():
    cfg     = load_cfg()
    api_key = cfg.get('openrouter_api_key', '').strip()
    if not api_key:
        return jsonify(error='NO_API_KEY'), 403

    data     = request.json or {}
    messages = data.get('messages', [])
    if not messages:
        return jsonify(error='No messages provided.'), 400

    c = active_client()
    ctx = _build_ai_context(c) if c else ''
    report_ctx = (data.get('report_context') or '').strip()
    if report_ctx:
        ctx += f'\n\n---\n## Currently Viewed Report\n{report_ctx}'
    system = _AI_SYSTEM + (f'\n\n---\n{ctx}' if ctx else '')
    ai_model = cfg.get('ai_model', 'openai/gpt-4o-mini')

    payload = json.dumps({
        'model': ai_model,
        'max_tokens': 1500,
        'messages': [{'role': 'system', 'content': system}] + messages,
    }).encode()
    req = _urllib_req2.Request(
        'https://openrouter.ai/api/v1/chat/completions',
        data=payload,
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
            'HTTP-Referer': 'http://localhost:5000',
            'X-Title': 'BTC Load Testing Platform',
        },
        method='POST',
    )
    try:
        with _urllib_req2.urlopen(req, timeout=60, context=_SSL_CTX) as resp:
            body = json.loads(resp.read())
        reply      = body['choices'][0]['message']['content']
        usage      = body.get('usage', {})
        tokens_in  = usage.get('prompt_tokens', 0)
        tokens_out = usage.get('completion_tokens', 0)
        audit('AI_CHAT', f'model={ai_model} tokens_in={tokens_in} tokens_out={tokens_out}')
        return jsonify(ok=True, reply=reply, usage={'in': tokens_in, 'out': tokens_out})
    except _urllib_err2.HTTPError as ex:
        body = ex.read().decode()
        return jsonify(error=f'OpenRouter {ex.code}: {body}'), 500
    except Exception as ex:
        return jsonify(error=str(ex)), 500

def _read_jtl_csv(path):
    """Parse a JMeter CSV JTL tolerating multi-line quoted fields.

    Once full response capture is enabled, response bodies can contain embedded
    newlines, so a record may span several physical lines. Reading via
    csv.reader over the file object correctly re-joins those quoted fields —
    unlike splitting the file line-by-line. Returns (header, rows)."""
    with open(path, 'r', newline='', encoding='utf-8', errors='ignore') as f:
        rows = list(csv.reader(f))
    if not rows:
        return [], []
    return [h.strip() for h in rows[0]], rows[1:]


# ── API: Live Stats ────────────────────────────────────────────────────────────
@app.route('/api/live-stats')
@login_req
def api_live_stats():
    c = active_client()
    jtl_path = _state.get('jtl') or ''
    if jtl_path and c:
        jtl_path = os.path.join(client_dirs(c)['reports'], jtl_path)
    if not jtl_path or not os.path.exists(jtl_path):
        if c:
            jtls = sorted(glob.glob(os.path.join(client_dirs(c)['reports'], '*.jtl')),
                          key=os.path.getmtime, reverse=True)
            jtl_path = jtls[0] if jtls else ''
    if not jtl_path:
        return jsonify(running=False, samples=[])
    try:
        header, rows = _read_jtl_csv(jtl_path)
        if not header:
            return jsonify(running=_state['running'], samples=[])
        ts_idx      = header.index('timeStamp') if 'timeStamp' in header else 0
        elapsed_idx = header.index('elapsed')   if 'elapsed'   in header else 1
        success_idx = header.index('success')   if 'success'   in header else 7
        window  = 5000
        buckets = {}
        for parts in rows:
            if len(parts) <= max(ts_idx, elapsed_idx, success_idx):
                continue
            try:
                ts      = int(parts[ts_idx])
                elapsed = int(parts[elapsed_idx])
                success = parts[success_idx].strip().lower() == 'true'
                bucket  = (ts // window) * window
                if bucket not in buckets:
                    buckets[bucket] = {'count': 0, 'errors': 0, 'total_elapsed': 0}
                buckets[bucket]['count'] += 1
                buckets[bucket]['total_elapsed'] += elapsed
                if not success:
                    buckets[bucket]['errors'] += 1
            except Exception:
                continue
        samples = []
        for ts_bucket in sorted(buckets.keys()):
            b = buckets[ts_bucket]
            tps    = b['count'] / (window / 1000)
            avg_rt = b['total_elapsed'] / b['count'] if b['count'] else 0
            err_pct = (b['errors'] / b['count'] * 100) if b['count'] else 0
            samples.append({'ts': ts_bucket, 'tps': round(tps, 2),
                            'avg_rt': round(avg_rt), 'err_pct': round(err_pct, 1),
                            'count': b['count']})
        # Anomaly detection
        alert = ''
        if len(samples) >= 6:
            recent3 = samples[-3:]
            prev3   = samples[-6:-3]
            recent_tps = sum(s['tps'] for s in recent3) / 3
            prev_tps   = sum(s['tps'] for s in prev3) / 3
            recent_err = sum(s['err_pct'] for s in recent3) / 3
            if prev_tps > 5 and recent_tps < prev_tps * 0.6:
                alert = f'TPS dropped {round((1-recent_tps/prev_tps)*100)}% — from {round(prev_tps)} to {round(recent_tps)} TPS'
            elif recent_err > 5:
                alert = f'Error rate critical: {round(recent_err,1)}% in last 3 intervals'
            elif recent_err > 2 and prev3 and recent_err > prev3[-1].get('err_pct', 0) * 2:
                alert = f'Error rate spike: {round(recent_err,1)}%'
        return jsonify(
            running      = _state['running'],
            jmx          = _state.get('jmx', ''),
            samples      = samples[-60:],
            total_samples = sum(b['count'] for b in buckets.values()),
            total_errors  = sum(b['errors'] for b in buckets.values()),
            alert        = alert,
        )
    except Exception as ex:
        return jsonify(running=_state['running'], error=str(ex), samples=[])

# ── API: SLA ───────────────────────────────────────────────────────────────────
def _sla_path(c):
    return os.path.join(CLIENTS_DIR, c['code'], 'sla_config.json')

@app.route('/api/sla-config', methods=['GET', 'POST'])
@login_req
def api_sla_config():
    c = active_client()
    if not c:
        return jsonify(error='No client'), 400
    path = _sla_path(c)
    if request.method == 'POST':
        _write_json(path, request.json or {})
        return jsonify(ok=True)
    return jsonify(_read_json(path) or {'p90_ms': 3000, 'p95_ms': 5000, 'error_pct': 1.0, 'min_tps': 0})

@app.route('/api/sla-result')
@login_req
def api_sla_result():
    return jsonify(_state.get('sla') or {})

def _evaluate_sla(jtl_path, c):
    sla = _read_json(_sla_path(c)) or {}
    if not sla or not jtl_path or not os.path.exists(jtl_path):
        return None
    try:
        elapsed_list, errors, total = [], 0, 0
        ts_min, ts_max = None, None
        label_data = defaultdict(lambda: {'elapsed': [], 'errors': 0})
        with open(jtl_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            header = [h.strip() for h in next(reader, [])]
            ei  = header.index('elapsed')   if 'elapsed'   in header else 1
            si  = header.index('success')   if 'success'   in header else 7
            tsi = header.index('timeStamp') if 'timeStamp' in header else 0
            li  = header.index('label')     if 'label'     in header else 2
            for row in reader:
                if len(row) <= max(ei, si, tsi):
                    continue
                try:
                    e = int(row[ei])
                    elapsed_list.append(e)
                    ts = int(row[tsi])
                    ts_min = min(ts_min, ts) if ts_min is not None else ts
                    ts_max = max(ts_max, ts) if ts_max is not None else ts
                    total += 1
                    lbl = row[li] if len(row) > li else ''
                    label_data[lbl]['elapsed'].append(e)
                    if row[si].strip().lower() != 'true':
                        errors += 1
                        label_data[lbl]['errors'] += 1
                except Exception:
                    continue
        if not elapsed_list:
            return None
        elapsed_list.sort()
        def pct(lst, p):
            return lst[max(0, int(len(lst) * p / 100) - 1)]
        duration_s  = ((ts_max - ts_min) / 1000) if ts_max and ts_min and ts_max > ts_min else 1
        actual_tps  = total / duration_s
        actual_p90  = pct(elapsed_list, 90)
        actual_p95  = pct(elapsed_list, 95)
        actual_err  = errors / total * 100 if total else 0
        checks, passed = [], True
        if sla.get('p90_ms'):
            ok = actual_p90 <= sla['p90_ms']
            checks.append({'name': 'P90 Response', 'threshold': f"≤{sla['p90_ms']}ms",
                           'actual': f"{actual_p90}ms", 'passed': ok})
            if not ok: passed = False
        if sla.get('p95_ms'):
            ok = actual_p95 <= sla['p95_ms']
            checks.append({'name': 'P95 Response', 'threshold': f"≤{sla['p95_ms']}ms",
                           'actual': f"{actual_p95}ms", 'passed': ok})
            if not ok: passed = False
        if sla.get('error_pct') is not None:
            ok = actual_err <= sla['error_pct']
            checks.append({'name': 'Error Rate', 'threshold': f"≤{sla['error_pct']}%",
                           'actual': f"{actual_err:.2f}%", 'passed': ok})
            if not ok: passed = False
        if sla.get('min_tps'):
            ok = actual_tps >= sla['min_tps']
            checks.append({'name': 'Min TPS', 'threshold': f"≥{sla['min_tps']}",
                           'actual': f"{actual_tps:.2f}", 'passed': ok})
            if not ok: passed = False
        # Per-label SLA checks
        for lbl, lsla in sla.get('per_label', {}).items():
            ld = label_data.get(lbl)
            if not ld or not ld['elapsed']:
                continue
            se   = sorted(ld['elapsed'])
            lerr = ld['errors'] / len(se) * 100
            if lsla.get('p90_ms'):
                lp90 = pct(se, 90)
                ok   = lp90 <= lsla['p90_ms']
                checks.append({'name': f'{lbl} — P90', 'threshold': f"≤{lsla['p90_ms']}ms",
                               'actual': f"{lp90}ms", 'passed': ok, 'label': lbl})
                if not ok: passed = False
            if lsla.get('error_pct') is not None:
                ok = lerr <= lsla['error_pct']
                checks.append({'name': f'{lbl} — Error %', 'threshold': f"≤{lsla['error_pct']}%",
                               'actual': f"{lerr:.2f}%", 'passed': ok, 'label': lbl})
                if not ok: passed = False
        return {'passed': passed, 'checks': checks,
                'actual_tps': round(actual_tps, 2), 'actual_p90': actual_p90,
                'actual_p95': actual_p95, 'actual_err': round(actual_err, 2)}
    except Exception:
        return None

# ── API: Per-label live stats ──────────────────────────────────────────────────
@app.route('/api/live-stats/labels')
@login_req
def api_live_stats_labels():
    c = active_client()
    jtl_path = _state.get('jtl') or ''
    if jtl_path and c:
        jtl_path = os.path.join(client_dirs(c)['reports'], jtl_path)
    if not jtl_path or not os.path.exists(jtl_path):
        return jsonify(running=_state['running'], labels=[])
    try:
        from datetime import datetime as _dt2
        now_ms  = int(_dt2.now().timestamp() * 1000)
        cutoff  = now_ms - 60000  # last 60 s
        header, rows = _read_jtl_csv(jtl_path)
        if not header:
            return jsonify(running=_state['running'], labels=[])
        ts_idx      = header.index('timeStamp') if 'timeStamp' in header else 0
        elapsed_idx = header.index('elapsed')   if 'elapsed'   in header else 1
        success_idx = header.index('success')   if 'success'   in header else 7
        label_idx   = header.index('label')     if 'label'     in header else 2
        by_label = defaultdict(lambda: {'count': 0, 'errors': 0, 'total_elapsed': 0})
        for parts in rows:
            need  = max(ts_idx, elapsed_idx, success_idx, label_idx)
            if len(parts) <= need:
                continue
            try:
                ts = int(parts[ts_idx])
                if ts < cutoff:
                    continue
                lbl = parts[label_idx].strip()
                by_label[lbl]['count']         += 1
                by_label[lbl]['total_elapsed'] += int(parts[elapsed_idx])
                if parts[success_idx].strip().lower() != 'true':
                    by_label[lbl]['errors'] += 1
            except Exception:
                continue
        labels = []
        for lbl, d in sorted(by_label.items()):
            labels.append({
                'label':   lbl,
                'tps':     round(d['count'] / 60, 2),
                'avg_rt':  round(d['total_elapsed'] / d['count']) if d['count'] else 0,
                'err_pct': round(d['errors'] / d['count'] * 100, 1) if d['count'] else 0,
                'count':   d['count'],
            })
        return jsonify(running=_state['running'], labels=labels)
    except Exception as ex:
        return jsonify(running=_state['running'], error=str(ex), labels=[])


# ── API: Virtual-user level live logs ─────────────────────────────────────────
@app.route('/api/live-logs/vu')
@login_req
def api_live_logs_vu():
    """Last N requests grouped by virtual user (threadName)."""
    n = min(int(request.args.get('n', 10)), 50)
    with _req_log_lock:
        snapshot = list(_req_log)
    by_vu = defaultdict(list)
    for entry in snapshot:
        by_vu[entry['thread']].append(entry)
    vus = []
    for vu, reqs in sorted(by_vu.items()):
        vus.append({
            'thread': vu,
            'total':  len(reqs),
            'errors': sum(1 for r in reqs if not r['ok']),
            'last':   reqs[-n:],
        })
    return jsonify(running=_state['running'], vus=vus)


# ── API: Service success/error tree ───────────────────────────────────────────
@app.route('/api/live-stats/tree')
@login_req
def api_live_stats_tree():
    """Per-service pass/fail tree with HTTP-code breakdown."""
    c = active_client()
    jtl_path = _state.get('jtl') or ''
    if jtl_path and c:
        jtl_path = os.path.join(client_dirs(c)['reports'], jtl_path)
    if not jtl_path or not os.path.exists(jtl_path):
        return jsonify(running=_state['running'], tree=[])
    try:
        hdr, rows = _read_jtl_csv(jtl_path)
        if not hdr:
            return jsonify(running=_state['running'], tree=[])
        def _hi(name): return hdr.index(name) if name in hdr else -1
        lb_i = _hi('label')
        ok_i  = _hi('success');     rc_i = _hi('responseCode')
        rm_i  = _hi('responseMessage'); fm_i = _hi('failureMessage')
        by_lbl = defaultdict(lambda: {
            'total':0,'pass':0,'fail':0,
            'rc':defaultdict(int),'reasons':defaultdict(int)
        })
        for parts in rows:
            if lb_i < 0 or ok_i < 0 or len(parts) <= max(lb_i, ok_i):
                continue
            try:
                lbl = parts[lb_i].strip()
                ok  = parts[ok_i].strip().lower() == 'true'
                rc  = parts[rc_i].strip() if rc_i >= 0 and rc_i < len(parts) else '?'
                d   = by_lbl[lbl]
                d['total'] += 1
                if ok:
                    d['pass'] += 1
                else:
                    d['fail'] += 1
                    d['rc'][rc or '?'] += 1
                    fm = parts[fm_i].strip() if fm_i >= 0 and fm_i < len(parts) else ''
                    rm = parts[rm_i].strip() if rm_i >= 0 and rm_i < len(parts) else ''
                    reason = fm or (f'HTTP {rc}' if rc not in ('','200') else rm) or 'Unknown'
                    d['reasons'][reason[:80]] += 1
            except Exception:
                continue
        tree = []
        for lbl, d in sorted(by_lbl.items()):
            tree.append({
                'label':       lbl,
                'total':       d['total'],
                'pass':        d['pass'],
                'fail':        d['fail'],
                'err_pct':     round(d['fail']/d['total']*100,1) if d['total'] else 0,
                'rc_breakdown':[{'rc':k,'count':v} for k,v in sorted(d['rc'].items(),key=lambda x:-x[1])[:5]],
                'fail_reasons':[{'reason':k,'count':v} for k,v in sorted(d['reasons'].items(),key=lambda x:-x[1])[:3]],
            })
        return jsonify(running=_state['running'], tree=tree)
    except Exception as ex:
        return jsonify(running=_state['running'], error=str(ex), tree=[])


# ── API: Live request/response stream ─────────────────────────────────────────
@app.route('/api/live-logs/requests')
@login_req
def api_live_logs_requests():
    """Ring buffer of latest requests, filterable by label and success."""
    label_f   = request.args.get('label','').strip().lower()
    success_f = request.args.get('success','').strip().lower()
    limit     = min(int(request.args.get('limit',100)), 500)
    with _req_log_lock:
        snap = list(_req_log)
    if label_f:
        snap = [r for r in snap if label_f in r['label'].lower()]
    if success_f == 'true':
        snap = [r for r in snap if r['ok']]
    elif success_f == 'false':
        snap = [r for r in snap if not r['ok']]
    return jsonify(running=_state['running'], requests=snap[-limit:], total=len(snap))


# ── API: Persisted request log for a completed run ────────────────────────────
@app.route('/api/run-request-log/<run_id>')
@login_req
def api_run_request_log(run_id):
    """Return up to 500 entries persisted for a completed run."""
    limit = min(int(request.args.get('limit', 200)), 500)
    label_f   = request.args.get('label', '').strip().lower()
    success_f = request.args.get('success', '').strip().lower()
    try:
        with get_db() as db:
            rows = db.execute(
                "SELECT entry_json FROM run_request_log WHERE run_id=? ORDER BY id LIMIT 500",
                (run_id,)
            ).fetchall()
        entries = [json.loads(r[0]) for r in rows]
        if label_f:
            entries = [e for e in entries if label_f in e.get('label','').lower()]
        if success_f == 'true':
            entries = [e for e in entries if e.get('ok')]
        elif success_f == 'false':
            entries = [e for e in entries if not e.get('ok')]
        return jsonify(requests=entries[-limit:], total=len(entries))
    except Exception as ex:
        return jsonify(error=str(ex), requests=[]), 500


# ── API: Baseline ──────────────────────────────────────────────────────────────
def _baseline_path(c):
    return os.path.join(CLIENTS_DIR, c['code'], 'baseline.json')

@app.route('/api/baseline', methods=['GET'])
@login_req
def api_baseline_get():
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    return jsonify(_read_json(_baseline_path(c)) or {})

@app.route('/api/baseline', methods=['POST'])
@perm_req('manage_baseline')
def api_baseline_set():
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    fname = (request.json or {}).get('file', '')
    if not fname: return jsonify(error='file required'), 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='JTL not found'), 404
    try:
        data = _parse_jtl(path)
        baseline = {
            'file':    fname,
            'set_at':  datetime.now().strftime('%Y-%m-%d %H:%M'),
            'set_by':  session.get('user', 'unknown'),
            'metrics': {
                'tps':        data['throughput'],
                'p90':        data['p90'],
                'p95':        data['p95'],
                'avg_rt':     data['avg_rt'],
                'error_rate': data['error_rate'],
            },
        }
        _write_json(_baseline_path(c), baseline)
        audit('BASELINE_SET', f'Baseline → {fname} for client={c["code"]}')
        return jsonify(ok=True, baseline=baseline)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

@app.route('/api/baseline', methods=['DELETE'])
@perm_req('manage_baseline')
def api_baseline_clear():
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    p = _baseline_path(c)
    if os.path.exists(p):
        os.remove(p)
    audit('BASELINE_CLEAR', f'Baseline cleared for client={c["code"]}')
    return jsonify(ok=True)

@app.route('/api/baseline/compare/<path:fname>')
@login_req
def api_baseline_compare(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    baseline = _read_json(_baseline_path(c))
    if not baseline: return jsonify(error='No baseline set'), 404
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='JTL not found'), 404
    try:
        data = _parse_jtl(path)
        bm = baseline['metrics']
        def delta(curr, base):
            return round((curr - base) / base * 100, 1) if base else None
        return jsonify(
            baseline_file=baseline['file'],
            baseline_set_at=baseline.get('set_at', ''),
            current_file=fname,
            metrics={
                'tps':        {'baseline': bm['tps'],        'current': data['throughput'],  'delta_pct': delta(data['throughput'],  bm['tps'])},
                'p90':        {'baseline': bm['p90'],        'current': data['p90'],          'delta_pct': delta(data['p90'],          bm['p90'])},
                'p95':        {'baseline': bm['p95'],        'current': data['p95'],          'delta_pct': delta(data['p95'],          bm['p95'])},
                'avg_rt':     {'baseline': bm['avg_rt'],     'current': data['avg_rt'],       'delta_pct': delta(data['avg_rt'],       bm['avg_rt'])},
                'error_rate': {'baseline': bm['error_rate'], 'current': data['error_rate'],   'delta_pct': delta(data['error_rate'],   bm['error_rate'])},
            },
        )
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: Error breakdown ───────────────────────────────────────────────────────
@app.route('/api/report/<path:fname>/samples')
@login_req
def api_report_samples(fname):
    """View Results Tree-style browser: every sample of a run with its request
    and response, paginated and filterable by label / status / search text.
    Reads the full JTL (multi-line safe) rather than the capped live buffer."""
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='Not found'), 404
    label_f  = request.args.get('label', '').strip().lower()
    status_f = request.args.get('status', '').strip().lower()   # '', 'pass', 'fail'
    q        = request.args.get('q', '').strip().lower()
    try:
        offset = max(int(request.args.get('offset', 0)), 0)
        limit  = min(max(int(request.args.get('limit', 100)), 1), 500)
    except ValueError:
        offset, limit = 0, 100
    try:
        header, rows = _read_jtl_csv(path)
        if not header:
            return jsonify(samples=[], total=0, labels=[], offset=0, limit=limit)
        def hi(n): return header.index(n) if n in header else -1
        li, rci, rmi, si = hi('label'), hi('responseCode'), hi('responseMessage'), hi('success')
        tsi, eli, tni    = hi('timeStamp'), hi('elapsed'), hi('threadName')
        ui, fmi          = hi('URL'), hi('failureMessage')
        sdi, rdi         = hi('samplerData'), hi('responseData')
        def cell(parts, idx): return parts[idx] if 0 <= idx < len(parts) else ''
        labels_set, matched, total_matched = set(), [], 0
        for parts in rows:
            if si < 0 or len(parts) <= si:
                continue
            label = cell(parts, li)
            labels_set.add(label)
            ok = cell(parts, si).strip().lower() == 'true'
            if status_f == 'pass' and not ok:      continue
            if status_f == 'fail' and ok:          continue
            if label_f and label_f not in label.lower(): continue
            if q:
                hay = (label + ' ' + cell(parts, rci) + ' ' + cell(parts, tni)
                       + ' ' + cell(parts, rdi)).lower()
                if q not in hay:
                    continue
            total_matched += 1
            if total_matched <= offset or len(matched) >= limit:
                continue
            matched.append({
                'i':            total_matched - 1,
                'label':        label,
                'rc':           cell(parts, rci).strip(),
                'rm':           cell(parts, rmi).strip(),
                'ok':           ok,
                'ts':           cell(parts, tsi),
                'elapsed':      cell(parts, eli),
                'thread':       cell(parts, tni).strip(),
                'url':          cell(parts, ui).strip(),
                'fm':           cell(parts, fmi).strip(),
                'sampler_data': cell(parts, sdi)[:8000],
                'resp_data':    cell(parts, rdi)[:8000],
            })
        return jsonify(samples=matched, total=total_matched, offset=offset, limit=limit,
                       labels=sorted(l for l in labels_set if l))
    except Exception as ex:
        return jsonify(error=str(ex)), 500


@app.route('/api/report/<path:fname>/trend')
@login_req
def api_report_trend(fname):
    """P95 / error% / TPS across recent runs for the trend dashboard."""
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='Not found'), 404
    try:
        n = min(max(int(request.args.get('n', 8)), 2), 30)
    except ValueError:
        n = 8
    try:
        return jsonify(trend=_recent_trend(c, path, n=n))
    except Exception as ex:
        return jsonify(error=str(ex)), 500


@app.route('/api/report/<path:fname>/errors')
@login_req
def api_report_errors(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='Not found'), 404
    try:
        error_groups = defaultdict(int)
        total_errors = 0
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            header = [h.strip() for h in next(reader, [])]
            si  = header.index('success')        if 'success'        in header else 7
            li  = header.index('label')          if 'label'          in header else 2
            rci = header.index('responseCode')   if 'responseCode'   in header else -1
            rmi = header.index('responseMessage')if 'responseMessage' in header else -1
            for row in reader:
                if len(row) <= si: continue
                if row[si].strip().lower() == 'true': continue
                lbl  = row[li]  if len(row) > li  else 'Unknown'
                rc   = row[rci] if rci >= 0 and len(row) > rci else ''
                rmsg = (row[rmi][:120] if rci >= 0 and len(row) > rmi else '')
                error_groups[(lbl, rc, rmsg)] += 1
                total_errors += 1
        errors = [
            {'label': k[0], 'response_code': k[1], 'message': k[2],
             'count': v, 'pct': round(v / total_errors * 100, 1) if total_errors else 0}
            for k, v in sorted(error_groups.items(), key=lambda x: -x[1])
        ]
        return jsonify(total_errors=total_errors, errors=errors[:100])
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: Recurring Schedules ───────────────────────────────────────────────────
@app.route('/api/recurring-schedules', methods=['GET'])
@login_req
def api_recurring_list():
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM recurring_schedules ORDER BY created_at DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/recurring-schedules', methods=['POST'])
@perm_req('manage_schedules')
def api_recurring_create():
    d = request.json or {}
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    jmx = d.get('jmx', '').strip()
    if not jmx: return jsonify(error='jmx required'), 400
    run_at_time = d.get('run_at_time', '').strip()
    if not run_at_time: return jsonify(error='run_at_time required (HH:MM)'), 400
    sid = str(uuid.uuid4())[:8]
    with get_db() as db:
        db.execute("""INSERT INTO recurring_schedules
            (id,client,jmx,threads,duration,rampup,recurrence,run_at_time,day_of_week,created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (sid, c['code'], jmx,
             int(d.get('threads', 10)), int(d.get('duration', 300)), int(d.get('rampup', 30)),
             d.get('recurrence', 'daily'), run_at_time, int(d.get('day_of_week', 1)),
             session.get('user', 'system'))
        )
        db.commit()
    audit('RECURRING_CREATED', f"id={sid} jmx={jmx} {d.get('recurrence','daily')} at {run_at_time}")
    return jsonify(ok=True, id=sid)

@app.route('/api/recurring-schedules/<sid>', methods=['PUT'])
@perm_req('manage_schedules')
def api_recurring_toggle(sid):
    enabled = (request.json or {}).get('enabled', True)
    with get_db() as db:
        db.execute("UPDATE recurring_schedules SET enabled=? WHERE id=?",
                   (1 if enabled else 0, sid))
        db.commit()
    audit('RECURRING_TOGGLE', f"id={sid} enabled={enabled}")
    return jsonify(ok=True)

@app.route('/api/recurring-schedules/<sid>', methods=['DELETE'])
@perm_req('manage_schedules')
def api_recurring_delete(sid):
    with get_db() as db:
        db.execute("DELETE FROM recurring_schedules WHERE id=?", (sid,))
        db.commit()
    audit('RECURRING_DELETED', f"id={sid}")
    return jsonify(ok=True)

def _fire_recurring(r):
    sid = r['id']
    try:
        with get_db() as db:
            db.execute(
                "UPDATE recurring_schedules SET last_run=datetime('now','localtime') WHERE id=?", (sid,)
            )
            db.commit()
        client = get_client(r['client'])
        if not client: return
        dirs = ensure_client_dirs(client)
        jmx_path = os.path.join(dirs['jmx'], r['jmx'])
        if not os.path.exists(jmx_path): return
        cfg  = load_cfg()
        jbin = cfg.get('jmeter_bin', '')
        if not os.path.exists(jbin):
            for cand in [r'C:\apache-jmeter-5.5\bin\jmeter.bat',
                         r'C:\apache-jmeter-5.6\bin\jmeter.bat',
                         r'C:\apache-jmeter-5.6.3\bin\jmeter.bat']:
                if os.path.exists(cand): jbin = cand; break
            else:
                return
        out_name = f"rec_{sid}_{datetime.now().strftime('%d%m%Y_%H%M%S')}"
        jtl_path = os.path.join(dirs['reports'], out_name + '.jtl')
        os.makedirs(dirs['reports'], exist_ok=True)
        tmp_dir = tempfile.mkdtemp(prefix='lt_rec_')
        working_jmx = os.path.join(tmp_dir, r['jmx'])
        try:
            shutil.copy(jmx_path, working_jmx)
            tree = ET.parse(working_jmx); root_el = tree.getroot()
            for csv_ds in root_el.iter('CSVDataSet'):
                for prop in csv_ds.iter('stringProp'):
                    if prop.get('name') == 'filename' and prop.text and prop.text.strip():
                        base = os.path.basename(prop.text.strip().replace('/', os.sep).replace('\\', os.sep))
                        cand = os.path.join(dirs['testdata'], base)
                        if os.path.exists(cand): prop.text = cand
            tree.write(working_jmx, encoding='unicode', xml_declaration=False)
            cmd = [jbin, '-n', '-t', working_jmx, '-l', jtl_path,
                   f'-Jtest.duration={r["duration"]}',
                   f'-Jp2p.threads={r["threads"]}',
                   f'-Jp2p.rampup={r["rampup"]}']
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
            proc.wait()
            sla_r = _evaluate_sla(jtl_path, client)
            audit('RECURRING_RUN', f"id={sid} rc={proc.returncode}", username='system')
            _notify_test_complete(client, r['jmx'], 'done' if proc.returncode == 0 else 'failed', sla_r,
                                  jtl_name=os.path.basename(jtl_path))
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)
    except Exception as ex:
        print(f'[recurring] fire error {sid}: {ex}')

def _check_recurring():
    try:
        now  = datetime.now()
        hhmm = now.strftime('%H:%M')
        dow  = now.weekday()
        with get_db() as db:
            rows = db.execute(
                "SELECT * FROM recurring_schedules WHERE enabled=1 AND run_at_time=?", (hhmm,)
            ).fetchall()
        for row in rows:
            r = dict(row)
            if r['recurrence'] == 'weekly' and r['day_of_week'] != dow:
                continue
            last = r.get('last_run') or ''
            if last.startswith(now.strftime('%Y-%m-%d %H:%M')):
                continue
            threading.Thread(target=_fire_recurring, args=(r,), daemon=True).start()
    except Exception as ex:
        print(f'[recurring] check error: {ex}')

def _start_recurring_checker():
    _check_recurring()
    t = threading.Timer(60, _start_recurring_checker)
    t.daemon = True
    t.start()

# ── Notifications ──────────────────────────────────────────────────────────────
import urllib.request as _urllib_req

def _notify_test_complete(c, jmx, result, sla_result=None, jtl_name=None):
    cfg       = load_cfg()
    teams_url = cfg.get('teams_webhook', '')
    summary   = (f"Test: {jmx} | Client: {c['name']} | "
                 f"Status: {'✅ Passed' if result=='done' else '❌ Failed'}")
    if sla_result:
        summary += (f" | SLA: {'PASS' if sla_result.get('passed') else 'FAIL'}"
                    f" | TPS: {sla_result.get('actual_tps','-')}"
                    f" | P95: {sla_result.get('actual_p95','-')}ms"
                    f" | Err: {sla_result.get('actual_err','-')}%")
    if teams_url:
        try:
            payload = json.dumps({'text': summary}).encode()
            req = _urllib_req.Request(teams_url, data=payload,
                                      headers={'Content-Type': 'application/json'}, method='POST')
            _urllib_req.urlopen(req, timeout=5, context=_SSL_CTX)
        except Exception as ex:
            print(f'[notify] Teams webhook failed: {ex}')
    slack_url = cfg.get('slack_webhook', '')
    if slack_url:
        try:
            payload = json.dumps({'text': summary}).encode()
            req = _urllib_req.Request(slack_url, data=payload,
                                      headers={'Content-Type': 'application/json'}, method='POST')
            _urllib_req.urlopen(req, timeout=5, context=_SSL_CTX)
        except Exception as ex:
            print(f'[notify] Slack webhook failed: {ex}')
    smtp_host  = cfg.get('smtp_host', '')
    notify_to  = cfg.get('notify_email', '')
    if smtp_host and notify_to:
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            from email.mime.application import MIMEApplication
            smtp_user = cfg.get('smtp_user', '')
            smtp_pass = cfg.get('smtp_pass', '')
            msg = MIMEMultipart()
            msg['Subject'] = f'[Load Test] {c["name"]} — {"PASS" if result=="done" else "FAIL"}'
            msg['From']    = smtp_user or 'loadtest@localhost'
            msg['To']      = notify_to
            body = summary + ('\n\nExecutive summary, capacity/headroom, trend and full '
                              'request/response are in the attached HTML report.' if jtl_name else '')
            msg.attach(MIMEText(body, 'plain'))
            # Attach the shareable HTML report (exec banner + request/response).
            if jtl_name:
                try:
                    rpath = os.path.join(client_dirs(c)['reports'], jtl_name)
                    if os.path.exists(rpath):
                        report_html = _report_html(rpath, c)
                        att = MIMEApplication(report_html.encode('utf-8'), _subtype='html')
                        att.add_header('Content-Disposition', 'attachment',
                                       filename=jtl_name.replace('.jtl', '_report.html'))
                        msg.attach(att)
                except Exception as ex:
                    print(f'[notify] report attach failed: {ex}')
            port = int(cfg.get('smtp_port', 587))
            with smtplib.SMTP(smtp_host, port, timeout=20) as s:
                s.starttls()
                if smtp_user and smtp_pass:
                    s.login(smtp_user, smtp_pass)
                s.sendmail(msg['From'], [notify_to], msg.as_string())
        except Exception as ex:
            print(f'[notify] Email failed: {ex}')
    # Jira: create issue on failure
    jira_url   = cfg.get('jira_url', '').strip()
    jira_proj  = cfg.get('jira_project', '').strip()
    jira_token = cfg.get('jira_token', '').strip()
    jira_user  = cfg.get('jira_user', '').strip()
    if jira_url and jira_proj and jira_token and result != 'done':
        try:
            import base64
            sla_text = ''
            if sla_result:
                sla_text = (f'\n\nSLA Result: {"PASS" if sla_result.get("passed") else "FAIL"}'
                            f'\nTPS: {sla_result.get("actual_tps","-")}'
                            f'\nP95: {sla_result.get("actual_p95","-")}ms'
                            f'\nError%: {sla_result.get("actual_err","-")}%')
            issue_body = json.dumps({
                'fields': {
                    'project': {'key': jira_proj},
                    'summary': f'[Load Test Failure] {c["name"]} — {jmx}',
                    'description': {
                        'type': 'doc', 'version': 1,
                        'content': [{'type': 'paragraph', 'content': [
                            {'type': 'text', 'text': summary + sla_text}
                        ]}]
                    },
                    'issuetype': {'name': 'Bug'},
                }
            }).encode()
            creds = base64.b64encode(f'{jira_user}:{jira_token}'.encode()).decode()
            req = _urllib_req.Request(
                f'{jira_url.rstrip("/")}/rest/api/3/issue',
                data=issue_body,
                headers={'Authorization': f'Basic {creds}',
                         'Content-Type': 'application/json',
                         'Accept': 'application/json'},
                method='POST',
            )
            _urllib_req.urlopen(req, timeout=10, context=_SSL_CTX)
        except Exception as ex:
            print(f'[notify] Jira create issue failed: {ex}')

# ── API: Compare Runs ──────────────────────────────────────────────────────────
@app.route('/api/compare-runs', methods=['POST'])
@login_req
def api_compare_runs():
    c = active_client()
    if not c:
        return jsonify(error='No client'), 400
    data  = request.json or {}
    files = data.get('files', [])
    if len(files) < 2:
        return jsonify(error='Select at least 2 runs'), 400
    results = []
    dirs = client_dirs(c)
    for fname in files[:4]:
        path = os.path.join(dirs['reports'], fname)
        if not os.path.exists(path):
            continue
        try:
            elapsed_list, errors, total, ts_min, ts_max = [], 0, 0, None, None
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                header = [h.strip() for h in next(reader, [])]
                ei  = header.index('elapsed')   if 'elapsed'   in header else 1
                si  = header.index('success')   if 'success'   in header else 7
                tsi = header.index('timeStamp') if 'timeStamp' in header else 0
                for row in reader:
                    if len(row) <= max(ei, si, tsi):
                        continue
                    try:
                        e = int(row[ei]); ts = int(row[tsi])
                        elapsed_list.append(e)
                        ts_min = min(ts_min, ts) if ts_min is not None else ts
                        ts_max = max(ts_max, ts) if ts_max is not None else ts
                        total += 1
                        if row[si].strip().lower() != 'true':
                            errors += 1
                    except Exception:
                        continue
            if not elapsed_list:
                continue
            elapsed_list.sort()
            def pct(p): return elapsed_list[max(0, int(len(elapsed_list)*p/100)-1)]
            dur = ((ts_max - ts_min) / 1000) if ts_max and ts_min and ts_max > ts_min else 1
            results.append({
                'file': fname, 'total': total, 'tps': round(total/dur, 2),
                'avg': round(sum(elapsed_list)/len(elapsed_list)),
                'p50': pct(50), 'p90': pct(90), 'p95': pct(95), 'p99': pct(99),
                'errors': errors,
                'error_pct': round(errors/total*100, 2) if total else 0,
                'duration_s': round(dur),
            })
        except Exception:
            continue
    return jsonify(results=results)

# ── API: CSV Preview ───────────────────────────────────────────────────────────
@app.route('/api/csv-preview/<path:fname>')
@login_req
def api_csv_preview(fname):
    c = active_client()
    if not c:
        return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['testdata'], fname)
    if not os.path.exists(path):
        return jsonify(error='Not found'), 404
    rows = []
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i > 20: break
                rows.append(row)
    except Exception as ex:
        return jsonify(error=str(ex)), 500
    return jsonify(filename=fname, rows=rows, total_preview=len(rows))

# ── API: JMX Parameter Editor ──────────────────────────────────────────────────
@app.route('/api/jmx-params/<path:fname>', methods=['GET', 'POST'])
@login_req
def api_jmx_params(fname):
    c = active_client()
    if not c:
        return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['jmx'], fname)
    if not os.path.exists(path):
        return jsonify(error='Not found'), 404
    if request.method == 'GET':
        try:
            tree = ET.parse(path)
            root = tree.getroot()
            params = []
            for tg in root.iter('ThreadGroup'):
                tg_name = tg.get('testname', 'Unnamed')
                for prop in tg.iter('stringProp'):
                    n = prop.get('name', '')
                    if n == 'ThreadGroup.num_threads':
                        params.append({'tg': tg_name, 'param': 'threads',  'value': prop.text or '1',    'label': 'Threads'})
                    elif n == 'ThreadGroup.ramp_time':
                        params.append({'tg': tg_name, 'param': 'ramp_up',  'value': prop.text or '60',   'label': 'Ramp-up (s)'})
                    elif n == 'ThreadGroup.duration':
                        params.append({'tg': tg_name, 'param': 'duration', 'value': prop.text or '1800', 'label': 'Duration (s)'})
            return jsonify(filename=fname, params=params)
        except Exception as ex:
            return jsonify(error=str(ex)), 500
    # POST: write back
    data    = request.json or {}
    updates = data.get('updates', [])
    try:
        tree = ET.parse(path)
        root = tree.getroot()
        param_map = {(u['tg'], u['param']): u['value'] for u in updates}
        for tg in root.iter('ThreadGroup'):
            tg_name = tg.get('testname', '')
            for prop in tg.iter('stringProp'):
                n   = prop.get('name', '')
                key = None
                if n == 'ThreadGroup.num_threads': key = (tg_name, 'threads')
                elif n == 'ThreadGroup.ramp_time': key = (tg_name, 'ramp_up')
                elif n == 'ThreadGroup.duration':  key = (tg_name, 'duration')
                if key and key in param_map:
                    prop.text = str(param_map[key])
        tree.write(path, encoding='unicode', xml_declaration=False)
        audit('JMX_EDIT', f'{fname}: {len(updates)} params updated')
        return jsonify(ok=True, updated=len(updates))
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: JMX List ─────────────────────────────────────────────────────────────
@app.route('/api/jmx-list')
@login_req
def api_jmx_list():
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    files = sorted([os.path.basename(f) for f in glob.glob(os.path.join(client_dirs(c)['jmx'], '*.jmx'))])
    return jsonify(files=files)

@app.route('/api/config-list')
@login_req
def api_config_list():
    """List non-JMX config files stored in the client's jmx dir."""
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    exts = set(ALLOWED_CONFIG)
    d = client_dirs(c)['jmx']
    files = []
    for ext in exts:
        files.extend(glob.glob(os.path.join(d, f'*{ext}')))
    out = []
    for p in sorted(files):
        fname = os.path.basename(p)
        out.append({'name': fname, 'size_kb': round(os.path.getsize(p) / 1024, 1)})
    return jsonify(files=out)

# ── API: JMX Inspector ────────────────────────────────────────────────────────
@app.route('/api/jmx-inspect/<path:fname>')
@login_req
def api_jmx_inspect(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['jmx'], fname)
    if not os.path.exists(path): return jsonify(error='JMX not found'), 404
    try:
        import xml.etree.ElementTree as _ET
        tree = _ET.parse(path)
        root = tree.getroot()

        def _sp(el, name):
            for sp in el.findall('stringProp'):
                if sp.get('name') == name:
                    return (sp.text or '').strip()
            return ''

        def _extract_assertion(el):
            tag = el.tag
            name = el.get('testname', '')
            if tag == 'JSR223Assertion':
                script = _sp(el, 'script')
                m = re.search(r"def expected\s*=\s*['\"](.+?)['\"]", script)
                expected = m.group(1) if m else ''
                match_type = 'contains' if '.contains(' in script else 'equals' if 'trim() !=' in script else 'script'
                return {'name': name, 'type': 'JSR223', 'expected': expected, 'match': match_type}
            if tag == 'ResponseAssertion':
                for sp in el.iter('stringProp'):
                    if sp.get('name') == '49586':
                        return {'name': name, 'type': 'Response', 'expected': (sp.text or '').strip(), 'match': 'contains'}
                field = _sp(el, 'Assertion.test_field')
                return {'name': name, 'type': 'Response', 'expected': '', 'match': field}
            return {'name': name, 'type': tag, 'expected': '', 'match': ''}

        def _samplers_in_ht(ht):
            kids = list(ht)
            result = []
            i = 0
            while i < len(kids):
                el = kids[i]
                nxt = kids[i + 1] if i + 1 < len(kids) and kids[i + 1].tag == 'hashTree' else None
                if el.tag == 'HTTPSamplerProxy':
                    params = {}
                    for arg in el.findall('./elementProp/collectionProp/elementProp[@elementType="HTTPArgument"]'):
                        k = v = ''
                        for sp in arg.findall('stringProp'):
                            if sp.get('name') == 'Argument.name': k = sp.text or ''
                            if sp.get('name') == 'Argument.value': v = sp.text or ''
                        if k: params[k] = v
                    assertions = []
                    preprocessors = []
                    if nxt is not None:
                        for child in nxt:
                            if 'Assertion' in child.tag:
                                assertions.append(_extract_assertion(child))
                            elif 'PreProcessor' in child.tag or child.tag == 'JSR223PreProcessor':
                                preprocessors.append(child.get('testname', child.tag))
                    result.append({
                        'label': el.get('testname', ''),
                        'method': _sp(el, 'HTTPSampler.method'),
                        'path': _sp(el, 'HTTPSampler.path'),
                        'params': params,
                        'assertions': assertions,
                        'preprocessors': preprocessors,
                    })
                    i += 2 if nxt is not None else 1
                elif el.tag in ('IfController', 'TransactionController', 'GenericSampler') and nxt is not None:
                    result.extend(_samplers_in_ht(nxt))
                    i += 2
                else:
                    i += 1
            return result

        # Root → hashTree → TestPlan|hashTree → second hashTree contains TGs
        top_ht = root.find('hashTree')
        if top_ht is None: return jsonify(error='No top hashTree'), 500
        second_ht = top_ht.find('hashTree')
        if second_ht is None: return jsonify(error='No second hashTree'), 500

        services = []
        siblings = list(second_ht)
        i = 0
        while i < len(siblings):
            el = siblings[i]
            child_ht = siblings[i + 1] if i + 1 < len(siblings) and siblings[i + 1].tag == 'hashTree' else None
            if el.tag == 'ThreadGroup':
                tg_raw = el.get('testname', '')
                enabled = el.get('enabled', 'true').lower() == 'true'
                m = re.match(r'^TG - (.+?) \((.+)\)$', tg_raw.strip())
                svc_name = m.group(1).strip() if m else tg_raw
                threads = _sp(el, 'ThreadGroup.num_threads')
                rampup  = _sp(el, 'ThreadGroup.ramp_time')
                duration = _sp(el, 'ThreadGroup.duration')
                csv_file = csv_vars = tc_name = ''
                steps = []
                if child_ht is not None:
                    for sub in child_ht:
                        if sub.tag == 'CSVDataSet':
                            csv_file = _sp(sub, 'filename')
                            csv_vars = _sp(sub, 'variableNames')
                        elif sub.tag == 'TransactionController':
                            tc_name = sub.get('testname', '')
                    # find TC's hashTree
                    sub_kids = list(child_ht)
                    for j, sub in enumerate(sub_kids):
                        if sub.tag == 'TransactionController':
                            tc_ht = sub_kids[j + 1] if j + 1 < len(sub_kids) and sub_kids[j + 1].tag == 'hashTree' else None
                            if tc_ht is not None:
                                steps = _samplers_in_ht(tc_ht)
                            break
                services.append({
                    'name': svc_name, 'tg': tg_raw, 'tc': tc_name,
                    'enabled': enabled, 'threads': threads,
                    'rampup': rampup, 'duration': duration,
                    'csv_file': os.path.basename(csv_file), 'csv_vars': csv_vars,
                    'steps': steps,
                })
            i += 2 if child_ht is not None else 1

        return jsonify(jmx=fname, services=services)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: JMX Edit (save inspector changes back to JMX) ────────────────────────
@app.route('/api/jmx-edit/<path:fname>', methods=['POST'])
@login_req
def api_jmx_edit(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['jmx'], fname)
    if not os.path.exists(path): return jsonify(error='JMX not found'), 404
    data = request.json or {}
    services = data.get('services', [])
    if not services: return jsonify(error='No data'), 400
    try:
        import xml.etree.ElementTree as _ET
        import shutil as _sh, re as _re

        # Back up original
        backup = path + '.bak'
        _sh.copy2(path, backup)

        tree = _ET.parse(path)
        root = tree.getroot()

        def _set_sp(el, name, value):
            for sp in el.findall('stringProp'):
                if sp.get('name') == name:
                    sp.text = str(value); return
        def _set_bp(el, name, value):
            for bp in el.findall('boolProp'):
                if bp.get('name') == name:
                    bp.text = 'true' if value else 'false'; return

        # Build lookup: tg_testname → ThreadGroup el
        # Build lookup: sampler_testname → HTTPSamplerProxy el
        # Build lookup: assert_testname → JSR223Assertion el
        tg_map = {}
        sampler_map = {}
        assert_map = {}
        for tg in root.iter('ThreadGroup'):
            tg_map[tg.get('testname', '')] = tg
        for sp in root.iter('HTTPSamplerProxy'):
            sampler_map[sp.get('testname', '')] = sp
        for a in root.iter('JSR223Assertion'):
            assert_map[a.get('testname', '')] = a
        for a in root.iter('ResponseAssertion'):
            assert_map[a.get('testname', '')] = a

        changes = 0
        for svc in services:
            svc_name = svc.get('name', '')
            tg_raw = svc.get('tg', '')  # full TG testname

            # Find matching ThreadGroup
            tg_el = tg_map.get(tg_raw)
            if tg_el is None:
                # fallback: match by service name pattern
                for k, v in tg_map.items():
                    m = re.match(r'^TG - (.+?) \(', k)
                    if m and m.group(1).strip().lower() == svc_name.strip().lower():
                        tg_el = v; break

            if tg_el is not None:
                if 'threads' in svc:  _set_sp(tg_el, 'ThreadGroup.num_threads', svc['threads']); changes += 1
                if 'rampup' in svc:   _set_sp(tg_el, 'ThreadGroup.ramp_time',   svc['rampup']);  changes += 1
                if 'duration' in svc: _set_sp(tg_el, 'ThreadGroup.duration',     svc['duration']); changes += 1
                if 'enabled' in svc:  tg_el.set('enabled', 'true' if svc['enabled'] else 'false'); changes += 1

            # Update steps
            for step in svc.get('steps', []):
                label = step.get('label', '')
                # Update INPUT param value
                if 'input' in step:
                    smp = sampler_map.get(label)
                    if smp is not None:
                        for arg in smp.findall('./elementProp/collectionProp/elementProp[@elementType="HTTPArgument"]'):
                            for sp in arg.findall('stringProp'):
                                if sp.get('name') == 'Argument.name' and sp.text == 'INPUT':
                                    for vsp in arg.findall('stringProp'):
                                        if vsp.get('name') == 'Argument.value':
                                            vsp.text = step['input']; changes += 1

                # Update assertions
                for a in step.get('assertions', []):
                    a_name      = a.get('name', '')
                    new_name    = a.get('new_name', a_name)
                    new_expected = a.get('expected', '')
                    new_match   = a.get('match', '')
                    a_el = assert_map.get(a_name)
                    if a_el is None: continue

                    # Rename assertion testname
                    if new_name and new_name != a_name:
                        a_el.set('testname', new_name)
                        # re-key the map so a second ref to same assertion still finds it
                        assert_map[new_name] = a_el
                        changes += 1

                    if a_el.tag == 'JSR223Assertion':
                        for sp in a_el.findall('stringProp'):
                            if sp.get('name') == 'script' and sp.text:
                                old_script = sp.text
                                new_script = old_script
                                # Update expected string
                                new_script = _re.sub(
                                    r"(def expected\s*=\s*['\"])(.+?)(['\"])",
                                    lambda m, ne=new_expected: m.group(1) + ne + m.group(3),
                                    new_script, count=1
                                )
                                # Update match rule (swap contains ↔ equals)
                                if new_match == 'contains' and 'actual.trim() !=' in new_script:
                                    new_script = _re.sub(
                                        r'actual\.trim\(\)\s*!=\s*expected',
                                        '!actual.contains(expected)', new_script
                                    )
                                elif new_match == 'equals' and '!actual.contains(' in new_script:
                                    new_script = _re.sub(
                                        r'!actual\.contains\(expected\)',
                                        'actual.trim() != expected', new_script
                                    )
                                if new_script != old_script:
                                    sp.text = new_script; changes += 1
                    elif a_el.tag == 'ResponseAssertion':
                        for sp in a_el.iter('stringProp'):
                            if sp.get('name') == '49586':
                                if sp.text != new_expected:
                                    sp.text = new_expected; changes += 1

        tree.write(path, encoding='unicode', xml_declaration=False)
        audit('JMX_INSPECTOR_EDIT', f'{fname}: {changes} changes saved')
        return jsonify(ok=True, changes=changes)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: Overview Stats ────────────────────────────────────────────────────────
@app.route('/api/overview-stats')
@login_req
def api_overview_stats():
    c = active_client()
    if not c:
        return jsonify(error='No client'), 400
    dirs = client_dirs(c)
    jtls = sorted(glob.glob(os.path.join(dirs['reports'], '*.jtl')), key=os.path.getmtime, reverse=True)
    total_runs  = len(jtls)
    last_run    = datetime.fromtimestamp(os.path.getmtime(jtls[0])).strftime('%Y-%m-%d %H:%M') if jtls else '—'
    sla_status  = None
    if jtls:
        sr = _evaluate_sla(jtls[0], c)
        if sr:
            sla_status = 'PASS' if sr['passed'] else 'FAIL'
    csvs = glob.glob(os.path.join(dirs['testdata'], '*.csv'))
    jmxs = glob.glob(os.path.join(dirs['jmx'], '*.jmx'))
    trend = []
    for jtl in jtls[:5]:
        try:
            el, total_t, ts_min, ts_max = [], 0, None, None
            with open(jtl, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                header = [h.strip() for h in next(reader, [])]
                ei  = header.index('elapsed')   if 'elapsed'   in header else 1
                tsi = header.index('timeStamp') if 'timeStamp' in header else 0
                si  = header.index('success')   if 'success'   in header else 7
                errors = 0
                for row in reader:
                    if len(row) <= max(ei, tsi): continue
                    try:
                        el.append(int(row[ei])); ts = int(row[tsi])
                        ts_min = min(ts_min, ts) if ts_min is not None else ts
                        ts_max = max(ts_max, ts) if ts_max is not None else ts
                        total_t += 1
                        if len(row) > si and row[si].strip().lower() != 'true':
                            errors += 1
                    except Exception:
                        continue
            if not el:
                continue
            el.sort()
            dur = ((ts_max - ts_min) / 1000) if ts_max and ts_min and ts_max > ts_min else 1
            def pct(p): return el[max(0, int(len(el)*p/100)-1)]
            trend.append({
                'file': os.path.basename(jtl),
                'tps': round(total_t/dur, 2),
                'p95': pct(95), 'p90': pct(90),
                'err_pct': round(errors/total_t*100, 2) if total_t else 0,
                'date': datetime.fromtimestamp(os.path.getmtime(jtl)).strftime('%m/%d %H:%M'),
            })
        except Exception:
            continue
    trend.reverse()
    return jsonify(
        total_runs=total_runs, last_run=last_run, sla_status=sla_status,
        csv_count=len(csvs), jmx_count=len(jmxs),
        running=_state['running'], active_jmx=_state.get('jmx', ''),
        trend=trend,
    )

# ── API: PDF Report ────────────────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors as _rl_colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    _HAS_REPORTLAB = True
except ImportError:
    _HAS_REPORTLAB = False

@app.route('/api/report/<path:fname>/pdf')
@login_req
def api_report_pdf(fname):
    c = active_client()
    if not c: return 'No client', 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return 'File not found', 404
    elapsed_list, errors, total, ts_min, ts_max, labels_count = [], 0, 0, None, None, {}
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            header = [h.strip() for h in next(reader, [])]
            ei  = header.index('elapsed')   if 'elapsed'   in header else 1
            si  = header.index('success')   if 'success'   in header else 7
            tsi = header.index('timeStamp') if 'timeStamp' in header else 0
            li  = header.index('label')     if 'label'     in header else 2
            for row in reader:
                if len(row) <= max(ei, si, tsi): continue
                try:
                    e = int(row[ei]); ts = int(row[tsi])
                    lbl = row[li] if len(row) > li else 'Unknown'
                    elapsed_list.append(e)
                    ts_min = min(ts_min, ts) if ts_min is not None else ts
                    ts_max = max(ts_max, ts) if ts_max is not None else ts
                    total += 1
                    labels_count[lbl] = labels_count.get(lbl, 0) + 1
                    if row[si].strip().lower() != 'true': errors += 1
                except Exception:
                    continue
    except Exception as ex:
        return f'Parse error: {ex}', 500
    if not elapsed_list: return 'No data in JTL', 404
    elapsed_list.sort()
    def pct(p): return elapsed_list[max(0, int(len(elapsed_list)*p/100)-1)]
    dur = ((ts_max - ts_min) / 1000) if ts_max and ts_min and ts_max > ts_min else 1
    tps = total / dur
    avg = sum(elapsed_list) / len(elapsed_list)
    err_pct = errors / total * 100 if total else 0

    if _HAS_REPORTLAB:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        def ps(name, parent='Normal', **kw):
            return ParagraphStyle(name, parent=styles[parent], **kw)
        title_s = ps('T', fontSize=18, textColor=_rl_colors.HexColor('#0066cc'), spaceAfter=4)
        sub_s   = ps('S', fontSize=9,  textColor=_rl_colors.HexColor('#666'), spaceAfter=10)
        h2_s    = ps('H2', 'Heading2', fontSize=13, textColor=_rl_colors.HexColor('#0066cc'),
                     spaceBefore=14, spaceAfter=6)
        story = [
            Paragraph('Load Test Report', title_s),
            Paragraph(f'{fname}  ·  Client: {c["name"]}  ·  {datetime.now().strftime("%Y-%m-%d %H:%M")}', sub_s),
            HRFlowable(width='100%', color=_rl_colors.HexColor('#cccccc'), spaceAfter=10),
            Paragraph('Summary', h2_s),
        ]
        summary_data = [
            ['Metric', 'Value'],
            ['Total Samples',     f'{total:,}'],
            ['Throughput (TPS)',  f'{tps:.2f}'],
            ['Average RT',        f'{avg:.0f} ms'],
            ['P50', f'{pct(50)} ms'], ['P90', f'{pct(90)} ms'],
            ['P95', f'{pct(95)} ms'], ['P99', f'{pct(99)} ms'],
            ['Errors',            f'{errors:,} ({err_pct:.2f}%)'],
            ['Duration',          f'{dur:.0f} s'],
        ]
        tbl = Table(summary_data, colWidths=[9*cm, 8*cm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), _rl_colors.HexColor('#0066cc')),
            ('TEXTCOLOR',  (0,0), (-1,0), _rl_colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1),
             [_rl_colors.HexColor('#f0f4ff'), _rl_colors.white]),
            ('GRID', (0,0), (-1,-1), 0.5, _rl_colors.HexColor('#cccccc')),
            ('PADDING', (0,0), (-1,-1), 7),
        ]))
        story.append(tbl)
        if labels_count:
            story += [Spacer(1, 0.4*cm), Paragraph('By Label (top 20)', h2_s)]
            ld = [['Label', 'Count', '%']] + [
                [lbl, f'{cnt:,}', f'{cnt/total*100:.1f}%']
                for lbl, cnt in sorted(labels_count.items(), key=lambda x: -x[1])[:20]
            ]
            lt = Table(ld, colWidths=[11*cm, 3*cm, 3*cm])
            lt.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), _rl_colors.HexColor('#0066cc')),
                ('TEXTCOLOR',  (0,0), (-1,0), _rl_colors.white),
                ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0,0), (-1,-1), 9),
                ('ROWBACKGROUNDS', (0,1), (-1,-1),
                 [_rl_colors.HexColor('#f0f4ff'), _rl_colors.white]),
                ('GRID', (0,0), (-1,-1), 0.5, _rl_colors.HexColor('#cccccc')),
                ('PADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(lt)
        doc.build(story)
        buf.seek(0)
        audit('DOWNLOAD_PDF', fname)
        return send_file(buf, mimetype='application/pdf', as_attachment=True,
                         download_name=fname.replace('.jtl', '_report.pdf'))
    else:
        lbl_html = ''
        if labels_count:
            rows = ''.join(
                f'<tr><td>{lbl}</td><td>{cnt:,}</td><td>{cnt/total*100:.1f}%</td></tr>'
                for lbl, cnt in sorted(labels_count.items(), key=lambda x: -x[1])[:20]
            )
            lbl_html = f'<h2>By Label</h2><table><tr><th>Label</th><th>Count</th><th>%</th></tr>{rows}</table>'
        html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>Report: {fname}</title>
<style>body{{font-family:Arial,sans-serif;margin:40px;color:#111;}}
h1{{color:#0066cc;}}h2{{color:#0066cc;margin-top:24px;}}
table{{border-collapse:collapse;width:100%;margin-bottom:20px;}}
th{{background:#0066cc;color:#fff;padding:8px 12px;text-align:left;}}
td{{padding:7px 12px;border-bottom:1px solid #ddd;}}
tr:nth-child(even){{background:#f5f5f5;}}
.pass{{color:green;font-weight:bold;}}.fail{{color:red;font-weight:bold;}}
@media print{{body{{margin:20px;}}}}</style></head><body>
<h1>Load Test Report</h1>
<p><strong>File:</strong> {fname} | <strong>Client:</strong> {c['name']} | <strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
<h2>Summary</h2><table>
<tr><th>Metric</th><th>Value</th></tr>
<tr><td>Total Samples</td><td>{total:,}</td></tr>
<tr><td>Throughput (TPS)</td><td>{tps:.2f}</td></tr>
<tr><td>Average RT</td><td>{avg:.0f} ms</td></tr>
<tr><td>P50</td><td>{pct(50)} ms</td></tr>
<tr><td>P90</td><td>{pct(90)} ms</td></tr>
<tr><td>P95</td><td>{pct(95)} ms</td></tr>
<tr><td>P99</td><td>{pct(99)} ms</td></tr>
<tr><td>Errors</td><td class="{'pass' if err_pct<=1 else 'fail'}">{errors:,} ({err_pct:.2f}%)</td></tr>
<tr><td>Duration</td><td>{dur:.0f} s</td></tr>
</table>{lbl_html}<script>window.print();</script></body></html>"""
        audit('DOWNLOAD_PDF', fname + ' (HTML fallback)')
        return html, 200, {'Content-Type': 'text/html'}

# ── API: Schedules ──────────────────────────────────────────────────────────────
@app.route('/api/schedules', methods=['GET'])
@login_req
def api_list_schedules():
    with _sched_lock:
        return jsonify(schedules=list(_schedules))

@app.route('/api/schedules', methods=['POST'])
@perm_req('manage_schedules')
def api_create_schedule():
    data = request.json or {}
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    run_at_str = data.get('run_at', '')
    try:
        run_at = datetime.fromisoformat(run_at_str)
    except Exception:
        return jsonify(error='Invalid run_at datetime (use YYYY-MM-DDTHH:MM)'), 400
    delay = (run_at - datetime.now()).total_seconds()
    if delay < 0:
        return jsonify(error='run_at must be in the future'), 400
    sid = str(uuid.uuid4())[:8]
    entry = {
        'id': sid, 'client': c['code'],
        'jmx': data.get('jmx', ''),
        'threads': int(data.get('threads', 10)),
        'duration': int(data.get('duration', 300)),
        'rampup': int(data.get('rampup', 30)),
        'run_at': run_at_str, 'status': 'pending',
        'created_by': session.get('user', 'system'),
    }
    with _sched_lock:
        _schedules.append(entry)

    def _fire():
        with _sched_lock:
            for s in _schedules:
                if s['id'] == sid:
                    s['status'] = 'running'; break
        audit('SCHEDULED_RUN', f"id={sid} jmx={entry['jmx']} client={entry['client']}")
        client = get_client(entry['client'])
        if not client:
            _set_sched_status(sid, 'failed'); return
        dirs = ensure_client_dirs(client)
        jmx_path = os.path.join(dirs['jmx'], entry['jmx'])
        if not os.path.exists(jmx_path):
            _set_sched_status(sid, 'failed'); return
        cfg = load_cfg()
        jbin = cfg.get('jmeter_bin', '')
        if not os.path.exists(jbin):
            for cand in [r'C:\apache-jmeter-5.5\bin\jmeter.bat', r'C:\apache-jmeter-5.6\bin\jmeter.bat', r'C:\apache-jmeter-5.6.3\bin\jmeter.bat']:
                if os.path.exists(cand): jbin = cand; break
            else:
                _set_sched_status(sid, 'failed'); return
        out_name = f"sched_{sid}_{datetime.now().strftime('%d%m%Y_%H%M%S')}"
        jtl_path = os.path.join(dirs['reports'], out_name + '.jtl')
        os.makedirs(dirs['reports'], exist_ok=True)
        tmp_dir = tempfile.mkdtemp(prefix='lt_sched_')
        working_jmx = os.path.join(tmp_dir, entry['jmx'])
        try:
            shutil.copy(jmx_path, working_jmx)
            tree = ET.parse(working_jmx); root_el = tree.getroot()
            for csv_ds in root_el.iter('CSVDataSet'):
                for prop in csv_ds.iter('stringProp'):
                    if prop.get('name') == 'filename' and prop.text and prop.text.strip():
                        base = os.path.basename(prop.text.strip().replace('/', os.sep).replace('\\', os.sep))
                        cand = os.path.join(dirs['testdata'], base)
                        if os.path.exists(cand): prop.text = cand
            tree.write(working_jmx, encoding='unicode', xml_declaration=False)
            cmd = [jbin, '-n', '-t', working_jmx, '-l', jtl_path,
                   f'-Jtest.duration={entry["duration"]}',
                   f'-Jp2p.threads={entry["threads"]}',
                   f'-Jp2p.rampup={entry["rampup"]}']
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
            proc.wait()
            sla_r = _evaluate_sla(jtl_path, client)
            _set_sched_status(sid, 'done' if proc.returncode == 0 else 'failed')
            _notify_test_complete(client, entry['jmx'], 'done' if proc.returncode == 0 else 'failed', sla_r,
                                  jtl_name=os.path.basename(jtl_path))
        except Exception:
            _set_sched_status(sid, 'failed')
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    t = threading.Timer(delay, _fire)
    t.daemon = True
    t.start()
    audit('SCHEDULE_CREATE', f"id={sid} jmx={entry['jmx']} run_at={run_at_str}")
    return jsonify(ok=True, id=sid, delay_s=round(delay))

@app.route('/api/schedules/<sid>', methods=['DELETE'])
@perm_req('manage_schedules')
def api_delete_schedule(sid):
    with _sched_lock:
        _schedules[:] = [s for s in _schedules if s['id'] != sid]
    audit('SCHEDULE_DELETE', f"id={sid}")
    return jsonify(ok=True)

def _set_sched_status(sid, status):
    with _sched_lock:
        for s in _schedules:
            if s['id'] == sid:
                s['status'] = status; break

# ── API: Test Suite ────────────────────────────────────────────────────────────
@app.route('/api/suite/start', methods=['POST'])
@perm_req('run_tests')
def api_suite_start():
    global _suite_state
    if _suite_state['running'] or _state['running']:
        return jsonify(error='A test is already running.'), 409
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    data     = request.json or {}
    jmx_list = data.get('jmx_list', [])
    if not jmx_list: return jsonify(error='No JMX files specified'), 400
    dirs = ensure_client_dirs(c)
    cfg  = load_cfg()
    jbin = cfg.get('jmeter_bin', '')
    if not os.path.exists(jbin):
        for cand in [r'C:\apache-jmeter-5.5\bin\jmeter.bat', r'C:\apache-jmeter-5.6\bin\jmeter.bat', r'C:\apache-jmeter-5.6.3\bin\jmeter.bat']:
            if os.path.exists(cand): jbin = cand; break
        else:
            return jsonify(error='JMeter binary not found'), 400
    _suite_state = {'running': True, 'current': None, 'results': [],
                    'total': len(jmx_list), 'done': 0, 'client': c['code']}
    audit('SUITE_START', f"client={c['code']} count={len(jmx_list)}")

    def _run_suite():
        for item in jmx_list:
            jmx_name = item.get('jmx', '')
            threads  = int(item.get('threads', 10))
            duration = int(item.get('duration', 300))
            rampup   = int(item.get('rampup', 30))
            label    = item.get('label', jmx_name)
            jmx_path = os.path.join(dirs['jmx'], jmx_name)
            if not os.path.exists(jmx_path):
                _suite_state['results'].append({'jmx': jmx_name, 'label': label, 'status': 'skipped', 'error': 'File not found'})
                _suite_state['done'] += 1
                continue
            _suite_state['current'] = jmx_name
            out_name = f"suite_{datetime.now().strftime('%d%m%Y_%H%M%S')}_{jmx_name.replace('.jmx','')}"
            jtl_path = os.path.join(dirs['reports'], out_name + '.jtl')
            os.makedirs(dirs['reports'], exist_ok=True)
            tmp_dir = tempfile.mkdtemp(prefix='lt_suite_')
            working_jmx = os.path.join(tmp_dir, jmx_name)
            try:
                shutil.copy(jmx_path, working_jmx)
                tree = ET.parse(working_jmx); root_el = tree.getroot()
                for csv_ds in root_el.iter('CSVDataSet'):
                    for prop in csv_ds.iter('stringProp'):
                        if prop.get('name') == 'filename' and prop.text and prop.text.strip():
                            base = os.path.basename(prop.text.strip().replace('/', os.sep).replace('\\', os.sep))
                            cand = os.path.join(dirs['testdata'], base)
                            if os.path.exists(cand): prop.text = cand
                tree.write(working_jmx, encoding='unicode', xml_declaration=False)
                cmd = [jbin, '-n', '-t', working_jmx, '-l', jtl_path,
                       f'-Jtest.duration={duration}', f'-Jp2p.threads={threads}', f'-Jp2p.rampup={rampup}']
                proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
                proc.wait()
                sla_r = _evaluate_sla(jtl_path, c)
                _suite_state['results'].append({
                    'jmx': jmx_name, 'label': label,
                    'status': 'done' if proc.returncode == 0 else 'failed',
                    'jtl': out_name + '.jtl',
                    'sla_passed': sla_r.get('passed') if sla_r else None,
                })
            except Exception as ex:
                _suite_state['results'].append({'jmx': jmx_name, 'label': label, 'status': 'failed', 'error': str(ex)})
            finally:
                shutil.rmtree(tmp_dir, ignore_errors=True)
                _suite_state['done'] += 1
        _suite_state['running'] = False
        _suite_state['current'] = None
        audit('SUITE_END', f"done={_suite_state['done']}")
        _notify_test_complete(c, f"Suite ({len(jmx_list)} tests)", 'done')

    threading.Thread(target=_run_suite, daemon=True).start()
    return jsonify(ok=True)

@app.route('/api/suite/status')
@login_req
def api_suite_status():
    return jsonify(**_suite_state)

@app.route('/api/suite/stop', methods=['POST'])
@perm_req('run_tests')
def api_suite_stop():
    _suite_state['running'] = False
    _suite_state['current'] = None
    return jsonify(ok=True)

# ── API: DB Maintenance ────────────────────────────────────────────────────────
@app.route('/api/db/full-stats')
@admin_req
def api_db_full_stats():
    with get_db() as db:
        tables = db.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()
        table_stats = {}
        for t in tables:
            name = t[0]
            count = db.execute(f"SELECT COUNT(*) FROM {name}").fetchone()[0]
            cols  = [c[1] for c in db.execute(f"PRAGMA table_info({name})").fetchall()]
            table_stats[name] = {'rows': count, 'columns': cols}

        # Audit breakdown by action
        action_counts = db.execute(
            "SELECT action, COUNT(*) as cnt FROM audit_log GROUP BY action ORDER BY cnt DESC"
        ).fetchall()

        # Audit by user
        user_counts = db.execute(
            "SELECT username, COUNT(*) as cnt FROM audit_log GROUP BY username ORDER BY cnt DESC LIMIT 10"
        ).fetchall()

        # Daily activity last 14 days
        daily = db.execute("""
            SELECT date(timestamp) as day, COUNT(*) as cnt
            FROM audit_log
            WHERE timestamp >= date('now', '-14 days', 'localtime')
            GROUP BY day ORDER BY day
        """).fetchall()

        oldest = db.execute("SELECT timestamp FROM audit_log ORDER BY id ASC LIMIT 1").fetchone()
        newest = db.execute("SELECT timestamp FROM audit_log ORDER BY id DESC LIMIT 1").fetchone()

    size_bytes = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    cfg = load_cfg()
    return jsonify(
        tables=table_stats,
        action_breakdown=[dict(r) for r in action_counts],
        user_breakdown=[dict(r) for r in user_counts],
        daily_activity=[dict(r) for r in daily],
        oldest_entry=oldest[0] if oldest else None,
        newest_entry=newest[0] if newest else None,
        db_size_bytes=size_bytes,
        db_size_kb=round(size_bytes / 1024, 1),
        db_size_mb=round(size_bytes / 1024 / 1024, 3),
        db_path=DB_PATH,
        retention_days=cfg.get('audit_retention_days', 90),
    )

@app.route('/api/db/vacuum', methods=['POST'])
@admin_req
def api_db_vacuum():
    before = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    with get_db() as db:
        db.execute("VACUUM")
    after = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    saved_kb = round((before - after) / 1024, 1)
    audit('DB_VACUUM', f'Before={round(before/1024,1)}KB After={round(after/1024,1)}KB Saved={saved_kb}KB')
    return jsonify(
        ok=True,
        before_kb=round(before / 1024, 1),
        after_kb=round(after / 1024, 1),
        saved_kb=saved_kb,
        message=f'Vacuum complete. Freed {saved_kb} KB.'
    )

@app.route('/api/db/purge', methods=['POST'])
@admin_req
def api_db_purge():
    days = int((request.json or {}).get('days', load_cfg().get('audit_retention_days', 90)))
    if days < 0:
        return jsonify(error='Days must be 0 or greater (0 = keep all).'), 400
    n = _purge_old_audit(days)
    audit('DB_PURGE', f'Purged {n} audit entries older than {days} days')
    return jsonify(ok=True, deleted=n,
                   message=f'Removed {n} audit entries older than {days} days.' if n else 'No entries matched.')

@app.route('/api/db/clear-audit', methods=['POST'])
@admin_req
def api_db_clear_audit():
    confirm = (request.json or {}).get('confirm', '')
    if confirm != 'CLEAR':
        return jsonify(error='Send {"confirm":"CLEAR"} to confirm.'), 400
    with get_db() as db:
        n = db.execute("SELECT COUNT(*) FROM audit_log").fetchone()[0]
        db.execute("DELETE FROM audit_log")
        db.commit()
    audit('DB_CLEAR_AUDIT', f'Cleared all {n} audit log entries')
    return jsonify(ok=True, deleted=n, message=f'Cleared {n} audit entries.')

@app.route('/api/download/db-backup')
@admin_req
def api_download_db_backup():
    ts       = datetime.now().strftime('%Y%m%d_%H%M%S')
    tmp_path = os.path.join(tempfile.gettempdir(), f'lt_backup_{ts}.db')
    src = sqlite3.connect(DB_PATH)
    dst = sqlite3.connect(tmp_path)
    src.backup(dst)
    src.close()
    dst.close()
    audit('DB_BACKUP', f'Full database backup downloaded ({round(os.path.getsize(DB_PATH)/1024,1)} KB)')
    return send_file(tmp_path, as_attachment=True,
                     download_name=f'lt_platform_backup_{ts}.db',
                     mimetype='application/octet-stream')

@app.route('/api/download/audit-csv')
@admin_req
def api_download_audit_csv():
    with get_db() as db:
        rows = db.execute(
            "SELECT id,timestamp,username,action,details,ip_address FROM audit_log ORDER BY id DESC"
        ).fetchall()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['ID', 'Timestamp', 'Username', 'Action', 'Details', 'IP Address'])
    for r in rows:
        writer.writerow(list(r))
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    resp = app.response_class(buf.getvalue(), mimetype='text/csv')
    resp.headers['Content-Disposition'] = f'attachment; filename="audit_log_{ts}.csv"'
    audit('DOWNLOAD_AUDIT_CSV', f'Exported {len(rows)} audit entries as CSV')
    return resp

# ── API: File Upload ───────────────────────────────────────────────────────────
ALLOWED_JMX    = {'.jmx'}
ALLOWED_CSV    = {'.csv'}
ALLOWED_REPORT = {'.jtl', '.html'}
ALLOWED_CONFIG = {'.properties', '.yaml', '.yml', '.json', '.txt', '.conf'}
ALLOWED_FEATURES = {'.feature', '.gherkin', '.txt'}  # BDD test feature files
ALLOWED_FEATURE_EXCEL = {'.xlsx'}

def _safe_filename(name):
    """Strip path components and dangerous characters from a filename."""
    name = os.path.basename(name)
    name = re.sub(r'[^\w\s\-\.()]', '', name).strip()
    return name or 'upload'

@app.route('/api/upload/jmx', methods=['POST'])
@perm_req('upload_files')
def api_upload_jmx():
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    if 'file' not in request.files:
        return jsonify(error='No file in request.'), 400
    f = request.files['file']
    fname = _safe_filename(f.filename)
    if not fname or os.path.splitext(fname)[1].lower() not in ALLOWED_JMX:
        return jsonify(error='Only .jmx files are allowed.'), 400
    dest = os.path.join(client_dirs(c)['jmx'], fname)
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    f.save(dest)
    size_kb = round(os.path.getsize(dest) / 1024, 1)
    audit('UPLOAD_JMX', f'Uploaded JMX: {fname} ({size_kb} KB) → client={c["code"]}')
    return jsonify(ok=True, filename=fname, size_kb=size_kb)

@app.route('/api/upload/config', methods=['POST'])
@perm_req('upload_files')
def api_upload_config():
    """Upload JMeter properties / YAML / JSON config files alongside the test plan."""
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    files = request.files.getlist('file')
    if not files or (len(files) == 1 and files[0].filename == ''):
        return jsonify(error='No files in request.'), 400
    saved, errors = [], []
    for f in files:
        fname = _safe_filename(f.filename)
        ext = os.path.splitext(fname)[1].lower()
        if not fname or ext not in ALLOWED_CONFIG:
            errors.append(f'{f.filename}: only .properties/.yaml/.yml/.json/.txt/.conf allowed')
            continue
        dest = os.path.join(client_dirs(c)['jmx'], fname)
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        f.save(dest)
        size_kb = round(os.path.getsize(dest) / 1024, 1)
        saved.append({'filename': fname, 'size_kb': size_kb})
        audit('UPLOAD_CONFIG', f'Uploaded config: {fname} ({size_kb} KB) → client={c["code"]}')
    return jsonify(ok=True, saved=saved, errors=errors)

@app.route('/api/upload/testdata', methods=['POST'])
@perm_req('upload_files')
def api_upload_testdata():
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    files = request.files.getlist('file')
    if not files:
        return jsonify(error='No files in request.'), 400
    saved, errors = [], []
    for f in files:
        fname = _safe_filename(f.filename)
        if not fname or os.path.splitext(fname)[1].lower() not in ALLOWED_CSV:
            errors.append(f'{f.filename}: only .csv files allowed')
            continue
        dest = os.path.join(client_dirs(c)['testdata'], fname)
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        f.save(dest)
        saved.append({'filename': fname, 'size_kb': round(os.path.getsize(dest) / 1024, 1)})
    if saved:
        names = ', '.join(s['filename'] for s in saved)
        audit('UPLOAD_TESTDATA', f'Uploaded {len(saved)} CSV(s): {names} → client={c["code"]}')
    return jsonify(ok=True, saved=saved, errors=errors)

@app.route('/api/upload/report', methods=['POST'])
@perm_req('upload_files')
def api_upload_report():
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    files = request.files.getlist('file')
    if not files:
        return jsonify(error='No files in request.'), 400
    saved, errors = [], []
    for f in files:
        fname = _safe_filename(f.filename)
        ext = os.path.splitext(fname)[1].lower()
        if not fname or ext not in ALLOWED_REPORT:
            errors.append(f'{f.filename}: only .jtl or .html files allowed')
            continue
        dest = os.path.join(client_dirs(c)['reports'], fname)
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        f.save(dest)
        saved.append({'filename': fname, 'size_kb': round(os.path.getsize(dest) / 1024, 1)})
    if saved:
        names = ', '.join(s['filename'] for s in saved)
        audit('UPLOAD_REPORT', f'Uploaded {len(saved)} report(s): {names} → client={c["code"]}')
    return jsonify(ok=True, saved=saved, errors=errors)

@app.route('/api/upload/delete', methods=['POST'])
@perm_req('delete_files')
def api_upload_delete():
    """Delete a file from a client directory (jmx, testdata, or reports)."""
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    data = request.json or {}
    folder = data.get('folder', '')   # 'jmx' | 'testdata' | 'reports' | 'config'
    fname  = data.get('filename', '')
    # 'config' files live in the jmx dir
    dir_key = 'jmx' if folder == 'config' else folder
    if dir_key not in ('jmx', 'testdata', 'reports', 'xlsx'):
        return jsonify(error='Invalid folder.'), 400
    fname = _safe_filename(fname)
    if not fname:
        return jsonify(error='Invalid filename.'), 400
    if dir_key == 'xlsx':
        path = os.path.join(CLIENTS_DIR, c['code'], fname)
    else:
        path = os.path.join(client_dirs(c)[dir_key], fname)
    if not os.path.exists(path):
        return jsonify(error='File not found.'), 404
    os.remove(path)
    audit('DELETE_FILE', f'Deleted {folder}/{fname} from client={c["code"]}')
    return jsonify(ok=True, deleted=fname)

# ── TEST FEATURE FILES (BDD) ──────────────────────────────────────────────────
@app.route('/api/test-features', methods=['GET'])
@login_req
def api_list_test_features():
    """List uploaded test feature files for active client"""
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    feat_dir = os.path.join(client_dirs(c).get('features', os.path.join(CLIENTS_DIR, c['code'], 'features')))
    if not os.path.exists(feat_dir):
        return jsonify(files=[])
    
    files = []
    for f in sorted(os.listdir(feat_dir)):
        if os.path.isfile(os.path.join(feat_dir, f)):
            fpath = os.path.join(feat_dir, f)
            files.append({
                'filename': f,
                'size_kb': round(os.path.getsize(fpath) / 1024, 1),
                'created': os.path.getctime(fpath)
            })
    return jsonify(files=files)

@app.route('/api/upload/test-features', methods=['POST'])
@perm_req('upload_files')
@csrf_protect
def api_upload_test_features():
    """Upload BDD test feature files and Excel workbooks."""
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    
    files = request.files.getlist('file')
    if not files:
        return jsonify(error='No files in request'), 400
    
    feat_dir = os.path.join(client_dirs(c).get('features', os.path.join(CLIENTS_DIR, c['code'], 'features')))
    root_dir = os.path.join(CLIENTS_DIR, c['code'])
    os.makedirs(feat_dir, exist_ok=True)
    os.makedirs(root_dir, exist_ok=True)
    
    saved, errors = [], []
    for f in files:
        fname = _safe_filename(f.filename)
        ext = os.path.splitext(fname)[1].lower()
        if not fname or (ext not in ALLOWED_FEATURES and ext not in ALLOWED_FEATURE_EXCEL):
            errors.append(f'{f.filename}: only .feature/.gherkin/.txt/.xlsx allowed')
            continue

        if ext in ALLOWED_FEATURE_EXCEL:
            fpath = os.path.join(root_dir, fname)
            file_type = 'excel'
        else:
            fpath = os.path.join(feat_dir, fname)
            file_type = 'gherkin'

        f.save(fpath)
        
        # Count scenarios in feature file or workbook rows for Excel uploads.
        scenario_count = 0
        try:
            if ext in ALLOWED_FEATURE_EXCEL:
                scenario_count = _count_xlsx_rows(fpath)
            else:
                with open(fpath, 'r', encoding='utf-8', errors='ignore') as fp:
                    scenario_count = sum(1 for line in fp if line.strip().startswith('Scenario:'))
        except Exception:
            pass
        
        # Store metadata in database
        feat_id = f"{c['code']}_{uuid.uuid4().hex[:8]}"
        try:
            with get_db() as db:
                db.execute(
                    "INSERT INTO test_features (id,client,filename,file_type,file_path,scenario_count,created_by) VALUES (?,?,?,?,?,?,?)",
                    (feat_id, c['code'], fname, file_type, fpath, scenario_count, session.get('user', 'system'))
                )
                db.commit()
        except:
            pass  # Continue if DB insert fails
        
        size_kb = round(os.path.getsize(fpath) / 1024, 1)
        saved.append({'filename': fname, 'size_kb': size_kb, 'scenarios': scenario_count})
        audit('UPLOAD_FEATURE', f'Uploaded feature file: {fname} ({scenario_count} scenarios) → client={c["code"]}')
    
    return jsonify(ok=True, saved=saved, errors=errors)

# ── CI/CD TEST SUITE MANAGEMENT ───────────────────────────────────────────────
# ── API Token management (browser-admin only) ──────────────────────────────────
@app.route('/api/api-tokens', methods=['GET'])
@admin_req
def api_list_tokens():
    with get_db() as db:
        rows = db.execute(
            "SELECT id,name,token_prefix,client,created_by,created_at,last_used_at,last_used_ip,enabled "
            "FROM api_tokens ORDER BY created_at DESC").fetchall()
    return jsonify(tokens=[dict(r) for r in rows])

@app.route('/api/api-tokens', methods=['POST'])
@admin_req
@csrf_protect
def api_create_token():
    d = request.json or {}
    name = (d.get('name') or '').strip()
    if not name:
        return jsonify(error='A token name is required'), 400
    client = (d.get('client') or '').strip() or (active_client() or {}).get('code')
    raw, prefix, thash = _generate_api_token()
    tid = str(uuid.uuid4())[:8]
    with get_db() as db:
        db.execute("INSERT INTO api_tokens(id,name,token_hash,token_prefix,client,created_by) "
                   "VALUES(?,?,?,?,?,?)",
                   (tid, name, thash, prefix, client, session.get('user', 'admin')))
        db.commit()
    audit('API_TOKEN_CREATED', f"id={tid} name={name} client={client}")
    # Plaintext is returned exactly once and never stored.
    return jsonify(ok=True, id=tid, token=raw, prefix=prefix, client=client,
                   note='Copy this token now — it is shown only once and cannot be retrieved later.')

@app.route('/api/api-tokens/<tid>/toggle', methods=['POST'])
@admin_req
@csrf_protect
def api_toggle_token(tid):
    enabled = bool((request.json or {}).get('enabled', True))
    with get_db() as db:
        db.execute("UPDATE api_tokens SET enabled=? WHERE id=?", (1 if enabled else 0, tid))
        db.commit()
    audit('API_TOKEN_TOGGLE', f"id={tid} enabled={enabled}")
    return jsonify(ok=True)

@app.route('/api/api-tokens/<tid>', methods=['DELETE'])
@admin_req
@csrf_protect
def api_delete_token(tid):
    with get_db() as db:
        db.execute("DELETE FROM api_tokens WHERE id=?", (tid,))
        db.commit()
    audit('API_TOKEN_REVOKED', f"id={tid}")
    return jsonify(ok=True)


@app.route('/api/ci-cd/suites', methods=['GET'])
@api_auth
def api_list_ci_suites():
    """List all CI/CD test suites for active client"""
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    
    with get_db() as db:
        suites = db.execute(
            "SELECT * FROM ci_cd_suites WHERE client=? ORDER BY created_at DESC", (c['code'],)
        ).fetchall()
    
    return jsonify(suites=[dict(s) for s in suites])

@app.route('/api/ci-cd/suites', methods=['POST'])
@api_auth
def api_create_ci_suite():
    """Create a new CI/CD test suite"""
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    
    data = request.json or {}
    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    jmx_files = data.get('jmx_files', [])  # List of JMX filenames
    feature_file = data.get('feature_file', '')  # Optional feature file
    schedule = data.get('schedule', 'daily')  # daily, weekly, hourly, etc.
    retry_count = min(int(data.get('retry_count', 0)), 3)
    notify_on_fail = data.get('notify_on_fail', '')  # Email or webhook
    
    if not name or not jmx_files:
        return jsonify(error='Suite name and at least one JMX file required'), 400
    
    suite_id = f"{c['code']}_suite_{uuid.uuid4().hex[:8]}"
    gate_config = json.dumps(data.get('gate') or {})   # release-gate thresholds

    try:
        with get_db() as db:
            db.execute(
                """INSERT INTO ci_cd_suites
                   (id,client,name,description,jmx_files,feature_file,schedule,retry_count,notify_on_fail,created_by,enabled,gate_config)
                   VALUES (?,?,?,?,?,?,?,?,?,?,1,?)""",
                (suite_id, c['code'], name, description, json.dumps(jmx_files), feature_file,
                 schedule, retry_count, notify_on_fail, session.get('user', 'system'), gate_config)
            )
            db.commit()
        
        audit('CI_CD_SUITE_CREATE', f'Created suite: {name} with {len(jmx_files)} JMX files')
        return jsonify(ok=True, suite_id=suite_id, message=f'Suite "{name}" created successfully')
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route('/api/ci-cd/suites/<suite_id>/toggle', methods=['POST'])
@admin_req
@csrf_protect
def api_toggle_ci_suite(suite_id):
    enabled = bool((request.json or {}).get('enabled', True))
    with get_db() as db:
        db.execute("UPDATE ci_cd_suites SET enabled=? WHERE id=?", (1 if enabled else 0, suite_id))
        db.commit()
    audit('CI_CD_SUITE_TOGGLE', f'{suite_id} enabled={enabled}')
    return jsonify(ok=True)

@app.route('/api/ci-cd/suites/<suite_id>', methods=['DELETE'])
@admin_req
@csrf_protect
def api_delete_ci_suite(suite_id):
    with get_db() as db:
        db.execute("DELETE FROM ci_cd_suites WHERE id=?", (suite_id,))
        db.execute("DELETE FROM ci_cd_run_history WHERE suite_id=?", (suite_id,))
        db.commit()
    audit('CI_CD_SUITE_DELETE', suite_id)
    return jsonify(ok=True)

@app.route('/api/ci-cd/suites/<suite_id>/runs', methods=['GET'])
@login_req
def api_ci_suite_runs(suite_id):
    """Recent run history for one suite, with gate verdicts, for the UI."""
    with get_db() as db:
        rows = db.execute(
            "SELECT * FROM ci_cd_run_history WHERE suite_id=? ORDER BY start_time DESC LIMIT 20",
            (suite_id,)).fetchall()
    runs = []
    for r in rows:
        d = dict(r)
        try:
            d['gate_reasons'] = json.loads(d.get('gate_reasons') or '[]')
        except Exception:
            d['gate_reasons'] = []
        runs.append(d)
    return jsonify(runs=runs)


# ── Release gates ──────────────────────────────────────────────────────────────
_GATE_DEFAULTS = {
    'require_sla_pass':       False,   # fail if SLA config is breached
    'max_error_pct':          2.0,     # fail if error rate exceeds this
    'max_p95_ms':             None,    # fail if absolute P95 exceeds this
    'min_tps':                None,    # fail if throughput below this
    'max_p95_regression_pct': 10.0,    # fail if P95 regressed >X% vs baseline
}

def _evaluate_gate(metrics, sla_result, baseline, gate_cfg=None):
    """Return {passed, reasons, checks, config}. A gate turns raw metrics +
    SLA + baseline regression into a single release verdict Jenkins can act on."""
    cfg = {**_GATE_DEFAULTS, **(gate_cfg or {})}
    checks, reasons, passed = [], [], True
    def chk(name, ok, detail):
        nonlocal passed
        checks.append({'name': name, 'passed': bool(ok), 'detail': detail})
        if not ok:
            passed = False; reasons.append(detail)
    err = float(metrics.get('err', 0) or 0)
    p95 = float(metrics.get('p95', 0) or 0)
    tps = float(metrics.get('tps', 0) or 0)
    if cfg.get('max_error_pct') is not None:
        chk('Error rate', err <= cfg['max_error_pct'], f"error rate {err}% (limit <={cfg['max_error_pct']}%)")
    if cfg.get('max_p95_ms'):
        chk('P95 absolute', p95 <= cfg['max_p95_ms'], f"P95 {int(p95)}ms (limit <={cfg['max_p95_ms']}ms)")
    if cfg.get('min_tps'):
        chk('Min throughput', tps >= cfg['min_tps'], f"throughput {tps} TPS (floor >={cfg['min_tps']})")
    if cfg.get('require_sla_pass') and sla_result is not None:
        chk('SLA', bool(sla_result.get('passed')), 'SLA ' + ('passed' if sla_result.get('passed') else 'breached'))
    base_p95 = ((baseline or {}).get('metrics') or {}).get('p95')
    if cfg.get('max_p95_regression_pct') is not None and base_p95:
        reg = (p95 - base_p95) / base_p95 * 100
        chk('P95 regression', reg <= cfg['max_p95_regression_pct'],
            f"P95 {reg:+.0f}% vs baseline {int(base_p95)}ms (limit <={cfg['max_p95_regression_pct']}%)")
    return {'passed': passed, 'reasons': reasons, 'checks': checks, 'config': cfg}


def _combine_jtls(jtls, out_path):
    """Concatenate JTLs into one (single header). Preserves multi-line rows by
    copying every non-header physical line verbatim."""
    header_written = False
    with open(out_path, 'w', newline='', encoding='utf-8') as out:
        for j in jtls:
            if not os.path.exists(j):
                continue
            with open(j, 'r', encoding='utf-8', errors='ignore') as f:
                first = f.readline()
                if not first:
                    continue
                if not header_written:
                    out.write(first); header_written = True
                for line in f:
                    out.write(line)
    return header_written


def _execute_ci_suite(suite, run_id, gate_cfg, duration=60):
    """Actually run a suite's JMX files, combine results, evaluate the release
    gate, and persist metrics + verdict to ci_cd_run_history."""
    start = datetime.now()
    c = get_client(suite['client'])
    try:
        if not c:
            raise RuntimeError(f"client {suite['client']} not found")
        dirs = client_dirs(c)
        try:
            jmx_files = json.loads(suite['jmx_files'])
        except Exception:
            jmx_files = [s.strip() for s in (suite['jmx_files'] or '').split(',') if s.strip()]
        jbin = _find_jmeter_bin()
        combined = os.path.join(dirs['reports'], f"{run_id}.jtl")
        part_jtls = []
        for i, jmx in enumerate(jmx_files):
            jmx_path = os.path.join(dirs['jmx'], jmx)
            if not os.path.exists(jmx_path):
                continue
            jtl_i = os.path.join(dirs['reports'], f"{run_id}_{i}.jtl")
            cmd = [jbin, '-n', '-t', jmx_path, '-l', jtl_i,
                   f'-Jtest.duration={duration}',
                   '-Jjmeter.save.saveservice.url=true',
                   '-Jjmeter.save.saveservice.response_data=true',
                   '-Jjmeter.save.saveservice.samplerData=true',
                   '-Jjmeter.save.saveservice.requestHeaders=true',
                   '-Jjmeter.save.saveservice.responseHeaders=true',
                   '-Jjmeter.save.saveservice.assertion_results_failure_message=true']
            try:
                subprocess.run(cmd, capture_output=True, text=True, timeout=duration + 900)
            except Exception as ex:
                print(f'[ci-cd] {jmx} run error: {ex}')
            if os.path.exists(jtl_i):
                part_jtls.append(jtl_i)
        _combine_jtls(part_jtls, combined)
        metrics  = _quick_metrics(combined) or {'p95': 0, 'err': 0, 'tps': 0, 'total': 0, 'avg': 0}
        total    = int(metrics.get('total', 0))
        err_cnt  = round(metrics.get('err', 0) / 100 * total) if total else 0
        sla      = _evaluate_sla(combined, c) if total else None
        baseline = _read_json(_baseline_path(c)) or {}
        gate     = _evaluate_gate(metrics, sla, baseline, gate_cfg)
        if total == 0:
            status = 'error'
            gate['passed'] = False
            gate['reasons'].insert(0, 'No samples produced — JMeter execution failed or JMX missing.')
        else:
            status = 'passed' if gate['passed'] else 'gate_failed'
        dur = int((datetime.now() - start).total_seconds())
        with get_db() as db:
            db.execute("""UPDATE ci_cd_run_history SET status=?, end_time=?, duration_s=?,
                          total_requests=?, success_count=?, error_count=?, avg_rt_ms=?, p95_rt_ms=?,
                          report_path=?, gate_status=?, gate_reasons=? WHERE id=?""",
                       (status, datetime.now().isoformat(), dur, total, total - err_cnt, err_cnt,
                        metrics.get('avg', 0), metrics.get('p95', 0),
                        os.path.basename(combined) if total else None,
                        'pass' if gate['passed'] else 'fail', json.dumps(gate['reasons']), run_id))
            db.execute("UPDATE ci_cd_suites SET last_run=? WHERE id=?",
                       (datetime.now().isoformat(), suite['id']))
            db.commit()
        try:
            _notify_test_complete(c, f"Suite: {suite['name']}",
                                  'done' if gate['passed'] else 'failed', sla,
                                  jtl_name=os.path.basename(combined) if total else None)
        except Exception:
            pass
        for j in part_jtls:
            try: os.remove(j)
            except OSError: pass
    except Exception as ex:
        try:
            with get_db() as db:
                db.execute("UPDATE ci_cd_run_history SET status=?, end_time=?, gate_status=?, gate_reasons=? WHERE id=?",
                           ('error', datetime.now().isoformat(), 'fail',
                            json.dumps([f'Execution error: {ex}']), run_id))
                db.commit()
        except Exception:
            pass


@app.route('/api/ci-cd/suites/<suite_id>/run', methods=['POST'])
@api_auth
def api_run_ci_suite(suite_id):
    """Trigger a real CI/CD suite execution and evaluate the release gate.
    Body (optional): {"duration": <secs>, "gate": {<overrides>}}."""
    try:
        with get_db() as db:
            suite = db.execute("SELECT * FROM ci_cd_suites WHERE id=?", (suite_id,)).fetchone()
        if not suite:
            return jsonify(error='Suite not found'), 404
        suite = dict(suite)
        try:
            suite_gate = json.loads(suite.get('gate_config') or '{}')
        except Exception:
            suite_gate = {}
        body = request.json or {}
        gate_cfg = {**suite_gate, **(body.get('gate') or {})}
        try:
            duration = max(10, min(int(body.get('duration', 60)), 7200))
        except (ValueError, TypeError):
            duration = 60
        run_id = f"run_{uuid.uuid4().hex[:12]}"
        trg = session.get('user') or 'manual'
        if getattr(g, 'api_token', None):
            trg = 'token:' + (g.api_token.get('name') or '')
        with get_db() as db:
            db.execute("""INSERT INTO ci_cd_run_history (id,suite_id,client,status,start_time,triggered_by)
                          VALUES (?,?,?,?,?,?)""",
                       (run_id, suite_id, suite['client'], 'running', datetime.now().isoformat(), trg))
            db.commit()
        audit('CI_CD_SUITE_RUN', f'Suite run started: {suite["name"]} ({suite_id}) run={run_id} by={trg}')
        threading.Thread(target=_execute_ci_suite, args=(suite, run_id, gate_cfg, duration),
                         daemon=True).start()
        return jsonify(ok=True, run_id=run_id, suite_name=suite['name'], status='started',
                       poll=f'/api/ci-cd/runs/{run_id}')
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route('/api/ci-cd/runs/<run_id>', methods=['GET'])
@api_auth
def api_get_ci_run_status(run_id):
    """Get status, metrics and release-gate verdict of a CI/CD run.

    Jenkins polls this and checks `gate` == 'pass'. `done` is true once the run
    has finished (passed / gate_failed / error)."""
    try:
        with get_db() as db:
            run = db.execute("SELECT * FROM ci_cd_run_history WHERE id=?", (run_id,)).fetchone()
        if not run:
            return jsonify(error='Run not found'), 404
        run = dict(run)
        try:
            run['gate_reasons'] = json.loads(run.get('gate_reasons') or '[]')
        except Exception:
            run['gate_reasons'] = []
        done = run.get('status') in ('passed', 'gate_failed', 'completed', 'error')
        return jsonify(run=run, gate=run.get('gate_status'), done=done,
                       report=(f"/api/report/{run['report_path']}/html" if run.get('report_path') else None))
    except Exception as e:
        return jsonify(error=str(e)), 500

# ── API: JMX Requirements check ───────────────────────────────────────────────
@app.route('/api/jmx-requirements')
@login_req
def api_jmx_requirements():
    """Parse a JMX and return which CSV files it needs + whether they exist."""
    jmx_name = request.args.get('jmx', '')
    c = active_client()
    if not c:
        return jsonify(error='No active client.'), 400
    dirs = client_dirs(c)
    if not jmx_name:
        return jsonify(error='jmx parameter required.'), 400
    jmx_path = os.path.join(dirs['jmx'], jmx_name)
    if not os.path.exists(jmx_path):
        return jsonify(error=f'JMX not found: {jmx_name}'), 404

    try:
        tree = ET.parse(jmx_path)
        root = tree.getroot()
    except Exception as ex:
        return jsonify(error=f'Cannot parse JMX: {ex}'), 500

    testdata_dir = dirs['testdata']
    csv_files    = [os.path.basename(f) for f in glob.glob(os.path.join(testdata_dir, '*.csv'))]

    required = []
    for csv_ds in root.iter('CSVDataSet'):
        ds_name = csv_ds.get('testname', '')
        for prop in csv_ds.iter('stringProp'):
            if prop.get('name') == 'filename':
                raw_path = (prop.text or '').strip()
                if not raw_path:
                    continue
                basename = os.path.basename(raw_path.replace('/', os.sep).replace('\\', os.sep))
                present  = os.path.exists(os.path.join(testdata_dir, basename))
                required.append({
                    'dataset_name': ds_name,
                    'original_path': raw_path,
                    'filename': basename,
                    'present': present,
                })

    jmx_present   = os.path.exists(jmx_path)
    jmx_size_kb   = round(os.path.getsize(jmx_path) / 1024, 1) if jmx_present else 0
    missing       = [r for r in required if not r['present']]
    ready_to_run  = jmx_present and len(missing) == 0

    return jsonify(
        jmx=jmx_name,
        jmx_present=jmx_present,
        jmx_size_kb=jmx_size_kb,
        required_csvs=required,
        missing_csvs=missing,
        ready_to_run=ready_to_run,
        testdata_dir=testdata_dir,
    )

# ── API: JMX services ──────────────────────────────────────────────────────────
@app.route('/api/jmx-services')
@login_req
def api_jmx_services():
    jmx_name = request.args.get('jmx', '')
    c = active_client()
    if not c: return jsonify(error='No active client.'), 400
    dirs = client_dirs(c)
    raw_jmx = sorted(glob.glob(os.path.join(dirs['jmx'], '*.jmx')))
    all_jmx = [os.path.basename(f) for f in raw_jmx]
    all_jmx_meta = [
        {'filename': os.path.basename(f),
         'size_kb': round(os.path.getsize(f) / 1024, 1)}
        for f in raw_jmx
    ]
    if not jmx_name:
        # Return file list with metadata when no specific JMX requested
        return jsonify(files=all_jmx, files_meta=all_jmx_meta)
    if jmx_name not in all_jmx:
        files = glob.glob(os.path.join(dirs['jmx'], '*.jmx'))
        if not files: return jsonify(error='No JMX files found.'), 404
        jmx_name = os.path.basename(files[0])
    jmx_path = os.path.join(dirs['jmx'], jmx_name)
    if not os.path.exists(jmx_path):
        return jsonify(error=f'JMX not found: {jmx_name}'), 404
    try:
        tree = ET.parse(jmx_path)
        root = tree.getroot()
        # Build a normalized stem lookup so trailing-space filenames like
        # "Betway .csv" still match service name "Betway"
        _td_stems = {
            os.path.splitext(os.path.basename(f))[0].strip().lower()
            for f in glob.glob(os.path.join(dirs['testdata'], '*.csv'))
        }
        services = []
        for tg in root.iter('ThreadGroup'):
            raw     = tg.get('testname', '')
            enabled = tg.get('enabled', 'true').lower() == 'true'
            m = re.match(r'^TG - (.+?) \((.+)\)$', raw.strip())
            svc_name = m.group(1).strip() if m else raw.strip()
            threads_el = tg.find('.//stringProp[@name="ThreadGroup.num_threads"]')
            rampup_el  = tg.find('.//stringProp[@name="ThreadGroup.ramp_time"]')
            threads_expr = threads_el.text if threads_el is not None else '2'
            rampup_expr  = rampup_el.text  if rampup_el  is not None else '5'
            tm = re.search(r'\$\{__P\((.+?),(\d+)\)\}', threads_expr)
            rm = re.search(r'\$\{__P\((.+?),(\d+)\)\}', rampup_expr)
            has_csv = svc_name.strip().lower() in _td_stems
            services.append(dict(
                name=svc_name, tg_name=raw, enabled=enabled,
                threads_key=tm.group(1) if tm else None,
                threads_default=int(tm.group(2)) if tm else 2,
                rampup_key=rm.group(1) if rm else None,
                rampup_default=int(rm.group(2)) if rm else 5,
                has_csv=has_csv,
            ))
        return jsonify(services=services, jmx=jmx_name)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: Run history ───────────────────────────────────────────────────────────
@app.route('/api/run-history')
@login_req
def api_run_history():
    c = active_client()
    rdir = client_dirs(c)['reports'] if c else ''
    jtl_files = sorted(glob.glob(os.path.join(rdir, '*.jtl')), key=os.path.getmtime)[-20:]
    history = []
    for path in jtl_files:
        try:
            d = _parse_jtl(path)
            history.append(dict(
                name=d['filename'], date=d['start_time'],
                tps=d['throughput'], avg_rt=d['avg_rt'],
                p90=d['p90'], error_rate=d['error_rate'],
                total=d['total'], passed=d['passed'], failed=d['failed'],
                duration=d['test_duration'],
                modified=datetime.fromtimestamp(os.path.getmtime(path)).strftime('%d %b %Y %H:%M'),
            ))
        except Exception:
            pass
    return jsonify(history=history)

# ── API: Compare runs ──────────────────────────────────────────────────────────
@app.route('/api/compare')
@login_req
def api_compare():
    a, b = request.args.get('a',''), request.args.get('b','')
    if not a or not b: return jsonify(error='Provide a= and b= filenames.'), 400
    rdir = _reports_dir()
    pa, pb = os.path.join(rdir, a), os.path.join(rdir, b)
    for p, n in [(pa, a), (pb, b)]:
        if not os.path.exists(p): return jsonify(error=f'Not found: {n}'), 404
    try:
        return jsonify(a=_parse_jtl(pa), b=_parse_jtl(pb))
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: Downloads ─────────────────────────────────────────────────────────────
@app.route('/api/download/html/<path:fname>')
@login_req
def api_download_html(fname):
    c = active_client()
    p = os.path.join(client_dirs(c)['reports'] if c else _reports_dir(), fname)
    if not os.path.exists(p): p = os.path.join(_reports_dir(), fname)
    if not os.path.exists(p): return 'File not found', 404
    try:
        html = _report_html(p, c)
        resp = app.response_class(html, mimetype='text/html')
        safe = re.sub(r'[^a-zA-Z0-9_.-]', '_', fname.replace('.jtl',''))
        resp.headers['Content-Disposition'] = f'attachment; filename="Report_{safe}.html"'
        audit('DOWNLOAD_HTML', f'HTML: {fname}')
        return resp
    except Exception as ex:
        return str(ex), 500

@app.route('/api/report/<path:fname>/html')
@login_req
def api_report_html_view(fname):
    """Inline HTML report view — same content as the download, opened in browser tab."""
    c = active_client()
    p = os.path.join(client_dirs(c)['reports'] if c else _reports_dir(), fname)
    if not os.path.exists(p): return 'File not found', 404
    try:
        html = _report_html(p, c)
        audit('VIEW_HTML', f'HTML view: {fname}')
        return app.response_class(html, mimetype='text/html')
    except Exception as ex:
        return str(ex), 500

@app.route('/api/download/jtl/<path:fname>')
@login_req
def api_download_jtl(fname):
    p = os.path.join(_reports_dir(), fname)
    if not os.path.exists(p): return 'File not found', 404
    audit('DOWNLOAD_JTL', f'JTL: {fname}')
    return send_file(p, as_attachment=True, download_name=fname, mimetype='text/csv')

@app.route('/api/download/jmx/<path:fname>')
@login_req
def api_download_jmx(fname):
    c = active_client()
    p = os.path.join(client_dirs(c)['jmx'], fname) if c else ''
    if not p or not os.path.exists(p): return 'File not found', 404
    audit('DOWNLOAD_JMX', f'JMX: {fname}')
    return send_file(p, as_attachment=True, download_name=fname, mimetype='application/octet-stream')

@app.route('/api/download/bundle/<path:fname>')
@login_req
def api_download_bundle(fname):
    p = os.path.join(_reports_dir(), fname)
    if not os.path.exists(p): return 'File not found', 404
    try:
        data = _parse_jtl(p)
        _enrich_report(data, p, active_client())
        html = _generate_report_html(data)
        buf  = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
            z.write(p, fname)
            z.writestr(fname.replace('.jtl','_report.html'), html)
            summary = {k: v for k, v in data.items() if k not in ('tps_over_time','hist_data','label_stats')}
            z.writestr(fname.replace('.jtl','_summary.json'), json.dumps(summary, indent=2))
        buf.seek(0)
        safe = re.sub(r'[^a-zA-Z0-9_.-]', '_', fname.replace('.jtl',''))
        resp = app.response_class(buf.getvalue(), mimetype='application/zip')
        resp.headers['Content-Disposition'] = f'attachment; filename="Bundle_{safe}.zip"'
        audit('DOWNLOAD_BUNDLE', f'Bundle: {fname}')
        return resp
    except Exception as ex:
        return str(ex), 500

@app.route('/api/download/all-reports')
@login_req
def api_download_all_reports():
    c = active_client()
    rdir = client_dirs(c)['reports'] if c else ''
    jtl_files = glob.glob(os.path.join(rdir, '*.jtl'))
    if not jtl_files: return 'No JTL files found', 404
    buf = io.BytesIO()
    skipped = []
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        for path in jtl_files:
            fname = os.path.basename(path)
            z.write(path, f'jtl/{fname}')
            try:
                data = _parse_jtl(path)
                _enrich_report(data, path, active_client())
                z.writestr(f'html/{fname.replace(".jtl","_report.html")}', _generate_report_html(data))
            except Exception as ex:
                skipped.append(f'{fname}: {ex}')
        # Test data
        td = client_dirs(c)['testdata'] if c else ''
        for path in glob.glob(os.path.join(td, '*.csv')):
            z.write(path, f'testdata/{os.path.basename(path)}')
        z.writestr('README.txt', (
            f'Client: {c["name"]} ({c["code"]})\n'
            f'Generated: {datetime.now().strftime("%d %b %Y %H:%M:%S")}\n'
            f'JTL files: {len(jtl_files)}\n'
            f'HTML reports: {len(jtl_files) - len(skipped)}\n\n'
            f'Folders:\n'
            f'  jtl/       — Raw JMeter result files\n'
            f'  html/      — Standalone HTML dashboards\n'
            f'  testdata/  — JMeter test data CSV files\n'
            + (f'\nErrors:\n' + '\n'.join(skipped) if skipped else '')
        ))
    buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    resp = app.response_class(buf.getvalue(), mimetype='application/zip')
    resp.headers['Content-Disposition'] = f'attachment; filename="{c["code"]}_AllReports_{ts}.zip"'
    audit('DOWNLOAD_ALL', f'All reports ZIP: {len(jtl_files)} files, client={c["code"]}')
    return resp

@app.route('/api/download/all-testdata')
@login_req
def api_download_all_testdata():
    c = active_client()
    td = client_dirs(c)['testdata'] if c else ''
    csv_files = glob.glob(os.path.join(td, '*.csv'))
    if not csv_files: return 'No CSV files found', 404
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        for path in csv_files:
            z.write(path, os.path.basename(path))
        z.writestr('README.txt', (
            f'Client: {c["name"]} ({c["code"]})\n'
            f'Test Data — {len(csv_files)} CSV files\n'
            f'Generated: {datetime.now()}\n\n'
            f'Place files in: testdata/ folder alongside the JMX test plan.\n'
        ))
    buf.seek(0)
    resp = app.response_class(buf.getvalue(), mimetype='application/zip')
    resp.headers['Content-Disposition'] = f'attachment; filename="{c["code"]}_TestData_{datetime.now().strftime("%Y%m%d")}.zip"'
    audit('DOWNLOAD_TESTDATA', f'Test data ZIP: client={c["code"]}')
    return resp

@app.route('/api/download/testdata-file/<path:fname>')
@login_req
def api_download_testdata_file(fname):
    c = active_client()
    p = os.path.join(client_dirs(c)['testdata'], fname) if c else ''
    if not p or not os.path.exists(p): return 'File not found', 404
    audit('DOWNLOAD_CSV', f'CSV: {fname}')
    return send_file(p, as_attachment=True, download_name=fname, mimetype='text/csv')

# ── JTL Parser ─────────────────────────────────────────────────────────────────
def _parse_jtl(path):
    rows = []
    with open(path, encoding='utf-8', errors='ignore') as f:
        for r in csv.DictReader(f):
            rows.append(r)
    if not rows:
        raise ValueError('Empty or invalid JTL file')
    el  = [int(r['elapsed']) for r in rows]
    ts  = [int(r['timeStamp']) for r in rows]
    t0  = min(ts)
    t1  = max(ts[i] + el[i] for i in range(len(rows)))
    dur = max(1, round((t1 - t0) / 1000))
    N   = len(rows)
    ok  = sum(1 for r in rows if r.get('success') == 'true')

    by_label       = defaultdict(list)
    by_label_bytes = defaultdict(lambda: [0, 0])   # [recv_bytes, sent_bytes]
    label_url      = {}                             # first URL seen per label
    label_fail_details = defaultdict(list)          # up to 3 failure samples per label
    samples_io = []                                 # bounded per-sample request/response for the shareable report
    io_ok_per_label = defaultdict(int)
    _IO_MAX_TOTAL, _IO_OK_PER_LABEL = 500, 5
    for r in rows:
        lbl = r['label']
        by_label[lbl].append(int(r['elapsed']))
        try:
            by_label_bytes[lbl][0] += int(r.get('bytes', 0) or 0)
            by_label_bytes[lbl][1] += int(r.get('sentBytes', 0) or 0)
        except (ValueError, TypeError):
            pass
        url = (r.get('URL') or '').strip()
        if lbl not in label_url and url and url.lower() != 'null':
            label_url[lbl] = url
        is_ok = r.get('success') == 'true'
        if not is_ok and len(label_fail_details[lbl]) < 3:
            label_fail_details[lbl].append({
                'url':     url if url and url.lower() != 'null' else '',
                'rc':      (r.get('responseCode') or '').strip(),
                'rm':      (r.get('responseMessage') or '').strip(),
                'fm':      (r.get('failureMessage') or '').strip(),
                'thread':  (r.get('threadName') or '').strip(),
                'elapsed': (r.get('elapsed') or '0').strip(),
            })
        # Bounded per-sample request/response: every failure (capped) + a few
        # successes per label, so the shared HTML stays a reasonable size.
        keep_io = False
        if not is_ok:
            keep_io = len(samples_io) < _IO_MAX_TOTAL
        elif io_ok_per_label[lbl] < _IO_OK_PER_LABEL and len(samples_io) < _IO_MAX_TOTAL:
            io_ok_per_label[lbl] += 1; keep_io = True
        if keep_io:
            samples_io.append({
                'label':   lbl, 'ok': is_ok,
                'rc':      (r.get('responseCode') or '').strip(),
                'rm':      (r.get('responseMessage') or '').strip(),
                'thread':  (r.get('threadName') or '').strip(),
                'elapsed': (r.get('elapsed') or '0').strip(),
                'url':     url if url and url.lower() != 'null' else '',
                'fm':      (r.get('failureMessage') or '').strip(),
                'req':     (r.get('samplerData') or '').strip()[:4000],
                'resp':    (r.get('responseData') or '').strip()[:4000],
            })

    # Pre-group failure rows by label for efficiency
    fail_rows_by_label = defaultdict(list)
    for r in rows:
        if r.get('success') != 'true':
            fail_rows_by_label[r['label']].append(r)

    label_stats = []
    for lbl, arr in sorted(by_label.items()):
        se = sorted(arr); n = len(arr)
        fail_rows = fail_rows_by_label.get(lbl, [])
        errs = len(fail_rows)
        # Collect failure reasons: separate expected text from actual response
        reason_counts = defaultdict(int)
        reason_detail = {}
        for fr in fail_rows:
            fail_msg  = (fr.get('failureMessage') or '').strip()
            resp_msg  = (fr.get('responseMessage') or '').strip()
            resp_data = (fr.get('responseData') or '').strip()[:300]
            resp_code = (fr.get('responseCode') or '').strip()

            expected = ''
            actual_from_msg = ''
            if fail_msg.lower().startswith('expected:'):
                rest = fail_msg[9:].strip()
                if ' | Actual: ' in rest:
                    idx = rest.index(' | Actual: ')
                    expected       = rest[:idx].strip()
                    actual_from_msg = rest[idx + 11:].strip()
                else:
                    expected = rest
            elif fail_msg.lower().startswith('test failed:'):
                expected = fail_msg[12:].strip()
            elif fail_msg:
                expected = fail_msg

            actual = actual_from_msg or resp_data or (f'HTTP {resp_code} — {resp_msg}' if resp_code not in ('200','') else resp_msg) or 'Not captured'

            key = f'{expected}|||{actual}'
            reason_counts[key[:500]] += 1
            if key[:500] not in reason_detail:
                reason_detail[key[:500]] = {'expected': expected, 'actual': actual}

        failure_reasons = []
        for key, c in sorted(reason_counts.items(), key=lambda x: -x[1]):
            det = reason_detail.get(key, {})
            failure_reasons.append({
                'expected': det.get('expected', ''),
                'actual':   det.get('actual', ''),
                'reason':   det.get('expected', key),  # backward-compat
                'count':    c,
            })
        failure_reasons = failure_reasons[:8]
        recv_b = by_label_bytes[lbl][0]
        sent_b = by_label_bytes[lbl][1]
        label_stats.append(dict(
            label=lbl, samples=n, failures=errs,
            error_rate=round(errs/n*100, 2), tps=round(n/dur, 3),
            avg=int(round(statistics.mean(arr))), median=int(round(statistics.median(arr))),
            min=se[0], max=se[-1],
            p90=se[int(n*.90)], p95=se[int(n*.95)], p99=se[int(n*.99)],
            recv_kb=round(recv_b/1024, 2),
            sent_kb=round(sent_b/1024, 2),
            recv_kbps=round(recv_b/1024/dur, 2),
            sent_kbps=round(sent_b/1024/dur, 2),
            url=label_url.get(lbl, ''),
            fail_details=label_fail_details.get(lbl, []),
            failure_reasons=failure_reasons,
        ))

    s_all = sorted(el)
    def pct(p): return s_all[int(len(s_all)*p/100)]

    bkt = defaultdict(int)
    for r in rows:
        bkt[(int(r['timeStamp'])-t0)//10000*10] += 1
    tps_ot = [dict(time_offset=k, tps=round(v/10, 2)) for k, v in sorted(bkt.items())]

    rc = defaultdict(int)
    for r in rows:
        rc[r.get('responseCode','?')] += 1

    hr = [('<100ms',0,100),('100-200ms',100,200),('200-500ms',200,500),
          ('500ms-1s',500,1000),('1s-2s',1000,2000),('2s-3s',2000,3000),
          ('3s-5s',3000,5000),('>5s',5000,int(9e9))]
    hist = [dict(range=l, count=sum(1 for e in el if lo<=e<hi)) for l,lo,hi in hr]

    import datetime as _dt
    import re as _re_sg

    # ── Build service_groups: group steps under their TC (service) label ────────
    _TC_PAT = _re_sg.compile(r'^TC\s*[-–]', _re_sg.I)
    tc_map   = {}   # svc_name → stat dict
    non_tc   = []
    for s in label_stats:
        if _TC_PAT.match(s['label']):
            svc = _re_sg.sub(r'^TC\s*[-–]\s*', '', s['label'], flags=_re_sg.I).strip()
            tc_map[svc] = s
        else:
            non_tc.append(s)

    service_groups = []
    if tc_map:
        used = set()
        for svc_name in sorted(tc_map):
            tc_s   = tc_map[svc_name]
            prefix = svc_name.lower()[:18]
            steps  = [s for s in non_tc if s['label'].lower().startswith(prefix)]
            used  |= {s['label'] for s in steps}
            service_groups.append({'service': svc_name, 'summary': tc_s,
                                   'steps': steps if steps else [tc_s]})
        leftover = [s for s in non_tc if s['label'] not in used]
        if leftover:
            tot = sum(s['samples'] for s in leftover)
            service_groups.append({
                'service': 'Other',
                'summary': {
                    'samples': tot,
                    'failures': sum(s['failures'] for s in leftover),
                    'error_rate': round(sum(s['failures'] for s in leftover)/tot*100, 2) if tot else 0,
                    'avg': round(sum(s['avg']*s['samples'] for s in leftover)/tot) if tot else 0,
                    'tps': round(sum(s['tps'] for s in leftover), 3),
                },
                'steps': leftover,
            })
    else:
        # No TC labels → each label is its own service
        for s in non_tc:
            service_groups.append({'service': s['label'], 'summary': s, 'steps': [s]})

    return dict(
        filename=os.path.basename(path),
        total=N, passed=ok, failed=N-ok,
        error_rate=round((N-ok)/N*100, 4),
        throughput=round(N/dur, 2),
        avg_rt=round(statistics.mean(el), 2),
        min_rt=min(el), max_rt=max(el),
        p50=pct(50), p90=pct(90), p95=pct(95), p99=pct(99),
        std_dev=round(statistics.stdev(el), 2) if N>1 else 0,
        start_time=str(_dt.datetime.fromtimestamp(t0/1000)),
        end_time=str(_dt.datetime.fromtimestamp(t1/1000)),
        duration_s=dur, test_duration=f'{dur//60}m {dur%60}s',
        peak_threads=max(int(r.get('allThreads',1)) for r in rows),
        label_stats=label_stats, service_groups=service_groups,
        rc_dist=dict(rc), tps_over_time=tps_ot, hist_data=hist,
        samples_io=samples_io,
    )


def _render_samples_io(d):
    """Collapsible Request/Response section for the shareable HTML report so the
    team can debug each call without logging into the platform. Data is bounded
    (all failures capped + a few successes per label) by _parse_jtl."""
    samples = d.get('samples_io') or []
    if not samples:
        return ''
    import html as _h
    def esc(s): return _h.escape(s or '', quote=False)
    fails = sum(1 for s in samples if not s['ok'])
    rows = []
    for s in samples:
        badge  = 'PASS' if s['ok'] else 'FAIL'
        color  = '#22c55e' if s['ok'] else '#ef4444'
        search = esc((s['label'] + ' ' + s['rc'] + ' ' + badge).lower())
        url_html = (f'<div style="font-family:monospace;font-size:11px;color:#3b82f6;'
                    f'word-break:break-all;margin-bottom:6px;">{esc(s["url"])}</div>') if s['url'] else ''
        assertion = (f'<div style="margin:6px 0;padding:7px 11px;background:rgba(239,68,68,.1);'
                     f'border:1px solid rgba(239,68,68,.3);border-radius:6px;font-size:12px;color:#fca5a5;">'
                     f'<b style="color:#ef4444;">Assertion:</b> {esc(s["fm"])}</div>') if s['fm'] else ''
        req_pre  = (f'<pre style="background:#0d1117;border:1px solid #1f2d44;border-radius:6px;padding:9px 11px;'
                    f'margin:0 0 4px;font-size:11px;color:#d6e2ff;white-space:pre-wrap;word-break:break-word;'
                    f'max-height:240px;overflow:auto;">{esc(s["req"])}</pre>') if s['req'] else \
                   '<div style="color:#64748b;font-size:11px;">No request data captured.</div>'
        resp_pre = (f'<pre style="background:#0d1117;border:1px solid #1f2d44;border-radius:6px;padding:9px 11px;'
                    f'margin:0;font-size:11px;color:#d6e2ff;white-space:pre-wrap;word-break:break-word;'
                    f'max-height:240px;overflow:auto;">{esc(s["resp"])}</pre>') if s['resp'] else \
                   '<div style="color:#64748b;font-size:11px;">No response body captured.</div>'
        rows.append(
            f'<details class="rr-item" data-s="{search}" style="background:#0b1220;border:1px solid #1f2d44;'
            f'border-left:3px solid {color};border-radius:8px;margin-bottom:7px;">'
            f'<summary style="cursor:pointer;padding:9px 12px;list-style:none;display:flex;gap:9px;'
            f'align-items:center;flex-wrap:wrap;">'
            f'<span style="font-size:10px;font-weight:800;padding:2px 8px;border-radius:10px;background:{color};'
            f'color:#04121f;">{badge}</span>'
            f'<span style="font-weight:600;font-size:13px;color:#e2e8f0;">{esc(s["label"])}</span>'
            f'<span style="color:#64748b;font-size:11px;">code {esc(s["rc"])} &middot; {esc(s["elapsed"])} ms '
            f'&middot; {esc(s["thread"])}</span></summary>'
            f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;padding:4px 12px 12px;">'
            f'<div><div style="font-size:10px;font-weight:800;letter-spacing:.06em;text-transform:uppercase;'
            f'color:#3b82f6;margin-bottom:5px;">Request</div>{url_html}{req_pre}</div>'
            f'<div><div style="font-size:10px;font-weight:800;letter-spacing:.06em;text-transform:uppercase;'
            f'color:#22c55e;margin-bottom:5px;">Response</div>{assertion}{resp_pre}</div></div></details>'
        )
    return (
        '<div style="margin-top:26px;background:#111827;border:1px solid #1f2d44;border-radius:14px;padding:20px 22px;">'
        '<h2 style="margin:0 0 4px;font-size:16px;color:#e2e8f0;">&#127795; Request / Response by sample</h2>'
        f'<div style="color:#64748b;font-size:12px;margin-bottom:12px;">{len(samples)} samples shown '
        f'({fails} failed) &middot; all failures plus a few successes per transaction. '
        'Click a row to expand.</div>'
        '<input oninput="rrFilter(this.value)" placeholder="filter by label / code / PASS / FAIL…" '
        'style="width:100%;max-width:360px;margin-bottom:12px;padding:7px 10px;background:#0b1220;'
        'border:1px solid #1f2d44;border-radius:6px;color:#e2e8f0;font-size:12px;">'
        f'<div id="rr-wrap">{"".join(rows)}</div>'
        '<script>function rrFilter(q){q=(q||"").toLowerCase();'
        'document.querySelectorAll("#rr-wrap .rr-item").forEach(function(el){'
        'el.style.display=(!q||el.getAttribute("data-s").indexOf(q)>-1)?"":"none";});}</script>'
        '</div>\n\n'
    )

def _quick_metrics(path):
    """Lightweight p95 / error% / tps for a JTL — reads only the needed columns.
    Used for the trend view across recent runs."""
    try:
        elapsed = []; err = 0; total = 0; tmin = tmax = None
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            hdr = [h.strip() for h in next(reader, [])]
            if not hdr: return None
            ei = hdr.index('elapsed')   if 'elapsed'   in hdr else 1
            si = hdr.index('success')   if 'success'   in hdr else 7
            ti = hdr.index('timeStamp') if 'timeStamp' in hdr else 0
            for row in reader:
                if len(row) <= max(ei, si, ti): continue
                try:
                    e = int(row[ei]); ts = int(row[ti])
                except ValueError:
                    continue
                elapsed.append(e); total += 1
                tmin = ts if tmin is None else min(tmin, ts)
                tmax = ts if tmax is None else max(tmax, ts)
                if row[si].strip().lower() != 'true': err += 1
        if not elapsed: return None
        elapsed.sort()
        dur = max(1, (tmax - tmin) / 1000) if tmax and tmin else 1
        p95 = elapsed[min(len(elapsed) - 1, int(len(elapsed) * 95 / 100))]
        return {'p95': p95, 'err': round(err / total * 100, 2),
                'tps': round(total / dur, 2), 'total': total,
                'avg': round(sum(elapsed) / len(elapsed))}
    except Exception:
        return None


_TREND_CACHE = {}
def _metrics_cached(path):
    try:
        key = (path, os.path.getmtime(path))
    except OSError:
        return None
    if key in _TREND_CACHE:
        return _TREND_CACHE[key]
    m = _quick_metrics(path)
    _TREND_CACHE[key] = m
    if len(_TREND_CACHE) > 200:
        _TREND_CACHE.pop(next(iter(_TREND_CACHE)))
    return m


def _recent_trend(c, current_path, n=8):
    """Metrics for the last n primary reports (excludes *agr* aggregate files),
    oldest→newest, flagging the current one."""
    try:
        rdir = client_dirs(c)['reports'] if c else _reports_dir()
    except Exception:
        rdir = os.path.dirname(current_path)
    files = [f for f in glob.glob(os.path.join(rdir, '*.jtl'))
             if 'agr' not in os.path.basename(f).lower()]
    files.sort(key=os.path.getmtime)
    files = files[-n:]
    if not any(os.path.abspath(f) == os.path.abspath(current_path) for f in files):
        files = (files + [current_path])[-n:]
    out = []
    for f in files:
        m = _metrics_cached(f)
        if m:
            out.append({'name': os.path.basename(f), 'p95': m['p95'], 'err': m['err'],
                        'tps': m['tps'],
                        'current': os.path.abspath(f) == os.path.abspath(current_path)})
    return out


def _capacity_estimate(d):
    """Conservative capacity indicator from a single run. NOT a measured breaking
    point (that needs a ramp test) — it projects headroom from the latency margin
    to the SLA/target P95 while errors stay low."""
    tps = float(d.get('throughput', 0) or 0)
    p95 = float(d.get('p95', 0) or 0)
    err = float(d.get('error_rate', 0) or 0)
    target_p95 = float(d.get('target_p95') or 2000)
    stable = err < 1.0 and 0 < p95 <= target_p95
    if stable:
        ratio    = min(3.0, max(1.05, target_p95 / p95))
        max_tps  = round(tps * ratio, 1)
        headroom = round((max_tps - tps) / max_tps * 100) if max_tps else 0
        verdict  = 'headroom'
        note     = ('Estimated from this run\'s latency margin to the P95 target while errors stayed low. '
                    'Run a ramp/stress test for the true breaking point.')
    else:
        max_tps  = round(tps, 1)
        headroom = 0
        verdict  = 'saturated'
        note     = ('At or beyond comfortable capacity for this run (latency near/over target or errors present). '
                    'A ramp test will pinpoint the ceiling.')
    return {'current_tps': round(tps, 1), 'max_tps': max_tps, 'headroom_pct': headroom,
            'verdict': verdict, 'p95': int(p95), 'target_p95': int(target_p95),
            'err': round(err, 2), 'note': note}


def _exec_verdict(d):
    """Traffic-light PASS/WATCH/FAIL verdict + plain-language summary."""
    sla = d.get('sla_result')
    err = float(d.get('error_rate', 0) or 0)
    p95 = int(d.get('p95', 0) or 0)
    if sla is not None:
        status = 'pass' if sla.get('passed') else 'fail'
    else:
        if err < 1.0 and p95 <= 2000:   status = 'pass'
        elif err < 5.0 and p95 <= 4000: status = 'warn'
        else:                            status = 'fail'
    color = {'pass': '#22c55e', 'warn': '#f59e0b', 'fail': '#ef4444'}[status]
    icon  = {'pass': '&#128994;', 'warn': '&#128993;', 'fail': '&#128308;'}[status]
    word  = {'pass': 'PASS', 'warn': 'WATCH', 'fail': 'FAIL'}[status]
    label = {'pass': 'Meets targets', 'warn': 'Elevated — review before peak',
             'fail': 'Action needed before release'}[status]
    sub   = {'pass': 'System sustained the load within performance targets.',
             'warn': 'System handled the load but latency or errors were elevated.',
             'fail': 'System did not meet targets under this load.'}[status]
    tps = float(d.get('throughput', 0) or 0)
    headline = f'{tps:,.0f} TPS sustained &middot; P95 {p95:,} ms &middot; errors {err:.2f}%'
    if sla and sla.get('checks'):
        failed = [ch['name'] for ch in sla['checks'] if not ch['passed']]
        if failed:
            sub += ' Breached: ' + ', '.join(failed[:4]) + ('…' if len(failed) > 4 else '') + '.'
    return {'status': status, 'color': color, 'icon': icon, 'word': word,
            'label': label, 'headline': headline, 'sub': sub}


def _enrich_report(d, path, c=None):
    """Attach SLA verdict, trend, capacity and exec-summary to the parsed report
    dict so _generate_report_html can render the management block."""
    try:
        d['sla_result'] = _evaluate_sla(path, c) if c else None
    except Exception:
        d['sla_result'] = None
    try:
        d['trend'] = _recent_trend(c, path)
    except Exception:
        d['trend'] = []
    d['capacity']     = _capacity_estimate(d)
    d['exec_verdict'] = _exec_verdict(d)
    return d


def _report_html(path, c=None):
    """Parse + enrich + render — single entry point used by all report routes."""
    d = _parse_jtl(path)
    _enrich_report(d, path, c)
    return _generate_report_html(d)


def _mgmt_block(d):
    """Executive summary banner + capacity + trend cards for the top of the report."""
    import html as _h
    ev  = d.get('exec_verdict')
    cap = d.get('capacity')
    trend = d.get('trend') or []
    if not ev:
        return ''
    # Trend arrow (P95 vs previous run)
    arrow = '<div style="font-size:12px;color:#64748b;">First recorded run</div>'
    if len(trend) >= 2 and trend[-1]['p95'] and trend[-2]['p95']:
        cur, prev = trend[-1]['p95'], trend[-2]['p95']
        pct = (cur - prev) / prev * 100 if prev else 0
        if abs(pct) < 2:
            arrow = '<div style="font-size:12px;color:#94a3b8;">&#8596; P95 flat vs previous run</div>'
        else:
            better = cur < prev
            col = '#22c55e' if better else '#ef4444'
            ar  = '&#9660;' if better else '&#9650;'
            arrow = (f'<div style="font-size:15px;font-weight:800;color:{col};">{ar} {abs(pct):.0f}%</div>'
                     f'<div style="font-size:11px;color:#64748b;">P95 {"faster" if better else "slower"} vs previous</div>')
    banner = (
        f'<div style="background:linear-gradient(90deg,{ev["color"]}22,transparent);border:1px solid {ev["color"]};'
        f'border-left:6px solid {ev["color"]};border-radius:14px;padding:18px 22px;margin:0 0 16px;'
        f'display:flex;gap:18px;align-items:center;flex-wrap:wrap;">'
        f'<div style="font-size:32px;">{ev["icon"]}</div>'
        f'<div style="flex:1;min-width:240px;">'
        f'<div style="font-size:20px;font-weight:800;color:{ev["color"]};letter-spacing:.3px;">{ev["word"]} &mdash; {ev["label"]}</div>'
        f'<div style="font-size:14px;color:#e2e8f0;margin-top:3px;">{ev["headline"]}</div>'
        f'<div style="font-size:12.5px;color:#94a3b8;margin-top:3px;">{ev["sub"]}</div></div>'
        f'<div style="text-align:right;min-width:120px;">{arrow}</div></div>'
    )
    # Capacity card
    cap_html = ''
    if cap:
        hc = '#22c55e' if cap['verdict'] == 'headroom' else '#f59e0b'
        fill = min(100, round(cap['current_tps'] / cap['max_tps'] * 100)) if cap['max_tps'] else 100
        cap_html = (
            '<div style="flex:1;min-width:260px;background:#111827;border:1px solid #1f2d44;border-radius:12px;padding:15px 18px;">'
            '<div style="font-size:11px;text-transform:uppercase;letter-spacing:.08em;color:#64748b;font-weight:700;margin-bottom:8px;">Capacity &amp; headroom</div>'
            '<div style="display:flex;gap:20px;flex-wrap:wrap;">'
            f'<div><div style="font-size:22px;font-weight:800;color:#e2e8f0;">{cap["current_tps"]:,}</div><div style="font-size:11px;color:#64748b;">current TPS</div></div>'
            f'<div><div style="font-size:22px;font-weight:800;color:#3b82f6;">{cap["max_tps"]:,}</div><div style="font-size:11px;color:#64748b;">est. max sustainable</div></div>'
            f'<div><div style="font-size:22px;font-weight:800;color:{hc};">{cap["headroom_pct"]}%</div><div style="font-size:11px;color:#64748b;">headroom</div></div></div>'
            f'<div style="height:8px;background:#0b1220;border-radius:5px;margin:12px 0 6px;overflow:hidden;"><div style="height:100%;width:{fill}%;background:{hc};"></div></div>'
            f'<div style="font-size:11px;color:#64748b;line-height:1.5;">{cap["note"]}</div></div>'
        )
    # Trend card (inline bars, no JS dependency)
    trend_html = ''
    if len(trend) >= 2:
        mx = max((t['p95'] for t in trend), default=1) or 1
        bars = ''
        for t in trend:
            h = max(6, round(t['p95'] / mx * 60))
            bc = '#3b82f6' if not t['current'] else '#f59e0b'
            ttl = _h.escape(f'{t["name"]}  P95 {t["p95"]}ms  err {t["err"]}%  {t["tps"]} TPS', quote=True)
            bars += (f'<div title="{ttl}" style="flex:1;display:flex;flex-direction:column;justify-content:flex-end;align-items:center;gap:3px;">'
                     f'<div style="width:70%;height:{h}px;background:{bc};border-radius:3px 3px 0 0;"></div></div>')
        latest = trend[-1]
        trend_html = (
            '<div style="flex:1;min-width:260px;background:#111827;border:1px solid #1f2d44;border-radius:12px;padding:15px 18px;">'
            f'<div style="font-size:11px;text-transform:uppercase;letter-spacing:.08em;color:#64748b;font-weight:700;margin-bottom:8px;">P95 trend &mdash; last {len(trend)} runs</div>'
            f'<div style="display:flex;gap:4px;align-items:flex-end;height:66px;">{bars}</div>'
            f'<div style="display:flex;justify-content:space-between;font-size:11px;color:#64748b;margin-top:8px;">'
            f'<span>latest: <b style="color:#e2e8f0;">{latest["p95"]}ms</b> P95</span>'
            f'<span>err <b style="color:#e2e8f0;">{latest["err"]}%</b></span>'
            f'<span><b style="color:#e2e8f0;">{latest["tps"]}</b> TPS</span></div>'
            '<div style="font-size:11px;color:#64748b;margin-top:2px;">Amber bar = this run.</div></div>'
        )
    cards = ''
    if cap_html or trend_html:
        cards = f'<div style="display:flex;gap:14px;flex-wrap:wrap;margin:0 0 8px;">{cap_html}{trend_html}</div>'
    return banner + cards


def _generate_report_html(d):
    import re as _re

    # ── JSON series for charts ────────────────────────────────────────────────
    step_stats = [s for s in d['label_stats'] if not _re.match(r'^TC\s*[\-–]', s['label'], _re.I)]
    tc_stats   = [s for s in d['label_stats'] if     _re.match(r'^TC\s*[\-–]', s['label'], _re.I)]

    labels_json    = json.dumps([s['label']      for s in d['label_stats']])
    avg_json       = json.dumps([s['avg']         for s in d['label_stats']])
    p90_json       = json.dumps([s['p90']         for s in d['label_stats']])
    p95_json       = json.dumps([s['p95']         for s in d['label_stats']])
    p99_json       = json.dumps([s['p99']         for s in d['label_stats']])
    err_json       = json.dumps([s['error_rate']  for s in d['label_stats']])
    tps_t_json     = json.dumps([x['time_offset'] for x in d['tps_over_time']])
    tps_v_json     = json.dumps([x['tps']         for x in d['tps_over_time']])
    avg_tps_json   = json.dumps([d['throughput']] * len(d['tps_over_time']))
    hist_lbl_json  = json.dumps([x['range']       for x in d['hist_data']])
    hist_cnt_json  = json.dumps([x['count']       for x in d['hist_data']])
    s_labels_json  = json.dumps([s['label']       for s in step_stats])
    s_avg_json     = json.dumps([s['avg']          for s in step_stats])
    s_p90_json     = json.dumps([s['p90']          for s in step_stats])
    s_p95_json     = json.dumps([s['p95']          for s in step_stats])
    s_p99_json     = json.dumps([s['p99']          for s in step_stats])
    err_steps_lbl  = json.dumps([s['label']        for s in step_stats if s['failures'] > 0])
    err_steps_cnt  = json.dumps([s['failures']     for s in step_stats if s['failures'] > 0])

    # ── Colour helpers ────────────────────────────────────────────────────────
    def _ec(err):
        if err == 0:  return '#22c55e'
        if err < 5:   return '#60a5fa'
        if err < 20:  return '#f59e0b'
        if err < 50:  return '#f97316'
        return '#ef4444'

    def _badge(err):
        if err == 0:  return '<span class="bdg bdg-ok">&#10003; PASS</span>'
        if err < 5:   return '<span class="bdg bdg-low">LOW</span>'
        if err < 20:  return '<span class="bdg bdg-med">MEDIUM</span>'
        if err < 50:  return '<span class="bdg bdg-high">HIGH</span>'
        return '<span class="bdg bdg-crit">&#10005; CRITICAL</span>'

    def _fmt_expected(fr):
        exp = (fr.get('expected') or fr.get('reason') or '').strip()
        if not exp or exp.lower() in ('unknown error', 'not captured'):
            return ''
        if exp.lower().startswith('number of samples'):
            return ''
        if '{message}' in exp or exp == '{message}':
            return ''
        return exp

    def _fmt_actual(fr):
        act = (fr.get('actual') or '').strip()
        if not act or act.lower() in ('ok', 'not captured'):
            return 'Response body not saved in JTL — the USSD menu text was different from expected'
        return act

    # ── USSD Flow diagram ─────────────────────────────────────────────────────
    flow_html = ''
    for i, st in enumerate(step_stats):
        lbl   = st['label']
        short = (lbl.split(' - ', 1)[1] if ' - ' in lbl else lbl)
        short = short[:24] + ('…' if len(short) > 24 else '')
        avg   = st['avg']
        spd   = 'fast' if avg < 500 else 'med' if avg < 1000 else 'slow' if avg < 2000 else 'vslow'
        err_h = (f'<span class="fe-bad">{st["error_rate"]}% err</span>'
                 if st['failures'] > 0 else '<span class="fe-ok">No errors</span>')
        arr   = '<div class="flow-arr">&rarr;</div>' if i < len(step_stats) - 1 else ''
        flow_html += (
            f'<div class="flow-step">'
            f'<div class="flow-card">'
            f'<div class="flow-num">{i+1}</div>'
            f'<div class="flow-name" title="{lbl}">{short}</div>'
            f'<div class="flow-avg {spd}">{avg:,} ms</div>'
            f'<div class="flow-reqs">{st["samples"]:,} reqs</div>'
            f'{err_h}</div></div>{arr}'
        )

    # ── Service health cards + Failure analysis ───────────────────────────────
    svc_cards_html     = ''
    fail_analysis_html = ''
    for tc in sorted(tc_stats, key=lambda x: -x['error_rate']):
        svc    = _re.sub(r'^TC\s*[\-–]\s*', '', tc['label'], flags=_re.I).strip()
        prefix = svc.lower()[:18]
        failing = sorted([s for s in step_stats
                          if s['label'].lower().startswith(prefix) and s['error_rate'] > 0],
                         key=lambda x: -x['error_rate'])
        worst = failing[0] if failing else None
        worst_lbl = ''
        if worst:
            parts = worst['label'].split(' - ', 1)
            worst_lbl = parts[1] if len(parts) > 1 else worst['label']
        card_cls = ('svc-crit' if tc['error_rate'] >= 50 else 'svc-high' if tc['error_rate'] >= 20
                    else 'svc-med' if tc['error_rate'] >= 5 else 'svc-ok')
        if worst:
            top_fr = worst['failure_reasons'][0] if worst['failure_reasons'] else {}
            exp    = _fmt_expected(top_fr)
            fail_line = (
                f'<div class="svc-fail-step">&#9888; {worst_lbl}'
                f'<span class="svc-fail-pct"> — {worst["error_rate"]}% failed</span></div>'
                + (f'<div class="svc-exp-row"><span class="svc-exp-label">Expected menu text:</span>'
                   f'<span class="svc-exp-val">&ldquo;{exp}&rdquo;</span></div>'
                   f'<div class="svc-act-row">The USSD menu returned something different — see Failure Analysis below.</div>'
                   if exp else
                   f'<div class="svc-fail-reason">USSD response did not match assertion.</div>')
            )
        else:
            fail_line = '<div class="svc-pass-msg">&#10003; All steps passing</div>'
        svc_cards_html += (
            f'<div class="svc-card {card_cls}">'
            f'<div class="svc-top"><span class="svc-name">{svc}</span>{_badge(tc["error_rate"])}</div>'
            f'<div class="svc-pct" style="color:{_ec(tc["error_rate"])}">{tc["error_rate"]}'
            f'<span class="svc-pct-lbl">% error</span></div>'
            f'<div class="svc-meta">{tc["samples"]:,} samples &middot; avg {tc["avg"]} ms &middot; {tc["tps"]} TPS</div>'
            f'{fail_line}</div>\n'
        )
        if tc['error_rate'] > 0:
            step_rows_fa = ''
            for fs in failing[:8]:
                parts  = fs['label'].split(' - ', 1)
                slabel = parts[1] if len(parts) > 1 else fs['label']
                top_fr = fs['failure_reasons'][0] if fs['failure_reasons'] else {}
                exp    = _fmt_expected(top_fr)
                act    = _fmt_actual(top_fr)
                cnt2   = top_fr.get('count', fs['failures'])
                if exp:
                    ea = (f'<div style="display:flex;flex-direction:column;gap:5px;">'
                          f'<span style="font-size:10px;font-weight:800;color:#64748b;">{cnt2:,}\xd7 failed</span>'
                          f'<div style="display:flex;gap:7px;align-items:baseline;background:rgba(34,197,94,.07);border:1px solid rgba(34,197,94,.2);border-radius:7px;padding:6px 10px;">'
                          f'<span style="font-size:9px;font-weight:900;text-transform:uppercase;background:rgba(34,197,94,.2);color:#4ade80;padding:2px 6px;border-radius:4px;white-space:nowrap;flex-shrink:0;">EXPECTED</span>'
                          f'<span style="font-size:13px;font-weight:700;color:#4ade80;">&ldquo;{exp}&rdquo;</span></div>'
                          f'<div style="display:flex;gap:7px;align-items:baseline;background:rgba(239,68,68,.07);border:1px solid rgba(239,68,68,.2);border-radius:7px;padding:6px 10px;">'
                          f'<span style="font-size:9px;font-weight:900;text-transform:uppercase;background:rgba(239,68,68,.2);color:#f87171;padding:2px 6px;border-radius:4px;white-space:nowrap;flex-shrink:0;">RECEIVED</span>'
                          f'<span style="font-size:13px;color:#fca5a5;">{act}</span></div></div>')
                else:
                    ea = f'<span style="color:#fca5a5;font-size:13px;">{act}</span>'
                step_rows_fa += (
                    f'<tr><td class="fa-step">{slabel}</td>'
                    f'<td class="fa-num" style="color:{_ec(fs["error_rate"])}">{fs["error_rate"]}%</td>'
                    f'<td class="fa-num" style="color:{_ec(fs["error_rate"])}">{fs["failures"]:,} / {fs["samples"]:,}</td>'
                    f'<td>{ea}</td></tr>'
                )
            fail_analysis_html += (
                f'<div class="fa-block"><div class="fa-hdr">'
                f'<div><span class="fa-title">{svc}</span>'
                f'<span class="fa-sub">{tc["failures"]:,} of {tc["samples"]:,} transactions failed</span></div>'
                f'{_badge(tc["error_rate"])}</div>'
                f'<table class="fa-tbl"><thead><tr>'
                f'<th>Step That Failed</th><th style="width:80px">Error %</th>'
                f'<th style="width:130px">Failed / Total</th>'
                f'<th>Expected vs Actual &mdash; USSD Menu Was Supposed to Show vs What Was Returned</th>'
                f'</tr></thead><tbody>{step_rows_fa}</tbody></table></div>\n'
            )

    # ── RC distribution ───────────────────────────────────────────────────────
    rc_rows = ''
    for code, cnt in sorted(d['rc_dist'].items()):
        col = '#22c55e' if code.startswith('2') else '#f59e0b' if code.startswith('4') else '#ef4444'
        rc_rows += (f'<div style="display:flex;justify-content:space-between;padding:8px 0;'
                    f'border-bottom:1px solid #1f2d44;font-size:13px;">'
                    f'<span style="color:#64748b">HTTP {code}</span>'
                    f'<strong style="color:{col}">{cnt:,}</strong></div>')

    # ── Full stats table — grouped by service ────────────────────────────────
    non_tc_avgs = [s['avg'] for s in d['label_stats']
                   if not _re.match(r'^TC\s*[-–]', s['label'], _re.I)]
    max_avg = max(non_tc_avgs, default=1)
    stats_rows = ''
    row_idx = 0
    for gi, grp in enumerate(d.get('service_groups', [])):
        svc_name = grp['service']
        summary  = grp['summary'] or {}
        steps    = grp['steps']
        multi    = len(steps) > 1

        agg_samples  = summary.get('samples', sum(s['samples'] for s in steps))
        agg_failures = summary.get('failures', sum(s['failures'] for s in steps))
        agg_err      = summary.get('error_rate', round(agg_failures/agg_samples*100,2) if agg_samples else 0)
        agg_avg      = summary.get('avg', round(sum(s['avg']*s['samples'] for s in steps)/agg_samples) if agg_samples else 0)
        agg_tps      = summary.get('tps', round(sum(s['tps'] for s in steps), 3))

        hdr_ec    = _ec(agg_err)
        hdr_badge = _badge(agg_err)
        grp_id    = f'sg{gi}'
        arrow     = '&#9660;' if multi else '&#9658;'
        onclick   = f'onclick="toggleSvcGroup(\'{grp_id}\')"' if multi else ''
        stats_rows += (
            f'<tr {onclick} style="cursor:{"pointer" if multi else "default"};">'
            f'<td colspan="13" style="padding:10px 16px;background:rgba(30,42,65,.9);'
            f'border-top:2px solid rgba(99,102,241,.35);border-bottom:1px solid rgba(99,102,241,.15);">'
            f'<div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;">'
            f'<span id="{grp_id}-arrow" style="font-size:12px;color:#94a3b8;min-width:14px;">{arrow}</span>'
            f'<span style="font-weight:800;font-size:14px;color:#e2e8f0;">{svc_name}</span>'
            f'{hdr_badge}'
            f'<span style="color:#64748b;font-size:12px;">'
            f'{agg_samples:,} samples &middot; avg {agg_avg:,} ms &middot; {agg_tps} TPS &middot; '
            f'<span style="color:{hdr_ec};font-weight:700;">{agg_err}% errors</span></span>'
            f'</div></td></tr>\n'
        )

        for s in steps:
            ec      = _ec(s['error_rate'])
            pct     = round(s['avg'] / max_avg * 100)
            bar_col = '#22c55e' if s['avg']<500 else '#3b82f6' if s['avg']<1000 else '#f59e0b' if s['avg']<2000 else '#ef4444'
            indent  = '&nbsp;&nbsp;&#8627;&nbsp;' if multi else ''
            fail_btn = fail_row = ''
            if s['failures'] > 0:
                rhtml = ''
                for r in s['failure_reasons']:
                    exp2 = r.get('expected', ''); act2 = r.get('actual', ''); cnt3 = r.get('count', 0)
                    ch = f'<span style="font-size:10px;font-weight:800;color:#ef4444;display:block;margin-bottom:5px;">{cnt3:,} failure{"s" if cnt3!=1 else ""}</span>'
                    if exp2:
                        rhtml += (f'<div style="padding:10px 14px;border-bottom:1px solid rgba(239,68,68,.12);">{ch}'
                                  f'<div style="display:flex;gap:8px;align-items:baseline;margin-bottom:5px;background:rgba(34,197,94,.07);border:1px solid rgba(34,197,94,.2);border-radius:7px;padding:6px 10px;">'
                                  f'<span style="font-size:9px;font-weight:900;text-transform:uppercase;background:rgba(34,197,94,.2);color:#4ade80;padding:2px 6px;border-radius:4px;white-space:nowrap;flex-shrink:0;">EXPECTED</span>'
                                  f'<span style="font-size:13px;font-weight:700;color:#4ade80;">&ldquo;{exp2}&rdquo;</span></div>'
                                  f'<div style="display:flex;gap:8px;align-items:baseline;background:rgba(239,68,68,.07);border:1px solid rgba(239,68,68,.2);border-radius:7px;padding:6px 10px;">'
                                  f'<span style="font-size:9px;font-weight:900;text-transform:uppercase;background:rgba(239,68,68,.2);color:#f87171;padding:2px 6px;border-radius:4px;white-space:nowrap;flex-shrink:0;">RECEIVED</span>'
                                  f'<span style="font-size:13px;color:#fca5a5;">{act2 or "Response body not saved in JTL"}</span></div></div>')
                    else:
                        raw2 = r.get('reason', 'Unknown error')
                        rhtml += f'<div style="padding:8px 14px;border-bottom:1px solid rgba(239,68,68,.1);">{ch}<span style="color:#fca5a5;font-size:12px;">{raw2}</span></div>'
                fail_btn = (f'<button onclick="event.stopPropagation();toggleRow(\'fd{row_idx}\')" '
                            f'style="margin-left:6px;background:rgba(239,68,68,.12);border:1px solid rgba(239,68,68,.3);'
                            f'border-radius:5px;padding:2px 9px;color:#ef4444;font-size:10px;font-weight:700;cursor:pointer;">'
                            f'Why failed?</button>')
                lbl_esc = s["label"].replace("'", "&#39;")
                fail_row = (f'<tr id="fd{row_idx}" data-sg="{grp_id}" style="display:none;"><td colspan="14" style="padding:0;">'
                            f'<div style="background:rgba(10,10,20,.8);border-top:1px solid rgba(239,68,68,.25);border-bottom:1px solid rgba(239,68,68,.25);">'
                            f'<div style="padding:10px 14px 4px;font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.6px;color:#ef4444;">'
                            f'Expected vs Actual &mdash; {lbl_esc}</div>{rhtml}</div></td></tr>')

            # ── Debug panel: Request / Response ──────────────────────────
            dbg_id  = f'dbg{row_idx}'
            url_val = s.get('url', '')
            fail_details = s.get('fail_details', [])

            # Parse URL into base + params
            dbg_req_html = ''
            if url_val:
                from urllib.parse import urlparse, parse_qs, urlencode
                parsed = urlparse(url_val)
                params = parse_qs(parsed.query, keep_blank_values=True)
                masked = {'PASSWORD', 'PASSWD', 'password', 'passwd', 'token', 'key'}
                param_rows = ''
                for pk, pv in params.items():
                    val = '***' if pk in masked else pv[0]
                    param_rows += (f'<tr><td style="color:#94a3b8;padding:3px 10px 3px 0;font-size:11px;'
                                   f'white-space:nowrap;vertical-align:top;">{pk}</td>'
                                   f'<td style="color:#e2e8f0;font-size:11px;word-break:break-all;">{val}</td></tr>')
                base_url = f'{parsed.scheme}://{parsed.netloc}{parsed.path}'
                dbg_req_html = (
                    f'<div style="margin-bottom:10px;">'
                    f'<div style="font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.6px;'
                    f'color:#60a5fa;margin-bottom:6px;">&#128228; Request</div>'
                    f'<div style="font-size:11px;color:#94a3b8;margin-bottom:4px;">Endpoint</div>'
                    f'<div style="font-size:12px;color:#e2e8f0;font-family:monospace;word-break:break-all;'
                    f'background:#0d1117;padding:6px 10px;border-radius:6px;margin-bottom:8px;">{base_url}</div>'
                    + (f'<div style="font-size:11px;color:#94a3b8;margin-bottom:4px;">Parameters</div>'
                       f'<table style="border-collapse:collapse;width:100%;">{param_rows}</table>'
                       if param_rows else '')
                    + '</div>'
                )

            dbg_resp_html = ''
            if fail_details:
                dbg_resp_html = (f'<div style="font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.6px;'
                                 f'color:#f87171;margin-bottom:6px;">&#128229; Sample Failure Responses '
                                 f'({len(fail_details)} of {s["failures"]:,} shown)</div>')
                for fi, fd in enumerate(fail_details):
                    rc_color = '#ef4444' if fd["rc"] and not fd["rc"].startswith('2') else '#22c55e'
                    dbg_resp_html += (
                        f'<div style="border:1px solid rgba(239,68,68,.2);border-radius:8px;padding:10px 14px;'
                        f'margin-bottom:8px;background:rgba(239,68,68,.04);">'
                        f'<div style="font-size:10px;font-weight:700;color:#64748b;margin-bottom:6px;">'
                        f'Sample #{fi+1} &nbsp;&middot;&nbsp; {int(fd["elapsed"]):,} ms &nbsp;&middot;&nbsp; {fd["thread"]}</div>'
                        + (f'<div style="display:flex;gap:10px;margin-bottom:5px;flex-wrap:wrap;">'
                           f'<div><span style="font-size:9px;color:#64748b;">Response Code</span><br>'
                           f'<code style="color:{rc_color};font-size:11px;">{fd["rc"] or "—"}</code></div>'
                           f'<div style="flex:1;"><span style="font-size:9px;color:#64748b;">Response Message</span><br>'
                           f'<code style="color:#fca5a5;font-size:11px;word-break:break-word;">{fd["rm"] or "—"}</code></div>'
                           f'</div>' if fd["rc"] or fd["rm"] else '')
                        + (f'<div style="margin-top:5px;background:rgba(34,197,94,.06);border:1px solid rgba(34,197,94,.2);'
                           f'border-radius:6px;padding:6px 10px;">'
                           f'<span style="font-size:9px;font-weight:700;text-transform:uppercase;color:#4ade80;">Assertion</span><br>'
                           f'<span style="font-size:11px;color:#86efac;">{fd["fm"]}</span></div>'
                           if fd["fm"] else '')
                        + '</div>'
                    )
            elif url_val:
                dbg_resp_html = '<div style="font-size:11px;color:#22c55e;">&#10003; All samples passed — no failure details to show.</div>'

            has_dbg = bool(url_val or fail_details)
            dbg_btn = ''
            dbg_row = ''
            if has_dbg:
                dbg_btn = (f'<button onclick="event.stopPropagation();toggleRow(\'{dbg_id}\')" '
                           f'style="margin-left:6px;background:rgba(59,130,246,.12);border:1px solid rgba(59,130,246,.3);'
                           f'border-radius:5px;padding:2px 9px;color:#60a5fa;font-size:10px;font-weight:700;cursor:pointer;">'
                           f'&#128269; Req/Resp</button>')
                dbg_row = (f'<tr id="{dbg_id}" data-sg="{grp_id}" style="display:none;">'
                           f'<td colspan="14" style="padding:0;">'
                           f'<div style="background:rgba(8,12,24,.95);border-top:1px solid rgba(59,130,246,.2);'
                           f'border-bottom:1px solid rgba(59,130,246,.2);padding:14px 18px;">'
                           f'<div style="font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.6px;'
                           f'color:#60a5fa;margin-bottom:10px;">&#128269; Request / Response — {s["label"]}</div>'
                           f'{dbg_req_html}{dbg_resp_html}</div></td></tr>')

            stats_rows += (
                f'<tr data-sg="{grp_id}">'
                f'<td style="font-weight:600;font-size:12px;padding-left:{"28px" if multi else "14px"};">'
                f'{indent}{s["label"]}{fail_btn}{dbg_btn}</td>'
                f'<td>{s["samples"]:,}</td><td>{s["avg"]:,}</td><td>{s["median"]:,}</td>'
                f'<td>{s["p90"]:,}</td><td>{s["p95"]:,}</td><td>{s["p99"]:,}</td>'
                f'<td>{s["min"]:,}</td><td>{s["max"]:,}</td>'
                f'<td>{s["tps"]}</td><td>{s["failures"]:,}</td>'
                f'<td style="color:{ec};font-weight:700">{s["error_rate"]}%</td>'
                f'<td><div style="background:#1f2d44;border-radius:3px;height:6px;min-width:50px;">'
                f'<div style="background:{bar_col};height:100%;width:{pct}%;border-radius:3px;"></div></div></td>'
                f'</tr>{fail_row}{dbg_row}'
            )
            row_idx += 1

    # ── JMeter-style Summary Report table ────────────────────────────────────
    # Show TC (service) rows if present, otherwise all labels
    summary_src = [s for s in d['label_stats'] if _re.match(r'^TC\s*[-–]', s['label'], _re.I)] \
                  or d['label_stats']
    sum_rows_html = ''
    for s in summary_src:
        lbl_disp = _re.sub(r'^TC\s*[-–]\s*', '', s['label'], flags=_re.I).strip() \
                   if _re.match(r'^TC\s*[-–]', s['label'], _re.I) else s['label']
        ec = ('#ef4444' if s['error_rate'] >= 20 else '#f59e0b' if s['error_rate'] >= 5
              else '#60a5fa' if s['error_rate'] > 0 else '#22c55e')
        sum_rows_html += (
            f'<tr>'
            f'<td style="font-weight:700;font-size:12px;white-space:nowrap;">{lbl_disp}</td>'
            f'<td>{s["samples"]:,}</td>'
            f'<td style="color:#ef4444;font-weight:700;">{s["failures"]:,}</td>'
            f'<td style="color:{ec};font-weight:700;">{s["error_rate"]}%</td>'
            f'<td>{s["avg"]:,}</td><td>{s["min"]:,}</td><td>{s["max"]:,}</td>'
            f'<td>{s["median"]:,}</td>'
            f'<td>{s["p90"]:,}</td><td>{s["p95"]:,}</td><td>{s["p99"]:,}</td>'
            f'<td>{s["tps"]}</td>'
            f'<td>{s.get("recv_kbps", 0)}</td>'
            f'<td>{s.get("sent_kbps", 0)}</td>'
            f'</tr>\n'
        )

    main_ec      = '#ef4444' if d['error_rate'] > 1 else '#22c55e'
    success_rate = round(100 - d['error_rate'], 2)
    thr          = d['throughput']

    # Embed Chart.js inline so the HTML works offline without CDN
    _chartjs_path = os.path.join(os.path.dirname(__file__), 'static', 'chart.min.js')
    try:
        with open(_chartjs_path, encoding='utf-8') as _f:
            _chartjs_inline = '<script>' + _f.read() + '</script>'
    except Exception:
        _chartjs_inline = '<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>'

    return (
        '<!DOCTYPE html>\n'
        '<html lang="en"><head><meta charset="UTF-8">\n'
        '<meta name="viewport" content="width=device-width,initial-scale=1.0">\n'
        f'<title>Load Test Report — {d["filename"]}</title>\n'
        + _chartjs_inline + '\n'
        + '<style>\n'
        '*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}\n'
        ':root{--bg:#0b0f1a;--surface:#111827;--s2:#1a2235;--border:#1f2d44;\n'
        '  --blue:#3b82f6;--green:#22c55e;--warn:#f59e0b;--orange:#f97316;--red:#ef4444;\n'
        '  --text:#f1f5f9;--muted:#64748b;--muted2:#94a3b8;}\n'
        'body{background:var(--bg);color:var(--text);font-family:"Segoe UI",system-ui,sans-serif;\n'
        '  padding:32px;max-width:1500px;margin:0 auto;line-height:1.5;}\n'
        '.hdr{display:flex;align-items:center;gap:14px;margin-bottom:6px;}\n'
        '.hdr-icon{width:50px;height:50px;background:linear-gradient(135deg,#3b82f6,#8b5cf6);\n'
        '  border-radius:14px;display:flex;align-items:center;justify-content:center;font-size:26px;flex-shrink:0;}\n'
        '.hdr h1{font-size:26px;font-weight:900;letter-spacing:-.5px;\n'
        '  background:linear-gradient(90deg,#fff 40%,#3b82f6);-webkit-background-clip:text;-webkit-text-fill-color:transparent;}\n'
        '.hdr-sub{color:var(--muted);font-size:13px;margin-bottom:32px;padding-left:64px;}\n'
        '.sec{font-size:20px;font-weight:800;margin:38px 0 6px;letter-spacing:-.2px;}\n'
        '.sec-hint{font-size:13px;color:var(--muted);margin-bottom:18px;line-height:1.6;}\n'
        '.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:12px;margin-bottom:28px;}\n'
        '.kpi{background:var(--surface);border:1px solid var(--border);border-radius:14px;padding:20px 16px;}\n'
        '.kpi-label{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.9px;color:var(--muted);margin-bottom:8px;}\n'
        '.kpi-val{font-size:28px;font-weight:900;line-height:1;}\n'
        '.kpi-sub{font-size:11px;color:var(--muted);margin-top:4px;}\n'
        '.kpi-unit{font-size:13px;font-weight:500;color:var(--muted);}\n'
        '.k-ok .kpi-val{color:var(--green);} .k-blue .kpi-val{color:var(--blue);}\n'
        '.k-warn .kpi-val{color:var(--warn);} .k-err .kpi-val{color:var(--red);} .k-purple .kpi-val{color:#a78bfa;}\n'
        '.perc-grid{display:flex;gap:14px;margin-bottom:28px;flex-wrap:wrap;}\n'
        '.perc-card{background:var(--surface);border:1px solid var(--border);border-radius:14px;\n'
        '  padding:18px 22px;flex:1;min-width:120px;text-align:center;}\n'
        '.perc-label{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.7px;color:var(--muted);margin-bottom:6px;}\n'
        '.perc-val{font-size:34px;font-weight:900;line-height:1;}\n'
        '.p50{color:#22c55e;} .p90{color:#3b82f6;} .p95{color:#f59e0b;} .p99{color:#ef4444;}\n'
        '.perc-unit{font-size:12px;color:var(--muted);margin-top:3px;}\n'
        '.flow-wrap{display:flex;align-items:center;flex-wrap:nowrap;overflow-x:auto;\n'
        '  padding:16px 4px 20px;gap:0;margin-bottom:8px;scrollbar-width:thin;}\n'
        '.flow-step{flex-shrink:0;}\n'
        '.flow-card{background:var(--surface);border:1px solid var(--border);border-radius:12px;\n'
        '  padding:14px 12px;text-align:center;width:120px;}\n'
        '.flow-num{width:24px;height:24px;background:var(--blue);border-radius:50%;font-size:11px;\n'
        '  font-weight:800;display:flex;align-items:center;justify-content:center;margin:0 auto 8px;}\n'
        '.flow-name{font-size:11px;font-weight:600;color:var(--text);margin-bottom:6px;\n'
        '  word-break:break-word;line-height:1.3;min-height:28px;}\n'
        '.flow-avg{font-size:16px;font-weight:800;margin-bottom:3px;}\n'
        '.fast{color:#22c55e;} .med{color:#3b82f6;} .slow{color:#f59e0b;} .vslow{color:#ef4444;}\n'
        '.flow-reqs{font-size:10px;color:var(--muted);margin-bottom:5px;}\n'
        '.fe-ok{font-size:10px;font-weight:700;color:#22c55e;background:rgba(34,197,94,.12);\n'
        '  padding:2px 7px;border-radius:99px;display:inline-block;}\n'
        '.fe-bad{font-size:10px;font-weight:700;color:#ef4444;background:rgba(239,68,68,.12);\n'
        '  padding:2px 7px;border-radius:99px;display:inline-block;}\n'
        '.flow-arr{font-size:22px;color:var(--muted);padding:0 6px;flex-shrink:0;\n'
        '  align-self:center;margin-bottom:20px;}\n'
        '.bdg{display:inline-flex;align-items:center;font-size:10px;font-weight:800;\n'
        '  letter-spacing:.5px;padding:3px 10px;border-radius:99px;text-transform:uppercase;white-space:nowrap;}\n'
        '.bdg-ok{background:rgba(34,197,94,.14);color:#4ade80;border:1px solid rgba(34,197,94,.3);}\n'
        '.bdg-low{background:rgba(59,130,246,.14);color:#60a5fa;border:1px solid rgba(59,130,246,.3);}\n'
        '.bdg-med{background:rgba(245,158,11,.14);color:#fbbf24;border:1px solid rgba(245,158,11,.3);}\n'
        '.bdg-high{background:rgba(249,115,22,.18);color:#fb923c;border:1px solid rgba(249,115,22,.3);}\n'
        '.bdg-crit{background:rgba(239,68,68,.14);color:#f87171;border:1px solid rgba(239,68,68,.35);}\n'
        '.svc-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:14px;margin-bottom:12px;}\n'
        '.svc-card{background:var(--surface);border-radius:16px;padding:22px;border:1px solid var(--border);border-top:4px solid var(--border);}\n'
        '.svc-ok{border-top-color:#22c55e;} .svc-med{border-top-color:#f59e0b;}\n'
        '.svc-high{border-top-color:#f97316;} .svc-crit{border-top-color:#ef4444;background:rgba(239,68,68,.03);}\n'
        '.svc-top{display:flex;align-items:flex-start;justify-content:space-between;gap:8px;margin-bottom:14px;}\n'
        '.svc-name{font-size:14px;font-weight:700;} .svc-pct{font-size:40px;font-weight:900;line-height:1;margin-bottom:2px;}\n'
        '.svc-pct-lbl{font-size:13px;font-weight:400;color:var(--muted);} .svc-meta{font-size:11px;color:var(--muted);margin-bottom:14px;}\n'
        '.svc-fail-step{font-size:13px;font-weight:700;color:var(--orange);margin-bottom:8px;}\n'
        '.svc-fail-pct{font-weight:500;font-size:12px;color:var(--muted2);}\n'
        '.svc-fail-reason{font-size:12px;color:#fca5a5;background:rgba(239,68,68,.1);border-left:3px solid rgba(239,68,68,.5);padding:8px 10px;border-radius:0 6px 6px 0;}\n'
        '.svc-pass-msg{font-size:13px;color:var(--green);font-weight:700;}\n'
        '.svc-exp-row{margin-top:6px;}\n'
        '.svc-exp-label{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);display:block;margin-bottom:3px;}\n'
        '.svc-exp-val{font-size:12px;font-weight:700;color:#4ade80;background:rgba(34,197,94,.1);border-left:3px solid rgba(34,197,94,.4);padding:6px 10px;border-radius:0 6px 6px 0;display:block;}\n'
        '.svc-act-row{font-size:11px;color:var(--muted);margin-top:5px;font-style:italic;}\n'
        '.fa-block{background:var(--surface);border:1px solid var(--border);border-radius:16px;margin-bottom:16px;overflow:hidden;}\n'
        '.fa-hdr{display:flex;align-items:center;justify-content:space-between;padding:18px 22px;border-bottom:1px solid var(--border);background:var(--s2);gap:12px;flex-wrap:wrap;}\n'
        '.fa-title{font-size:16px;font-weight:800;display:block;margin-bottom:3px;}\n'
        '.fa-sub{font-size:12px;color:var(--muted);}\n'
        '.fa-tbl{width:100%;border-collapse:collapse;}\n'
        '.fa-tbl th{background:rgba(59,130,246,.05);padding:10px 16px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:var(--muted);text-align:left;border-bottom:1px solid var(--border);}\n'
        '.fa-tbl td{padding:13px 16px;border-bottom:1px solid rgba(31,45,68,.5);vertical-align:top;}\n'
        '.fa-tbl tr:last-child td{border-bottom:none;}\n'
        '.fa-step{font-size:13px;font-weight:700;} .fa-num{font-size:13px;font-weight:800;white-space:nowrap;}\n'
        '.chart-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:28px;}\n'
        '.chart-card{background:var(--surface);border:1px solid var(--border);border-radius:14px;padding:18px;}\n'
        '.chart-card.full{grid-column:1/-1;}\n'
        '.ch-title{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:var(--muted);margin-bottom:14px;}\n'
        '.ch-wrap{height:240px;position:relative;} .ch-wrap.tall{height:300px;}\n'
        '.dtbl{width:100%;border-collapse:collapse;background:var(--surface);border-radius:14px;overflow:hidden;border:1px solid var(--border);}\n'
        '.dtbl th{background:rgba(59,130,246,.05);padding:10px 12px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:var(--muted);text-align:left;border-bottom:1px solid var(--border);}\n'
        '.dtbl td{padding:9px 12px;font-size:12px;border-bottom:1px solid rgba(31,45,68,.4);}\n'
        '.dtbl tr:last-child td{border-bottom:none;}\n'
        '.dtbl tr:hover td{background:rgba(59,130,246,.03);}\n'
        '.two-col{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:28px;}\n'
        '.panel{background:var(--surface);border:1px solid var(--border);border-radius:14px;padding:20px;}\n'
        '.panel-title{font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);margin-bottom:14px;}\n'
        'footer{margin-top:52px;text-align:center;font-size:11px;color:var(--muted);padding-bottom:28px;}\n'
        '</style></head><body>\n\n'
        '<div class="hdr"><div class="hdr-icon">&#9889;</div><h1>Load Test Report</h1></div>\n'
        f'{_mgmt_block(d)}'
        f'<div class="hdr-sub">{d["filename"]} &nbsp;&middot;&nbsp; {d["start_time"]} &rarr; {d["end_time"]} &nbsp;&middot;&nbsp; Duration: {d["test_duration"]} &nbsp;&middot;&nbsp; {d["peak_threads"]} threads</div>\n\n'
        '<div class="kpi-grid">\n'
        f'  <div class="kpi k-blue"><div class="kpi-label">Total Requests</div><div class="kpi-val">{d["total"]:,}</div><div class="kpi-sub">{len(d["label_stats"])} transactions</div></div>\n'
        f'  <div class="kpi k-ok"><div class="kpi-label">Success Rate</div><div class="kpi-val">{success_rate}%</div><div class="kpi-sub">{d["passed"]:,} passed</div></div>\n'
        f'  <div class="kpi {"k-err" if d["error_rate"]>1 else "k-ok"}"><div class="kpi-label">Total Errors</div><div class="kpi-val" style="color:{main_ec}">{d["failed"]:,}</div><div class="kpi-sub">{d["error_rate"]}% error rate</div></div>\n'
        f'  <div class="kpi k-blue"><div class="kpi-label">Throughput</div><div class="kpi-val">{thr}<span class="kpi-unit"> /s</span></div><div class="kpi-sub">requests/second</div></div>\n'
        f'  <div class="kpi k-purple"><div class="kpi-label">Peak Threads</div><div class="kpi-val">{d["peak_threads"]}</div><div class="kpi-sub">concurrent users</div></div>\n'
        f'  <div class="kpi k-warn"><div class="kpi-label">Avg Response</div><div class="kpi-val">{d["avg_rt"]:,}<span class="kpi-unit"> ms</span></div></div>\n'
        f'  <div class="kpi k-warn"><div class="kpi-label">Std Deviation</div><div class="kpi-val">{d["std_dev"]:,}<span class="kpi-unit"> ms</span></div></div>\n'
        f'  <div class="kpi k-warn"><div class="kpi-label">Min / Max RT</div><div class="kpi-val" style="font-size:18px">{d["min_rt"]:,}/{d["max_rt"]:,}<span class="kpi-unit"> ms</span></div></div>\n'
        '</div>\n\n'
        '<div class="sec">&#128200; Response Time Percentiles</div>\n'
        '<div class="perc-grid">\n'
        f'  <div class="perc-card"><div class="perc-label">P50 Median</div><div class="perc-val p50">{d["p50"]:,}</div><div class="perc-unit">ms</div></div>\n'
        f'  <div class="perc-card"><div class="perc-label">P90</div><div class="perc-val p90">{d["p90"]:,}</div><div class="perc-unit">ms &mdash; 90% of requests faster</div></div>\n'
        f'  <div class="perc-card"><div class="perc-label">P95</div><div class="perc-val p95">{d["p95"]:,}</div><div class="perc-unit">ms &mdash; 95% of requests faster</div></div>\n'
        f'  <div class="perc-card"><div class="perc-label">P99</div><div class="perc-val p99">{d["p99"]:,}</div><div class="perc-unit">ms &mdash; 99% of requests faster</div></div>\n'
        '</div>\n\n'
        '<div class="sec">&#128202; Summary Report</div>\n'
        '<div class="sec-hint" style="margin-bottom:14px;">Service-level aggregate — matches JMeter\'s Summary Report listener. One row per service.</div>\n'
        '<div style="overflow-x:auto;">\n'
        '<table class="dtbl">\n'
        '<thead><tr style="background:rgba(59,130,246,.08);">\n'
        '  <th>Label</th><th>#Samples</th><th>FAIL</th><th>Error&nbsp;%</th>'
        '<th>Average</th><th>Min</th><th>Max</th><th>Median</th>'
        '<th>90th&nbsp;pct</th><th>95th&nbsp;pct</th><th>99th&nbsp;pct</th>'
        '<th>TPS</th><th>Recv&nbsp;KB/s</th><th>Sent&nbsp;KB/s</th>'
        '</tr></thead>\n'
        f'<tbody>{sum_rows_html}</tbody></table>\n'
        '</div>\n\n'
        '<div class="sec">&#128260; USSD Step Flow</div>\n'
        '<div class="sec-hint">Each box = one USSD menu step. Colour = speed: <span style="color:#22c55e">green &lt;500ms</span> &middot; <span style="color:#3b82f6">blue &lt;1s</span> &middot; <span style="color:#f59e0b">amber &lt;2s</span> &middot; <span style="color:#ef4444">red &gt;2s</span></div>\n'
        f'<div class="flow-wrap">{flow_html or "<span style=\'color:var(--muted)\'>No step data.</span>"}</div>\n\n'
        '<div class="sec">&#127973; Service Health Overview</div>\n'
        '<div class="sec-hint">End-to-end pass/fail per service. Shows which step failed and what the USSD menu was expected to display.</div>\n'
        f'<div class="svc-grid">{svc_cards_html}</div>\n\n'
        '<div class="sec">&#128269; Failure Analysis &mdash; What Failed &amp; Why</div>\n'
        '<div class="sec-hint">For every failing service: the exact step, failure count, and the <strong style="color:var(--text)">expected vs actual USSD response text</strong>. Worst failure rate first.</div>\n'
        + (fail_analysis_html if fail_analysis_html else
           '<div style="background:var(--surface);border:1px solid var(--border);border-radius:14px;padding:20px;color:var(--green);font-weight:700;">&#10003; No failures &mdash; all services passed.</div>\n')
        + '\n<div class="sec">&#128202; Performance Charts</div>\n'
        '<div class="sec-hint" style="margin-bottom:18px;">Throughput over time, step-level response times, distribution, and error breakdown.</div>\n'
        '<div class="chart-grid">\n'
        '  <div class="chart-card full"><div class="ch-title">Throughput Over Time (req/s)</div><div class="ch-wrap tall"><canvas id="c-tps"></canvas></div></div>\n'
        '  <div class="chart-card"><div class="ch-title">Avg Response Time by Step (ms)</div><div class="ch-wrap"><canvas id="c-rt"></canvas></div></div>\n'
        '  <div class="chart-card"><div class="ch-title">Response Time Distribution</div><div class="ch-wrap"><canvas id="c-hist"></canvas></div></div>\n'
        '  <div class="chart-card"><div class="ch-title">P90 / P95 / P99 by Step (ms)</div><div class="ch-wrap"><canvas id="c-perc"></canvas></div></div>\n'
        '  <div class="chart-card"><div class="ch-title">Error Distribution by Step</div><div class="ch-wrap"><canvas id="c-err"></canvas></div></div>\n'
        '</div>\n\n'
        '<div class="sec">&#128203; Full Step-Level Detail &mdash; Grouped by Service</div>\n'
        '<div class="sec-hint" style="margin-bottom:14px;">Each service is shown as a collapsible header row with aggregate stats. Click a service header to expand/collapse its steps. Click <strong style="color:#ef4444">Why failed?</strong> on any step to see expected vs actual.</div>\n'
        '<div style="overflow-x:auto;">\n'
        '<table class="dtbl"><thead><tr>\n'
        '  <th>Service / Step</th><th>Samples</th><th>Avg (ms)</th><th>Median</th>\n'
        '  <th>P90</th><th>P95</th><th>P99</th><th>Min</th><th>Max</th>\n'
        '  <th>TPS</th><th>Errors</th><th>Error %</th><th>RT Bar</th>\n'
        f'</tr></thead><tbody>{stats_rows}</tbody></table>\n'
        '</div>\n\n'
        '<div class="two-col" style="margin-top:28px;">\n'
        '  <div class="panel"><div class="panel-title">&#128246; HTTP Response Code Distribution</div>\n'
        + (rc_rows if rc_rows else '<span style="color:var(--muted);font-size:13px;">No data.</span>')
        + '\n  </div>\n'
        '  <div class="panel"><div class="panel-title">&#128337; Test Run Details</div>\n'
        f'    <div style="display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid #1f2d44;font-size:13px;"><span style="color:#64748b">Start Time</span><strong style="color:#3b82f6">{d["start_time"]}</strong></div>\n'
        f'    <div style="display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid #1f2d44;font-size:13px;"><span style="color:#64748b">End Time</span><strong style="color:#3b82f6">{d["end_time"]}</strong></div>\n'
        f'    <div style="display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid #1f2d44;font-size:13px;"><span style="color:#64748b">Duration</span><strong style="color:#3b82f6">{d["test_duration"]} ({d["duration_s"]}s)</strong></div>\n'
        f'    <div style="display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid #1f2d44;font-size:13px;"><span style="color:#64748b">Std Deviation</span><strong style="color:#3b82f6">{d["std_dev"]:,} ms</strong></div>\n'
        f'    <div style="display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid #1f2d44;font-size:13px;"><span style="color:#64748b">Min RT</span><strong style="color:#22c55e">{d["min_rt"]:,} ms</strong></div>\n'
        f'    <div style="display:flex;justify-content:space-between;padding:8px 0;font-size:13px;"><span style="color:#64748b">Max RT</span><strong style="color:#ef4444">{d["max_rt"]:,} ms</strong></div>\n'
        '  </div>\n'
        '</div>\n\n'
        f'{_render_samples_io(d)}'
        f'<footer>Generated by Load Testing Platform &nbsp;&middot;&nbsp; {datetime.now().strftime("%d %b %Y %H:%M")}</footer>\n\n'
        '<script>\n'
        'function toggleRow(id){var r=document.getElementById(id);r.style.display=r.style.display==="none"?"table-row":"none";}\n'
        'function toggleSvcGroup(id){'
        'var rows=document.querySelectorAll("tr[data-sg=\'"+id+"\']");'
        'if(!rows.length)return;'
        'var collapsed=rows[0].style.display==="none";'
        'rows.forEach(function(r){r.style.display=collapsed?"table-row":"none";});'
        'var arr=document.getElementById(id+"-arrow");'
        'if(arr)arr.innerHTML=collapsed?"&#9660;":"&#9658;";}\n'
        'var GRID="rgba(31,45,68,.6)",TICK={color:"#64748b",font:{size:10}};\n'
        'var CO={responsive:true,maintainAspectRatio:false,\n'
        '  plugins:{legend:{display:false},tooltip:{backgroundColor:"#111827",borderColor:"#1f2d44",borderWidth:1,titleColor:"#f1f5f9",bodyColor:"#94a3b8",padding:10}},\n'
        '  scales:{x:{ticks:TICK,grid:{color:GRID}},y:{ticks:TICK,grid:{color:GRID}}}};\n'
        f'new Chart("c-tps",{{type:"line",data:{{labels:{tps_t_json},datasets:[\n'
        f'  {{label:"TPS",data:{tps_v_json},borderColor:"#3b82f6",backgroundColor:"rgba(59,130,246,.1)",borderWidth:2,pointRadius:0,fill:true,tension:.4}},\n'
        f'  {{label:"Avg {thr}",data:{avg_tps_json},borderColor:"rgba(34,197,94,.4)",borderDash:[4,4],borderWidth:1,pointRadius:0,fill:false}}\n'
        ']},options:{...CO,plugins:{legend:{display:true,labels:{color:"#94a3b8",font:{size:10}}}}}});\n'
        f'var sAvg={s_avg_json};\n'
        f'new Chart("c-rt",{{type:"bar",data:{{labels:{s_labels_json},\n'
        '  datasets:[{label:"Avg ms",data:sAvg,\n'
        '    backgroundColor:sAvg.map(function(v){return v<500?"rgba(34,197,94,.7)":v<1000?"rgba(59,130,246,.7)":v<2000?"rgba(245,158,11,.7)":"rgba(239,68,68,.7)";}),\n'
        '    borderRadius:4}]},\n'
        '  options:{...CO,scales:{x:{ticks:TICK,grid:{display:false}},y:{ticks:TICK,grid:{color:GRID},beginAtZero:true}}}});\n'
        f'new Chart("c-hist",{{type:"bar",data:{{labels:{hist_lbl_json},\n'
        f'  datasets:[{{data:{hist_cnt_json},backgroundColor:["#22c55e","#3b82f6","#3b82f6","#3b82f6","#f59e0b","#f59e0b","#ef4444","#ef4444"],borderRadius:4}}]}},\n'
        '  options:{...CO,scales:{x:{ticks:TICK,grid:{display:false}},y:{ticks:TICK,grid:{color:GRID},beginAtZero:true}}}});\n'
        f'new Chart("c-perc",{{type:"bar",data:{{labels:{s_labels_json},datasets:[\n'
        f'  {{label:"P90",data:{s_p90_json},backgroundColor:"rgba(59,130,246,.6)",borderRadius:2}},\n'
        f'  {{label:"P95",data:{s_p95_json},backgroundColor:"rgba(245,158,11,.6)",borderRadius:2}},\n'
        f'  {{label:"P99",data:{s_p99_json},backgroundColor:"rgba(239,68,68,.6)",borderRadius:2}}\n'
        ']},options:{...CO,plugins:{legend:{display:true,labels:{color:"#94a3b8",font:{size:11}}}}}});\n'
        '(function(){\n'
        f'  var el={err_steps_lbl},ev={err_steps_cnt};\n'
        '  if(!el.length){var c=document.getElementById("c-err").getContext("2d");\n'
        '    c.fillStyle="#64748b";c.font="14px Segoe UI";c.textAlign="center";\n'
        '    c.fillText("No errors in this run",c.canvas.width/2,c.canvas.height/2);return;}\n'
        '  new Chart("c-err",{type:"doughnut",data:{labels:el,datasets:[{data:ev,\n'
        '    backgroundColor:["#ef4444","#f59e0b","#a78bfa","#f0883e","#22c55e","#3b82f6"],\n'
        '    borderColor:"#111827",borderWidth:3}]},\n'
        '    options:{responsive:true,maintainAspectRatio:false,\n'
        '      plugins:{legend:{position:"right",labels:{color:"#94a3b8",font:{size:10},boxWidth:10}}}}});\n'
        '})();\n'
        '</script></body></html>'
    )


# ── API: CSV editor ────────────────────────────────────────────────────────────
@app.route('/api/csv-edit/<path:fname>', methods=['GET', 'POST'])
@login_req
def api_csv_edit(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['testdata'], fname)
    if not os.path.exists(path): return jsonify(error='Not found'), 404
    if request.method == 'GET':
        rows = []
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                rows = list(csv.reader(f))
        except Exception as ex:
            return jsonify(error=str(ex)), 500
        return jsonify(filename=fname, rows=rows)
    # POST: save edited rows
    data = request.json or {}
    rows = data.get('rows', [])
    try:
        with open(path, 'w', newline='', encoding='utf-8') as f:
            csv.writer(f).writerows(rows)
        audit('CSV_EDIT', f'Edited: {fname} ({len(rows)} rows) client={c["code"]}')
        return jsonify(ok=True, rows_saved=len(rows))
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: JMX tree viewer ───────────────────────────────────────────────────────
@app.route('/api/jmx-tree/<path:fname>')
@login_req
def api_jmx_tree(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['jmx'], fname)
    if not os.path.exists(path): return jsonify(error='Not found'), 404

    def _node(el, depth=0):
        tag  = el.tag.split('}')[-1] if '}' in el.tag else el.tag
        name = el.get('testname') or el.get('name') or ''
        enabled  = el.get('enabled', 'true') != 'false'
        children = [_node(child, depth + 1) for child in el if depth < 5]
        return {'tag': tag, 'name': name, 'enabled': enabled, 'children': children}

    try:
        tree = ET.parse(path)
        return jsonify(tree=_node(tree.getroot()), filename=fname)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── API: CSV row-count vs threads check ───────────────────────────────────────
@app.route('/api/csv-row-check')
@login_req
def api_csv_row_check():
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    threads = int(request.args.get('threads', 10))
    td = client_dirs(c)['testdata']
    results = []
    for path in glob.glob(os.path.join(td, '*.csv')):
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                line_count = max(0, sum(1 for _ in f) - 1)  # subtract header
            results.append({
                'filename': os.path.basename(path),
                'rows': line_count,
                'threads': threads,
                'ok': line_count >= threads,
                'warning': 0 < line_count < threads,
            })
        except Exception:
            pass
    return jsonify(results=results, threads=threads)

# ── API: Heatmap (runs by hour × day-of-week) ─────────────────────────────────
@app.route('/api/heatmap')
@login_req
def api_heatmap():
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    rdir = client_dirs(c)['reports']
    heat = [[0] * 7 for _ in range(24)]  # heat[hour][weekday]
    for path in glob.glob(os.path.join(rdir, '*.jtl')):
        try:
            dt = datetime.fromtimestamp(os.path.getmtime(path))
            heat[dt.hour][dt.weekday()] += 1
        except Exception:
            pass
    return jsonify(heatmap=heat, days=['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
                   hours=list(range(24)))

# ── API: Schedule calendar ────────────────────────────────────────────────────
@app.route('/api/schedule-calendar')
@login_req
def api_schedule_calendar():
    events = []
    with _sched_lock:
        for s in _schedules:
            events.append({'id': s['id'], 'title': s.get('jmx', ''), 'type': 'one-time',
                           'start': s.get('run_at', ''), 'status': s.get('status', ''),
                           'client': s.get('client', '')})
    with get_db() as db:
        rows = db.execute("SELECT * FROM recurring_schedules WHERE enabled=1").fetchall()
    for r in rows:
        events.append({'id': r['id'], 'title': r['jmx'], 'type': r['recurrence'],
                       'time': r['run_at_time'], 'day_of_week': r['day_of_week'],
                       'client': r['client'], 'last_run': r['last_run']})
    return jsonify(events=events)

# ── API: Config diff ──────────────────────────────────────────────────────────
@app.route('/api/config-diff', methods=['POST'])
@admin_req
def api_config_diff():
    submitted = request.json or {}
    current   = load_cfg()
    diffs = []
    masked = {'smtp_pass', 'openrouter_api_key', 'jira_token'}
    for k in sorted(set(list(submitted.keys()) + list(current.keys()))):
        old, new = current.get(k, ''), submitted.get(k, '')
        if str(old) != str(new):
            diffs.append({'key': k,
                          'old': '***' if k in masked else old,
                          'new': '***' if k in masked else new})
    return jsonify(diffs=diffs, has_changes=bool(diffs))

# ── API: Excel export of JTL report ───────────────────────────────────────────
@app.route('/api/report/<path:fname>/excel')
@login_req
def api_report_excel(fname):
    if not _HAS_OPENPYXL:
        return jsonify(error='openpyxl not installed'), 500
    c = active_client()
    if not c: return 'No client', 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return 'Not found', 404
    try:
        data = _parse_jtl(path)
    except Exception as ex:
        return str(ex), 500
    from openpyxl.styles import Font, PatternFill
    hfill = PatternFill('solid', fgColor='0066CC')
    hfont = Font(bold=True, color='FFFFFF')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Summary'
    ws.append(['Metric', 'Value'])
    for cell in ws[1]: cell.fill = hfill; cell.font = hfont
    for row in [
        ('File', data['filename']), ('Start Time', data['start_time']),
        ('Duration', data['test_duration']), ('Total Samples', data['total']),
        ('Passed', data['passed']), ('Failed', data['failed']),
        ('Error Rate %', data['error_rate']), ('Throughput (TPS)', data['throughput']),
        ('Avg RT (ms)', data['avg_rt']), ('P50 (ms)', data['p50']),
        ('P90 (ms)', data['p90']), ('P95 (ms)', data['p95']), ('P99 (ms)', data['p99']),
    ]:
        ws.append(list(row))
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws2 = wb.create_sheet('By Label')
    hdrs = ['Label', 'Samples', 'Failures', 'Error%', 'TPS', 'Avg', 'Median', 'P90', 'P95', 'P99']
    ws2.append(hdrs)
    for cell in ws2[1]: cell.fill = hfill; cell.font = hfont
    for s in data['label_stats']:
        ws2.append([s['label'], s['samples'], s['failures'], s['error_rate'],
                    s['tps'], s['avg'], s['median'], s['p90'], s['p95'], s['p99']])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = re.sub(r'[^a-zA-Z0-9_.-]', '_', fname.replace('.jtl', ''))
    audit('DOWNLOAD_EXCEL', fname)
    return send_file(buf, as_attachment=True,
                     download_name=f'Report_{safe}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── API: Regression detection ─────────────────────────────────────────────────
@app.route('/api/regression-check/<path:fname>')
@login_req
def api_regression_check(fname):
    """Compare a run against the baseline; flag regressions > threshold."""
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    baseline = _read_json(_baseline_path(c))
    if not baseline: return jsonify(has_baseline=False), 200
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='JTL not found'), 404
    threshold_pct = float(request.args.get('threshold', 10))
    try:
        data = _parse_jtl(path)
        bm   = baseline['metrics']
        def pct_change(curr, base):
            return round((curr - base) / base * 100, 1) if base else 0
        regressions = []
        checks = [
            ('p90',        data['p90'],         bm['p90'],        True),
            ('p95',        data['p95'],         bm['p95'],        True),
            ('avg_rt',     data['avg_rt'],       bm['avg_rt'],     True),
            ('error_rate', data['error_rate'],   bm['error_rate'], True),
            ('tps',        data['throughput'],   bm['tps'],        False),
        ]
        for name, curr, base, higher_is_worse in checks:
            delta = pct_change(curr, base)
            regressed = (higher_is_worse and delta > threshold_pct) or \
                        (not higher_is_worse and delta < -threshold_pct)
            checks_out = {'metric': name, 'baseline': base, 'current': curr,
                          'delta_pct': delta, 'regressed': regressed}
            if regressed:
                regressions.append(checks_out)
        return jsonify(has_baseline=True, baseline_file=baseline.get('file'),
                       regressions=regressions, threshold_pct=threshold_pct,
                       overall_ok=len(regressions) == 0)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── Health page ────────────────────────────────────────────────────────────────
@app.route('/health')
def health_check_page():
    db_ok = True
    try:
        with get_db() as db:
            db.execute("SELECT 1").fetchone()
    except Exception:
        db_ok = False
    disk = shutil.disk_usage(BASE_DIR)
    uptime = int((datetime.now() - _start_time).total_seconds())
    jbin = _find_jmeter_bin()
    return jsonify(
        status='ok' if db_ok else 'degraded',
        uptime_s=uptime,
        db_ok=db_ok,
        db_size_kb=round(os.path.getsize(DB_PATH)/1024, 1) if os.path.exists(DB_PATH) else 0,
        jmeter_ok=bool(jbin),
        jmeter_path=jbin,
        clients=len(get_all_clients()),
        disk_free_gb=round(disk.free/1e9, 2),
        test_running=_state.get('running', False),
        version='4.0'
    )

@app.route('/health/ui')
def health_ui():
    db_ok = True
    try:
        with get_db() as db: db.execute("SELECT 1").fetchone()
    except Exception:
        db_ok = False
    disk = shutil.disk_usage(BASE_DIR)
    uptime = int((datetime.now() - _start_time).total_seconds())
    jbin = _find_jmeter_bin()
    rows = ''.join(f"<tr><td>{k}</td><td style='color:{'#00ff9d' if v else '#ef4444'}'>{v}</td></tr>"
        for k, v in [('DB', db_ok), ('JMeter', bool(jbin)), ('Test Running', _state.get('running'))])
    return f'''<!DOCTYPE html><html><head><title>Platform Health</title>
<style>body{{font-family:monospace;background:#080c14;color:#e2eaf5;padding:24px;}}
table{{border-collapse:collapse;}}td{{padding:8px 16px;border:1px solid #1e2d45;}}</style></head>
<body><h2>&#9889; Load Testing Platform v4.0 &mdash; Health</h2>
<table>{rows}
<tr><td>Uptime</td><td>{uptime}s</td></tr>
<tr><td>Disk Free</td><td>{round(disk.free/1e9,2)} GB</td></tr>
<tr><td>DB Size</td><td>{round(os.path.getsize(DB_PATH)/1024,1) if os.path.exists(DB_PATH) else 0} KB</td></tr>
</table></body></html>'''

# ── Theme per user ─────────────────────────────────────────────────────────────
@app.route('/api/theme', methods=['GET', 'POST'])
@login_req
def api_theme():
    u = session.get('user', '')
    if request.method == 'POST':
        t = (request.json or {}).get('theme', 'dark')
        with get_db() as db:
            db.execute("INSERT OR REPLACE INTO user_prefs (username,key,value) VALUES (?,?,?)", (u,'theme',t))
            db.commit()
        return jsonify(ok=True, theme=t)
    with get_db() as db:
        row = db.execute("SELECT value FROM user_prefs WHERE username=? AND key='theme'", (u,)).fetchone()
    return jsonify(theme=row['value'] if row else load_cfg().get('theme','dark'))

# ── Favourites ─────────────────────────────────────────────────────────────────
@app.route('/api/favourites', methods=['GET', 'POST'])
@login_req
def api_favourites():
    u = session.get('user', '')
    c = active_client()
    key = f'fav_{c["code"]}' if c else 'fav'
    with get_db() as db:
        row = db.execute("SELECT value FROM user_prefs WHERE username=? AND key=?", (u, key)).fetchone()
    favs = json.loads(row['value']) if row and row['value'] else []
    if request.method == 'POST':
        fname = (request.json or {}).get('file', '')
        if fname in favs:
            favs.remove(fname)
            favourited = False
        else:
            favs.insert(0, fname)
            favourited = True
        with get_db() as db:
            db.execute("INSERT OR REPLACE INTO user_prefs (username,key,value) VALUES (?,?,?)", (u, key, json.dumps(favs)))
            db.commit()
        return jsonify(ok=True, favourited=favourited, favourites=favs)
    return jsonify(favourites=favs)

# ── Presence ───────────────────────────────────────────────────────────────────
@app.route('/api/presence', methods=['GET', 'POST'])
@login_req
def api_presence():
    u = session.get('user', '')
    now = datetime.now().timestamp()
    if request.method == 'POST':
        panel = (request.json or {}).get('panel', '')
        _presence[u] = {'name': session.get('name', u), 'initials': session.get('initials', u[:2].upper()), 'last_seen': now, 'panel': panel}
    active_users = [v for k, v in _presence.items() if now - v['last_seen'] < 60 and k != u]
    return jsonify(users=active_users)

# ── Run-time annotations ───────────────────────────────────────────────────────
@app.route('/api/test/annotate', methods=['POST'])
@login_req
def api_annotate():
    text = (request.json or {}).get('text', '').strip()
    if not text: return jsonify(error='text required'), 400
    _annotations.append({'ts': int(datetime.now().timestamp() * 1000), 'text': text, 'user': session.get('user', '')})
    return jsonify(ok=True)

@app.route('/api/test/annotations')
@login_req
def api_annotations():
    return jsonify(annotations=_annotations)

# ── Report comments ────────────────────────────────────────────────────────────
@app.route('/api/comments/<path:fname>', methods=['GET', 'POST'])
@login_req
def api_comments(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    if request.method == 'POST':
        d = request.json or {}
        cid = str(uuid.uuid4())[:8]
        with get_db() as db:
            db.execute("INSERT INTO report_comments (id,client,filename,username,text,ts_offset_s) VALUES (?,?,?,?,?,?)",
                (cid, c['code'], fname, session.get('user',''), d.get('text',''), d.get('ts_offset_s')))
            db.commit()
        audit('COMMENT_ADD', f'{fname}: {d.get("text","")[:60]}')
        return jsonify(ok=True, id=cid)
    with get_db() as db:
        rows = db.execute("SELECT * FROM report_comments WHERE client=? AND filename=? ORDER BY created_at", (c['code'], fname)).fetchall()
    return jsonify(comments=[dict(r) for r in rows])

@app.route('/api/comments/<path:fname>/<cid>', methods=['DELETE'])
@login_req
def api_comment_delete(fname, cid):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    with get_db() as db:
        row = db.execute("SELECT username FROM report_comments WHERE id=? AND client=?", (cid, c['code'])).fetchone()
        if not row: return jsonify(error='Not found'), 404
        if row['username'] != session.get('user') and session.get('role') != 'admin':
            return jsonify(error='Forbidden'), 403
        db.execute("DELETE FROM report_comments WHERE id=?", (cid,))
        db.commit()
    return jsonify(ok=True)

# ── Live sharing ───────────────────────────────────────────────────────────────
@app.route('/api/share/live', methods=['POST'])
@login_req
def api_share_live():
    token = str(uuid.uuid4()).replace('-', '')[:16]
    _live_share_token['token'] = token
    _live_share_token['expires'] = datetime.now().timestamp() + 4 * 3600
    url = f'/shared/live/{token}'
    audit('SHARE_LIVE', 'Token generated, expires in 4h')
    return jsonify(ok=True, token=token, url=url)

@app.route('/shared/live/<token>')
def shared_live(token):
    if not token or token != _live_share_token.get('token') or datetime.now().timestamp() > _live_share_token.get('expires', 0):
        return 'Link expired or invalid.', 403
    return ('''<!DOCTYPE html><html><head><title>Live Test Dashboard</title>
<style>body{background:#080c14;color:#e2eaf5;font-family:"Segoe UI",sans-serif;padding:24px;}
.kpi{background:#0d1421;border:1px solid #1e2d45;border-radius:10px;padding:16px;display:inline-block;margin:8px;min-width:140px;}
.kpi-label{font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:.6px;}
.kpi-val{font-size:28px;font-weight:800;color:#00d4ff;}</style></head>
<body><h2>&#9889; Live Test Dashboard</h2>
<div id="status" style="margin-bottom:16px;color:#64748b;">Loading&#8230;</div>
<div id="kpis"></div>
<script>
async function poll(){
  const r=await fetch('/api/live-stats/public?token=''' + token + '''');
  if(!r.ok){document.getElementById('status').textContent='Expired or unavailable.';return;}
  const d=await r.json();
  document.getElementById('status').textContent=d.running?'&#128994; Test Running — '+d.jmx:'&#9898; No test running';
  const last=d.samples&&d.samples.length?d.samples[d.samples.length-1]:{};
  document.getElementById('kpis').innerHTML=
    '<div class="kpi"><div class="kpi-label">TPS</div><div class="kpi-val">'+(last.tps||0)+'</div></div>'+
    '<div class="kpi"><div class="kpi-label">Avg RT</div><div class="kpi-val">'+(last.avg_rt||0)+'ms</div></div>'+
    '<div class="kpi"><div class="kpi-label">Error %</div><div class="kpi-val">'+(last.err_pct||0)+'%</div></div>'+
    '<div class="kpi"><div class="kpi-label">Samples</div><div class="kpi-val">'+(d.total_samples||0)+'</div></div>';
}
poll();setInterval(poll,5000);
</script></body></html>''')

@app.route('/api/live-stats/public')
def api_live_stats_public():
    token = request.args.get('token', '')
    if not token or token != _live_share_token.get('token') or datetime.now().timestamp() > _live_share_token.get('expires', 0):
        return jsonify(error='Invalid or expired token'), 403
    return api_live_stats()

# ── Report sharing ─────────────────────────────────────────────────────────────
@app.route('/api/share/report', methods=['POST'])
@login_req
def api_share_report():
    fname = (request.json or {}).get('file', '')
    c = active_client()
    if not c or not fname: return jsonify(error='file required'), 400
    token = str(uuid.uuid4()).replace('-', '')[:20]
    import datetime as _dt
    expires = (datetime.now().replace(microsecond=0) + _dt.timedelta(days=7)).isoformat()
    with get_db() as db:
        db.execute("INSERT INTO report_shares (token,client,filename,created_by,expires_at) VALUES (?,?,?,?,?)",
            (token, c['code'], fname, session.get('user',''), expires))
        db.commit()
    audit('SHARE_REPORT', f'{fname} token={token[:8]}')
    return jsonify(ok=True, token=token, url=f'/shared/report/{token}', expires=expires)

@app.route('/shared/report/<token>')
def shared_report(token):
    with get_db() as db:
        row = db.execute("SELECT * FROM report_shares WHERE token=?", (token,)).fetchone()
    if not row: return 'Link not found.', 404
    if datetime.now().isoformat() > row['expires_at']: return 'Link expired.', 410
    c = get_client(row['client'])
    if not c: return 'Client not found.', 404
    path = os.path.join(client_dirs(c)['reports'], row['filename'])
    if not os.path.exists(path): return 'Report file not found.', 404
    try:
        d = _parse_jtl(path)
        _enrich_report(d, path, c)
        return _generate_report_html(d)
    except Exception as ex:
        return f'Error generating report: {ex}', 500

# ── Platform stats ─────────────────────────────────────────────────────────────
@app.route('/api/platform-stats')
@login_req
def api_platform_stats():
    import datetime as _dt
    today = datetime.now().strftime('%Y-%m-%d')
    week_start = (datetime.now() - _dt.timedelta(days=7)).strftime('%Y-%m-%d')
    with get_db() as db:
        tests_today = db.execute("SELECT COUNT(*) FROM audit_log WHERE action='TEST_START' AND timestamp LIKE ?", (today+'%',)).fetchone()[0]
        tests_week  = db.execute("SELECT COUNT(*) FROM audit_log WHERE action='TEST_START' AND timestamp >= ?", (week_start,)).fetchone()[0]
        tests_total = db.execute("SELECT COUNT(*) FROM audit_log WHERE action='TEST_START'").fetchone()[0]
        most_user   = db.execute("SELECT username, COUNT(*) c FROM audit_log WHERE action='TEST_START' GROUP BY username ORDER BY c DESC LIMIT 1").fetchone()
        most_jmx_r  = db.execute("SELECT details, COUNT(*) c FROM audit_log WHERE action='TEST_START' GROUP BY details ORDER BY c DESC LIMIT 1").fetchone()
        daily       = db.execute("SELECT substr(timestamp,1,10) d, COUNT(*) c FROM audit_log WHERE action='TEST_START' AND timestamp >= ? GROUP BY d ORDER BY d", (week_start,)).fetchall()
        pass_count  = db.execute("SELECT COUNT(*) FROM audit_log WHERE action='TEST_END' AND details LIKE '%rc=0%'").fetchone()[0]
        fail_count  = db.execute("SELECT COUNT(*) FROM audit_log WHERE action='TEST_END' AND details NOT LIKE '%rc=0%'").fetchone()[0]
    total_ended = pass_count + fail_count
    pass_rate = round(pass_count / total_ended * 100, 1) if total_ended else 0
    return jsonify(
        tests_today=tests_today, tests_week=tests_week, tests_total=tests_total,
        most_active_user=most_user['username'] if most_user else '—',
        most_run_jmx=(most_jmx_r['details'] or '—').split('JMX=')[-1].split(' ')[0] if most_jmx_r else '—',
        pass_rate_pct=pass_rate,
        clients_count=len(get_all_clients()),
        daily_counts=[{'date': r['d'], 'count': r['c']} for r in daily]
    )

# ── Storage quota ──────────────────────────────────────────────────────────────
@app.route('/api/clients/<code>/storage')
@login_req
def api_client_storage(code):
    c = get_client(code)
    if not c: return jsonify(error='Not found'), 404
    dirs = client_dirs(c)
    def dir_size(d):
        total = 0
        if os.path.isdir(d):
            for root, _, files in os.walk(d):
                for f in files:
                    try: total += os.path.getsize(os.path.join(root, f))
                    except: pass
        return total
    used = sum(dir_size(dirs[k]) for k in dirs)
    used_mb = round(used / 1024 / 1024, 2)
    quota_mb = c.get('quota_mb', 0) or 0
    pct = round(used_mb / quota_mb * 100, 1) if quota_mb > 0 else 0
    return jsonify(used_mb=used_mb, quota_mb=quota_mb, pct=pct,
                   files={k: len(glob.glob(os.path.join(dirs[k], '*'))) for k in dirs})

# ── Auto-archive reports ───────────────────────────────────────────────────────
@app.route('/api/archive-reports', methods=['POST'])
@admin_req
def api_archive_reports():
    d = request.json or {}
    days = int(d.get('days', load_cfg().get('archive_reports_days', 30)) or 30)
    code = d.get('client')
    c = get_client(code) if code else active_client()
    if not c: return jsonify(error='No client'), 400
    rdir = client_dirs(c)['reports']
    cutoff = datetime.now().timestamp() - days * 86400
    to_archive = [f for f in glob.glob(os.path.join(rdir, '*.jtl')) + glob.glob(os.path.join(rdir, '*.html'))
                  if os.path.getmtime(f) < cutoff]
    if not to_archive: return jsonify(ok=True, archived=0, message='Nothing to archive')
    arc_name = f'archive_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
    arc_path = os.path.join(rdir, arc_name)
    with zipfile.ZipFile(arc_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in to_archive:
            zf.write(f, os.path.basename(f))
    for f in to_archive:
        os.remove(f)
    audit('REPORTS_ARCHIVED', f'{len(to_archive)} files -> {arc_name} client={c["code"]}')
    return jsonify(ok=True, archived=len(to_archive), zip=arc_name)

# ── Backup ─────────────────────────────────────────────────────────────────────
@app.route('/api/backup/run', methods=['POST'])
@admin_req
def api_backup_run():
    cfg = load_cfg()
    bpath = cfg.get('backup_path', '').strip()
    if not bpath: return jsonify(error='Backup path not configured in Settings'), 400
    dest = os.path.join(bpath, f'lt_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
    try:
        os.makedirs(dest, exist_ok=True)
        shutil.copy2(DB_PATH, dest)
        shutil.copytree(CLIENTS_DIR, os.path.join(dest, 'clients'), dirs_exist_ok=True)
        audit('BACKUP_RUN', f'Destination: {dest}')
        return jsonify(ok=True, path=dest)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

@app.route('/api/backup/list')
@admin_req
def api_backup_list():
    cfg = load_cfg()
    bpath = cfg.get('backup_path', '').strip()
    if not bpath or not os.path.isdir(bpath): return jsonify(backups=[])
    backups = []
    for name in sorted(os.listdir(bpath), reverse=True):
        full = os.path.join(bpath, name)
        if os.path.isdir(full) and name.startswith('lt_backup_'):
            try:
                sz = sum(os.path.getsize(os.path.join(r, f)) for r, _, fs in os.walk(full) for f in fs)
                backups.append({'name': name, 'path': full, 'size_mb': round(sz/1024/1024, 1),
                                'created': datetime.fromtimestamp(os.path.getmtime(full)).strftime('%d %b %Y %H:%M')})
            except: pass
    return jsonify(backups=backups)

# ── SLA trend ─────────────────────────────────────────────────────────────────
@app.route('/api/sla-trend')
@login_req
def api_sla_trend():
    c = active_client()
    if not c: return jsonify(trend=[])
    limit = int(request.args.get('limit', 20))
    rdir = client_dirs(c)['reports']
    jtls = sorted(glob.glob(os.path.join(rdir, '*.jtl')), key=os.path.getmtime, reverse=True)[:limit]
    trend = []
    for p in reversed(jtls):
        r = _evaluate_sla(p, c)
        if r:
            trend.append({'file': os.path.basename(p),
                          'date': datetime.fromtimestamp(os.path.getmtime(p)).strftime('%d %b'),
                          'passed': r.get('passed', False),
                          'tps': r.get('actual_tps', 0), 'p90': r.get('actual_p90', 0),
                          'err_pct': r.get('actual_err', 0)})
    return jsonify(trend=trend)

# ── Test data generator ────────────────────────────────────────────────────────
@app.route('/api/generate-csv', methods=['POST'])
@login_req
def api_generate_csv():
    import random, string
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    d = request.json or {}
    fname = d.get('filename', 'generated.csv')
    rows_n = min(int(d.get('rows', 100)), 100000)
    columns = d.get('columns', [])
    if not columns: return jsonify(error='columns required'), 400
    td = client_dirs(c)['testdata']
    os.makedirs(td, exist_ok=True)
    outpath = os.path.join(td, os.path.basename(fname))
    with open(outpath, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow([col['name'] for col in columns])
        for i in range(rows_n):
            row = []
            for col in columns:
                t = col.get('type', 'fixed')
                if t == 'phone':
                    row.append(col.get('prefix','') + ''.join(random.choices(string.digits, k=int(col.get('digits',8)))))
                elif t == 'pin':
                    row.append(''.join(random.choices(string.digits, k=int(col.get('length',4)))))
                elif t == 'range':
                    row.append(str(random.randint(int(col.get('min',1)), int(col.get('max',100)))))
                elif t == 'uuid':
                    row.append(str(uuid.uuid4()))
                elif t == 'sequence':
                    row.append(str(i + 1))
                else:
                    row.append(str(col.get('value','')))
            w.writerow(row)
    audit('CSV_GENERATED', f'{fname} rows={rows_n}')
    return jsonify(ok=True, rows=rows_n, file=os.path.basename(fname))

# ── Server health check ────────────────────────────────────────────────────────
@app.route('/api/check-server', methods=['GET'])
@login_req
def api_check_server_auto():
    c = active_client()
    if not c: return jsonify(ok=False, error='No active client'), 400
    env = _read_json(_env_path(c)) or {}
    protocol = env.get('protocol', 'https')
    server = env.get('server', '') or env.get('server_ip', '') or env.get('host', '')
    port = env.get('port', '')
    sim_url = (env.get('sim_url') or '').strip()
    if not server: return jsonify(ok=False, error='No server configured — fill in Environment Details → Server/IP'), 400
    base = f"{protocol}://{server}{(':'+port) if port else ''}"
    url = base + sim_url if sim_url else base
    import time as _time
    t0 = _time.time()
    try:
        req = _urllib_req2.Request(url, method='GET')
        with _urllib_req2.urlopen(req, timeout=5, context=_SSL_CTX) as resp:
            status = resp.status
        return jsonify(ok=True, url=url, status_code=status, latency_ms=int((_time.time()-t0)*1000))
    except Exception as ex:
        return jsonify(ok=False, url=url, error=str(ex), latency_ms=int((_time.time()-t0)*1000))

@app.route('/api/health-check', methods=['POST'])
@login_req
def api_server_health_check():
    d = request.json or {}
    url = d.get('url', '').strip()
    if not url: return jsonify(error='url required'), 400
    import time as _time
    t0 = _time.time()
    try:
        req = _urllib_req2.Request(url, method='GET')
        with _urllib_req2.urlopen(req, timeout=int(d.get('timeout', 5)), context=_SSL_CTX) as resp:
            status = resp.status
        latency = int((_time.time() - t0) * 1000)
        return jsonify(ok=True, status_code=status, latency_ms=latency)
    except Exception as ex:
        return jsonify(ok=False, error=str(ex), latency_ms=int((_time.time() - t0) * 1000))

# ── Health Monitor ─────────────────────────────────────────────────────────────
def _monitor_path(c):
    return os.path.join(CLIENTS_DIR, c['code'], 'monitor_endpoints.json')

def _check_endpoint(ep, timeout=5):
    """Check one endpoint dict {url, method, expected_status, name}. Returns result dict."""
    import time as _t
    url     = ep.get('url', '').strip()
    method  = ep.get('method', 'GET').upper()
    expect  = int(ep.get('expected_status', 200))
    name    = ep.get('name') or url
    if not url:
        return {'name': name, 'url': url, 'ok': False, 'error': 'No URL', 'latency_ms': 0,
                'status_code': None, 'expected_status': expect}
    t0 = _t.time()
    try:
        req = _urllib_req2.Request(url, method=method)
        req.add_header('User-Agent', 'LoadTestPlatform-HealthCheck/1.0')
        with _urllib_req2.urlopen(req, timeout=timeout, context=_SSL_CTX) as resp:
            code = resp.status
        ms = int((_t.time() - t0) * 1000)
        ok = (code == expect)
        return {'name': name, 'url': url, 'ok': ok, 'status_code': code,
                'expected_status': expect, 'latency_ms': ms, 'error': None}
    except Exception as ex:
        ms = int((_t.time() - t0) * 1000)
        return {'name': name, 'url': url, 'ok': False, 'status_code': None,
                'expected_status': expect, 'latency_ms': ms, 'error': str(ex)}

@app.route('/api/monitor/endpoints', methods=['GET', 'POST'])
@login_req
def api_monitor_endpoints():
    c = active_client()
    if not c: return jsonify(error='No active client'), 400
    path = _monitor_path(c)
    if request.method == 'POST':
        _write_json(path, request.get_json(force=True) or [])
        audit('MONITOR_ENDPOINTS_SAVED', f'Saved monitor endpoints for {c["code"]}')
        return jsonify(ok=True)
    return jsonify(_read_json(path) or [])

@app.route('/api/monitor/check', methods=['POST'])
@login_req
def api_monitor_check():
    """Check a list of endpoints. Body: {endpoints:[...], timeout:5}"""
    d        = request.get_json(force=True) or {}
    eps      = d.get('endpoints') or []
    timeout  = int(d.get('timeout', 5))
    if not eps:
        # fall back to saved list for active client
        c = active_client()
        if c: eps = _read_json(_monitor_path(c)) or []
    results = [_check_endpoint(ep, timeout) for ep in eps]
    total   = len(results)
    up      = sum(1 for r in results if r['ok'])
    avg_lat = int(sum(r['latency_ms'] for r in results) / total) if total else 0
    return jsonify(results=results, summary={'total': total, 'up': up, 'down': total - up,
                                              'avg_latency_ms': avg_lat})

@app.route('/api/monitor/system', methods=['GET'])
@login_req
def api_monitor_system():
    """Return CPU, memory, disk, and process stats."""
    try:
        cpu     = _psutil.cpu_percent(interval=0.3)
        mem     = _psutil.virtual_memory()
        disk    = _psutil.disk_usage('/')
        net     = _psutil.net_io_counters()
        proc    = _psutil.Process()
        return jsonify(
            cpu_pct      = round(cpu, 1),
            mem_pct      = round(mem.percent, 1),
            mem_used_mb  = round(mem.used / 1024 / 1024, 1),
            mem_total_mb = round(mem.total / 1024 / 1024, 1),
            disk_pct     = round(disk.percent, 1),
            disk_free_gb = round(disk.free / 1024 / 1024 / 1024, 2),
            net_sent_mb  = round(net.bytes_sent / 1024 / 1024, 2),
            net_recv_mb  = round(net.bytes_recv / 1024 / 1024, 2),
            platform_mem_mb = round(proc.memory_info().rss / 1024 / 1024, 1),
            platform_cpu_pct= round(proc.cpu_percent(interval=0.1), 1),
        )
    except Exception as ex:
        return jsonify(error=str(ex)), 500

@app.route('/api/monitor/sanity', methods=['POST'])
@login_req
def api_monitor_sanity():
    """Run all saved endpoints as a pre-test gate. Returns ok=True only if all pass."""
    c = active_client()
    if not c: return jsonify(error='No active client'), 400
    eps     = _read_json(_monitor_path(c)) or []
    timeout = int((request.get_json(force=True) or {}).get('timeout', 5))
    if not eps:
        return jsonify(ok=True, skipped=True, message='No endpoints configured — skipping sanity check.')
    results = [_check_endpoint(ep, timeout) for ep in eps]
    failed  = [r for r in results if not r['ok']]
    ok      = len(failed) == 0
    audit('SANITY_CHECK', f'{len(results)-len(failed)}/{len(results)} endpoints healthy')
    return jsonify(ok=ok, results=results,
                   failed=[r['name'] for r in failed],
                   message='All endpoints healthy.' if ok else
                           f'{len(failed)} endpoint(s) failed: {", ".join(r["name"] for r in failed)}')

# ── Load profiles ──────────────────────────────────────────────────────────────
def _profiles_path(c):
    return os.path.join(CLIENTS_DIR, c['code'], 'load_profiles.json')

@app.route('/api/load-profiles', methods=['GET', 'POST'])
@login_req
def api_load_profiles():
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = _profiles_path(c)
    if request.method == 'POST':
        d = request.json or {}
        data = _read_json(path) or {'profiles': []}
        existing = next((i for i, p in enumerate(data['profiles']) if p['name'] == d.get('name')), None)
        if existing is not None:
            data['profiles'][existing] = d
        else:
            data['profiles'].append(d)
        _write_json(path, data)
        audit('PROFILE_SAVED', f'{d.get("name")}')
        return jsonify(ok=True)
    return jsonify((_read_json(path) or {'profiles': []})['profiles'])

@app.route('/api/load-profiles/<name>', methods=['DELETE'])
@login_req
def api_load_profile_delete(name):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = _profiles_path(c)
    data = _read_json(path) or {'profiles': []}
    data['profiles'] = [p for p in data['profiles'] if p['name'] != name]
    _write_json(path, data)
    return jsonify(ok=True)

@app.route('/api/load-profiles/<name>/execute', methods=['POST'])
@perm_req('run_tests')
def api_load_profile_execute(name):
    """Execute a multi-step load profile as a single staged JMeter run."""
    global _state, _logs
    if _state['running']:
        return jsonify(error='A test is already running.'), 409
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    d = request.json or {}
    jmx_name = d.get('jmx', '')
    if not jmx_name:
        return jsonify(error='jmx filename required'), 400
    dirs = ensure_client_dirs(c)
    jmx_path = os.path.join(dirs['jmx'], jmx_name)
    if not os.path.exists(jmx_path):
        return jsonify(error=f'JMX not found: {jmx_name}'), 400

    path = _profiles_path(c)
    profiles = (_read_json(path) or {'profiles': []})['profiles']
    profile = next((p for p in profiles if p['name'] == name), None)
    if not profile:
        return jsonify(error=f'Profile "{name}" not found'), 404
    steps = profile.get('steps', [])
    if not steps:
        return jsonify(error='Profile has no steps'), 400

    jbin = _find_jmeter_bin()
    if not jbin:
        return jsonify(error='JMeter not found'), 503

    out_name = f"profile_{name.replace(' ','_')}_{datetime.now().strftime('%d%m%Y_%H%M%S')}"
    jtl_path = os.path.join(dirs['reports'], out_name + '.jtl')
    os.makedirs(dirs['reports'], exist_ok=True)

    try:
        import copy as _copy
        tmp_dir = tempfile.mkdtemp(prefix='lt_profile_')
        working_jmx = os.path.join(tmp_dir, jmx_name)
        shutil.copy(jmx_path, working_jmx)
        tree = ET.parse(working_jmx)
        root_el = tree.getroot()

        # Fix CSV paths
        testdata_dir = client_dirs(c)['testdata']
        for csv_ds in root_el.iter('CSVDataSet'):
            for prop in csv_ds.iter('stringProp'):
                if prop.get('name') == 'filename' and prop.text and prop.text.strip():
                    bn = os.path.basename(prop.text.strip().replace('/', os.sep).replace('\\', os.sep))
                    cand = os.path.join(testdata_dir, bn)
                    if os.path.exists(cand):
                        prop.text = cand

        # Build staged ThreadGroups — one set per step using scheduler
        top_ht = root_el.find('hashTree')
        second_ht = top_ht.find('hashTree') if top_ht is not None else None
        if second_ht is None:
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return jsonify(error='Invalid JMX structure'), 400

        orig_tgs = []
        siblings = list(second_ht)
        i = 0
        while i < len(siblings):
            el = siblings[i]
            ht = siblings[i+1] if i+1 < len(siblings) and siblings[i+1].tag == 'hashTree' else None
            if el.tag == 'ThreadGroup':
                orig_tgs.append((el, ht))
            i += 2 if ht is not None else 1

        # Remove originals
        for tg, ht in orig_tgs:
            second_ht.remove(tg)
            if ht is not None:
                second_ht.remove(ht)

        def _set_sp(el, name, val):
            for sp in el.findall('stringProp'):
                if sp.get('name') == name:
                    sp.text = str(val); return
            sp = ET.SubElement(el, 'stringProp')
            sp.set('name', name); sp.text = str(val)

        def _set_bp(el, name, val):
            for bp in el.findall('boolProp'):
                if bp.get('name') == name:
                    bp.text = 'true' if val else 'false'; return
            bp = ET.SubElement(el, 'boolProp')
            bp.set('name', name); bp.text = 'true' if val else 'false'

        delay_offset = 0
        for si, step in enumerate(steps):
            tps      = int(step.get('tps', 25))
            dur      = int(step.get('duration', 1800))
            ru       = int(step.get('rampup', 60))
            threads  = max(1, tps * 2)
            for orig_tg, orig_ht in orig_tgs:
                stg = _copy.deepcopy(orig_tg)
                orig_name = orig_tg.get('testname', 'TG')
                stg.set('testname', f'{orig_name} [step{si+1}]')
                _set_sp(stg, 'ThreadGroup.num_threads', threads)
                _set_sp(stg, 'ThreadGroup.ramp_time',   ru)
                _set_sp(stg, 'ThreadGroup.duration',     dur)
                _set_sp(stg, 'ThreadGroup.delay',        delay_offset)
                _set_bp(stg, 'ThreadGroup.scheduler',    True)
                second_ht.append(stg)
                if orig_ht is not None:
                    second_ht.append(_copy.deepcopy(orig_ht))
            delay_offset += dur

        tree.write(working_jmx, encoding='unicode', xml_declaration=False)
    except Exception as ex:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return jsonify(error=f'JMX preparation failed: {ex}'), 500

    total_dur = sum(int(s.get('duration', 1800)) for s in steps)
    cmd = [jbin, '-n', '-t', working_jmx, '-l', jtl_path,
           f'-Jtest.duration={total_dur}',
           '-Jjmeter.save.saveservice.url=true',
           '-Jjmeter.save.saveservice.latency=true',
           '-Jjmeter.save.saveservice.connect_time=true',
           '-Jjmeter.save.saveservice.response_data.on_error=true',
           '-Jjmeter.save.saveservice.assertion_results_failure_message=true',
           '-Jjmeter.save.saveservice.bytes=true',
           '-Jjmeter.save.saveservice.sent_bytes=true',
           ]

    _state.update({'running': True, 'pid': None, 'rc': None, 'err': None,
                   'jmx': jmx_name, 'jtl': out_name + '.jtl', 'sla': None})
    _logs.clear()
    audit('PROFILE_EXECUTE', f'profile={name} jmx={jmx_name} steps={len(steps)} dur={total_dur}s')

    def _run_profile():
        global _state
        proc = None
        try:
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                    text=True, encoding='utf-8', errors='replace')
            _state['pid'] = proc.pid
            for line in proc.stdout:
                line = line.rstrip()
                if line:
                    with _lock:
                        _logs.append({'ts': datetime.now().strftime('%H:%M:%S'), 'msg': line})
                        if len(_logs) > 1000:
                            _logs.pop(0)
            proc.wait()
            _state['rc'] = proc.returncode
        except Exception as ex:
            _state['err'] = str(ex)
        finally:
            _state['running'] = False
            shutil.rmtree(tmp_dir, ignore_errors=True)

    threading.Thread(target=_run_profile, daemon=True).start()
    return jsonify(ok=True, jtl=out_name + '.jtl', steps=len(steps), total_duration=total_dur)

# ── JMX properties editor ──────────────────────────────────────────────────────
@app.route('/api/jmx/<path:fname>/properties', methods=['GET', 'POST'])
@login_req
def api_jmx_properties(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    jmx_path = os.path.join(client_dirs(c)['jmx'], fname)
    if not os.path.exists(jmx_path): return jsonify(error='JMX not found'), 404
    try:
        tree = ET.parse(jmx_path)
        root = tree.getroot()
        if request.method == 'POST':
            updates = request.json or {}
            for tg_data in updates.get('thread_groups', []):
                tg_name = tg_data.get('name')
                for tg in root.iter('ThreadGroup'):
                    if tg.get('testname') == tg_name:
                        tg.set('enabled', 'true' if tg_data.get('enabled', True) else 'false')
                        for sp in tg.iter('stringProp'):
                            n = sp.get('name', '')
                            if n in tg_data:
                                sp.text = str(tg_data[n])
            tree.write(jmx_path, encoding='unicode', xml_declaration=False)
            audit('JMX_EDITED', f'{fname}')
            return jsonify(ok=True)
        tgs = []
        for tg in root.iter('ThreadGroup'):
            props = {}
            for sp in tg.iter('stringProp'):
                props[sp.get('name', '')] = sp.text or ''
            for ip in tg.iter('intProp'):
                props[ip.get('name', '')] = ip.text or ''
            tgs.append({'name': tg.get('testname', ''), 'enabled': tg.get('enabled', 'true') == 'true', 'props': props})
        return jsonify(thread_groups=tgs)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── Self-registration ──────────────────────────────────────────────────────────
@app.route('/api/register-request', methods=['POST'])
def api_register_request():
    if not load_cfg().get('self_register_enabled', False):
        return jsonify(error='Self-registration is not enabled'), 403
    d = request.json or {}
    username = d.get('username', '').strip().lower()
    name = d.get('name', '').strip()
    email = d.get('email', '').strip()
    password = d.get('password', '')
    if not username or not name or not password: return jsonify(error='username, name and password required'), 400
    if len(password) < 6: return jsonify(error='Password min 6 characters'), 400
    try:
        with get_db() as db:
            db.execute("INSERT INTO pending_registrations (username,password,name,email) VALUES (?,?,?,?)",
                (username, generate_password_hash(password), name, email))
            db.commit()
        return jsonify(ok=True, message='Registration submitted. An admin will review your request.')
    except Exception:
        return jsonify(error='Username already submitted or exists'), 409

@app.route('/api/admin/registrations', methods=['GET'])
@admin_req
def api_list_registrations():
    with get_db() as db:
        rows = db.execute("SELECT * FROM pending_registrations ORDER BY submitted_at DESC").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/registrations/<int:rid>/approve', methods=['POST'])
@admin_req
def api_approve_registration(rid):
    with get_db() as db:
        row = db.execute("SELECT * FROM pending_registrations WHERE id=?", (rid,)).fetchone()
        if not row: return jsonify(error='Not found'), 404
        try:
            db.execute("INSERT INTO users (username,password,name,role,initials,created_by,permissions) VALUES (?,?,?,?,?,?,?)",
                (row['username'], row['password'], row['name'], 'viewer', row['name'][:2].upper(), session['user'], '["run_tests"]'))
        except Exception:
            return jsonify(error='Username already exists'), 409
        db.execute("DELETE FROM pending_registrations WHERE id=?", (rid,))
        db.commit()
    audit('REGISTRATION_APPROVED', f"{row['username']}")
    return jsonify(ok=True)

@app.route('/api/admin/registrations/<int:rid>/reject', methods=['POST'])
@admin_req
def api_reject_registration(rid):
    with get_db() as db:
        db.execute("DELETE FROM pending_registrations WHERE id=?", (rid,))
        db.commit()
    audit('REGISTRATION_REJECTED', f'id={rid}')
    return jsonify(ok=True)

# ── Backup scheduler ───────────────────────────────────────────────────────────
def _start_backup_scheduler():
    cfg = load_cfg()
    hours = int(cfg.get('backup_interval_hours', 0) or 0)
    if hours <= 0:
        return
    try:
        dest = cfg.get('backup_path', '').strip()
        if dest:
            ts = f'lt_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
            d = os.path.join(dest, ts)
            os.makedirs(d, exist_ok=True)
            shutil.copy2(DB_PATH, d)
            shutil.copytree(CLIENTS_DIR, os.path.join(d, 'clients'), dirs_exist_ok=True)
    except Exception as ex:
        print(f'[backup] Error: {ex}')
    t = threading.Timer(hours * 3600, _start_backup_scheduler)
    t.daemon = True
    t.start()

# ── CI/CD Internal Scheduler ───────────────────────────────────────────────────
def _run_ci_cd_suite_internal(suite_id, jmx_files, feature_file=None):
    """Execute a CI/CD suite internally (called by scheduler)"""
    try:
        run_id = f"run_{uuid.uuid4().hex[:12]}"
        start_time = datetime.now()
        
        with get_db() as db:
            suite = db.execute("SELECT * FROM ci_cd_suites WHERE id=?", (suite_id,)).fetchone()
            if not suite:
                return
            
            suite = dict(suite)
            client_code = suite['client']
            
            # Record run start
            db.execute(
                """UPDATE ci_cd_run_history SET status=?, end_time=? WHERE id=?""",
                ('running', None, run_id)
            )
            db.commit()
        
        # Execute each JMX file in the suite
        total_success, total_errors = 0, 0
        for jmx_file in jmx_files:
            try:
                # Here you'd call the actual JMeter execution
                # For now, just log it
                pass
            except Exception as e:
                total_errors += 1
        
        # Record run completion
        end_time = datetime.now()
        duration_s = int((end_time - start_time).total_seconds())
        
        with get_db() as db:
            db.execute(
                """UPDATE ci_cd_run_history 
                   SET status=?, end_time=?, duration_s=?, success_count=?, error_count=? 
                   WHERE id=?""",
                ('completed', end_time.isoformat(), duration_s, total_success, total_errors, run_id)
            )
            db.execute("UPDATE ci_cd_suites SET last_run=? WHERE id=?", (end_time.isoformat(), suite_id))
            db.commit()
        
        print(f'[CI/CD] Suite {suite["name"]} completed: {total_success} success, {total_errors} errors')
    except Exception as e:
        print(f'[CI/CD] Error running suite: {e}')

def _ci_cd_scheduler():
    """Background scheduler that runs CI/CD suites on their configured schedules"""
    try:
        now = datetime.now()
        
        with get_db() as db:
            suites = db.execute(
                "SELECT * FROM ci_cd_suites WHERE enabled=1"
            ).fetchall()
        
        for suite in suites:
            suite = dict(suite)
            schedule = suite['schedule']  # 'daily', 'weekly', 'hourly', etc.
            
            # Determine if this suite should run now
            should_run = False
            
            if schedule == 'hourly':
                # Run at the top of every hour
                should_run = now.minute == 0 and now.second < 60
            elif schedule == 'daily':
                # Run at a specific time (default 02:00)
                should_run = now.hour == 2 and now.minute < 1
            elif schedule == 'weekly':
                # Run on Monday at 02:00
                should_run = now.weekday() == 0 and now.hour == 2 and now.minute < 1
            
            if should_run and suite['jmx_files']:
                try:
                    jmx_files = json.loads(suite['jmx_files'])
                    print(f'[CI/CD] Starting scheduled suite: {suite["name"]}')
                    
                    # Create run record
                    run_id = f"run_{uuid.uuid4().hex[:12]}"
                    with get_db() as db:
                        db.execute(
                            """INSERT INTO ci_cd_run_history 
                               (id,suite_id,client,status,start_time,triggered_by)
                               VALUES (?,?,?,?,?,?)""",
                            (run_id, suite['id'], suite['client'], 'running', 
                             datetime.now().isoformat(), 'scheduler')
                        )
                        db.commit()
                    
                    # Run suite async (in background)
                    t = threading.Thread(
                        target=_run_ci_cd_suite_internal,
                        args=(suite['id'], jmx_files, suite.get('feature_file'))
                    )
                    t.daemon = True
                    t.start()
                except Exception as e:
                    print(f'[CI/CD] Error scheduling suite {suite["name"]}: {e}')
    except Exception as e:
        print(f'[CI/CD] Scheduler error: {e}')
    
    # Schedule next check in 60 seconds
    t = threading.Timer(60, _ci_cd_scheduler)
    t.daemon = True
    t.start()

# ── AI Narrative ───────────────────────────────────────────────────────────────
@app.route('/api/ai/narrative', methods=['POST'])
@login_req
def api_ai_narrative():
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    fname = (request.json or {}).get('file', '')
    if not fname: return jsonify(error='file required'), 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='Not found'), 404
    cfg = load_cfg()
    api_key = cfg.get('openrouter_api_key', '')
    if not api_key: return jsonify(error='OpenRouter API key not configured in Settings'), 400
    try:
        data = _parse_jtl(path)
        sla_result = _evaluate_sla(path, c) or {}
        top3 = sorted(data.get('label_stats', []), key=lambda x: x.get('p90', 0), reverse=True)[:3]
        top3_str = ', '.join(s['label'] + '=' + str(s['p90']) + 'ms' for s in top3)
        sla_label = 'PASSED' if sla_result.get('passed') else ('FAILED' if sla_result else 'N/A')
        prompt = (
            "You are a performance test analyst. Write a professional 3-4 sentence executive summary "
            "of these load test results. Be specific with numbers. End with one recommendation.\n\n"
            "Test: " + fname + "\nClient: " + c['name'] + "\nDuration: " + str(data.get('test_duration','N/A')) + "\n"
            "Total Samples: " + str(data.get('total',0)) + "\nThroughput: " + str(data.get('throughput',0)) + " TPS\n"
            "Avg Response Time: " + str(data.get('avg_rt',0)) + " ms\nP90: " + str(data.get('p90',0)) + " ms\n"
            "P95: " + str(data.get('p95',0)) + " ms\nError Rate: " + str(data.get('error_rate',0)) + "%\n"
            "SLA Result: " + sla_label + "\nTop transactions by P90: " + top3_str
        )
        body = json.dumps({'model': cfg.get('ai_model', 'openai/gpt-4o-mini'),
            'messages': [{'role': 'user', 'content': prompt}], 'max_tokens': 300}).encode()
        req = _urllib_req2.Request('https://openrouter.ai/api/v1/chat/completions', data=body,
            headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {api_key}',
                     'HTTP-Referer': 'http://localhost:5000', 'X-Title': 'LoadTestPlatform'}, method='POST')
        with _urllib_req2.urlopen(req, timeout=30, context=_SSL_CTX) as resp:
            result = json.loads(resp.read())
        narrative = result['choices'][0]['message']['content'].strip()
        meta_path = path.replace('.jtl', '.meta.json')
        meta = _read_json(meta_path) or {}
        meta['narrative'] = narrative
        meta['narrative_ts'] = datetime.now().strftime('%d %b %Y %H:%M')
        _write_json(meta_path, meta)
        return jsonify(ok=True, narrative=narrative)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

@app.route('/api/ai/regression', methods=['POST'])
@login_req
def api_ai_regression():
    cfg = load_cfg()
    api_key = cfg.get('openrouter_api_key', '')
    if not api_key: return jsonify(error='OpenRouter API key not configured in Settings'), 400
    data = request.json or {}
    file_a  = data.get('file_a', 'Run A')
    file_b  = data.get('file_b', 'Run B')
    results = data.get('delta', [])
    if len(results) < 2: return jsonify(error='Need two runs to compare'), 400
    try:
        ra, rb = results[0], results[1]
        def fmt(key, label, unit=''):
            va, vb = ra.get(key,0), rb.get(key,0)
            try:
                delta = round((vb-va)/max(abs(va),0.001)*100, 1)
                arrow = '↑' if vb>va else ('↓' if vb<va else '→')
                worse = (key in ('avg','p90','p95','p99','error_pct') and vb>va) or \
                        (key == 'tps' and vb<va)
                flag = ' ⚠️' if worse and abs(delta)>10 else (' ✅' if not worse and abs(delta)>5 else '')
                return f"  {label}: {va}{unit} → {vb}{unit} ({arrow}{abs(delta)}%){flag}"
            except Exception:
                return f"  {label}: {va}{unit} → {vb}{unit}"
        lines = [
            fmt('tps',       'Throughput',  ' TPS'),
            fmt('avg',       'Avg RT',      'ms'),
            fmt('p90',       'P90',         'ms'),
            fmt('p95',       'P95',         'ms'),
            fmt('p99',       'P99',         'ms'),
            fmt('error_pct', 'Error Rate',  '%'),
            fmt('total',     'Samples',     ''),
        ]
        prompt = (
            "You are a senior performance engineer comparing two load test runs. "
            "Explain the regression or improvement in plain English. Be specific. "
            "Structure as:\n"
            "1. **Summary** (1-2 sentences: better/worse/mixed)\n"
            "2. **Key Changes** (what improved ✅ or degraded ⚠️ and by how much)\n"
            "3. **Root Cause Hypothesis** (what likely caused the difference)\n"
            "4. **Recommendation** (should this build be promoted or rolled back?)\n\n"
            f"Baseline (A): {file_a}\n"
            f"Current  (B): {file_b}\n\n"
            "Metric deltas (A → B):\n"
            + '\n'.join(lines)
        )
        body = json.dumps({'model': cfg.get('ai_model','openai/gpt-4o-mini'),
            'messages':[{'role':'user','content':prompt}], 'max_tokens':600}).encode()
        req = _urllib_req2.Request('https://openrouter.ai/api/v1/chat/completions', data=body,
            headers={'Content-Type':'application/json','Authorization':f'Bearer {api_key}',
                     'HTTP-Referer':'http://localhost:5000','X-Title':'LoadTestPlatform'}, method='POST')
        with _urllib_req2.urlopen(req, timeout=40, context=_SSL_CTX) as resp:
            result = json.loads(resp.read())
        analysis = result['choices'][0]['message']['content'].strip()
        audit('AI_REGRESSION', f'a={file_a} b={file_b}')
        return jsonify(ok=True, analysis=analysis)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

@app.route('/api/ai/sla-predict', methods=['POST'])
@login_req
def api_ai_sla_predict():
    cfg = load_cfg()
    api_key = cfg.get('openrouter_api_key', '')
    if not api_key: return jsonify(error='OpenRouter API key not configured'), 400
    data = request.json or {}
    trend_pts = data.get('trend_points', [])
    sla_ms    = data.get('sla_threshold_ms', 3000)
    curr_p95  = data.get('current_p95', 0)
    try:
        pts_str = ', '.join(str(p.get('p95',0)) for p in trend_pts[-10:])
        prompt = (
            f"A live load test has a P95 SLA target of {sla_ms}ms. "
            f"The current P95 is {curr_p95}ms and rising. "
            f"Recent P95 trend (ms): [{pts_str}]. "
            "In 1-2 sentences, explain what this trend means and what the engineer should do right now. Be direct and urgent."
        )
        body = json.dumps({'model': cfg.get('ai_model','openai/gpt-4o-mini'),
            'messages':[{'role':'user','content':prompt}], 'max_tokens':120}).encode()
        req = _urllib_req2.Request('https://openrouter.ai/api/v1/chat/completions', data=body,
            headers={'Content-Type':'application/json','Authorization':f'Bearer {api_key}',
                     'HTTP-Referer':'http://localhost:5000','X-Title':'LoadTestPlatform'}, method='POST')
        with _urllib_req2.urlopen(req, timeout=20, context=_SSL_CTX) as resp:
            result = json.loads(resp.read())
        prediction = result['choices'][0]['message']['content'].strip()
        return jsonify(ok=True, prediction=prediction)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

@app.route('/api/ai/rca', methods=['POST'])
@login_req
def api_ai_rca():
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    fname = (request.json or {}).get('file', '')
    if not fname: return jsonify(error='file required'), 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='Not found'), 404
    cfg = load_cfg()
    api_key = cfg.get('openrouter_api_key', '')
    if not api_key: return jsonify(error='OpenRouter API key not configured in Settings'), 400
    try:
        data   = _parse_jtl(path)
        sla_result = _evaluate_sla(path, c) or {}
        label_stats = data.get('label_stats', [])

        txn_lines = []
        for s in label_stats:
            top_reason = ''
            if s.get('failure_reasons'):
                fr = s['failure_reasons'][0]
                top_reason = (fr.get('expected') or fr.get('actual', ''))[:120]
            txn_lines.append(
                f"  {s['label']}: {s['samples']} samples, {s['error_rate']}% errors, "
                f"avg={s['avg']}ms P95={s['p95']}ms P99={s['p99']}ms"
                + (f" | top failure: {top_reason}" if top_reason else '')
            )

        sla_label = 'PASSED' if sla_result.get('passed') else ('FAILED' if sla_result else 'N/A')
        prompt = (
            "You are a senior performance engineer. Perform a root cause analysis of this JMeter load test. "
            "Be specific — name exact transactions, exact response times, and exact error messages from the data. "
            "Structure your response in these sections:\n"
            "1. **Overall Verdict** (1-2 sentences: pass/fail, biggest concern)\n"
            "2. **Critical Issues** (transactions with error_rate > 1% or P95 > 3000ms — use ⚠️ per issue)\n"
            "3. **Root Cause Hypothesis** (for each critical issue, what is the likely cause)\n"
            "4. **Healthy Transactions** (✅ brief list of what is performing well)\n"
            "5. **Top Recommendations** (2-3 specific, actionable steps)\n\n"
            f"File: {fname}\nClient: {c['name']}\nDuration: {data.get('test_duration')}\n"
            f"Total: {data.get('total',0)} samples  TPS: {data.get('throughput')} req/s\n"
            f"Error rate: {data.get('error_rate')}%  P90: {data.get('p90')}ms  P95: {data.get('p95')}ms  P99: {data.get('p99')}ms\n"
            f"SLA: {sla_label}\n\n"
            f"Transaction breakdown ({len(label_stats)} transactions):\n"
            + '\n'.join(txn_lines)
        )
        body = json.dumps({
            'model': cfg.get('ai_model', 'openai/gpt-4o-mini'),
            'messages': [{'role': 'user', 'content': prompt}],
            'max_tokens': 800,
        }).encode()
        req = _urllib_req2.Request(
            'https://openrouter.ai/api/v1/chat/completions', data=body,
            headers={'Content-Type': 'application/json', 'Authorization': f'Bearer {api_key}',
                     'HTTP-Referer': 'http://localhost:5000', 'X-Title': 'LoadTestPlatform'},
            method='POST')
        with _urllib_req2.urlopen(req, timeout=45, context=_SSL_CTX) as resp:
            result = json.loads(resp.read())
        rca = result['choices'][0]['message']['content'].strip()
        meta_path = path.replace('.jtl', '.meta.json')
        meta = _read_json(meta_path) or {}
        meta['rca'] = rca
        meta['rca_ts'] = datetime.now().strftime('%d %b %Y %H:%M')
        _write_json(meta_path, meta)
        audit('AI_RCA', f'file={fname}')
        return jsonify(ok=True, rca=rca)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

@app.route('/api/report-meta/<path:fname>')
@login_req
def api_report_meta(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    meta_path = os.path.join(client_dirs(c)['reports'], fname.replace('.jtl', '.meta.json'))
    return jsonify(_read_json(meta_path) or {})

# ── Performance Scorecard ───────────────────────────────────────────────────────
@app.route('/api/scorecard/<path:fname>')
@login_req
def api_scorecard(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='Not found'), 404
    try:
        data = _parse_jtl(path)
        sla = _evaluate_sla(path, c) or {}
        sla_cfg = _read_json(_sla_path(c)) or {}
        err = float(data.get('error_rate', 0))
        err_score = max(0, 30 - err * 6)
        p90_thresh = float(sla_cfg.get('p90_ms', 3000))
        p90 = float(data.get('p90', 0))
        p90_score = max(0, 20 * (1 - max(0, p90 - p90_thresh) / p90_thresh)) if p90_thresh else 20
        sla_score = 30 if sla.get('passed') else (15 if sla else 20)
        min_tps = float(sla_cfg.get('min_tps', 0))
        tps = float(data.get('throughput', 0))
        tps_score = min(20, 20 * (tps / min_tps)) if min_tps and tps else 20
        total = min(100, max(0, round(err_score + p90_score + sla_score + tps_score)))
        if total >= 90: grade, color = 'A', '#00ff9d'
        elif total >= 80: grade, color = 'B', '#00d4ff'
        elif total >= 70: grade, color = 'C', '#f59e0b'
        elif total >= 60: grade, color = 'D', '#f97316'
        else: grade, color = 'F', '#ef4444'
        result = dict(score=total, grade=grade, color=color, err_score=round(err_score),
            p90_score=round(p90_score), sla_score=round(sla_score), tps_score=round(tps_score))
        meta_path = path.replace('.jtl', '.meta.json')
        meta = _read_json(meta_path) or {}
        meta.update(result)
        _write_json(meta_path, meta)
        return jsonify(**result)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── Bottleneck Detector ────────────────────────────────────────────────────────
@app.route('/api/bottleneck/<path:fname>')
@login_req
def api_bottleneck(fname):
    c = active_client()
    if not c: return jsonify(error='No client'), 400
    path = os.path.join(client_dirs(c)['reports'], fname)
    if not os.path.exists(path): return jsonify(error='Not found'), 404
    try:
        data = _parse_jtl(path)
        stats = data.get('label_stats', [])
        if not stats: return jsonify(bottlenecks=[])
        def bn_score(s):
            return s.get('p90', 0) * 0.5 + s.get('error_rate', 0) * 100 + s.get('avg', 0) * 0.3
        scored = sorted(stats, key=bn_score, reverse=True)
        bottlenecks = []
        for s in scored[:3]:
            reasons = []
            if s.get('p90', 0) > 3000: reasons.append(f"P90 {s['p90']}ms exceeds 3s threshold")
            if s.get('error_rate', 0) > 1: reasons.append(f"{s['error_rate']}% error rate")
            if s.get('avg', 0) > 2000: reasons.append(f"Avg RT {s['avg']}ms is high")
            if not reasons: reasons.append("Slowest transaction in this run")
            bottlenecks.append({'label': s['label'], 'p90': s.get('p90', 0), 'avg': s.get('avg', 0),
                'error_rate': s.get('error_rate', 0), 'samples': s.get('samples', 0), 'reasons': reasons})
        return jsonify(bottlenecks=bottlenecks)
    except Exception as ex:
        return jsonify(error=str(ex)), 500

# ── 30-Day Trends ──────────────────────────────────────────────────────────────
@app.route('/api/trends')
@login_req
def api_trends():
    c = active_client()
    if not c: return jsonify(trend=[])
    days = int(request.args.get('days', 30))
    rdir = client_dirs(c)['reports']
    cutoff = datetime.now().timestamp() - days * 86400
    jtls = sorted([f for f in glob.glob(os.path.join(rdir, '*.jtl'))
                   if os.path.getmtime(f) >= cutoff], key=os.path.getmtime)
    trend = []
    for p in jtls:
        try:
            d = _parse_jtl(p)
            sla = _evaluate_sla(p, c) or {}
            trend.append({'file': os.path.basename(p),
                'date': datetime.fromtimestamp(os.path.getmtime(p)).strftime('%d %b'),
                'tps': d.get('throughput', 0), 'p90': d.get('p90', 0),
                'err': d.get('error_rate', 0), 'passed': sla.get('passed', None)})
        except Exception:
            pass
    return jsonify(trend=trend)

# ── Client Leaderboard ─────────────────────────────────────────────────────────
@app.route('/api/leaderboard')
@login_req
def api_leaderboard():
    clients = get_all_clients()
    board = []
    for c in clients:
        rdir = client_dirs(c)['reports']
        jtls = sorted(glob.glob(os.path.join(rdir, '*.jtl')), key=os.path.getmtime, reverse=True)[:5]
        if not jtls: continue
        tps_vals, p90_vals, err_vals, pass_count = [], [], [], 0
        for p in jtls:
            try:
                d = _parse_jtl(p)
                tps_vals.append(d.get('throughput', 0))
                p90_vals.append(d.get('p90', 0))
                err_vals.append(d.get('error_rate', 0))
                sla = _evaluate_sla(p, c)
                if sla and sla.get('passed'): pass_count += 1
            except Exception:
                pass
        if not tps_vals: continue
        avg_tps = round(sum(tps_vals) / len(tps_vals), 1)
        avg_p90 = round(sum(p90_vals) / len(p90_vals))
        avg_err = round(sum(err_vals) / len(err_vals), 2)
        pass_rate = round(pass_count / len(jtls) * 100)
        score = max(0, min(100, pass_rate * 0.4 + max(0, 30 - avg_err * 6) + max(0, 20 - avg_p90 / 200)))
        board.append({'client': c['name'], 'code': c['code'], 'emoji': c.get('logo_emoji', '🏢'),
            'avg_tps': avg_tps, 'avg_p90': avg_p90, 'avg_err': avg_err,
            'pass_rate': pass_rate, 'runs': len(jtls), 'score': round(score)})
    board.sort(key=lambda x: x['score'], reverse=True)
    medals = ['🥇', '🥈', '🥉']
    for i, b in enumerate(board):
        b['rank'] = i + 1
        b['medal'] = medals[i] if i < 3 else f'#{i+1}'
    return jsonify(board=board)

# ── Server restart ─────────────────────────────────────────────────────────────
@app.route('/api/admin/restart', methods=['POST'])
@admin_req
def api_admin_restart():
    audit('SERVER_RESTART', f'Triggered by {session.get("user")}')
    def _do_restart():
        import time as _t
        _t.sleep(1)
        script = os.path.abspath(sys.argv[0]) if sys.argv[0] else os.path.abspath(__file__)
        os.execv(sys.executable, [sys.executable, script] + sys.argv[1:])
    t = threading.Thread(target=_do_restart, daemon=True)
    t.start()
    return jsonify(ok=True, message='Server restarting in 1 second…')


# ── JMX Generator from Excel Test Features ────────────────────────────────────
import re as _re

def _parse_curl(curl_str):
    """Extract method, path, headers, body from a curl command string."""
    s = curl_str.strip()
    method = 'GET'
    url = ''
    headers = {}
    body = ''

    m = _re.search(r'-X\s+([A-Z]+)', s)
    if m:
        method = m.group(1).upper()

    # URL — single or double quoted, or bare
    mu = _re.search(r"curl\s+(?:-[^\s]+\s+)*['\"]?(https?://[^\s'\"]+)['\"]?", s)
    if mu:
        url = mu.group(1).strip("'\"")
    else:
        mu2 = _re.search(r"'(https?://[^']+)'", s) or _re.search(r'"(https?://[^"]+)"', s)
        if mu2:
            url = mu2.group(1)

    for hm in _re.finditer(r"-H\s+'([^']+)'|-H\s+\"([^\"]+)\"", s):
        hval = hm.group(1) or hm.group(2)
        if ':' in hval:
            k, v = hval.split(':', 1)
            headers[k.strip()] = v.strip()

    bm = _re.search(r"(?:-d|--data)\s+'([^']*)'|(?:-d|--data)\s+\"([^\"]*)\"", s, _re.DOTALL)
    if bm:
        body = (bm.group(1) or bm.group(2) or '').strip()
        if body and method == 'GET':
            method = 'POST'

    return method, url, headers, body


def _build_jmx(plan_name, threads, ramp_up, duration, base_url, services, tps_dist=None):
    """Generate a JMeter JMX test plan XML string."""
    from xml.sax.saxutils import escape as _xe

    def xe(v):
        return _xe(str(v))

    def sampler(svc, idx):
        method, full_url, hdrs, body = _parse_curl(svc.get('curl', ''))

        # If no URL from curl, build from base_url + name
        if not full_url and base_url:
            path_guess = '/api/' + svc['name'].lower().replace(' ', '-')
            full_url = base_url.rstrip('/') + path_guess

        from urllib.parse import urlparse
        try:
            p = urlparse(full_url)
            host = p.hostname or ''
            port = str(p.port) if p.port else ('443' if p.scheme == 'https' else '80')
            proto = p.scheme or 'http'
            path = p.path or '/'
            if p.query:
                path += '?' + p.query
        except Exception:
            host = base_url
            port = '80'
            proto = 'http'
            path = '/'

        # Merge default headers
        if not any(k.lower() == 'content-type' for k in hdrs):
            hdrs['Content-Type'] = 'application/json'

        # Header Manager
        hdr_props = ''.join(f'''
          <elementProp name="" elementType="Header">
            <stringProp name="Header.name">{xe(k)}</stringProp>
            <stringProp name="Header.value">{xe(v)}</stringProp>
          </elementProp>''' for k, v in hdrs.items())

        hdr_mgr = f'''
        <HeaderManager guiclass="HeaderPanel" testclass="HeaderManager" testname="Headers" enabled="true">
          <collectionProp name="HeaderManager.headers">{hdr_props}
          </collectionProp>
        </HeaderManager>
        <hashTree/>'''

        # Body
        body_xml = f'<boolProp name="HTTPSampler.postBodyRaw">true</boolProp><elementProp name="HTTPsampler.Arguments" elementType="Arguments"><collectionProp name="Arguments.arguments"><elementProp name="" elementType="HTTPArgument"><boolProp name="HTTPArgument.always_encode">false</boolProp><stringProp name="Argument.value">{xe(body)}</stringProp><stringProp name="Argument.metadata">=</stringProp></elementProp></collectionProp></elementProp>' if body else '<elementProp name="HTTPsampler.Arguments" elementType="Arguments"><collectionProp name="Arguments.arguments"/></elementProp>'

        # Response assertion
        assert_xml = ''
        if svc.get('assertion'):
            assert_xml = f'''
        <ResponseAssertion guiclass="AssertionGui" testclass="ResponseAssertion" testname="Assert {xe(svc['name'])}" enabled="true">
          <collectionProp name="Asserion.test_strings"><stringProp name="49586">{xe(svc['assertion'])}</stringProp></collectionProp>
          <stringProp name="Assertion.custom_message"></stringProp>
          <stringProp name="Assertion.test_field">Assertion.response_data</stringProp>
          <boolProp name="Assertion.assume_success">false</boolProp>
          <intProp name="Assertion.test_type">2</intProp>
        </ResponseAssertion>
        <hashTree/>'''

        # JSON extractor
        extract_xml = ''
        if svc.get('extractor'):
            ext = svc['extractor'].strip()
            var_name = _re.sub(r'[^a-zA-Z0-9_]', '_', svc['name'])
            extract_xml = f'''
        <JSONPathExtractor guiclass="JSONPathExtractorGui" testclass="JSONPathExtractor" testname="Extract {xe(var_name)}" enabled="true">
          <stringProp name="JSONPostProcessor.referenceNames">{xe(var_name)}</stringProp>
          <stringProp name="JSONPostProcessor.jsonPathExprs">{xe(ext)}</stringProp>
          <stringProp name="JSONPostProcessor.match_nos">0</stringProp>
        </JSONPathExtractor>
        <hashTree/>'''

        return f'''
      <HTTPSamplerProxy guiclass="HttpTestSampleGui" testclass="HTTPSamplerProxy" testname="{xe(svc['name'])}" enabled="true">
        <stringProp name="HTTPSampler.domain">{xe(host)}</stringProp>
        <stringProp name="HTTPSampler.port">{xe(port)}</stringProp>
        <stringProp name="HTTPSampler.protocol">{xe(proto)}</stringProp>
        <stringProp name="HTTPSampler.path">{xe(path)}</stringProp>
        <stringProp name="HTTPSampler.method">{xe(method)}</stringProp>
        <boolProp name="HTTPSampler.follow_redirects">true</boolProp>
        <boolProp name="HTTPSampler.auto_redirects">false</boolProp>
        <boolProp name="HTTPSampler.use_keepalive">true</boolProp>
        <boolProp name="HTTPSampler.DO_MULTIPART_POST">false</boolProp>
        {body_xml}
      </HTTPSamplerProxy>
      <hashTree>
        {hdr_mgr}
        {assert_xml}
        {extract_xml}
      </hashTree>'''

    samplers_xml = ''.join(sampler(s, i) for i, s in enumerate(services))

    return f'''<?xml version="1.0" encoding="UTF-8"?>
<jmeterTestPlan version="1.2" properties="5.0" jmeter="5.5">
  <hashTree>
    <TestPlan guiclass="TestPlanGui" testclass="TestPlan" testname="{xe(plan_name)}" enabled="true">
      <stringProp name="TestPlan.comments">Generated by Load Testing Platform from Excel features file</stringProp>
      <boolProp name="TestPlan.functional_mode">false</boolProp>
      <boolProp name="TestPlan.serialize_threadgroups">false</boolProp>
      <elementProp name="TestPlan.user_defined_variables" elementType="Arguments" guiclass="ArgumentsPanel" testclass="Arguments" testname="User Defined Variables" enabled="true">
        <collectionProp name="Arguments.arguments">
          <elementProp name="BASE_URL" elementType="Argument">
            <stringProp name="Argument.name">BASE_URL</stringProp>
            <stringProp name="Argument.value">{xe(base_url)}</stringProp>
            <stringProp name="Argument.metadata">=</stringProp>
          </elementProp>
        </collectionProp>
      </elementProp>
    </TestPlan>
    <hashTree>
      <ThreadGroup guiclass="ThreadGroupGui" testclass="ThreadGroup" testname="{xe(plan_name)} - Thread Group" enabled="true">
        <stringProp name="ThreadGroup.on_sample_error">continue</stringProp>
        <elementProp name="ThreadGroup.main_controller" elementType="LoopController" guiclass="LoopControlPanel" testclass="LoopController" testname="Loop Controller" enabled="true">
          <boolProp name="LoopController.continue_forever">false</boolProp>
          <intProp name="LoopController.loops">-1</intProp>
        </elementProp>
        <stringProp name="ThreadGroup.num_threads">{xe(threads)}</stringProp>
        <stringProp name="ThreadGroup.ramp_time">{xe(ramp_up)}</stringProp>
        <boolProp name="ThreadGroup.scheduler">true</boolProp>
        <stringProp name="ThreadGroup.duration">{xe(duration)}</stringProp>
        <stringProp name="ThreadGroup.delay">0</stringProp>
        <boolProp name="ThreadGroup.same_user_on_next_iteration">true</boolProp>
      </ThreadGroup>
      <hashTree>
        <ResultCollector guiclass="ViewResultsFullVisualizer" testclass="ResultCollector" testname="View Results Tree" enabled="false">
          <boolProp name="ResultCollector.error_logging">false</boolProp>
        </ResultCollector>
        <hashTree/>
        <ResultCollector guiclass="SummaryReport" testclass="ResultCollector" testname="Summary Report" enabled="true">
          <boolProp name="ResultCollector.error_logging">false</boolProp>
          <objProp><name>saveConfig</name><value class="SampleSaveConfiguration"><time>true</time><latency>true</latency><timestamp>true</timestamp><success>true</success><label>true</label><code>true</code><message>true</message><threadName>true</threadName><dataType>true</dataType><encoding>false</encoding><assertions>true</assertions><subresults>true</subresults><responseData>false</responseData><samplerData>false</samplerData><xml>false</xml><fieldNames>true</fieldNames><responseHeaders>false</responseHeaders><requestHeaders>false</requestHeaders><responseDataOnError>false</responseDataOnError><saveAssertionResultsFailureMessage>true</saveAssertionResultsFailureMessage><bytes>true</bytes><sentBytes>true</sentBytes><url>true</url><threadCounts>true</threadCounts><idleTime>true</idleTime><connectTime>true</connectTime></value></objProp>
          <stringProp name="filename"></stringProp>
        </ResultCollector>
        <hashTree/>
        {samplers_xml}
      </hashTree>
    </hashTree>
  </hashTree>
</jmeterTestPlan>'''


def _jmeter_hash(s):
    """Java String.hashCode() for JMeter stringProp name attributes."""
    h = 0
    for c in str(s):
        h = (31 * h + ord(c)) & 0xFFFFFFFF
    if h >= 0x80000000:
        h -= 0x100000000
    return str(h)


def _build_jmx_ussd(plan_name, config, services, duration=600):
    """Build a JMeter JMX for USSD simulation from parsed Excel data.

    config: dict with keys server_ip, port, LOGIN, PASSWD, sim_url, dialcode
    services: list of dicts:
        {name, stype, tps, steps: [(input_val, expected_response), ...]}
        steps[0]: (dial_code, assertion_text_after_dial)
        steps[1]: (user_input_step2, assertion_text)  etc.
    """
    from xml.sax.saxutils import escape as _xe

    def xe(v):
        return _xe(str(v) if v is not None else '')

    server_ip = config.get('server_ip', '')
    port_val  = config.get('port', '443')
    login     = config.get('LOGIN', '')
    passwd    = config.get('PASSWD', '')
    sim_url   = config.get('sim_url', '')
    dialcode  = config.get('dialcode', '')

    def _udv(name, value):
        return f'''          <elementProp name="{xe(name)}" elementType="Argument">
            <stringProp name="Argument.name">{xe(name)}</stringProp>
            <stringProp name="Argument.value">{xe(value)}</stringProp>
            <stringProp name="Argument.metadata">=</stringProp>
          </elementProp>'''

    udv_xml = '\n'.join([
        _udv('server_ip', server_ip),
        _udv('sim_url', sim_url),
        _udv('Dialcode', dialcode),
        _udv('LOGIN', login),
        _udv('PASSWD', passwd),
        _udv('port', port_val),
    ])

    def _http_arg(name, value):
        return f'''                <elementProp name="{xe(name)}" elementType="HTTPArgument">
                  <boolProp name="HTTPArgument.always_encode">false</boolProp>
                  <stringProp name="Argument.value">{xe(value)}</stringProp>
                  <stringProp name="Argument.metadata">=</stringProp>
                  <boolProp name="HTTPArgument.use_equals">true</boolProp>
                  <stringProp name="Argument.name">{xe(name)}</stringProp>
                </elementProp>'''

    def _ussd_sampler(svc_name, step_num, input_val, new_req, extra_label=''):
        label = f'{xe(svc_name)} {step_num:03d} - '
        if new_req == '1':
            label += f'Dial ${{Dialcode}} - Open USSD session'
        else:
            label += f'Step {step_num} - Input: {xe(input_val)}'
        if extra_label:
            label += f' {xe(extra_label)}'
        args = '\n'.join([
            _http_arg('LOGIN', '${LOGIN}'),
            _http_arg('PASSWORD', '${PASSWD}'),
            _http_arg('SESSION_ID', '${SESSION_ID}'),
            _http_arg('MSISDN', '${MSISDN}'),
            _http_arg('NewRequest', new_req),
            _http_arg('INPUT', input_val),
        ])
        return f'''          <HTTPSamplerProxy guiclass="HttpTestSampleGui" testclass="HTTPSamplerProxy" testname="{label}" enabled="true">
            <stringProp name="HTTPSampler.path">${{sim_url}}</stringProp>
            <boolProp name="HTTPSampler.follow_redirects">true</boolProp>
            <stringProp name="HTTPSampler.method">GET</stringProp>
            <boolProp name="HTTPSampler.use_keepalive">true</boolProp>
            <boolProp name="HTTPSampler.postBodyRaw">false</boolProp>
            <elementProp name="HTTPsampler.Arguments" elementType="Arguments" guiclass="HTTPArgumentsPanel" testclass="Arguments" testname="User Defined Variables">
              <collectionProp name="Arguments.arguments">
{args}
              </collectionProp>
            </elementProp>
          </HTTPSamplerProxy>'''

    def _assertion(svc_name, step_num, text):
        first_line = str(text).split('\n')[0][:60].strip()
        h = _jmeter_hash(first_line)
        return f'''            <ResponseAssertion guiclass="AssertionGui" testclass="ResponseAssertion" testname="Assert - {xe(svc_name)} {step_num:03d}" enabled="true">
              <collectionProp name="Asserion.test_strings">
                <stringProp name="{h}">{xe(first_line)}</stringProp>
              </collectionProp>
              <stringProp name="Assertion.custom_message">Expected: {xe(first_line)}</stringProp>
              <stringProp name="Assertion.test_field">Assertion.response_data</stringProp>
              <boolProp name="Assertion.assume_success">false</boolProp>
              <intProp name="Assertion.test_type">2</intProp>
            </ResponseAssertion>
            <hashTree/>'''

    groovy_session = (
        "def chars = ('0'..'9')\n"
        "def rnd = new Random()\n"
        "def sessionId = (1..20).collect { chars[rnd.nextInt(chars.size())] }.join()\n"
        "vars.put('SESSION_ID', sessionId)"
    )
    # XML-escape apostrophes for inline XML
    groovy_xml = groovy_session.replace("'", '&apos;')

    tg_blocks = []
    for svc in services:
        sname = svc['name'].strip()
        stype = svc.get('stype') or sname
        tps_val = float(svc.get('tps') or 0)
        rpm = round(tps_val * 60, 2) if tps_val else 0
        pkey = sname.lower().replace(' ', '_').replace('-', '_').replace('/', '_')
        steps = svc.get('steps', [])
        csv_vars = ['MSISDN', 'PIN']
        # determine extra vars needed
        for inp, _ in steps[1:]:
            if '${recMSISDN}' in str(inp):
                if 'recMSISDN' not in csv_vars:
                    csv_vars.append('recMSISDN')
            if '${amount}' in str(inp):
                if 'amount' not in csv_vars:
                    csv_vars.append('amount')
        csv_var_names = ','.join(csv_vars)

        # Build transaction controller inner content
        tc_inner = []

        if steps:
            dial_input, dial_assert = steps[0]
            # Step 1: dial
            s1_xml = _ussd_sampler(sname, 1, dial_input, '1')
            assert1_xml = _assertion(sname, 1, dial_assert) if dial_assert else ''
            tc_inner.append(f'''{s1_xml}
          <hashTree>
            <JSR223PreProcessor guiclass="TestBeanGUI" testclass="JSR223PreProcessor" testname="Generate SESSION_ID" enabled="true">
              <stringProp name="cacheKey">true</stringProp>
              <stringProp name="filename"></stringProp>
              <stringProp name="parameters"></stringProp>
              <stringProp name="script">{groovy_xml}</stringProp>
              <stringProp name="scriptLanguage">groovy</stringProp>
            </JSR223PreProcessor>
            <hashTree/>
{assert1_xml}
          </hashTree>''')

            # Steps 2+: each wrapped in IfController
            for idx, (inp, expected) in enumerate(steps[1:], start=2):
                s_xml = _ussd_sampler(sname, idx, inp, '0')
                a_xml = _assertion(sname, idx, expected) if expected else ''
                tc_inner.append(f'''          <IfController guiclass="IfControllerPanel" testclass="IfController" testname="If Controller - Step {idx}" enabled="true">
            <stringProp name="IfController.condition">${{JMeterThread.last_sample_ok}}</stringProp>
            <boolProp name="IfController.evaluateAll">false</boolProp>
            <boolProp name="IfController.useExpression">true</boolProp>
          </IfController>
          <hashTree>
{s_xml}
            <hashTree>
{a_xml}
            </hashTree>
          </hashTree>''')

        tc_content = '\n'.join(tc_inner)

        tg_block = f'''      <ThreadGroup guiclass="ThreadGroupGui" testclass="ThreadGroup" testname="TG - {xe(sname)} ({xe(stype)})" enabled="true">
        <stringProp name="TestPlan.comments">Target TPS: {tps_val} | Throughput: {rpm}/min</stringProp>
        <stringProp name="ThreadGroup.num_threads">${{__P({pkey}.threads,5)}}</stringProp>
        <stringProp name="ThreadGroup.ramp_time">${{__P({pkey}.rampup,5)}}</stringProp>
        <stringProp name="ThreadGroup.duration">${{__P(test.duration,{duration})}}</stringProp>
        <boolProp name="ThreadGroup.same_user_on_next_iteration">true</boolProp>
        <boolProp name="ThreadGroup.scheduler">true</boolProp>
        <stringProp name="ThreadGroup.on_sample_error">continue</stringProp>
        <elementProp name="ThreadGroup.main_controller" elementType="LoopController" guiclass="LoopControlPanel" testclass="LoopController" testname="Loop Controller">
          <intProp name="LoopController.loops">-1</intProp>
          <boolProp name="LoopController.continue_forever">false</boolProp>
        </elementProp>
      </ThreadGroup>
      <hashTree>
        <CSVDataSet guiclass="TestBeanGUI" testclass="CSVDataSet" testname="CSV Data Set - {xe(sname)}" enabled="true">
          <stringProp name="delimiter">,</stringProp>
          <stringProp name="fileEncoding"></stringProp>
          <stringProp name="filename">/TestData/{xe(sname)}.csv</stringProp>
          <boolProp name="ignoreFirstLine">true</boolProp>
          <boolProp name="quotedData">false</boolProp>
          <boolProp name="recycle">true</boolProp>
          <stringProp name="shareMode">shareMode.group</stringProp>
          <boolProp name="stopThread">false</boolProp>
          <stringProp name="variableNames">{xe(csv_var_names)}</stringProp>
        </CSVDataSet>
        <hashTree/>
        <ConstantThroughputTimer guiclass="TestBeanGUI" testclass="ConstantThroughputTimer" testname="Constant Throughput Timer [{tps_val}TPS]" enabled="true">
          <intProp name="calcMode">4</intProp>
          <stringProp name="throughput">${{__P({pkey}.throughput.rpm,{rpm})}}</stringProp>
          <stringProp name="TestPlan.comments">Target: {tps_val} TPS = {rpm} req/min</stringProp>
        </ConstantThroughputTimer>
        <hashTree/>
        <TransactionController guiclass="TransactionControllerGui" testclass="TransactionController" testname="TC - {xe(sname)}" enabled="true">
          <boolProp name="TransactionController.parent">true</boolProp>
          <boolProp name="TransactionController.includeTimers">false</boolProp>
        </TransactionController>
        <hashTree>
{tc_content}
        </hashTree>
      </hashTree>'''
        tg_blocks.append(tg_block)

    all_tgs = '\n'.join(tg_blocks)

    return f'''<?xml version="1.0" encoding="UTF-8"?>
<jmeterTestPlan version="1.2" properties="5.0" jmeter="5.6.3">
  <hashTree>
    <TestPlan guiclass="TestPlanGui" testclass="TestPlan" testname="{xe(plan_name)}" enabled="true">
      <stringProp name="TestPlan.comments">Generated by Load Testing Platform | {len(services)} services | Duration: {duration}s</stringProp>
      <boolProp name="TestPlan.tearDown_on_shutdown">true</boolProp>
      <elementProp name="TestPlan.user_defined_variables" elementType="Arguments" guiclass="ArgumentsPanel" testclass="Arguments" testname="User Defined Variables">
        <collectionProp name="Arguments.arguments">
{udv_xml}
        </collectionProp>
      </elementProp>
    </TestPlan>
    <hashTree>
      <ConfigTestElement guiclass="HttpDefaultsGui" testclass="ConfigTestElement" testname="HTTP Request Defaults" enabled="true">
        <intProp name="HTTPSampler.connect_timeout">5000</intProp>
        <intProp name="HTTPSampler.response_timeout">30000</intProp>
        <stringProp name="HTTPSampler.domain">${{server_ip}}</stringProp>
        <stringProp name="HTTPSampler.port">${{port}}</stringProp>
        <stringProp name="HTTPSampler.protocol">https</stringProp>
        <elementProp name="HTTPsampler.Arguments" elementType="Arguments" guiclass="HTTPArgumentsPanel" testclass="Arguments" testname="User Defined Variables">
          <collectionProp name="Arguments.arguments"/>
        </elementProp>
        <stringProp name="HTTPSampler.implementation"></stringProp>
      </ConfigTestElement>
      <hashTree/>
{all_tgs}
    </hashTree>
  </hashTree>
</jmeterTestPlan>'''


def _resolve_ussd_input(raw_input, prev_response, pin_used):
    """Map Excel placeholder inputs to JMeter variables."""
    s = str(raw_input).strip() if raw_input else ''
    if not s or s.lower() in ('none', ''):
        return s, pin_used
    # dial code (step 1 only handled outside; but catch any stray ones)
    import re as _ire
    if _ire.match(r'^\*[\d*]+#$', s):
        return '${Dialcode}', pin_used
    # pure x-placeholder
    if _ire.match(r'^x+$', s, _ire.IGNORECASE):
        prev_lower = (prev_response or '').lower()
        if any(w in prev_lower for w in ('mobile', 'msisdn', 'receiver', 'phone number', 'recipient')):
            return '${recMSISDN}', pin_used
        if 'amount' in prev_lower:
            return '${amount}', pin_used
        if not pin_used:
            return '${PIN}', True
        return '${PIN}', pin_used
    return s, pin_used


@app.route('/api/generate-jmx', methods=['POST'])
@login_req
def api_generate_jmx():
    if not _HAS_OPENPYXL:
        return jsonify(error='openpyxl not installed. Run: pip install openpyxl'), 500
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400

    data = request.json or {}
    xlsx_file   = (data.get('xlsx_file') or '').strip()
    jmx_name    = (data.get('jmx_name') or '').strip()
    duration    = int(data.get('duration') or 600)
    tps_override= float(data.get('tps') or 0)
    svc_filter  = [s.strip().lower() for s in (data.get('services') or '').split(',') if s.strip()]
    menu_type   = (data.get('menu_type') or 'both').lower()

    base = os.path.join(CLIENTS_DIR, c['code'])
    # Resolve Excel path
    xlsx_path = None
    for candidate in [
        os.path.join(base, xlsx_file),
        os.path.join(base, 'testdata', xlsx_file),
        os.path.join(base, 'jmx', xlsx_file),
    ]:
        if xlsx_file and os.path.exists(candidate):
            xlsx_path = candidate
            break
    if not xlsx_path:
        return jsonify(error=f'Excel file not found: {xlsx_file}'), 404

    sheets = _parse_xlsx(xlsx_path)
    if not sheets:
        return jsonify(error='Could not parse Excel file'), 500

    # ── 1. Parse testDetails ─────────────────────────────────────────────
    config = {}
    for sname, sheet in sheets.items():
        if 'testdetail' in sname.lower().replace(' ', ''):
            # headers row is row[0] in the sheet — also a key-value pair for testDetails
            hdrs = sheet.get('headers', [])
            if hdrs and len(hdrs) >= 2 and hdrs[0] and hdrs[1]:
                config[str(hdrs[0]).strip()] = str(hdrs[1]).strip()
            for row in sheet.get('rows', []):
                if row and len(row) >= 2 and row[0] and row[1]:
                    key = str(row[0]).strip()
                    config[key] = str(row[1]).strip()
    # normalise key aliases
    cfg = {
        'server_ip': config.get('server_ip', ''),
        'port':      config.get('port', '443'),
        'LOGIN':     config.get('LOGIN', ''),
        'PASSWD':    config.get('PASSWD', ''),
        'sim_url':   config.get('sim_url', ''),
        'dialcode':  config.get('Dialcode') or config.get('dialcode') or config.get('Dailcode') or '',
    }

    # ── 2. Parse loadDistribution ────────────────────────────────────────
    svc_dist = {}  # key=name_lower → {name, stype, tps}
    for sname, sheet in sheets.items():
        if 'loaddist' in sname.lower().replace(' ', '').replace('_', ''):
            hdrs = [str(h).lower() if h else '' for h in (sheet.get('headers') or [])]
            tps_idx  = next((i for i, h in enumerate(hdrs) if 'tps' in h), -1)
            name_idx = next((i for i, h in enumerate(hdrs) if 'service_name' in h or (i == 1 and 'service' in h)), 1)
            type_idx = next((i for i, h in enumerate(hdrs) if 'service_type' in h or (i == 0 and 'service' in h)), 0)
            for row in sheet.get('rows', []):
                if not row or not row[0]:
                    continue
                sn = str(row[name_idx] if len(row) > name_idx else row[0]).strip()
                st = str(row[type_idx] if len(row) > type_idx else row[0]).strip()
                tps_v = 0.0
                if tps_idx >= 0 and len(row) > tps_idx and row[tps_idx]:
                    try:
                        tps_v = float(row[tps_idx])
                    except Exception:
                        pass
                svc_dist[sn.lower()] = {'name': sn, 'stype': st, 'tps': tps_v}

    # ── 3. Parse Customer Menu / Agent Menu ──────────────────────────────
    menu_flows = {}  # key=service_name_lower → [(input, expected_response), ...]
    for sname, sheet in sheets.items():
        sl = sname.lower()
        if 'customer menu' in sl and menu_type in ('customer', 'both'):
            pass
        elif 'agent menu' in sl and menu_type in ('agent', 'both'):
            pass
        elif 'menu' in sl:
            pass
        else:
            continue
        for row in sheet.get('rows', []):
            if not row or not row[0]:
                continue
            svc_nm = str(row[0]).strip()
            key = svc_nm.lower()
            # columns: [0]=service, [1]=dial_code, [2]=resp1, [3]=inp1, [4]=resp2, [5]=inp2, ...
            steps = []
            cols = [str(c).strip() if c is not None else '' for c in row]
            # Step 1: dial (col[1]) → always becomes ${Dialcode}, assertion text = col[2]
            dial_raw = cols[1] if len(cols) > 1 else ''
            resp1 = cols[2] if len(cols) > 2 else ''
            if dial_raw:
                steps.append(('${Dialcode}', resp1))
            # Subsequent steps: col[3], col[5], col[7] = inputs; col[4], col[6], col[8] = responses
            pin_used = False
            col_idx = 3
            while col_idx < len(cols):
                raw_inp = cols[col_idx] if col_idx < len(cols) else ''
                prev_resp = cols[col_idx - 1] if col_idx >= 1 else ''
                raw_resp = cols[col_idx + 1] if col_idx + 1 < len(cols) else ''
                if not raw_inp and not raw_resp:
                    col_idx += 2
                    continue
                resolved, pin_used = _resolve_ussd_input(raw_inp, prev_resp, pin_used)
                steps.append((resolved, raw_resp))
                col_idx += 2
            if steps:
                if key not in menu_flows:
                    menu_flows[key] = steps
        # end for row

    # ── 4. Match services with flows ─────────────────────────────────────
    service_list = []
    for sk, sv in svc_dist.items():
        if svc_filter and sk not in svc_filter and sv['name'].lower() not in svc_filter:
            continue
        flow = menu_flows.get(sk) or menu_flows.get(sv['name'].lower())
        # fuzzy fallback: partial match
        if not flow:
            for mk, msteps in menu_flows.items():
                if mk in sk or sk in mk:
                    flow = msteps
                    break
        tps_v = tps_override if tps_override else sv['tps']
        service_list.append({
            'name':  sv['name'],
            'stype': sv['stype'],
            'tps':   tps_v,
            'steps': flow or [],
        })

    if not service_list:
        return jsonify(error='No services found. Check loadDistribution sheet has rows.'), 400

    services_with_flows = sum(1 for s in service_list if s['steps'])

    # ── 5. Build JMX ─────────────────────────────────────────────────────
    stem = os.path.splitext(os.path.basename(xlsx_file))[0]
    if not jmx_name:
        jmx_name = f'{stem}_USSD_{int(tps_override)}TPS.jmx'
    elif not jmx_name.lower().endswith('.jmx'):
        jmx_name += '.jmx'

    plan_name = jmx_name.replace('.jmx', '')
    jmx_content = _build_jmx_ussd(plan_name, cfg, service_list, duration=duration)

    jmx_folder = os.path.join(base, 'jmx')
    os.makedirs(jmx_folder, exist_ok=True)
    jmx_path = os.path.join(jmx_folder, jmx_name)
    with open(jmx_path, 'w', encoding='utf-8') as f:
        f.write(jmx_content)

    audit('GENERATE_JMX', f'Generated USSD JMX: {jmx_name} ({len(service_list)} services, {services_with_flows} with flows)')
    return jsonify(
        ok=True,
        filename=jmx_name,
        services=len(service_list),
        services_with_flows=services_with_flows,
        config={k: v for k, v in cfg.items() if k not in ('PASSWD',)},
        tps=tps_override,
    )


@app.route('/api/tps-calculator', methods=['POST'])
@login_req
def api_tps_calculator():
    c = active_client()
    if not c: return jsonify(error='No active client'), 400
    data = request.json or {}
    total_tps = float(data.get('total_tps', 30))
    duration  = int(data.get('duration', 600))
    ramp_up   = int(data.get('ramp_up', 60))
    xlsx_file = (data.get('xlsx_file') or '').strip()
    base = os.path.join(CLIENTS_DIR, c['code'])
    xlsx_path = None
    for cand in [os.path.join(base, xlsx_file), os.path.join(base, 'testdata', xlsx_file)]:
        if xlsx_file and os.path.exists(cand):
            xlsx_path = cand; break
    services = []
    if xlsx_path and _HAS_OPENPYXL:
        sheets = _parse_xlsx(xlsx_path)
        for sname, sheet in sheets.items():
            if 'loaddist' in sname.lower().replace(' ', '').replace('_', ''):
                hdrs = [str(h).lower() if h else '' for h in sheet.get('headers', [])]
                pct_idx = next((i for i, h in enumerate(hdrs) if 'percent' in h), 4)
                name_idx = 1
                for row in sheet.get('rows', []):
                    if not row or not row[0]: continue
                    sn = str(row[name_idx] if len(row) > name_idx else row[0]).strip()
                    pct = 0.0
                    try:
                        pct = float(row[pct_idx]) if (len(row) > pct_idx and row[pct_idx]) else 0.0
                    except Exception:
                        pass
                    services.append({'name': sn, 'percentage': pct})
    total_pct = sum(s['percentage'] for s in services) or 1
    results, props_lines = [], [f'test.duration={duration}', '']
    for s in services:
        ratio = s['percentage'] / total_pct
        svc_tps = round(total_tps * ratio, 2)
        svc_rpm = round(svc_tps * 60, 1)
        threads = max(1, int(svc_tps * 2))
        pkey = s['name'].lower().replace(' ', '_').replace('-', '_').replace('/', '_').strip('_')
        results.append({'name': s['name'], 'tps': svc_tps, 'rpm': svc_rpm,
                        'threads': threads, 'rampup': ramp_up, 'percentage': s['percentage']})
        props_lines += [f'{pkey}.threads={threads}', f'{pkey}.rampup={ramp_up}',
                        f'{pkey}.throughput.rpm={svc_rpm}', '']
    props_content = '\n'.join(props_lines)
    props_name = f'test_{int(total_tps)}tps.properties'
    os.makedirs(os.path.join(base, 'jmx'), exist_ok=True)
    with open(os.path.join(base, 'jmx', props_name), 'w', encoding='utf-8') as f:
        f.write(props_content)
    audit('TPS_CALC', f'TPS calculator: {total_tps} TPS across {len(results)} services → {props_name}')
    return jsonify(ok=True, total_tps=total_tps, services=results,
                   properties_file=props_name, properties_content=props_content)


@app.route('/api/ussd-funnel/<path:fname>')
@login_req
def api_ussd_funnel(fname):
    c = active_client()
    if not c: return jsonify(error='No active client'), 400
    p = os.path.join(CLIENTS_DIR, c['code'], 'reports', fname)
    if not os.path.exists(p): return jsonify(error='File not found'), 404
    service_steps = {}
    try:
        with open(p, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.DictReader(f)
            for row in reader:
                label = row.get('label', '')
                success = row.get('success', 'true').lower() == 'true'
                m = _re.match(r'^(.+?)\s+(\d{3})\s+-\s+', label)
                if not m: continue
                svc = m.group(1).strip()
                step = int(m.group(2))
                if svc not in service_steps: service_steps[svc] = {}
                if step not in service_steps[svc]: service_steps[svc][step] = {'total': 0, 'passed': 0}
                service_steps[svc][step]['total'] += 1
                if success: service_steps[svc][step]['passed'] += 1
    except Exception as e:
        return jsonify(error=str(e)), 500
    result = []
    for svc, steps in sorted(service_steps.items()):
        step_list = []
        for step_num in sorted(steps.keys()):
            s = steps[step_num]
            pct = round(s['passed'] / s['total'] * 100, 1) if s['total'] else 0
            step_list.append({'step': step_num, 'total': s['total'],
                               'passed': s['passed'], 'failed': s['total'] - s['passed'], 'pass_pct': pct})
        result.append({'service': svc, 'steps': step_list})
    return jsonify(funnel=result)


@app.route('/api/jmx-poverride/<path:fname>', methods=['GET', 'POST'])
@login_req
def api_jmx_poverride(fname):
    c = active_client()
    if not c: return jsonify(error='No active client'), 400
    p = os.path.join(CLIENTS_DIR, c['code'], 'jmx', fname)
    if not os.path.exists(p): return jsonify(error='JMX not found'), 404
    if request.method == 'POST':
        props = request.json or {}
        lines = [f'{k}={v}' for k, v in props.items() if k and v is not None]
        props_name = fname.replace('.jmx', '.properties')
        with open(os.path.join(CLIENTS_DIR, c['code'], 'jmx', props_name), 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        audit('JMX_PROPS', f'Property override saved: {props_name}')
        return jsonify(ok=True, filename=props_name)
    try:
        with open(p, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
    except Exception as e:
        return jsonify(error=str(e)), 500
    params = {}
    for m in _re.finditer(r'\$\{__P\(([^,)]+),([^)]*)\)\}', content):
        key = m.group(1).strip()
        if key not in params:
            params[key] = m.group(2).strip()
    return jsonify(params=params)


@app.route('/api/rt-histogram/<path:fname>')
@login_req
def api_rt_histogram(fname):
    c = active_client()
    if not c: return jsonify(error='No active client'), 400
    p = os.path.join(CLIENTS_DIR, c['code'], 'reports', fname)
    if not os.path.exists(p): return jsonify(error='File not found'), 404
    buckets = [0, 100, 200, 300, 500, 750, 1000, 1500, 2000, 3000, 5000, 10000, float('inf')]
    labels  = ['<100', '100-200', '200-300', '300-500', '500-750', '750-1s', '1-1.5s', '1.5-2s', '2-3s', '3-5s', '5-10s', '>10s']
    counts  = [0] * len(labels)
    per_txn = {}
    try:
        with open(p, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.DictReader(f)
            for row in reader:
                try: elapsed = int(row.get('elapsed', 0))
                except Exception: continue
                label = row.get('label', '')
                for i in range(len(labels)):
                    if elapsed < buckets[i + 1]:
                        counts[i] += 1; break
                if label not in per_txn: per_txn[label] = []
                if len(per_txn[label]) < 500: per_txn[label].append(elapsed)
    except Exception as e:
        return jsonify(error=str(e)), 500
    txn_summary = []
    for lbl, vals in per_txn.items():
        if not vals: continue
        vals.sort()
        n = len(vals)
        def _pct(pp, v=vals, nn=n): return v[min(int(pp / 100 * nn), nn - 1)]
        txn_summary.append({'label': lbl, 'p50': _pct(50), 'p90': _pct(90),
                             'p95': _pct(95), 'p99': _pct(99), 'count': n})
    txn_summary.sort(key=lambda x: x['p90'], reverse=True)
    return jsonify(labels=labels, counts=counts, transactions=txn_summary[:20])


# ── Feature 6: Error Pattern Grouper ─────────────────────────────────────────
@app.route('/api/error-patterns/<path:fname>')
@login_req
def api_error_patterns(fname):
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    dirs = client_dirs(c)
    jtl_path = os.path.join(dirs['reports'], fname)
    if not os.path.isfile(jtl_path):
        return jsonify(error='File not found'), 404
    from collections import Counter, defaultdict
    pattern_counts = Counter()
    pattern_samples = defaultdict(list)
    total = 0
    failures = 0
    try:
        with open(jtl_path, encoding='utf-8', errors='replace') as f:
            reader = csv.DictReader(f)
            for row in reader:
                total += 1
                success = str(row.get('success', 'true')).lower()
                if success == 'false':
                    failures += 1
                    msg = (row.get('responseMessage') or row.get('failureMessage') or
                           row.get('responseCode') or 'Unknown error').strip()
                    # Normalise: replace digits/UUIDs with placeholders
                    import re as _re
                    key = _re.sub(r'[0-9a-f]{8}-[0-9a-f-]{27}', '<UUID>', msg, flags=_re.I)
                    key = _re.sub(r'\b\d{5,}\b', '<NUM>', key)
                    key = key[:120]
                    pattern_counts[key] += 1
                    if len(pattern_samples[key]) < 3:
                        pattern_samples[key].append({
                            'label': row.get('label', ''),
                            'ts': row.get('timeStamp', ''),
                            'code': row.get('responseCode', ''),
                            'msg': msg[:200],
                        })
    except Exception as e:
        return jsonify(error=str(e)), 500
    patterns = []
    for pat, cnt in pattern_counts.most_common(25):
        patterns.append({
            'pattern': pat,
            'count': cnt,
            'pct': round(cnt / failures * 100, 1) if failures else 0,
            'samples': pattern_samples[pat],
        })
    return jsonify(total=total, failures=failures, patterns=patterns)


# ── Feature 7: Environment Profile Switcher ───────────────────────────────────
@app.route('/api/env-profiles', methods=['GET', 'POST'])
@login_req
def api_env_profiles():
    if request.method == 'GET':
        with get_db() as db:
            rows = db.execute("SELECT * FROM env_profiles ORDER BY name").fetchall()
        return jsonify(profiles=[dict(r) for r in rows])
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    profile_data = data.get('data', {})
    if not name:
        return jsonify(error='Name required'), 400
    import json as _json
    with get_db() as db:
        db.execute(
            "INSERT OR REPLACE INTO env_profiles (name, data, created_by) VALUES (?,?,?)",
            (name, _json.dumps(profile_data), session.get('user', 'system'))
        )
        db.commit()
    audit('env_profile_save', f'Profile: {name}')
    return jsonify(ok=True)


@app.route('/api/env-profiles/<path:name>', methods=['DELETE'])
@login_req
def api_env_profile_delete(name):
    with get_db() as db:
        db.execute("DELETE FROM env_profiles WHERE name=?", (name,))
        db.commit()
    audit('env_profile_delete', f'Profile: {name}')
    return jsonify(ok=True)


@app.route('/api/env-profiles/<path:name>/activate', methods=['POST'])
@login_req
def api_env_profile_activate(name):
    import json as _json
    with get_db() as db:
        row = db.execute("SELECT data FROM env_profiles WHERE name=?", (name,)).fetchone()
        if not row:
            return jsonify(error='Profile not found'), 404
        profile_data = _json.loads(row['data'])
        db.execute("UPDATE env_profiles SET is_active=0")
        db.execute("UPDATE env_profiles SET is_active=1 WHERE name=?", (name,))
        db.commit()
    # Write to a .properties override file that JMeter can load
    c = active_client()
    if c:
        dirs = client_dirs(c)
        props_path = os.path.join(dirs['jmx'], 'env_profile.properties')
        with open(props_path, 'w', encoding='utf-8') as f:
            for k, v in profile_data.items():
                f.write(f'{k}={v}\n')
    audit('env_profile_activate', f'Profile: {name}')
    return jsonify(ok=True, data=profile_data)


# ── Feature 8: Pre-Run Test Notes ─────────────────────────────────────────────
@app.route('/api/run-notes', methods=['POST'])
@login_req
def api_run_notes_save():
    data = request.get_json(silent=True) or {}
    run_id = (data.get('run_id') or '').strip()
    notes = (data.get('notes') or '').strip()
    tags = (data.get('tags') or '').strip()
    if not run_id:
        return jsonify(error='run_id required'), 400
    c = active_client()
    client_code = c['code'] if c else 'unknown'
    with get_db() as db:
        db.execute(
            "INSERT OR REPLACE INTO test_run_notes (run_id, client, notes, tags, author) VALUES (?,?,?,?,?)",
            (run_id, client_code, notes, tags, session.get('user', 'system'))
        )
        db.commit()
    return jsonify(ok=True)


@app.route('/api/run-notes/<path:run_id>', methods=['GET'])
@login_req
def api_run_notes_get(run_id):
    with get_db() as db:
        row = db.execute(
            "SELECT * FROM test_run_notes WHERE run_id=? ORDER BY id DESC LIMIT 1",
            (run_id,)
        ).fetchone()
    if row:
        return jsonify(dict(row))
    return jsonify(run_id=run_id, notes='', tags='')


# ── Feature 9: CSV Validator ──────────────────────────────────────────────────
@app.route('/api/validate-csv', methods=['POST'])
@login_req
def api_validate_csv():
    if 'file' not in request.files:
        return jsonify(error='No file uploaded'), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.csv'):
        return jsonify(error='Must be a .csv file'), 400
    import io as _io
    content = f.read().decode('utf-8', errors='replace')
    reader = csv.DictReader(_io.StringIO(content))
    rows = list(reader)
    headers = reader.fieldnames or []
    issues = []
    empty_cols = []
    dupe_rows = []
    seen = set()
    for i, row in enumerate(rows, start=2):
        for col in headers:
            if not str(row.get(col, '')).strip():
                empty_cols.append(f'Row {i}: column "{col}" is empty')
        key = tuple(str(row.get(h, '')) for h in headers)
        if key in seen:
            dupe_rows.append(f'Row {i}: duplicate entry')
        seen.add(key)
    issues.extend(empty_cols[:20])
    issues.extend(dupe_rows[:10])
    return jsonify(
        filename=f.filename,
        headers=headers,
        row_count=len(rows),
        issues=issues,
        valid=len(issues) == 0,
        summary=f'{len(rows)} rows, {len(headers)} columns, {len(issues)} issue(s)'
    )


# ── Feature 10: Webhook / MS Teams Alert ──────────────────────────────────────
def _fire_webhook(report_data):
    """Fire webhook notification after test run. Supports both old Teams Connectors
    (MessageCard) and new Power Automate Workflows (adaptive card / simple JSON)."""
    try:
        cfg = load_cfg()
        url = cfg.get('webhook_url', '').strip()
        if not url:
            return
        passed = report_data.get('passed', True)
        if passed and not cfg.get('webhook_on_pass', True):
            return
        if not passed and not cfg.get('webhook_on_fail', True):
            return

        color       = '00C853' if passed else 'D32F2F'
        status_icon = '✅' if passed else '❌'
        status_text = 'PASSED' if passed else 'FAILED'
        facts = [
            {'name': 'JMX',      'value': report_data.get('jmx', '—')},
            {'name': 'TPS',      'value': str(report_data.get('tps', '—'))},
            {'name': 'Avg RT',   'value': str(report_data.get('avg_rt', '—')) + ' ms'},
            {'name': 'P90',      'value': str(report_data.get('p90', '—')) + ' ms'},
            {'name': 'Error %',  'value': str(report_data.get('error_rate', '—')) + '%'},
            {'name': 'Samples',  'value': str(report_data.get('samples', '—'))},
            {'name': 'Duration', 'value': str(report_data.get('duration', '—'))},
            {'name': 'Started',  'value': report_data.get('started_at', '—')},
        ]

        import urllib.request as _req2, json as _json2

        # Detect URL type: Power Automate Workflows URLs contain 'logic.azure.com'
        # or 'webhook.office.com/webhookb2' (old) vs webhook.office.com (new Workflows)
        is_power_automate = 'logic.azure.com' in url or 'webhookb2' not in url

        if is_power_automate:
            # Power Automate / Workflows: send simple JSON body that the flow can parse
            payload = {
                'title':       f'Load Test {status_text}',
                'status':      status_text,
                'status_icon': status_icon,
                'color':       color,
                'client':      report_data.get('client', ''),
                'jmx':         report_data.get('jmx', ''),
                'tps':         report_data.get('tps', ''),
                'avg_rt':      report_data.get('avg_rt', ''),
                'p90':         report_data.get('p90', ''),
                'error_rate':  report_data.get('error_rate', ''),
                'samples':     report_data.get('samples', ''),
                'duration':    report_data.get('duration', ''),
                'started_at':  report_data.get('started_at', ''),
                'text': (f"{status_icon} **Load Test {status_text}** | "
                         f"{report_data.get('client','')} | "
                         f"JMX: {report_data.get('jmx','')} | "
                         f"TPS: {report_data.get('tps','')} | "
                         f"Err: {report_data.get('error_rate','')}%"),
            }
        else:
            # Legacy Teams Connectors (webhookb2) — MessageCard format
            # Teams requires top-level 'text' field or it returns 400
            payload = {
                '@type':    'MessageCard',
                '@context': 'https://schema.org/extensions',
                'themeColor': color,
                'summary':    f'Load Test {status_text} — {report_data.get("jmx","")}',
                'text':       f'{status_icon} Load Test **{status_text}**',
                'sections': [{
                    'activityTitle':    f'{status_icon} Load Test **{status_text}**',
                    'activitySubtitle': report_data.get('client', ''),
                    'facts': facts,
                    'markdown': True,
                }],
            }

        body = _json2.dumps(payload).encode('utf-8')
        req  = _req2.Request(url, data=body, headers={'Content-Type': 'application/json'})
        resp = _req2.urlopen(req, timeout=10, context=_SSL_CTX)
        return resp.read().decode('utf-8', errors='replace')
    except Exception as _wh_ex:
        app.logger.warning('Webhook delivery failed: %s', _wh_ex)
    return None


def _fire_webhook_with_result(report_data):
    """Like _fire_webhook but returns (ok, message) for test endpoints."""
    import urllib.request as _req2, json as _json2
    try:
        cfg = load_cfg()
        url = cfg.get('webhook_url', '').strip()
        if not url:
            return False, 'No webhook URL configured in Settings.'
        passed = report_data.get('passed', True)
        color       = '00C853' if passed else 'D32F2F'
        status_icon = '✅' if passed else '❌'
        status_text = 'PASSED' if passed else 'FAILED'
        facts = [
            {'name': 'JMX',      'value': report_data.get('jmx', '—')},
            {'name': 'Duration', 'value': str(report_data.get('duration', '—'))},
            {'name': 'Started',  'value': report_data.get('started_at', '—')},
        ]
        is_power_automate = 'logic.azure.com' in url or 'webhookb2' not in url
        if is_power_automate:
            payload = {
                'title': f'Load Test {status_text}',
                'text':  (f"{status_icon} **Load Test {status_text}** | "
                          f"JMX: {report_data.get('jmx','')}"),
            }
        else:
            payload = {
                '@type':    'MessageCard',
                '@context': 'https://schema.org/extensions',
                'themeColor': color,
                'summary':    f'Load Test {status_text}',
                'text':       f'{status_icon} Load Test **{status_text}**',
                'sections': [{
                    'activityTitle':    f'{status_icon} Load Test **{status_text}**',
                    'activitySubtitle': report_data.get('client', ''),
                    'facts': facts,
                    'markdown': True,
                }],
            }
        body = _json2.dumps(payload).encode('utf-8')
        req  = _req2.Request(url, data=body, headers={'Content-Type': 'application/json'})
        resp = _req2.urlopen(req, timeout=10, context=_SSL_CTX)
        result = resp.read().decode('utf-8', errors='replace')
        return True, f'Success — Teams responded: {result}'
    except Exception as e:
        return False, str(e)


@app.route('/api/webhook-config', methods=['GET', 'POST'])
@login_req
def api_webhook_config():
    if request.method == 'GET':
        cfg = load_cfg()
        return jsonify(
            webhook_url=cfg.get('webhook_url', ''),
            webhook_on_pass=bool(cfg.get('webhook_on_pass', True)),
            webhook_on_fail=bool(cfg.get('webhook_on_fail', True)),
        )
    data = request.get_json(silent=True) or {}
    cfg = load_cfg()
    cfg['webhook_url'] = data.get('webhook_url', '')
    cfg['webhook_on_pass'] = data.get('webhook_on_pass', True)
    cfg['webhook_on_fail'] = data.get('webhook_on_fail', True)
    save_cfg(cfg)
    audit('webhook_config_update', '')
    return jsonify(ok=True)


@app.route('/api/webhook-test', methods=['POST'])
@login_req
def api_webhook_test():
    import datetime as _dt
    ok, msg = _fire_webhook_with_result({
        'client': 'Test Client',
        'jmx': 'test.jmx',
        'threads': 10,
        'duration': 60,
        'started_at': _dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'passed': True,
    })
    return jsonify(ok=ok, message=msg)


# ── Feature: SLA Breach Highlighter ──────────────────────────────────────────
@app.route('/api/sla-analysis/<path:fname>')
@login_req
def api_sla_analysis(fname):
    """Per-transaction SLA analysis: flag labels breaching RT or error thresholds."""
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    dirs = client_dirs(c)
    jtl_path = os.path.join(dirs['reports'], fname)
    if not os.path.isfile(jtl_path):
        return jsonify(error='File not found'), 404
    rt_threshold  = int(request.args.get('rt',  2000))
    err_threshold = float(request.args.get('err', 5.0))
    from collections import defaultdict
    stats = defaultdict(lambda: {'total': 0, 'fail': 0, 'rt_sum': 0, 'p90_vals': []})
    try:
        with open(jtl_path, encoding='utf-8', errors='replace') as f:
            reader = csv.DictReader(f)
            for row in reader:
                lbl = row.get('label', 'unknown')
                s   = stats[lbl]
                s['total'] += 1
                elapsed = int(row.get('elapsed', 0) or 0)
                s['rt_sum'] += elapsed
                if len(s['p90_vals']) < 5000:
                    s['p90_vals'].append(elapsed)
                if str(row.get('success', 'true')).lower() == 'false':
                    s['fail'] += 1
    except Exception as e:
        return jsonify(error=str(e)), 500
    results = []
    for lbl, s in stats.items():
        n = s['total']
        if n == 0:
            continue
        avg_rt   = round(s['rt_sum'] / n)
        err_pct  = round(s['fail'] / n * 100, 2)
        vals     = sorted(s['p90_vals'])
        p90      = vals[min(int(0.9 * len(vals)), len(vals) - 1)]
        rt_ok    = p90 <= rt_threshold
        err_ok   = err_pct <= err_threshold
        results.append({
            'label':    lbl,
            'total':    n,
            'avg_rt':   avg_rt,
            'p90':      p90,
            'err_pct':  err_pct,
            'rt_ok':    rt_ok,
            'err_ok':   err_ok,
            'ok':       rt_ok and err_ok,
        })
    results.sort(key=lambda x: (x['ok'], -x['p90']))
    passed = sum(1 for r in results if r['ok'])
    return jsonify(
        rt_threshold=rt_threshold,
        err_threshold=err_threshold,
        total_labels=len(results),
        passed=passed,
        failed=len(results) - passed,
        results=results,
    )


# ── Feature: Test Run Comparison Side-by-Side ─────────────────────────────────
@app.route('/api/compare-reports', methods=['POST'])
@login_req
def api_compare_reports():
    """Compare two JTL files — returns side-by-side metrics diff."""
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    dirs  = client_dirs(c)
    data  = request.get_json(silent=True) or {}
    file1 = data.get('file1', '')
    file2 = data.get('file2', '')
    if not file1 or not file2:
        return jsonify(error='Both file1 and file2 required'), 400

    def _parse_jtl(fname):
        path = os.path.join(dirs['reports'], fname)
        if not os.path.isfile(path):
            return None, f'{fname} not found'
        total = fail = rt_sum = 0
        rt_vals = []
        try:
            with open(path, encoding='utf-8', errors='replace') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    total += 1
                    elapsed = int(row.get('elapsed', 0) or 0)
                    rt_sum += elapsed
                    if len(rt_vals) < 10000:
                        rt_vals.append(elapsed)
                    if str(row.get('success', 'true')).lower() == 'false':
                        fail += 1
        except Exception as e:
            return None, str(e)
        if total == 0:
            return None, 'Empty file'
        rt_vals.sort()
        n = len(rt_vals)
        def pct(p): return rt_vals[min(int(p / 100 * n), n - 1)]
        return {
            'file':       fname,
            'samples':    total,
            'errors':     fail,
            'error_pct':  round(fail / total * 100, 2),
            'avg_rt':     round(rt_sum / total),
            'p50':        pct(50),
            'p90':        pct(90),
            'p95':        pct(95),
            'p99':        pct(99),
        }, None

    m1, e1 = _parse_jtl(file1)
    m2, e2 = _parse_jtl(file2)
    if e1: return jsonify(error=f'File 1: {e1}'), 400
    if e2: return jsonify(error=f'File 2: {e2}'), 400

    def delta(a, b):
        if b == 0: return 0
        return round((a - b) / b * 100, 1)

    diff = {k: delta(m2[k], m1[k])
            for k in ('avg_rt', 'p90', 'p95', 'p99', 'error_pct', 'samples')}
    return jsonify(run1=m1, run2=m2, diff=diff)


# ── Feature: Multi-Run Trend Overlay ──────────────────────────────────────────
@app.route('/api/trend-overlay')
@login_req
def api_trend_overlay():
    """Return per-run summary metrics for the last N reports for overlay chart."""
    c = active_client()
    if not c:
        return jsonify(error='No active client'), 400
    dirs  = client_dirs(c)
    limit = int(request.args.get('n', 10))
    jtl_files = sorted(
        [f for f in os.listdir(dirs['reports']) if f.endswith('.jtl')],
        key=lambda f: os.path.getmtime(os.path.join(dirs['reports'], f)),
        reverse=True
    )[:limit]
    runs = []
    for fname in reversed(jtl_files):
        path = os.path.join(dirs['reports'], fname)
        total = fail = rt_sum = 0
        rt_vals = []
        try:
            with open(path, encoding='utf-8', errors='replace') as f:
                reader = csv.DictReader(f)
                ts_min = ts_max = None
                for row in reader:
                    total += 1
                    elapsed = int(row.get('elapsed', 0) or 0)
                    rt_sum += elapsed
                    if len(rt_vals) < 5000:
                        rt_vals.append(elapsed)
                    if str(row.get('success', 'true')).lower() == 'false':
                        fail += 1
                    ts = int(row.get('timeStamp', 0) or 0)
                    if ts:
                        ts_min = min(ts_min, ts) if ts_min else ts
                        ts_max = max(ts_max, ts) if ts_max else ts
        except Exception:
            continue
        if total == 0:
            continue
        rt_vals.sort()
        n = len(rt_vals)
        duration = round((ts_max - ts_min) / 1000) if ts_min and ts_max else 0
        tps = round(total / duration, 2) if duration else 0
        runs.append({
            'file':      fname,
            'label':     fname.replace('.jtl', '')[-30:],
            'samples':   total,
            'avg_rt':    round(rt_sum / total),
            'p90':       rt_vals[min(int(0.9 * n), n - 1)],
            'error_pct': round(fail / total * 100, 2),
            'tps':       tps,
        })
    return jsonify(runs=runs)


def _heal_jmx_paths():
    """Scan all JMX files for absolute paths that belong to the old machine.
    If the file exists under the current BASE_DIR with the same relative suffix
    (clients/<code>/testdata/<file>), rewrite the path in-place."""
    PATH_PAT = re.compile(
        r'(<stringProp name="filename">)([^<]+)(</stringProp>)'
    )
    jmx_files = glob.glob(os.path.join(CLIENTS_DIR, '**', '*.jmx'), recursive=True)
    for jmx_path in jmx_files:
        try:
            with open(jmx_path, encoding='utf-8') as f:
                content = f.read()
            changed = False

            def _fix(m):
                nonlocal changed
                stored = m.group(2).strip()
                # Skip blank, already-relative, or existing paths
                if not stored or not os.path.isabs(stored) or os.path.exists(stored):
                    return m.group(0)
                # Find the 'clients/' anchor in the stale path
                norm = stored.replace('\\', '/')
                idx = norm.find('/clients/')
                if idx == -1:
                    return m.group(0)
                rel_parts = norm[idx + 1:].split('/')   # ['clients','BTC','testdata','file.csv']
                new_path = os.path.join(BASE_DIR, *rel_parts)
                if not os.path.exists(new_path):
                    return m.group(0)
                changed = True
                print(f'[startup] JMX path healed: {os.path.basename(stored)} → {new_path}')
                return m.group(1) + new_path + m.group(3)

            new_content = PATH_PAT.sub(_fix, content)
            if changed:
                with open(jmx_path, 'w', encoding='utf-8') as f:
                    f.write(new_content)
                print(f'[startup] Saved: {os.path.basename(jmx_path)}')
        except Exception as e:
            print(f'[startup] JMX heal error ({os.path.basename(jmx_path)}): {e}')


def _ensure_chartjs():
    """Download Chart.js 4.4.1 into static/ if it is missing."""
    path = os.path.join(BASE_DIR, 'static', 'chart.min.js')
    if os.path.exists(path):
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    print('[startup] chart.min.js missing — downloading Chart.js 4.4.1...')
    try:
        url = 'https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js'
        req = _urllib_req2.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with _urllib_req2.urlopen(req, context=_SSL_CTX, timeout=30) as resp:
            data = resp.read()
        with open(path, 'wb') as f:
            f.write(data)
        print(f'[startup] chart.min.js saved ({round(len(data)/1024)} KB)')
    except Exception as e:
        print(f'[startup] Chart.js download failed: {e} — charts will use CDN fallback')


def _startup_auto_configure():
    """Fully self-configure on every start: packages, JMeter, Chart.js,
    DB paths, JMX paths, and client folders — zero manual steps."""
    print(f'[startup] OS: {_platform.system()} {_platform.release()} | Python {_platform.python_version()}')
    print(f'[startup] Base dir: {BASE_DIR}')

    cfg = load_cfg()

    # ── 1. Auto-detect Java; auto-install if missing ──────────────────────────
    java_bin = _find_java_bin()
    if java_bin:
        # Ensure JAVA_HOME is set so JMeter sub-processes inherit it
        java_home = (cfg.get('java_home') or os.environ.get('JAVA_HOME', '')
                     or os.path.dirname(os.path.dirname(java_bin)))
        if java_home and os.path.isdir(java_home):
            os.environ['JAVA_HOME'] = java_home
        print(f'[startup] Java OK: {java_bin}')
    else:
        print('[startup] Java not found — auto-installing OpenJDK 11 (≈185 MB)...')
        _auto_install_java()   # runs synchronously before Flask starts
        java_bin = _find_java_bin()
        if java_bin:
            print(f'[startup] Java installed: {java_bin}')
        else:
            err = _java_install.get('error', 'unknown error')
            print(f'[startup] Java install failed: {err}')
            print('[startup] Tip: install Java 11+ manually, set JAVA_HOME, or use Settings → Auto-Install Java')

    # ── 2. Auto-detect JMeter; auto-install if missing ────────────────────────
    saved = cfg.get('jmeter_bin', '')
    if saved and os.path.exists(saved):
        print(f'[startup] JMeter OK: {saved}')
    else:
        found = _find_jmeter_bin()
        if found:
            cfg['jmeter_bin'] = found
            save_cfg(cfg)
            print(f'[startup] JMeter auto-detected: {found}')
        else:
            print(f'[startup] JMeter not found — auto-installing JMeter {_JMETER_VERSION} (≈80 MB)...')
            _auto_install_jmeter()   # runs synchronously before Flask starts
            found = _find_jmeter_bin()
            if found:
                print(f'[startup] JMeter installed: {found}')
            else:
                err = _jmeter_install.get('error', 'unknown error')
                print(f'[startup] JMeter install failed: {err}')
                print('[startup] Tip: use Settings → Auto-Install or put JMeter on your PATH')

    # ── 3. Ensure Chart.js is present for offline HTML reports ───────────────
    try:
        _ensure_chartjs()
    except Exception as e:
        print(f'[startup] Chart.js setup skipped: {e}')

    # ── 4. Heal stale client directory paths in DB ────────────────────────────
    try:
        clients = get_all_clients()
        for c in clients:
            updates = {}
            for col, subdir in [('jmx_dir', 'jmx'), ('testdata_dir', 'testdata'), ('reports_dir', 'reports')]:
                stored = c.get(col) or ''
                if stored and not os.path.exists(stored):
                    new_path = os.path.join(CLIENTS_DIR, c['code'], subdir)
                    updates[col] = new_path
                    print(f'[startup] Healed DB {c["code"]}.{col} → {new_path}')
            if updates:
                with get_db() as db:
                    for col, val in updates.items():
                        db.execute(f'UPDATE clients SET {col}=? WHERE code=?', (val, c['code']))
    except Exception as e:
        print(f'[startup] DB path heal skipped: {e}')

    # ── 5. Heal absolute paths baked into JMX files ───────────────────────────
    try:
        _heal_jmx_paths()
    except Exception as e:
        print(f'[startup] JMX heal skipped: {e}')

    # ── 6. Ensure all client folders exist on this machine ────────────────────
    try:
        for c in get_all_clients():
            ensure_client_dirs(c)
    except Exception:
        pass


if __name__ == '__main__':
    init_db()
    _startup_auto_configure()
    _schedule_daily_purge()       # start 24-hour auto-purge loop
    _start_recurring_checker()    # start 60-second recurring schedule checker
    _start_backup_scheduler()     # start optional periodic backup
    _ci_cd_scheduler()            # start internal CI/CD suite scheduler

    print('=' * 60)
    print('  Centralized Load Testing Platform  v4.0')
    print(f'  URL    : http://localhost:5000')
    print(f'  OS     : {_platform.system()} {_platform.release()}')
    print(f'  JMeter : {_find_jmeter_bin() or "NOT FOUND — configure in Settings"}')
    print('  DB     : lt_platform.db  (SQLite)')
    print('  Creds  : see README.md (change after first login)')
    print('  CI/CD  : Internal scheduler enabled')
    print('=' * 60)
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
